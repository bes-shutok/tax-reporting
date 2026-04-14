"""Crypto rewards sheet writer for the Excel tax report."""

from __future__ import annotations

from typing import TYPE_CHECKING

import openpyxl
from openpyxl.styles import Font

if TYPE_CHECKING:
    from ..crypto_reporting import AggregatedRewardIncomeEntry, CryptoTaxReport

from ..crypto_reporting import ZERO, RewardTaxClassification
from .excel_utils import auto_column_width


def write_crypto_rewards_sheet(  # noqa: PLR0912, PLR0915
    workbook: openpyxl.Workbook,
    crypto_tax_report: CryptoTaxReport,
    aggregated_rewards: list[AggregatedRewardIncomeEntry],
) -> None:
    """Create and populate a 'Crypto Rewards' worksheet with reward income and reconciliation.

    Writes:
    - Section 2: IRS-ready filing summary (taxable_now aggregated)
    - Section 2a2: Taxable-now support detail
    - Section 2b: Deferred by law support detail
    - Section 2c: Rewards classification reconciliation

    Args:
        workbook: The Excel workbook to add the sheet to.
        crypto_tax_report: The crypto tax report data.
        aggregated_rewards: Pre-computed aggregated taxable-now rewards.
    """
    worksheet = workbook.create_sheet("Crypto Rewards")

    taxable_now_entries = [
        e for e in crypto_tax_report.reward_entries if e.tax_classification == RewardTaxClassification.TAXABLE_NOW
    ]
    deferred_entries = [
        e for e in crypto_tax_report.reward_entries if e.tax_classification == RewardTaxClassification.DEFERRED_BY_LAW
    ]

    row_no = 1

    # 2. REWARDS INCOME - IRS-READY SUMMARY
    worksheet.cell(row_no, 1, "2. REWARDS INCOME - IRS-READY FILING SUMMARY").font = Font(bold=True)
    row_no += 1

    summary_note = worksheet.cell(
        row_no,
        1,
        "This table shows only income taxable immediately in Category E. "
        "Crypto-denominated rewards (deferred until disposal) are in section 2b below.",
    )
    summary_note.font = Font(italic=True, size=9)
    row_no += 1

    aggregated_headers = [
        "Income code",
        "Source country",
        "Reward chain",
        "Gross income (EUR)",
        "Foreign tax (EUR)",
        "Net income (EUR)",
        "Raw rows",
    ]
    for idx, header in enumerate(aggregated_headers, start=1):
        header_cell = worksheet.cell(row_no, idx, header)
        header_cell.font = Font(bold=True)
    row_no += 1

    if aggregated_rewards:
        for entry in aggregated_rewards:
            worksheet.cell(row_no, 1, entry.income_code)
            worksheet.cell(row_no, 2, entry.source_country)
            worksheet.cell(row_no, 3, ", ".join(entry.chains) if entry.chains else "Unknown")
            worksheet.cell(row_no, 4, float(entry.gross_income_eur))
            worksheet.cell(row_no, 5, float(entry.foreign_tax_eur))
            worksheet.cell(row_no, 6, float(entry.gross_income_eur - entry.foreign_tax_eur))
            worksheet.cell(row_no, 7, entry.raw_row_count)
            row_no += 1
    else:
        worksheet.cell(
            row_no, 1, "No immediately taxable rewards (all rewards are crypto-denominated and deferred by law)"
        )
        row_no += 1

    # 2a2. TAXABLE-NOW SUPPORT DETAIL
    row_no += 1
    worksheet.cell(row_no, 1, "2a2. TAXABLE-NOW - SUPPORT DETAIL").font = Font(bold=True)
    row_no += 1

    taxable_note = worksheet.cell(
        row_no,
        1,
        "Individual rows that contribute to the IRS-ready filing table above. "
        "Use this section to trace each aggregated line back to its source Koinly rows.",
    )
    taxable_note.font = Font(italic=True, size=9)
    row_no += 1

    taxable_detail_headers = [
        "Date",
        "Asset",
        "Value (EUR)",
        "Income type",
        "Wallet",
        "Platform",
        "Reward chain",
        "Country",
        "Foreign tax (EUR)",
        "Review flag",
        "Description",
    ]
    for idx, header in enumerate(taxable_detail_headers, start=1):
        header_cell = worksheet.cell(row_no, idx, header)
        header_cell.font = Font(bold=True)
    row_no += 1

    if taxable_now_entries:
        for entry in taxable_now_entries:
            worksheet.cell(row_no, 1, entry.date)
            worksheet.cell(row_no, 2, entry.asset)
            worksheet.cell(row_no, 3, float(entry.value_eur))
            worksheet.cell(row_no, 4, entry.source_type)
            worksheet.cell(row_no, 5, entry.wallet)
            worksheet.cell(row_no, 6, entry.platform)
            worksheet.cell(row_no, 7, entry.chain)
            worksheet.cell(row_no, 8, entry.operator_origin.operator_country)
            worksheet.cell(row_no, 9, float(entry.foreign_tax_eur))
            worksheet.cell(row_no, 10, f"YES: {entry.review_reason}" if entry.review_required else "NO")
            worksheet.cell(row_no, 11, entry.description)
            row_no += 1
    else:
        worksheet.cell(row_no, 1, "No taxable-now rewards")
        row_no += 1

    # 2b. DEFERRED BY LAW SUPPORT SECTION
    row_no += 1
    worksheet.cell(row_no, 1, "2b. DEFERRED BY LAW - SUPPORT DETAIL").font = Font(bold=True)
    row_no += 1

    deferred_note = worksheet.cell(
        row_no,
        1,
        "These rewards are crypto-denominated and taxation is deferred until disposal per CIRS art. 5(11). "
        "They are shown here for auditability but NOT included in the IRS-ready filing table above.",
    )
    deferred_note.font = Font(italic=True, size=9)
    row_no += 1

    deferred_headers = [
        "Date",
        "Asset",
        "Value (EUR)",
        "Income type",
        "Wallet",
        "Platform",
        "Reward chain",
        "Country",
        "Foreign tax (EUR)",
        "Review flag",
        "Description",
    ]
    for idx, header in enumerate(deferred_headers, start=1):
        header_cell = worksheet.cell(row_no, idx, header)
        header_cell.font = Font(bold=True)
    row_no += 1

    if deferred_entries:
        for entry in deferred_entries:
            worksheet.cell(row_no, 1, entry.date)
            worksheet.cell(row_no, 2, entry.asset)
            worksheet.cell(row_no, 3, float(entry.value_eur))
            worksheet.cell(row_no, 4, entry.source_type)
            worksheet.cell(row_no, 5, entry.wallet)
            worksheet.cell(row_no, 6, entry.platform)
            worksheet.cell(row_no, 7, entry.chain)
            worksheet.cell(row_no, 8, entry.operator_origin.operator_country)
            worksheet.cell(row_no, 9, float(entry.foreign_tax_eur))
            worksheet.cell(row_no, 10, f"YES: {entry.review_reason}" if entry.review_required else "NO")
            worksheet.cell(row_no, 11, entry.description)
            row_no += 1
    else:
        worksheet.cell(row_no, 1, "No deferred rewards")
        row_no += 1

    # 2c. REWARDS CLASSIFICATION RECONCILIATION
    row_no += 1
    worksheet.cell(row_no, 1, "2c. REWARDS CLASSIFICATION RECONCILIATION").font = Font(bold=True)
    row_no += 1

    taxable_now_total_eur = sum((e.value_eur for e in taxable_now_entries), start=ZERO)
    deferred_total_eur = sum((e.value_eur for e in deferred_entries), start=ZERO)

    reconciliation_rewards_rows = [
        ("Total reward rows (raw)", len(crypto_tax_report.reward_entries)),
        ("Taxable-now rows (immediately taxable)", len(taxable_now_entries)),
        ("Deferred-by-law rows (taxation deferred)", len(deferred_entries)),
        ("Taxable-now total value (EUR)", float(taxable_now_total_eur)),
        ("Deferred total value (EUR)", float(deferred_total_eur)),
        ("Filing-ready lines after aggregation", len(aggregated_rewards)),
    ]

    for key, value in reconciliation_rewards_rows:
        worksheet.cell(row_no, 1, key)
        worksheet.cell(row_no, 2, value)
        row_no += 1

    auto_column_width(worksheet)
