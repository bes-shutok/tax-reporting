"""Crypto capital gains sheet writer for the Excel tax report."""

from __future__ import annotations

from typing import TYPE_CHECKING

import openpyxl
from openpyxl.styles import Font

if TYPE_CHECKING:
    from ..crypto_reporting import CryptoTaxReport

from .excel_utils import auto_column_width


def write_crypto_gains_sheet(workbook: openpyxl.Workbook, crypto_tax_report: CryptoTaxReport) -> None:  # noqa: PLR0915
    """Create and populate a 'Crypto Gains' worksheet with capital gains entries and statistics.

    Writes:
    - Title and tax year metadata
    - PDF summary (if present)
    - Section 1: Capital gain entries in 17 columns
    - Section 1b: Capital gain statistics by holding period

    Args:
        workbook: The Excel workbook to add the sheet to.
        crypto_tax_report: The crypto tax report data.
    """
    worksheet = workbook.create_sheet("Crypto Gains")
    worksheet.cell(1, 1, "CRYPTO TAX REPORT - PORTUGAL").font = Font(bold=True)
    worksheet.cell(2, 1, "Tax year")
    worksheet.cell(2, 2, crypto_tax_report.tax_year)
    if crypto_tax_report.pdf_summary:
        worksheet.cell(3, 1, "PDF period")
        worksheet.cell(3, 2, crypto_tax_report.pdf_summary.period or "N/A")
        worksheet.cell(3, 3, "PDF timezone")
        worksheet.cell(3, 4, crypto_tax_report.pdf_summary.timezone or "N/A")
        worksheet.cell(3, 5, "PDF extracted tokens")
        worksheet.cell(3, 6, crypto_tax_report.pdf_summary.extracted_tokens)

    row_no = 5 if crypto_tax_report.pdf_summary else 4
    worksheet.cell(row_no, 1, "1. CAPITAL GAINS").font = Font(bold=True)
    row_no += 1

    capital_headers = [
        "Disposal date",
        "Acquisition date",
        "Asset",
        "Amount",
        "Cost (EUR)",
        "Proceeds (EUR)",
        "Gain/Loss (EUR)",
        "Holding period",
        "Wallet",
        "Platform",
        "Disposal chain",
        "Operator entity",
        "Operator country",
        "Annex hint",
        "Review flag",
        "Notes",
        "Token origin",
    ]
    for idx, header in enumerate(capital_headers, start=1):
        worksheet.cell(row_no, idx, header)
    row_no += 1

    for entry in crypto_tax_report.capital_entries:
        worksheet.cell(row_no, 1, entry.disposal_date)
        worksheet.cell(row_no, 2, entry.acquisition_date)
        worksheet.cell(row_no, 3, entry.asset)
        worksheet.cell(row_no, 4, float(entry.amount))
        worksheet.cell(row_no, 5, float(entry.cost_eur))
        worksheet.cell(row_no, 6, float(entry.proceeds_eur))
        worksheet.cell(row_no, 7, float(entry.gain_loss_eur))
        worksheet.cell(row_no, 8, entry.holding_period)
        worksheet.cell(row_no, 9, entry.wallet)
        worksheet.cell(row_no, 10, entry.platform)
        worksheet.cell(row_no, 11, entry.chain)
        worksheet.cell(row_no, 12, entry.operator_origin.operator_entity)
        worksheet.cell(row_no, 13, entry.operator_origin.operator_country)
        worksheet.cell(row_no, 14, entry.annex_hint)
        worksheet.cell(row_no, 15, f"YES: {entry.review_reason}" if entry.review_required else "NO")
        worksheet.cell(row_no, 16, entry.notes)
        worksheet.cell(row_no, 17, entry.token_swap_history or "")
        row_no += 1

    # 1b. CAPITAL GAINS STATISTICS
    row_no += 1
    worksheet.cell(row_no, 1, "1b. CAPITAL GAINS STATISTICS").font = Font(bold=True)
    row_no += 1

    stats_headers = [
        "Holding Period",
        "Count",
        "Cost Total (EUR)",
        "Proceeds Total (EUR)",
        "Gain/Loss Total (EUR)",
    ]
    for idx, header in enumerate(stats_headers, start=1):
        header_cell = worksheet.cell(row_no, idx, header)
        header_cell.font = Font(bold=True)
    row_no += 1

    stats = crypto_tax_report.capital_gain_stats
    period_rows = [
        ("Short-term", stats.short_term),
        ("Long-term", stats.long_term),
        ("Mixed", stats.mixed),
        ("Unknown", stats.unknown),
        ("Grand Total", stats.grand_total),
    ]
    for label, period_stats in period_rows:
        worksheet.cell(row_no, 1, label)
        worksheet.cell(row_no, 2, period_stats.count)
        worksheet.cell(row_no, 3, float(period_stats.cost_total_eur))
        worksheet.cell(row_no, 4, float(period_stats.proceeds_total_eur))
        worksheet.cell(row_no, 5, float(period_stats.gain_loss_total_eur))
        row_no += 1

    auto_column_width(worksheet)
