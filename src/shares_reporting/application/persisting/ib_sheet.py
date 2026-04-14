"""IB reporting sheet writer for the Excel tax report."""

from __future__ import annotations

import openpyxl
from openpyxl.comments import Comment
from openpyxl.styles import Font, PatternFill

from ...domain.collections import (
    CapitalGainLinesPerCompany,
    DividendIncomePerCompany,
)
from ...domain.constants import (
    EXCEL_COLUMN_OFFSET,
    EXCEL_HEADER_ROW_1,
    EXCEL_HEADER_ROW_2,
    EXCEL_NUMBER_FORMAT,
    EXCEL_START_COLUMN,
    EXCEL_START_ROW,
    PLACEHOLDER_YEAR,
    ZERO_QUANTITY,
)
from ...domain.exceptions import ReportGenerationError
from ...infrastructure.config import Config, ConversionRate
from ...infrastructure.logging_config import create_module_logger
from .excel_utils import auto_column_width


def write_ib_reporting_sheet(  # noqa: PLR0912, PLR0915
    worksheet: openpyxl.worksheet.worksheet.Worksheet,
    config: Config,
    capital_gain_lines_per_company: CapitalGainLinesPerCompany,
    dividend_income_per_company: DividendIncomePerCompany | None = None,
) -> None:
    """Write the IB Reporting sheet with capital gains and dividend income.

    Args:
        worksheet: The openpyxl worksheet to populate (must already be named).
        config: Application configuration with exchange rates.
        capital_gain_lines_per_company: Calculated capital gains grouped by company.
        dividend_income_per_company: Dividend income data grouped by company (optional).
    """
    logger = create_module_logger(__name__)

    first_header = [
        "Beneficiary",
        "Country of Source",
        "SALE",
        "",
        "",
        "",
        "PURCHASE",
        "",
        "",
        "",
        "WITHOLDING TAX",
        "",
        "Expenses incurred with obtaining the capital gains",
        "",
        "Symbol",
        "Currency",
        "Sale amount",
        "Buy amount",
        "Expenses amount",
    ]
    second_header = [
        "",
        "",
        "Day ",
        "Month ",
        "Year",
        "Amount",
        "Day ",
        "Month ",
        "Year",
        "Amount",
        "Country",
        "Amount",
        "",
        "",
        "",
        "",
        "",
        "",
        "",
    ]

    last_column: int = max(len(first_header), len(second_header))
    exchange_rates: dict[str, str] = create_currency_table(worksheet, last_column + 2, 1, config)
    logger.debug("Created currency exchange table with %s rates", len(config.rates) + 1)

    for i in range(len(first_header)):
        _ = worksheet.cell(EXCEL_HEADER_ROW_1, i + 1, first_header[i])
        _ = worksheet.cell(EXCEL_HEADER_ROW_2, i + 1, second_header[i])

    start_column = EXCEL_START_COLUMN
    line_number = EXCEL_START_ROW
    processed_lines = ZERO_QUANTITY

    for currency_company, capital_gain_lines in capital_gain_lines_per_company.items():
        currency = currency_company.currency
        company = currency_company.company
        logger.debug("Processing capital gain lines for %s (%s)", company.ticker, currency.currency)

        for line in capital_gain_lines:
            if currency != line.get_currency():
                raise ReportGenerationError(f"Currency mismatch in line: {currency} != {line.get_currency()}")
            processed_lines += 1
            idx = start_column

            _ = worksheet.cell(line_number, start_column, line.get_sell_date().day)
            idx += 1
            _ = worksheet.cell(line_number, idx, line.get_sell_date().get_month_name())
            idx += 1
            _ = worksheet.cell(line_number, idx, line.get_sell_date().year)
            idx += 1
            _ = worksheet.cell(
                line_number,
                idx,
                "=" + exchange_rates[currency.currency] + "*(" + line.get_sell_amount() + ")",
            )

            idx += 1
            _ = worksheet.cell(line_number, idx, line.get_buy_date().day)
            idx += 1
            _ = worksheet.cell(line_number, idx, line.get_buy_date().get_month_name())
            idx += 1
            _ = worksheet.cell(line_number, idx, line.get_buy_date().year)
            idx += 1
            _ = worksheet.cell(
                line_number,
                idx,
                "=" + exchange_rates[currency.currency] + "*(" + line.get_buy_amount() + ")",
            )

            idx += EXCEL_COLUMN_OFFSET

            expense_cell = worksheet.cell(
                line_number,
                idx,
                "=" + exchange_rates[currency.currency] + "*(" + line.get_expense_amount() + ")",
            )
            expense_cell.number_format = EXCEL_NUMBER_FORMAT  # type: ignore[assignment]
            idx += 2

            _ = worksheet.cell(line_number, idx, company.ticker)
            idx += 1
            _ = worksheet.cell(line_number, idx, currency.currency)
            idx += 1

            sell_amount_cell = worksheet.cell(line_number, idx, "=" + line.get_sell_amount())
            sell_amount_cell.number_format = EXCEL_NUMBER_FORMAT  # type: ignore[assignment]
            idx += 1
            buy_amount_cell = worksheet.cell(line_number, idx, "=" + line.get_buy_amount())
            buy_amount_cell.number_format = EXCEL_NUMBER_FORMAT  # type: ignore[assignment]
            idx += 1
            expense_amount_cell = worksheet.cell(line_number, idx, "=" + line.get_expense_amount())
            expense_amount_cell.number_format = EXCEL_NUMBER_FORMAT  # type: ignore[assignment]

            if line.get_buy_date().year == PLACEHOLDER_YEAR:
                red_fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")
                for col_idx in range(start_column, idx + 1):
                    cell = worksheet.cell(line_number, col_idx)
                    cell.fill = red_fill  # type: ignore[assignment]

            line_number += 1

    logger.debug("Processed %s capital gain lines", processed_lines)

    line_number = EXCEL_START_ROW
    for currency_company, capital_gain_lines in capital_gain_lines_per_company.items():
        company = currency_company.company
        for _ in capital_gain_lines:
            _ = worksheet.cell(line_number, 2, company.country_of_issuance)
            _ = worksheet.cell(line_number, 11, company.country_of_issuance)
            line_number += 1

    if dividend_income_per_company:
        logger.info("Adding CAPITAL INVESTMENT INCOME section with %s securities", len(dividend_income_per_company))

        line_number += 1

        section_title_cell = worksheet.cell(line_number, 1, "5. CAPITAL INVESTMENT INCOME:")
        section_title_cell.font = Font(bold=True)  # type: ignore[assignment]
        line_number += 1

        line_number += 1

        dividend_headers = [
            "Beneficiary\n(choose one)",
            "Type of capital income\n(choose one)",
            "Country of source",
            "ISIN",
            "Gross amount",
            "Withholding tax at source",
            "Withholding tax in Portugal\n(if any)",
            "",
            "Symbol",
            "Currency",
            "Original gross amount",
            "Original tax amount",
            "Net amount",
        ]

        for i, header in enumerate(dividend_headers):
            _ = worksheet.cell(line_number, i + 1, header)

        line_number += 1

        for symbol, dividend_data in dividend_income_per_company.items():
            _ = worksheet.cell(line_number, 1, "")
            _ = worksheet.cell(line_number, 2, "Dividends")

            if dividend_data.isin == "MISSING_ISIN_REQUIRES_ATTENTION":
                country_cell = worksheet.cell(line_number, 3, "\u26a0\ufe0f MISSING DATA")
                country_cell.fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")  # type: ignore[assignment]

                isin_cell = worksheet.cell(line_number, 4, f"\u26a0\ufe0f {symbol}")
                isin_cell.fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")  # type: ignore[assignment]

                isin_cell.comment = Comment(  # type: ignore[assignment]
                    f"Security information missing for {symbol}. Please verify this symbol in your IB account.",
                    "Shares Reporting",
                )
            else:
                _ = worksheet.cell(line_number, 3, dividend_data.country)
                _ = worksheet.cell(line_number, 4, dividend_data.isin)

            gross_amount_cell = worksheet.cell(
                line_number,
                5,
                "=" + exchange_rates[dividend_data.currency.currency] + "*(" + str(dividend_data.gross_amount) + ")",
            )
            gross_amount_cell.number_format = EXCEL_NUMBER_FORMAT  # type: ignore[assignment]

            tax_amount_cell = worksheet.cell(
                line_number,
                6,
                "=" + exchange_rates[dividend_data.currency.currency] + "*(" + str(dividend_data.total_taxes) + ")",
            )
            tax_amount_cell.number_format = EXCEL_NUMBER_FORMAT  # type: ignore[assignment]

            _ = worksheet.cell(line_number, 7, "")

            _ = worksheet.cell(line_number, 9, symbol)
            _ = worksheet.cell(line_number, 10, dividend_data.currency.currency)

            original_gross_cell = worksheet.cell(line_number, 11, str(dividend_data.gross_amount))
            original_gross_cell.number_format = EXCEL_NUMBER_FORMAT

            original_tax_cell = worksheet.cell(line_number, 12, str(dividend_data.total_taxes))
            original_tax_cell.number_format = EXCEL_NUMBER_FORMAT

            net_amount = dividend_data.gross_amount - dividend_data.total_taxes
            net_amount_cell = worksheet.cell(line_number, 13, str(net_amount))
            net_amount_cell.number_format = EXCEL_NUMBER_FORMAT

            logger.debug(
                "Added dividend income row for %s: %s gross, %s tax, %s net (%s)",
                symbol,
                dividend_data.gross_amount,
                dividend_data.total_taxes,
                net_amount,
                dividend_data.currency.currency,
            )
            line_number += 1

    auto_column_width(worksheet)


def create_currency_table(
    worksheet: openpyxl.worksheet.worksheet.Worksheet,
    column_no: int,
    row_no: int,
    config: Config,
) -> dict[str, str]:
    """Create a currency configuration table in the excel worksheet.

    Args:
        worksheet: The target excel worksheet.
        column_no: The starting column number (1-based).
        row_no: The starting row number (1-based).
        config: The application configuration object.

    Returns:
        A dictionary mapping currency codes to their cell coordinate strings.
    """
    logger = create_module_logger(__name__)
    currency_header = ["Base/target", "Rate"]
    rates: list[ConversionRate] = config.rates

    logger.debug("Creating currency table starting at column %s, row %s", column_no, row_no)

    _ = worksheet.cell(row_no, column_no, "Currency exchange rate")
    row_no += 1
    for i in range(len(currency_header)):
        _ = worksheet.cell(row_no, column_no + i, currency_header[i])
    row_no += 1

    coordinates: dict[str, str] = {}
    for j in range(len(rates)):
        _ = worksheet.cell(row_no + j, column_no, rates[j].base + "/" + rates[j].calculated)
        cell = worksheet.cell(row_no + j, column_no + 1, str(rates[j].rate))
        coordinates[rates[j].calculated] = cell.coordinate
        logger.debug("Added currency rate %s/%s = %s", rates[j].base, rates[j].calculated, rates[j].rate)

    _ = worksheet.cell(row_no + len(rates), column_no, config.base + "/" + config.base)
    cell = worksheet.cell(row_no + len(rates), column_no + 1, "1")
    coordinates[config.base] = cell.coordinate
    logger.debug("Added base currency %s/%s = 1", config.base, config.base)

    logger.debug("Created currency table with %s exchange rates", len(coordinates))
    return coordinates
