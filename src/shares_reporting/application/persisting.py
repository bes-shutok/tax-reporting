"""Persistence layer containing logic for generating reports and saving data."""

from __future__ import annotations

import csv
from pathlib import Path
from typing import TYPE_CHECKING

import openpyxl
from openpyxl.comments import Comment
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

if TYPE_CHECKING:
    from os import PathLike

    from .crypto_reporting import CryptoTaxReport

from ..domain.collections import (
    CapitalGainLinesPerCompany,
    DividendIncomePerCompany,
    TradeCyclePerCompany,
)
from ..domain.constants import (
    EXCEL_COLUMN_OFFSET,
    EXCEL_HEADER_ROW_1,
    EXCEL_HEADER_ROW_2,
    EXCEL_NUMBER_FORMAT,
    EXCEL_START_COLUMN,
    EXCEL_START_ROW,
    PLACEHOLDER_YEAR,
    ZERO_QUANTITY,
)
from ..domain.exceptions import FileProcessingError, ReportGenerationError
from ..domain.value_objects import TradeType
from ..infrastructure.config import Config, ConversionRate, load_configuration_from_file
from ..infrastructure.logging_config import create_module_logger

# Import crypto reporting helpers for aggregation
from .crypto_reporting import (
    ZERO,
    AggregatedRewardIncomeEntry,
    RewardTaxClassification,
    aggregate_taxable_rewards,
)


def export_rollover_file(leftover: str | PathLike[str], leftover_trades: TradeCyclePerCompany) -> None:
    """Export unmatched securities rollover file for next year's FIFO calculations.

    This function creates a CSV file containing all trades that couldn't be matched
    during the FIFO capital gains calculation process. These unmatched securities
    are rolled over to the next tax year to maintain FIFO continuity.

    Args:
        leftover: Output file path for the unmatched securities rollover file
        leftover_trades: Dictionary of trades to be rolled over to next year's calculations
    """
    logger = create_module_logger(__name__)
    logger.info("Generating unmatched securities rollover file: %s", Path(leftover).name)

    safe_remove_file(leftover)
    processed_companies = ZERO_QUANTITY

    with Path(leftover).open("w", newline="") as right_obj:
        writer = csv.DictWriter(
            right_obj,
            fieldnames=[
                "Trades",
                "Header",
                "DataDiscriminator",
                "Asset Category",
                "Currency",
                "Symbol",
                "Date/Time",
                "Quantity",
                "T. Price",
                "C. Price",
                "Proceeds",
                "Comm/Fee",
                "Basis",
                "Realized P/L",
            ],
        )
        writer.writeheader()
        for currency_company, trade_cycle in leftover_trades.items():
            processed_companies += 1
            row = {
                "Trades": "Trades",
                "Header": "Data",
                "DataDiscriminator": "Order",
                "Asset Category": "Stocks",
                "Currency": currency_company.currency.currency,
                "Symbol": currency_company.company.ticker,
            }

            logger.debug("Processing leftover trades for %s (%s)", row["Symbol"], row["Currency"])

            # we are not expecting any sold shares in the leftover file
            if trade_cycle.has_bought():
                bought_trades = trade_cycle.get(TradeType.BUY)
                logger.debug("Writing %s leftover buy trades for %s", len(bought_trades), row["Symbol"])

                for bought_trade in bought_trades:
                    row["Quantity"] = str(bought_trade.quantity)
                    action = bought_trade.action
                    row["Date/Time"] = str(action.date_time.date()) + ", " + str(action.date_time.time())
                    row["T. Price"] = str(action.price)
                    row["Proceeds"] = str(action.price * bought_trade.quantity)
                    proportional_fee = action.fee * (bought_trade.quantity / action.quantity)
                    row["Comm/Fee"] = str(proportional_fee)
                    row["Basis"] = ""  # Empty for unmatched trades
                    row["Realized P/L"] = ""  # Empty for unmatched trades
                    writer.writerow(row)

    logger.info("Generated unmatched securities rollover file for %s companies", processed_companies)


def generate_tax_report(  # noqa: PLR0912, PLR0915
    extract: str | PathLike[str],
    capital_gain_lines_per_company: CapitalGainLinesPerCompany,
    dividend_income_per_company: DividendIncomePerCompany | None = None,
    crypto_tax_report: CryptoTaxReport | None = None,
) -> bool:
    """Generate comprehensive Excel tax report with capital gains and dividend income.

    This function creates a professional Excel report containing all tax-relevant
    information for capital gains and dividend income reporting, including currency
    exchange rate tables and proper formatting for submission to tax authorities.

    Args:
        extract: Output file path for the Excel tax report
        capital_gain_lines_per_company: Calculated capital gains grouped by company
        dividend_income_per_company: Dividend income data grouped by company (optional)
        crypto_tax_report: Crypto tax data parsed from Koinly exports (optional)

    Returns:
        True if the Crypto sheet was successfully created, False otherwise.
        Returns False when crypto_tax_report is None or crypto sheet rendering fails.
    """
    logger = create_module_logger(__name__)
    logger.info("Generating capital gains report: %s", Path(extract).name)

    total_gain_lines = sum(len(lines) for lines in capital_gain_lines_per_company.values())
    logger.debug(
        "Processing %s capital gain lines across %s companies",
        total_gain_lines,
        len(capital_gain_lines_per_company),
    )

    first_header = [
        "Beneficiary",
        "Country of Source",
        "SALE",
        "",
        "",
        "",
        "PURCHASE",
        "",
        "",
        "",
        "WITHOLDING TAX",
        "",
        "Expenses incurred with obtaining the capital gains",
        "",
        "Symbol",
        "Currency",
        "Sale amount",
        "Buy amount",
        "Expenses amount",
    ]
    second_header = [
        "",
        "",
        "Day ",
        "Month ",
        "Year",
        "Amount",
        "Day ",
        "Month ",
        "Year",
        "Amount",
        "Country",
        "Amount",
        "",
        "",
        "",
        "",
        "",
        "",
        "",
    ]

    last_column: int = max(len(first_header), len(second_header))
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    if worksheet is None:
        raise ReportGenerationError("Failed to create worksheet in workbook")
    worksheet.title = "Reporting"

    try:
        config: Config = load_configuration_from_file()
        exchange_rates: dict[str, str] = create_currency_table(worksheet, last_column + 2, 1, config)
        logger.debug("Created currency exchange table with %s rates", len(config.rates) + 1)
    except Exception as e:
        raise ReportGenerationError(f"Failed to read configuration for currency exchange: {e}") from e

    for i in range(len(first_header)):
        _ = worksheet.cell(EXCEL_HEADER_ROW_1, i + 1, first_header[i])
        _ = worksheet.cell(EXCEL_HEADER_ROW_2, i + 1, second_header[i])

    start_column = EXCEL_START_COLUMN
    line_number = EXCEL_START_ROW
    processed_lines = ZERO_QUANTITY

    for currency_company, capital_gain_lines in capital_gain_lines_per_company.items():
        currency = currency_company.currency
        company = currency_company.company
        logger.debug("Processing capital gain lines for %s (%s)", company.ticker, currency.currency)

        for line in capital_gain_lines:
            if currency != line.get_currency():
                raise ReportGenerationError(f"Currency mismatch in line: {currency} != {line.get_currency()}")
            processed_lines += 1
            idx = start_column

            # SALE information
            _ = worksheet.cell(line_number, start_column, line.get_sell_date().day)
            idx += 1
            _ = worksheet.cell(line_number, idx, line.get_sell_date().get_month_name())
            idx += 1
            _ = worksheet.cell(line_number, idx, line.get_sell_date().year)
            idx += 1
            _ = worksheet.cell(
                line_number,
                idx,
                "=" + exchange_rates[currency.currency] + "*(" + line.get_sell_amount() + ")",
            )

            # PURCHASE information
            idx += 1
            _ = worksheet.cell(line_number, idx, line.get_buy_date().day)
            idx += 1
            _ = worksheet.cell(line_number, idx, line.get_buy_date().get_month_name())
            idx += 1
            _ = worksheet.cell(line_number, idx, line.get_buy_date().year)
            idx += 1
            _ = worksheet.cell(
                line_number,
                idx,
                "=" + exchange_rates[currency.currency] + "*(" + line.get_buy_amount() + ")",
            )

            # WITHOLDING TAX information (skip Country and Amount columns for now)
            idx += EXCEL_COLUMN_OFFSET

            # EXPENSES information
            expense_cell = worksheet.cell(
                line_number,
                idx,
                "=" + exchange_rates[currency.currency] + "*(" + line.get_expense_amount() + ")",
            )
            expense_cell.number_format = EXCEL_NUMBER_FORMAT  # type: ignore[assignment]
            idx += 2

            # Symbol and Currency
            _ = worksheet.cell(line_number, idx, company.ticker)
            idx += 1
            _ = worksheet.cell(line_number, idx, currency.currency)
            idx += 1

            # Amounts section
            sell_amount_cell = worksheet.cell(line_number, idx, "=" + line.get_sell_amount())
            sell_amount_cell.number_format = EXCEL_NUMBER_FORMAT  # type: ignore[assignment]
            idx += 1
            buy_amount_cell = worksheet.cell(line_number, idx, "=" + line.get_buy_amount())
            buy_amount_cell.number_format = EXCEL_NUMBER_FORMAT  # type: ignore[assignment]
            idx += 1
            expense_amount_cell = worksheet.cell(line_number, idx, "=" + line.get_expense_amount())
            expense_amount_cell.number_format = EXCEL_NUMBER_FORMAT  # type: ignore[assignment]

            # Highlight placeholder buy transactions in red
            if line.get_buy_date().year == PLACEHOLDER_YEAR:
                red_fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")
                # Apply red fill to entire row
                for col_idx in range(start_column, idx + 1):
                    cell = worksheet.cell(line_number, col_idx)
                    cell.fill = red_fill  # type: ignore[assignment]

            line_number += 1

    logger.debug("Processed %s capital gain lines", processed_lines)

    # Populate Country of Source column for all rows
    # This is done after the main loop to ensure we have all data
    line_number = EXCEL_START_ROW
    for currency_company, capital_gain_lines in capital_gain_lines_per_company.items():
        company = currency_company.company
        for _ in capital_gain_lines:
            # Column 2 is "Country of Source" (according to first_header array)
            _ = worksheet.cell(line_number, 2, company.country_of_issuance)

            # Populate WITHOLDING TAX Country (Column 11 according to first_header array)
            _ = worksheet.cell(line_number, 11, company.country_of_issuance)

            line_number += 1

    # Add CAPITAL INVESTMENT INCOME section if dividend data is provided
    if dividend_income_per_company:
        logger.info("Adding CAPITAL INVESTMENT INCOME section with %s securities", len(dividend_income_per_company))

        # Add empty row for spacing
        line_number += 1

        # Section title "5. CAPITAL INVESTMENT INCOME:"
        section_title_cell = worksheet.cell(line_number, 1, "5. CAPITAL INVESTMENT INCOME:")
        section_title_cell.font = Font(bold=True)  # type: ignore[assignment]
        line_number += 1

        # Empty row
        line_number += 1

        # Dividend income headers
        dividend_headers = [
            "Beneficiary\n(choose one)",
            "Type of capital income\n(choose one)",
            "Country of source",
            "ISIN",
            "Gross amount",
            "Withholding tax at source",
            "Withholding tax in Portugal\n(if any)",
            "",  # One empty column separator
            "Symbol",
            "Currency",
            "Original gross amount",
            "Original tax amount",
            "Net amount",
        ]

        for i, header in enumerate(dividend_headers):
            _ = worksheet.cell(line_number, i + 1, header)

        line_number += 1

        # Dividend income data rows
        for symbol, dividend_data in dividend_income_per_company.items():
            _ = worksheet.cell(line_number, 1, "")  # Beneficiary column
            _ = worksheet.cell(line_number, 2, "Dividends")  # Type of capital income

            # Handle missing security information with error highlighting
            if dividend_data.isin == "MISSING_ISIN_REQUIRES_ATTENTION":
                # Highlight missing ISIN entries with red background
                country_cell = worksheet.cell(line_number, 3, "⚠️ MISSING DATA")
                country_cell.fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")  # type: ignore[assignment]

                isin_cell = worksheet.cell(line_number, 4, f"⚠️ {symbol}")
                isin_cell.fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")  # type: ignore[assignment]

                # Add comment explaining the issue
                isin_cell.comment = Comment(  # type: ignore[assignment]
                    f"Security information missing for {symbol}. Please verify this symbol in your IB account.",
                    "Shares Reporting",
                )
            else:
                _ = worksheet.cell(line_number, 3, dividend_data.country)  # Country of source
                _ = worksheet.cell(line_number, 4, dividend_data.isin)  # ISIN

            # Convert amounts using exchange rates and add Excel formulas
            gross_amount_cell = worksheet.cell(
                line_number,
                5,
                "=" + exchange_rates[dividend_data.currency.currency] + "*(" + str(dividend_data.gross_amount) + ")",
            )
            gross_amount_cell.number_format = EXCEL_NUMBER_FORMAT  # type: ignore[assignment]

            tax_amount_cell = worksheet.cell(
                line_number,
                6,
                "=" + exchange_rates[dividend_data.currency.currency] + "*(" + str(dividend_data.total_taxes) + ")",
            )
            tax_amount_cell.number_format = EXCEL_NUMBER_FORMAT  # type: ignore[assignment]

            _ = worksheet.cell(line_number, 7, "")  # Withholding tax in Portugal (empty for now)

            # Symbol and Currency columns (new columns)
            _ = worksheet.cell(line_number, 9, symbol)  # Symbol column
            _ = worksheet.cell(line_number, 10, dividend_data.currency.currency)  # Currency column

            # Original amounts in original currency (new columns)
            original_gross_cell = worksheet.cell(line_number, 11, str(dividend_data.gross_amount))
            original_gross_cell.number_format = EXCEL_NUMBER_FORMAT

            original_tax_cell = worksheet.cell(line_number, 12, str(dividend_data.total_taxes))
            original_tax_cell.number_format = EXCEL_NUMBER_FORMAT

            # Net amount (gross - tax) in original currency (new column)
            net_amount = dividend_data.gross_amount - dividend_data.total_taxes
            net_amount_cell = worksheet.cell(line_number, 13, str(net_amount))
            net_amount_cell.number_format = EXCEL_NUMBER_FORMAT

            logger.debug(
                "Added dividend income row for %s: %s gross, %s tax, %s net (%s)",
                symbol,
                dividend_data.gross_amount,
                dividend_data.total_taxes,
                net_amount,
                dividend_data.currency.currency,
            )
            line_number += 1

    # auto width for the populated cells
    logger.debug("Auto-adjusting column widths")
    for column_cells in worksheet.columns:
        length = max((len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells), default=0)
        first_cell = column_cells[0]
        # Use cell column property to get column index
        try:
            column_idx = first_cell.column
            if column_idx is not None:
                column_letter = get_column_letter(column_idx)
                worksheet.column_dimensions[column_letter].width = length + 2
        except (AttributeError, TypeError):
            # Skip if we can't determine column letter
            pass

    crypto_sheet_created = False
    if crypto_tax_report:
        logger.info(
            "Adding Crypto worksheet with %s capital and %s reward rows",
            len(crypto_tax_report.capital_entries),
            len(crypto_tax_report.reward_entries),
        )
        crypto_worksheet = None
        try:
            # Aggregate taxable-now rewards before sheet creation to ensure validation runs first.
            # If aggregation fails due to validation (e.g., invalid Tabela X country), the entire
            # report generation fails with a clear error per plan requirement (Task 2).
            aggregated_rewards = aggregate_taxable_rewards(crypto_tax_report.reward_entries)
            add_crypto_report_sheet(workbook, crypto_tax_report, aggregated_rewards)
            crypto_sheet_created = True
        except FileProcessingError:
            # Remove partially-created worksheet if it exists, clean up resources,
            # remove old output file, then re-raise. FileProcessingError from
            # aggregate_taxable_rewards indicates missing mandatory IRS fields for
            # immediately taxable income; per plan, generation must fail.
            crypto_worksheet = workbook["Crypto"] if "Crypto" in workbook.sheetnames else None
            if crypto_worksheet is not None:
                workbook.remove(crypto_worksheet)
                logger.debug("Removed partially-created Crypto worksheet due to validation error")
            workbook.close()
            safe_remove_file(extract)
            raise
        except Exception:
            # Any other exception also requires cleanup: remove partial worksheet,
            # close workbook, remove old output file, then re-raise.
            crypto_worksheet = workbook["Crypto"] if "Crypto" in workbook.sheetnames else None
            if crypto_worksheet is not None:
                workbook.remove(crypto_worksheet)
                logger.debug("Removed partially-created Crypto worksheet due to unexpected error")
            workbook.close()
            safe_remove_file(extract)
            raise

    safe_remove_file(extract)

    try:
        workbook.save(extract)
        report_type = "capital gains and dividend income" if dividend_income_per_company else "capital gains"
        logger.info("Successfully generated %s report with %s capital gain lines", report_type, processed_lines)
    except Exception as e:
        raise ReportGenerationError(f"Failed to save Excel report: {e}") from e
    finally:
        workbook.close()

    return crypto_sheet_created


def add_crypto_report_sheet(  # noqa: PLR0912, PLR0915
    workbook: openpyxl.Workbook,
    crypto_tax_report: CryptoTaxReport,
    aggregated_rewards: list[AggregatedRewardIncomeEntry],
) -> None:
    """Append a dedicated crypto worksheet with capital, rewards, and reconciliation data.

    Args:
        workbook: The Excel workbook to add the sheet to.
        crypto_tax_report: The crypto tax report data.
        aggregated_rewards: Pre-computed aggregated taxable-now rewards. Must be validated
            before calling this function to ensure the sheet is never partially created.
    """
    worksheet = workbook.create_sheet("Crypto")
    worksheet.cell(1, 1, "CRYPTO TAX REPORT - PORTUGAL").font = Font(bold=True)  # type: ignore[assignment]
    worksheet.cell(2, 1, "Tax year")
    worksheet.cell(2, 2, crypto_tax_report.tax_year)
    if crypto_tax_report.pdf_summary:
        worksheet.cell(3, 1, "PDF period")
        worksheet.cell(3, 2, crypto_tax_report.pdf_summary.period or "N/A")
        worksheet.cell(3, 3, "PDF timezone")
        worksheet.cell(3, 4, crypto_tax_report.pdf_summary.timezone or "N/A")
        worksheet.cell(3, 5, "PDF extracted tokens")
        worksheet.cell(3, 6, crypto_tax_report.pdf_summary.extracted_tokens)

    row_no = 5 if crypto_tax_report.pdf_summary else 4
    worksheet.cell(row_no, 1, "1. CAPITAL GAINS").font = Font(bold=True)  # type: ignore[assignment]
    row_no += 1

    capital_headers = [
        "Disposal date",
        "Acquisition date",
        "Asset",
        "Amount",
        "Cost (EUR)",
        "Proceeds (EUR)",
        "Gain/Loss (EUR)",
        "Holding period",
        "Wallet",
        "Platform",
        "Disposal chain",
        "Operator entity",
        "Operator country",
        "Annex hint",
        "Review flag",
        "Notes",
        "Token origin",
    ]
    for idx, header in enumerate(capital_headers, start=1):
        worksheet.cell(row_no, idx, header)
    row_no += 1

    for entry in crypto_tax_report.capital_entries:
        worksheet.cell(row_no, 1, entry.disposal_date)
        worksheet.cell(row_no, 2, entry.acquisition_date)
        worksheet.cell(row_no, 3, entry.asset)
        worksheet.cell(row_no, 4, float(entry.amount))
        worksheet.cell(row_no, 5, float(entry.cost_eur))
        worksheet.cell(row_no, 6, float(entry.proceeds_eur))
        worksheet.cell(row_no, 7, float(entry.gain_loss_eur))
        worksheet.cell(row_no, 8, entry.holding_period)
        worksheet.cell(row_no, 9, entry.wallet)
        worksheet.cell(row_no, 10, entry.platform)
        worksheet.cell(row_no, 11, entry.chain)
        worksheet.cell(row_no, 12, entry.operator_origin.operator_entity)
        worksheet.cell(row_no, 13, entry.operator_origin.operator_country)
        worksheet.cell(row_no, 14, entry.annex_hint)
        worksheet.cell(row_no, 15, f"YES: {entry.review_reason}" if entry.review_required else "NO")
        worksheet.cell(row_no, 16, entry.notes)
        worksheet.cell(row_no, 17, entry.token_swap_history or "")
        row_no += 1

    # Split reward entries for detailed sections (aggregated rewards already computed)
    taxable_now_entries = [
        e for e in crypto_tax_report.reward_entries if e.tax_classification == RewardTaxClassification.TAXABLE_NOW
    ]
    deferred_entries = [
        e for e in crypto_tax_report.reward_entries if e.tax_classification == RewardTaxClassification.DEFERRED_BY_LAW
    ]

    # 2. REWARDS INCOME - IRS-READY SUMMARY (taxable_now aggregated)
    row_no += 2
    worksheet.cell(row_no, 1, "2. REWARDS INCOME - IRS-READY FILING SUMMARY").font = Font(bold=True)  # type: ignore[assignment]
    row_no += 1

    summary_note = worksheet.cell(
        row_no,
        1,
        "This table shows only income taxable immediately in Category E. "
        "Crypto-denominated rewards (deferred until disposal) are in section 2b below.",
    )
    summary_note.font = Font(italic=True, size=9)  # type: ignore[assignment]
    row_no += 1

    # IRS-ready aggregated rewards table headers
    aggregated_headers = [
        "Income code",
        "Source country",
        "Reward chain",
        "Gross income (EUR)",
        "Foreign tax (EUR)",
        "Net income (EUR)",
        "Raw rows",
    ]
    for idx, header in enumerate(aggregated_headers, start=1):
        header_cell = worksheet.cell(row_no, idx, header)
        header_cell.font = Font(bold=True)  # type: ignore[assignment]
    row_no += 1

    # Write aggregated taxable-now rows
    if aggregated_rewards:
        for entry in aggregated_rewards:
            worksheet.cell(row_no, 1, entry.income_code)
            worksheet.cell(row_no, 2, entry.source_country)
            worksheet.cell(row_no, 3, ", ".join(entry.chains) if entry.chains else "Unknown")
            worksheet.cell(row_no, 4, float(entry.gross_income_eur))
            worksheet.cell(row_no, 5, float(entry.foreign_tax_eur))
            worksheet.cell(row_no, 6, float(entry.gross_income_eur - entry.foreign_tax_eur))
            worksheet.cell(row_no, 7, entry.raw_row_count)
            row_no += 1
    else:
        worksheet.cell(
            row_no, 1, "No immediately taxable rewards (all rewards are crypto-denominated and deferred by law)"
        )
        row_no += 1

    # 2a2. TAXABLE-NOW SUPPORT DETAIL (reconciliation trail per plan requirement)
    row_no += 1
    worksheet.cell(row_no, 1, "2a2. TAXABLE-NOW - SUPPORT DETAIL").font = Font(bold=True)  # type: ignore[assignment]
    row_no += 1
    taxable_note = worksheet.cell(
        row_no,
        1,
        "Individual rows that contribute to the IRS-ready filing table above. "
        "Use this section to trace each aggregated line back to its source Koinly rows.",
    )
    taxable_note.font = Font(italic=True, size=9)  # type: ignore[assignment]
    row_no += 1

    taxable_detail_headers = [
        "Date",
        "Asset",
        "Value (EUR)",
        "Income type",
        "Wallet",
        "Platform",
        "Reward chain",
        "Country",
        "Foreign tax (EUR)",
        "Review flag",
        "Description",
    ]
    for idx, header in enumerate(taxable_detail_headers, start=1):
        header_cell = worksheet.cell(row_no, idx, header)
        header_cell.font = Font(bold=True)  # type: ignore[assignment]
    row_no += 1

    if taxable_now_entries:
        for entry in taxable_now_entries:
            worksheet.cell(row_no, 1, entry.date)
            worksheet.cell(row_no, 2, entry.asset)
            worksheet.cell(row_no, 3, float(entry.value_eur))
            worksheet.cell(row_no, 4, entry.source_type)
            worksheet.cell(row_no, 5, entry.wallet)
            worksheet.cell(row_no, 6, entry.platform)
            worksheet.cell(row_no, 7, entry.chain)
            worksheet.cell(row_no, 8, entry.operator_origin.operator_country)
            worksheet.cell(row_no, 9, float(entry.foreign_tax_eur))
            worksheet.cell(row_no, 10, f"YES: {entry.review_reason}" if entry.review_required else "NO")
            worksheet.cell(row_no, 11, entry.description)
            row_no += 1
    else:
        worksheet.cell(row_no, 1, "No taxable-now rewards")
        row_no += 1

    # 2b. DEFERRED BY LAW SUPPORT SECTION
    row_no += 1
    worksheet.cell(row_no, 1, "2b. DEFERRED BY LAW - SUPPORT DETAIL").font = Font(bold=True)  # type: ignore[assignment]
    row_no += 1

    deferred_note = worksheet.cell(
        row_no,
        1,
        "These rewards are crypto-denominated and taxation is deferred until disposal per CIRS art. 5(11). "
        "They are shown here for auditability but NOT included in the IRS-ready filing table above.",
    )
    deferred_note.font = Font(italic=True, size=9)  # type: ignore[assignment]
    row_no += 1

    # Deferred rewards support table headers
    deferred_headers = [
        "Date",
        "Asset",
        "Value (EUR)",
        "Income type",
        "Wallet",
        "Platform",
        "Reward chain",
        "Country",
        "Foreign tax (EUR)",
        "Review flag",
        "Description",
    ]
    for idx, header in enumerate(deferred_headers, start=1):
        header_cell = worksheet.cell(row_no, idx, header)
        header_cell.font = Font(bold=True)  # type: ignore[assignment]
    row_no += 1

    # Write deferred reward rows (support detail for auditability)
    if deferred_entries:
        for entry in deferred_entries:
            worksheet.cell(row_no, 1, entry.date)
            worksheet.cell(row_no, 2, entry.asset)
            worksheet.cell(row_no, 3, float(entry.value_eur))
            worksheet.cell(row_no, 4, entry.source_type)
            worksheet.cell(row_no, 5, entry.wallet)
            worksheet.cell(row_no, 6, entry.platform)
            worksheet.cell(row_no, 7, entry.chain)
            worksheet.cell(row_no, 8, entry.operator_origin.operator_country)
            worksheet.cell(row_no, 9, float(entry.foreign_tax_eur))
            worksheet.cell(row_no, 10, f"YES: {entry.review_reason}" if entry.review_required else "NO")
            worksheet.cell(row_no, 11, entry.description)
            row_no += 1
    else:
        worksheet.cell(row_no, 1, "No deferred rewards")
        row_no += 1

    # 2c. REWARDS CLASSIFICATION RECONCILIATION
    row_no += 1
    worksheet.cell(row_no, 1, "2c. REWARDS CLASSIFICATION RECONCILIATION").font = Font(bold=True)  # type: ignore[assignment]
    row_no += 1

    # Calculate totals for reconciliation
    taxable_now_total_eur = sum((e.value_eur for e in taxable_now_entries), start=ZERO)
    deferred_total_eur = sum((e.value_eur for e in deferred_entries), start=ZERO)

    reconciliation_rewards_rows = [
        ("Total reward rows (raw)", len(crypto_tax_report.reward_entries)),
        ("Taxable-now rows (immediately taxable)", len(taxable_now_entries)),
        ("Deferred-by-law rows (taxation deferred)", len(deferred_entries)),
        ("Taxable-now total value (EUR)", float(taxable_now_total_eur)),
        ("Deferred total value (EUR)", float(deferred_total_eur)),
        ("Filing-ready lines after aggregation", len(aggregated_rewards)),
    ]

    for key, value in reconciliation_rewards_rows:
        worksheet.cell(row_no, 1, key)
        worksheet.cell(row_no, 2, value)
        row_no += 1

    row_no += 2
    worksheet.cell(row_no, 1, "3. RECONCILIATION").font = Font(bold=True)  # type: ignore[assignment]
    row_no += 1

    reconciliation_rows = [
        ("Capital sale events (aggregated)", crypto_tax_report.reconciliation.capital_rows),
        ("Reward rows", crypto_tax_report.reconciliation.reward_rows),
        ("Short term rows", crypto_tax_report.reconciliation.short_term_rows),
        ("Long term rows", crypto_tax_report.reconciliation.long_term_rows),
        ("Mixed holding period rows", crypto_tax_report.reconciliation.mixed_rows),
        ("Unknown holding period rows", crypto_tax_report.reconciliation.unknown_rows),
        ("Capital cost total (EUR)", float(crypto_tax_report.reconciliation.capital_cost_total_eur)),
        ("Capital proceeds total (EUR)", float(crypto_tax_report.reconciliation.capital_proceeds_total_eur)),
        ("Capital gain total (EUR)", float(crypto_tax_report.reconciliation.capital_gain_total_eur)),
        ("Rewards total (EUR)", float(crypto_tax_report.reconciliation.reward_total_eur)),
    ]

    opening_holdings = crypto_tax_report.reconciliation.opening_holdings
    if opening_holdings:
        reconciliation_rows.extend(
            [
                ("Opening holdings rows", opening_holdings.asset_rows),
                ("Opening holdings cost (EUR)", float(opening_holdings.total_cost_eur)),
                ("Opening holdings value (EUR)", float(opening_holdings.total_value_eur)),
            ]
        )

    closing_holdings = crypto_tax_report.reconciliation.closing_holdings
    if closing_holdings:
        reconciliation_rows.extend(
            [
                ("Closing holdings rows", closing_holdings.asset_rows),
                ("Closing holdings cost (EUR)", float(closing_holdings.total_cost_eur)),
                ("Closing holdings value (EUR)", float(closing_holdings.total_value_eur)),
            ]
        )

    for key, value in reconciliation_rows:
        worksheet.cell(row_no, 1, key)
        worksheet.cell(row_no, 2, value)
        row_no += 1

    row_no += 2
    worksheet.cell(row_no, 1, "4. SKIPPED ZERO VALUE TOKENS").font = Font(bold=True)  # type: ignore[assignment]
    row_no += 1
    worksheet.cell(row_no, 1, "Source section")
    worksheet.cell(row_no, 2, "Asset")
    worksheet.cell(row_no, 3, "Skipped rows")
    row_no += 1

    if crypto_tax_report.skipped_zero_value_tokens:
        for skipped in crypto_tax_report.skipped_zero_value_tokens:
            worksheet.cell(row_no, 1, skipped.source_section)
            worksheet.cell(row_no, 2, skipped.asset)
            worksheet.cell(row_no, 3, skipped.count)
            row_no += 1
    else:
        worksheet.cell(row_no, 1, "none")
        worksheet.cell(row_no, 2, "")
        worksheet.cell(row_no, 3, 0)

    for column_cells in worksheet.columns:
        length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
        first_cell = column_cells[0]
        try:
            column_idx = first_cell.column
            if column_idx is not None:
                column_letter = get_column_letter(column_idx)
                worksheet.column_dimensions[column_letter].width = length + 2
        except (AttributeError, TypeError):
            pass


def safe_remove_file(path: str | PathLike[str]) -> None:
    """Safely remove a file if it exists, logging any errors.

    Args:
        path: File path to remove
    """
    logger = create_module_logger(__name__)
    try:
        p = Path(path)
        if p.exists():
            p.unlink()
            logger.debug("Removed existing file: %s", p.name)
    except OSError as e:
        logger.warning("Failed to remove file %s: %s", path, e)
        # Non-critical error, continue processing


# https://openpyxl.readthedocs.io/en/latest/tutorial.html
def create_currency_table(worksheet: Worksheet, column_no: int, row_no: int, config: Config) -> dict[str, str]:
    """Create a currency configuration table in the excel worksheet.

    Args:
        worksheet: The target excel worksheet.
        column_no: The starting column number (1-based).
        row_no: The starting row number (1-based).
        config: The application configuration object.

    Returns:
        A dictionary mapping cell coordinates to their formatted values.
    """
    logger = create_module_logger(__name__)
    currency_header = ["Base/target", "Rate"]
    rates: list[ConversionRate] = config.rates

    logger.debug("Creating currency table starting at column %s, row %s", column_no, row_no)

    _ = worksheet.cell(row_no, column_no, "Currency exchange rate")
    row_no += 1
    for i in range(len(currency_header)):
        _ = worksheet.cell(row_no, column_no + i, currency_header[i])
    row_no += 1

    coordinates: dict[str, str] = {}
    for j in range(len(rates)):
        _ = worksheet.cell(row_no + j, column_no, rates[j].base + "/" + rates[j].calculated)
        cell = worksheet.cell(row_no + j, column_no + 1, str(rates[j].rate))
        coordinates[rates[j].calculated] = cell.coordinate
        logger.debug("Added currency rate %s/%s = %s", rates[j].base, rates[j].calculated, rates[j].rate)

    # Add base currency rate (1:1)
    _ = worksheet.cell(row_no + len(rates), column_no, config.base + "/" + config.base)
    cell = worksheet.cell(row_no + len(rates), column_no + 1, "1")
    coordinates[config.base] = cell.coordinate
    logger.debug("Added base currency %s/%s = 1", config.base, config.base)

    logger.debug("Created currency table with %s exchange rates", len(coordinates))
    return coordinates
