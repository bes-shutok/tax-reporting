"""Crypto reconciliation sheet writer for the Excel tax report."""

from __future__ import annotations

from typing import TYPE_CHECKING

import openpyxl
from openpyxl.styles import Font

if TYPE_CHECKING:
    from ..crypto_reporting import CryptoTaxReport

from .excel_utils import auto_column_width


def write_crypto_reconciliation_sheet(workbook: openpyxl.Workbook, crypto_tax_report: CryptoTaxReport) -> None:
    """Create and populate a 'Crypto Reconciliation' worksheet with reconciliation and skipped token data.

    Writes:
    - Section 3: Reconciliation key-value rows (capital/reward totals, holding period counts, holdings)
    - Section 4: Skipped zero value tokens table

    Args:
        workbook: The Excel workbook to add the sheet to.
        crypto_tax_report: The crypto tax report data.
    """
    worksheet = workbook.create_sheet("Crypto Reconciliation")

    row_no = 1

    # 3. RECONCILIATION
    worksheet.cell(row_no, 1, "3. RECONCILIATION").font = Font(bold=True)
    row_no += 1

    reconciliation_rows = [
        ("Capital sale events (aggregated)", crypto_tax_report.reconciliation.capital_rows),
        ("Reward rows", crypto_tax_report.reconciliation.reward_rows),
        # Cross-sheet audit count (CRG-022 / r3 review finding #2). The
        # ``reward_rows`` count above drops under the parse-time skip
        # (zero-value DEFERRED_BY_LAW rows leave ``reward_entries``). Without this sibling line, the Reconciliation
        # sheet's ``Reward rows`` count would silently disagree with the
        # Supplementary sheet's deferred subtotal. This is the audit line r1
        # review dropped as tautological on the Supplementary sheet (dust +
        # unpriced == total by construction there) - here it is load-bearing
        # because it makes the cross-sheet count auditable.
        ("Skipped zero-value deferred rewards (audit)", len(crypto_tax_report.skipped_zero_value_deferred_rewards)),
        ("Short term rows", crypto_tax_report.reconciliation.short_term_rows),
        ("Long term rows", crypto_tax_report.reconciliation.long_term_rows),
        ("Mixed holding period rows", crypto_tax_report.reconciliation.mixed_rows),
        ("Unknown holding period rows", crypto_tax_report.reconciliation.unknown_rows),
        ("Capital cost total (EUR)", float(crypto_tax_report.reconciliation.capital_cost_total_eur)),
        ("Capital proceeds total (EUR)", float(crypto_tax_report.reconciliation.capital_proceeds_total_eur)),
        ("Capital gain total (EUR)", float(crypto_tax_report.reconciliation.capital_gain_total_eur)),
        ("Rewards total (EUR)", float(crypto_tax_report.reconciliation.reward_total_eur)),
    ]

    opening_holdings = crypto_tax_report.reconciliation.opening_holdings
    if opening_holdings:
        reconciliation_rows.extend(
            [
                ("Opening holdings rows", opening_holdings.asset_rows),
                ("Opening holdings cost (EUR)", float(opening_holdings.total_cost_eur)),
                ("Opening holdings value (EUR)", float(opening_holdings.total_value_eur)),
            ]
        )

    closing_holdings = crypto_tax_report.reconciliation.closing_holdings
    if closing_holdings:
        reconciliation_rows.extend(
            [
                ("Closing holdings rows", closing_holdings.asset_rows),
                ("Closing holdings cost (EUR)", float(closing_holdings.total_cost_eur)),
                ("Closing holdings value (EUR)", float(closing_holdings.total_value_eur)),
            ]
        )

    for key, value in reconciliation_rows:
        worksheet.cell(row_no, 1, key)
        worksheet.cell(row_no, 2, value)
        row_no += 1

    # 4. SKIPPED ZERO VALUE TOKENS
    row_no += 2
    worksheet.cell(row_no, 1, "4. SKIPPED ZERO VALUE TOKENS").font = Font(bold=True)
    row_no += 1
    worksheet.cell(row_no, 1, "Source section")
    worksheet.cell(row_no, 2, "Asset")
    worksheet.cell(row_no, 3, "Skipped rows")
    worksheet.cell(row_no, 4, "Suspicious")
    row_no += 1

    if crypto_tax_report.skipped_zero_value_tokens:
        for skipped in crypto_tax_report.skipped_zero_value_tokens:
            worksheet.cell(row_no, 1, skipped.source_section)
            worksheet.cell(row_no, 2, skipped.asset)
            worksheet.cell(row_no, 3, skipped.count)
            # Mark suspicious assets (potential homoglyph scam tokens)
            if skipped.suspicious:
                suspicious_cell = worksheet.cell(row_no, 4, "YES - non-Latin characters, probable scam")
                suspicious_cell.font = Font(color="FF0000")  # Red text
                suspicious_cell = worksheet.cell(row_no, 2, skipped.asset)
                suspicious_cell.font = Font(color="FF0000", bold=True)  # Also highlight asset in red
            row_no += 1
    else:
        worksheet.cell(row_no, 1, "none")
        worksheet.cell(row_no, 2, "")
        worksheet.cell(row_no, 3, 0)
        worksheet.cell(row_no, 4, "")

    auto_column_width(worksheet)
