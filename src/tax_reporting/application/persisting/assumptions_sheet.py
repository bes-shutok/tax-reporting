"""Assumptions & Methodology sheet writer for platform-level verification notes.

This sheet provides:
1. A complete manifest of every platform seen in the tax report, showing operator
   entity, country, confidence level, any verification assumption, and whether
   the platform requires manual review before filing.
2. Methodology assumptions documenting the legal basis and rationale for reporting
   decisions (aggregation approach, FIFO methodology, holding period classification,
   materiality threshold, data sources).

Platforms with platform_review_required=True are highlighted in red and sorted
to the top so they are immediately visible.
"""

from __future__ import annotations

from collections.abc import Iterable
from dataclasses import dataclass
from typing import TYPE_CHECKING

import openpyxl

if TYPE_CHECKING:
    from tax_reporting.application.crypto.entities import CryptoDecisionCounts
    from tax_reporting.application.crypto.wallet_kind import (
        RegistrySnapshot,
        WalletClassification,
    )
    from tax_reporting.domain.transaction import TransactionHistoryRow

    from ..crypto_reporting import CryptoCapitalGainEntry, CryptoRewardIncomeEntry

from .assumptions_kind_column import classify_platforms_for_summaries
from .excel_utils import REVIEW_ROW_FILL, auto_column_width, safe_cell_value

_KIND_LOW_CONFIDENCE_NOTE_PREFIX = "Kind classification below threshold"
"""Prefix for the Note-column append when kind classification is low-confidence.

CLAUDE.md: review reasons must be specific and actionable; this prefix is
joined with the resolver's evidence reason (e.g. vote tallies or 'no rows').
"""

_NOTE_SEPARATOR = "; "
"""Separator between an existing assumption and the kind-low-confidence note."""


@dataclass
class _PlatformSummary:
    """Aggregated per-platform data for the Assumptions & Methodology sheet."""

    platform: str
    operator_entity: str
    operator_country: str
    confidence: str
    platform_review_required: bool
    assumption: str | None
    transaction_count: int
    classification: WalletClassification | None = None
    """Two-tier WalletKind resolver result for this platform. None when the
    caller supplied NEITHER ``th_rows`` NOR ``registry`` - in that case the
    Kind column renders empty and no kind-low-confidence signal fires."""


def _collect_platform_summaries(
    capital_entries: list[CryptoCapitalGainEntry] | None = None,
    reward_entries: list[CryptoRewardIncomeEntry] | None = None,
    th_rows: Iterable[TransactionHistoryRow] | None = None,
    registry: RegistrySnapshot | None = None,
) -> list[_PlatformSummary]:
    """Collect one summary row per unique platform seen in the data.

    Aggregates counts and unions review flags across all entries for the same platform.
    If any entry for a platform has platform_review_required=True, the summary row
    is marked as requiring review.

    When ``th_rows`` is supplied, each summary is also decorated with a
    ``WalletClassification`` from the two-tier resolver. The kind-low-confidence
    signal (``not classification.is_high_probability()``) is OR'd into the
    existing ``platform_review_required`` flag (Design Invariant 15); a Note
    fragment is appended so the user sees a specific reason. When ``th_rows``
    is omitted but ``registry`` is supplied, the Kind column resolves via
    tier-1 registry lookup; when BOTH are omitted, no classification runs and
    the existing columns render byte-identically to the pre-Phase-A baseline.
    """
    summaries: dict[tuple[str, str], _PlatformSummary] = {}

    def _accumulate(entry: CryptoCapitalGainEntry | CryptoRewardIncomeEntry) -> None:
        origin = entry.operator_origin
        key = (origin.platform, origin.operator_entity)
        if key in summaries:
            s = summaries[key]
            s.transaction_count += 1
            if origin.platform_review_required:
                s.platform_review_required = True
            if origin.platform_assumption and not s.assumption:
                s.assumption = origin.platform_assumption
        else:
            summaries[key] = _PlatformSummary(
                platform=origin.platform,
                operator_entity=origin.operator_entity,
                operator_country=origin.operator_country,
                confidence=origin.confidence,
                platform_review_required=origin.platform_review_required,
                assumption=origin.platform_assumption,
                transaction_count=1,
            )

    for entry in capital_entries or []:
        _accumulate(entry)
    for entry in reward_entries or []:
        _accumulate(entry)

    if th_rows is not None or registry is not None:
        # Normalize th_rows to () when None so the registry-only call path
        # (workbook_builder.py production caller, which has no th_rows to
        # pass) does not crash aggregate_platform_evidence (r7 Blocker #1).
        effective_th_rows = th_rows or ()
        platform_names = [s.platform for s in summaries.values()]
        classifications = classify_platforms_for_summaries(
            platform_names, effective_th_rows, registry
        )
        # `classify_platforms_for_summaries` always returns an entry for every
        # requested platform (UNKNOWN at confidence 0.0 when TH-row evidence is
        # absent), so capital-vs-TH platform-name drift (e.g., capital entries
        # see "ByBit" while TH rows are attributed to "ByBit (2)") surfaces via
        # the user-visible Kind-column UNKNOWN + red Review Required + Note
        # text below, not via a logger.warning. Family G + H.
        for s in summaries.values():
            classification = classifications[s.platform]
            s.classification = classification
            if not classification.is_high_probability():
                # OR the kind-low-confidence signal into the existing flag.
                s.platform_review_required = True
                kind_note = f"{_KIND_LOW_CONFIDENCE_NOTE_PREFIX} ({classification.reason})"
                if s.assumption:
                    s.assumption = f"{s.assumption}{_NOTE_SEPARATOR}{kind_note}"
                else:
                    s.assumption = kind_note

    return list(summaries.values())


def write_assumptions_and_methodology_sheet(  # noqa: PLR0912, PLR0913, PLR0915
    workbook: openpyxl.Workbook,
    capital_entries: list[CryptoCapitalGainEntry] | None = None,
    reward_entries: list[CryptoRewardIncomeEntry] | None = None,
    th_rows: Iterable[TransactionHistoryRow] | None = None,
    registry: RegistrySnapshot | None = None,
    decision_counts: CryptoDecisionCounts | None = None,
) -> None:
    """Create and populate an 'Assumptions & Methodology' worksheet.

    Contains two sections:
    1. Platform Assumptions: complete manifest of platforms with operator metadata
    2. Methodology Assumptions: legal basis and rationale for reporting decisions

    Args:
        workbook: The Excel workbook to add the sheet to.
        capital_entries: Capital gain entries used to discover platform metadata.
            Each entry's operator_origin is inspected for platform name, entity,
            country, confidence, and review flags.
        reward_entries: Reward income entries used to discover platform metadata.
            Combined with capital_entries to build the complete platform manifest.
        th_rows: Optional iterable of ``TransactionHistoryRow`` used to classify
            each platform's WalletKind (CEX/DEX/UNKNOWN) via the two-tier
            resolver. When ``th_rows`` is omitted but ``registry`` is supplied,
            the Kind column resolves via tier-1 registry lookup; when BOTH are
            omitted, the Kind column is left blank and no kind-low-confidence
            signal fires.
        registry: Optional ``RegistrySnapshot`` for tier-1 WalletKind lookup.
            Phase D Task 1 constructs a ``ProductionWalletKindRegistry`` backed
            by ``operator_origin`` so each platform's kind is sourced from the
            operator-entity classification (closing the 540-row Binance gap).
        decision_counts: Optional ``CryptoDecisionCounts`` carrying per-run
            crypto-pipeline decision counts. When supplied, the PT-C-028
            "Materiality Threshold" item renders a ``[This run: filtered N
            entries, M retained.]`` suffix. ``None`` for IB-only runs (no
            crypto) - the static description renders with no suffix.
    """
    summaries = _collect_platform_summaries(capital_entries, reward_entries, th_rows, registry)

    worksheet = workbook.create_sheet("Assumptions & Methodology")
    row_no = 1

    # Title
    worksheet.cell(row_no, 1, "Assumptions & Methodology")
    worksheet.cell(row_no, 1).font = openpyxl.styles.Font(bold=True, size=14)
    row_no += 2

    # Description
    description = (
        "Complete manifest of all platforms in this report. "
        "Red rows (Review Required = YES) must be resolved before submitting the tax return. "
        "Informational assumptions are shown in the last column for all platforms."
    )
    worksheet.cell(row_no, 1, description)
    worksheet.cell(row_no, 1).font = openpyxl.styles.Font(italic=True, size=10)
    row_no += 2

    if not summaries:
        worksheet.cell(row_no, 1, "No platform data found.")
        row_no += 2  # Skip past the "no data" message, but don't return
    else:
        # Headers for Platform Assumptions section
        headers = [
            "Platform",
            "Operator Entity",
            "Country",
            "Confidence",
            "Review Required",
            "Assumption / Verification Note",
            "Transaction Count",
            "Kind",
        ]
        for col_idx, header in enumerate(headers, 1):
            cell = worksheet.cell(row_no, col_idx, header)
            cell.font = openpyxl.styles.Font(bold=True)
        row_no += 1

        # Sort: review-required first, then alphabetical by platform name
        sorted_summaries = sorted(summaries, key=lambda s: (not s.platform_review_required, s.platform.lower()))

        for s in sorted_summaries:
            worksheet.cell(row_no, 1, safe_cell_value(s.platform))
            worksheet.cell(row_no, 2, safe_cell_value(s.operator_entity))
            worksheet.cell(row_no, 3, safe_cell_value(s.operator_country))
            worksheet.cell(row_no, 4, safe_cell_value(s.confidence))
            worksheet.cell(row_no, 5, "YES" if s.platform_review_required else "NO")
            worksheet.cell(row_no, 6, safe_cell_value(s.assumption or ""))
            worksheet.cell(row_no, 7, s.transaction_count)
            # Kind column (col 8). Empty when no TH rows were supplied.
            kind_value = s.classification.kind.name if s.classification is not None else ""
            worksheet.cell(row_no, 8, kind_value)
            if s.platform_review_required:
                for col_idx in range(1, 9):
                    worksheet.cell(row_no, col_idx).fill = REVIEW_ROW_FILL  # type: ignore[assignment]
            row_no += 1

        # After platform data section, add spacing before methodology
        row_no += 2

    # After platform data section, add methodology section
    row_no += 2
    worksheet.cell(row_no, 1, "Methodology Assumptions")
    worksheet.cell(row_no, 1).font = openpyxl.styles.Font(bold=True, size=12)
    row_no += 2

    # Grouped methodology sections: each section has a title and list of items
    # Item format: (label, rule_ids, description)
    methodology_items: list[tuple[str, list[tuple[str, str, str]]]] = [
        # Section 1: Taxable Events
        (
            "Taxable Events",
            [
                (
                    "Loan Repayment Exclusion",
                    "DP-001",
                    (
                        "Returning borrowed crypto is NOT a taxable disposal. Under CIRS art. 10(20), "
                        "loan repayments do not constitute 'alienação onerosa' (onerous disposal). "
                        "Koinly incorrectly treats loan repayments as disposals by default; "
                        "this tool excludes them per Portuguese law. "
                        "Source: CIRS art. 10(20), confirmed by AT folheto 2026-01-12."
                    ),
                ),
                (
                    "Crypto-to-Crypto Deferral",
                    "DP-002, PT-C-005",
                    (
                        "Crypto-to-crypto swaps are NOT taxable at the time of the swap. "
                        "When proceeds take the form of another crypto asset, taxation is deferred "
                        "until the replacement asset is disposed of for fiat (CIRS art. 10(20)). "
                        "The replacement asset takes the acquisition cost of the surrendered asset "
                        "(carry-over cost). Koinly setting: 'Realize gains on crypto → crypto trades?' = OFF. "
                        "Exception: swaps with non-EU/EEA counterparties are immediately taxable "
                        "(CIRS art. 10(21)). Source: AT folheto 2026-01-12."
                    ),
                ),
                (
                    "Liquidity Provision",
                    "DP-005",
                    (
                        "Providing liquidity to DEX pools is treated as crypto-to-crypto exchange. "
                        "Depositing crypto to receive LP tokens = deferral under CIRS art. 10(20). "
                        "LP tokens are crypto-assets, so the swap is deferred. "
                        "Koinly setting: 'Realize gains on liquidity transactions?' = OFF. "
                        "Source: CIRS art. 10(20)."
                    ),
                ),
                (
                    "Transfer Fees",
                    "DP-006",
                    (
                        "Transfer fees (gas/network fees paid in crypto to move assets between wallets) are NOT "
                        "taxable disposals under Portuguese law. Paying a network fee yields no received "
                        "consideration, so it does not constitute 'alienação onerosa' under CIRS art. 10(1)(k). "
                        "Koinly's default (ON) incorrectly realizes a capital gain on the fee token (phantom gain). "
                        "Setting 'Realize gains on transfer fees?' = OFF causes Koinly to instead add the fee value "
                        "to the cost basis of the transferred asset, which is the correct "
                        "treatment per CIRS art. 10(4)(a) "
                        "('despesas necessárias e efetivamente praticadas, inerentes à aquisição e alienação'). "
                        "Verified against FY2025 wallet-to-wallet transfer data. "
                        "Source: CIRS art. 10(1)(k) (narrow taxable-event definition), "
                        "CIRS art. 10(4)(a) (expense deductibility)."
                    ),
                ),

            ],
        ),
        # Section 2: Holding Period & Exemptions
        (
            "Holding Period & Exemptions",
            [
                (
                    "365-Day Exemption",
                    "PT-C-011",
                    (
                        "Gains and losses on disposal of crypto assets held for ≥365 days are EXCLUDED "
                        "from taxation (CIRS art. 10(19)). These must be declared in Anexo G1, Quadro 7 "
                        "(not Anexo J Quadro 9.4). Calendar-year arithmetic is used "
                        "(e.g., 2024-02-29 + 365 days = 2025-03-01, not 2025-02-28). "
                        "Source: AT folheto 2026-01-12; Ofício Circulado 20269/2024, section 9."
                    ),
                ),
                (
                    "Transitional Rule",
                    "PT-C-012",
                    (
                        "For crypto acquired before 01/01/2023, the holding period counts from the actual "
                        "acquisition date (not from 01/01/2023). Assets bought before 2023 that were not "
                        "disposed of before 2023 may qualify for the 365-day exemption immediately. "
                        "Source: Art. 220 Lei n.º 24-D/2022 (LOE 2023), 30/12/2022."
                    ),
                ),
                (
                    "Blacklisted Jurisdiction Exception",
                    "PT-C-013, PT-C-017",
                    (
                        "The 365-day exemption does NOT apply when the counterparty is in a jurisdiction "
                        "without a double-tax treaty or information-exchange agreement with Portugal "
                        "(CIRS art. 10(21)). Losses from blacklisted jurisdictions are NOT deductible "
                        "(CIRS art. 43(5)). Source: AT folheto 2026-01-12."
                    ),
                ),
            ],
        ),
        # Section 3: Capital Gains Calculation
        (
            "Capital Gains Calculation",
            [
                (
                    "Cost Basis Method: FIFO",
                    "DP-003, PT-C-008",
                    (
                        "First-In-First-Out is MANDATORY for crypto disposals (CIRS art. 43(6)(g)). "
                        "Assets acquired earliest are considered disposed of first. "
                        "Koinly setting: 'Cost basis method' = FIFO. Source: AT folheto 2026-01-12."
                    ),
                ),
                (
                    "Per-Wallet FIFO",
                    "DP-004, PT-C-009",
                    (
                        "FIFO is applied per wallet/exchange independently, NOT globally across all wallets. "
                        "When the same asset is held on multiple platforms, FIFO is applied to each separately "
                        "(CIRS art. 43(7), renumbered from n.7 by Lei n.º 31/2024; "
                        "AT folheto cites old numbering). "
                        "Koinly setting: 'Wallet based cost-tracking' = ON. "
                        "Source: AT folheto 2026-01-12, page 6; CIRS art. 43(9) consolidated."
                    ),
                ),
                (
                    "Capital Gain Formula",
                    "PT-C-007",
                    (
                        "Capital gain = valor de realização − valor de aquisição − despesas necessárias. "
                        "Expenses must be actually incurred and directly related to acquisition or disposal "
                        "(exchange fees, gas fees). Source: AT folheto 2026-01-12."
                    ),
                ),
                (
                    "Fee Deductibility",
                    "DP-007",
                    (
                        "Transfer fees are deductible from gains under CIRS art. 10(4)(a) "
                        "('despesas necessárias e efetivamente praticadas, inerentes à aquisição e alienação'). "
                        "With DP-006 = OFF (Koinly setting 'Realize gains on transfer fees?' disabled), "
                        "Koinly achieves this automatically by folding the fee value into the cost basis "
                        "of the transferred asset. The 'Treat transfer fees as deductible costs?' option "
                        "is greyed out in the Koinly UI when the realize-gains setting is OFF: "
                        "it is not needed because deductibility is handled implicitly via cost-basis adjustment. "
                        "Source: CIRS art. 10(4)(a)."
                    ),
                ),
                (
                    "Aggregation Approach",
                    "PT-C-027",
                    (
                        "FIFO lots from the same sale event are aggregated into one line. "
                        "Quadro 9.4 (Anexo J) has one 'Data de aquisição' field per line. "
                        "Multiple FIFO lots matched to the same sale event (same disposal date, asset, platform) "
                        "are reported as one aggregated line per holding period. "
                        "Multi-lot sales show full acquisition date detail in Notes "
                        "('Acquired: date1, date2 (N lots)') with blue fill. "
                        "Aggregation uses day-level dates matching Anexo J's Ano/Mês/Dia precision. "
                        "Source: PT-C-025 (CryptoBooks secondary guidance); PT-C-020 (official form structure)."
                    ),
                ),
            ],
        ),
        # Section 4: Losses
        (
            "Losses",
            [
                (
                    "Losses Carry-Forward",
                    "PT-C-016",
                    (
                        "Capital losses may be carried forward for 5 years and offset against future gains "
                        "from the same category, but ONLY if the taxpayer opts for englobamento "
                        "(CIRS art. 55(1)(d)). Source: AT folheto 2026-01-12."
                    ),
                ),
                (
                    "Blacklisted Jurisdictions",
                    "PT-C-017",
                    (
                        "Losses from transactions where the counterparty is in a blacklisted jurisdiction "
                        "(regime fiscal claramente mais favorável) are NOT deductible (CIRS art. 43(5)). "
                        "Source: AT folheto 2026-01-12."
                    ),
                ),
                (
                    "Futures/Derivatives Losses",
                    "DP-010, DP-012, PT-C-031, PT-C-034",
                    (
                        "Futures and derivatives are 'instrumentos financeiros derivados' under "
                        "CIRS art. 10(1)(e). Liquidations are taxable disposals (alienação onerosa), "
                        "NOT withdrawals. There is no 365-day exemption for derivatives: CIRS art. 10(19) "
                        "excludes gains/losses only for 'operações previstas na alínea k)' (spot "
                        "criptoativos); derivatives fall under alínea e), so both gains and losses are "
                        "always recognized, regardless of holding period. Losses can be carried forward "
                        "5 years (PT-C-016; CIRS art. 55(1)(d)). Routing is by counterparty residency "
                        "per AT binding ruling Processo 28298/2025: resident counterparty -> Anexo G "
                        "Quadro 13, operation code G51; non-resident counterparty (the usual case for "
                        "foreign exchanges) -> Anexo J Quadro 9.2.B, operation code G30. Report the loss "
                        "with a negative gain/loss amount under the matching counterparty route. "
                        "Source: CIRS art. 10(1)(e), art. 10(19), art. 55(1)(d); "
                        "AT PIV 28298/2025 (docs/maintenance/tax/laws/pt/crypto-tax/official/at_piv_28298_2025.pdf); "
                        "AT portal rendering cirs_art10_portal_2026-04-01.html."
                    ),
                ),
                (
                    "OGR Usage for Derivatives",
                    "DP-011, PT-C-033",
                    (
                        "For futures/derivatives reporting, prefer Koinly Other Gains Report (OGR) values "
                        "over Capital Gains Report values. OGR provides explicit Type classification "
                        "('Loss' or 'Profit') and better collateral flow handling. "
                        "When OGR contains a futures/derivatives entry, use OGR as the authoritative source. "
                        "Source: CIRS art. 10(1)(e); AT folheto 2026-01-12."
                    ),
                ),
            ],
        ),
        # Section 5: Tax Rates
        (
            "Tax Rates",
            [
                (
                    "28% Flat Rate and Englobamento",
                    "PT-C-014",
                    (
                        "Taxable crypto capital gains are subject to a 28% flat rate (taxa autónoma). "
                        "Tax residents may opt to englobar (add to total income for progressive rates), "
                        "which may be beneficial at lower income levels (CIRS art. 72(1)(c), (13)). "
                        "Source: AT folheto 2026-01-12."
                    ),
                ),
                (
                    "Mandatory Englobamento",
                    "PT-C-015",
                    (
                        "Englobamento is MANDATORY when: (1) net short-term balance "
                        "(gains − losses on assets held <365 days) is positive AND "
                        "(2) taxpayer's taxable income reaches the top bracket of CIRS art. 68(1) "
                        "(CIRS art. 72(14)). Source: Ofício Circulado 20269/2024, section 8.6."
                    ),
                ),
            ],
        ),
        # Section 6: Other Gains
        (
            "Other Gains",
            [
                (
                    "Other Gains Classification",
                    "DP-008, PT-C-005",
                    (
                        "Crypto rewards, staking income, mining income, and airdrops are NOT capital gains. "
                        "They are Category E income under CIRS art. 5(11). Taxed separately under different rules; "
                        "some are taxed on receipt, some on disposal. "
                        "Koinly setting: 'Treat other gains as capital gains' = OFF. "
                        "Source: CIRS art. 5(11); AT PIV 22065 (2023-11-06)."
                    ),
                ),
            ],
        ),
        # Section 7: Derivatives & Crypto Gains Classification
        (
            "Derivatives & Crypto Gains Classification",
            [
                (
                    "Derivatives P&L Tab Legal Basis",
                    "DP-012",
                    (
                        "The Derivatives P&L tab reports realized P&L, funding fees, and futures fees "
                        "from futures and perpetuals as Category G income under CIRS art. 10(1)(e) "
                        "(instrumentos financeiros derivados). The tab is rendered only when the "
                        "separate_derivatives_reporting flag is set; entries are aggregated by "
                        "(date, asset, platform, event_type). The Crypto Gains tab continues to report "
                        "cryptoasset capital gains under CIRS art. 10(1)(k), including the 365-day "
                        "holding-period exemption. Losses on derivatives are deductible against other "
                        "Category G gains and carry forward 5 years under PT-C-016. "
                        "Source: CIRS art. 10(1)(e), art. 10(1)(k); AT folheto 2026-01-12."
                    ),
                ),
            ],
        ),
        # Section 8: Implementation
        (
            "Implementation",
            [
                (
                    "Materiality Threshold",
                    "PT-C-028, PT-C-029",
                    (
                        "Lines where |gain/loss| < 1 EUR are excluded from the report. "
                        "No de minimis threshold exists in Portuguese law (PT-C-024), but sub-1-EUR lines "
                        "have no material tax impact and create impractical manual filing burden. "
                        "The 1 EUR filter applies symmetrically to gains and losses. "
                        "Source: Implementation decision; absence of threshold confirmed in "
                        "AT folheto 2026-01-12, OC 20269/2024, OC 20278/2025."
                    ),
                ),
                (
                    "OGR-vs-CG and Fee Lot Dedup",
                    "DP-015, PT-C-034, PT-C-036",
                    (
                        "Capital-gains (CG) FIFO lots matched against OGR/derivatives disposal events "
                        "(PT-C-034) and against standalone transaction/network fee events (DP-015 / "
                        "PT-C-036) are removed from the Crypto Gains tab to avoid double-counting the "
                        "same disposal. OGR-routed derivatives gains are reported once on the "
                        "Derivatives P&L tab; Koinly-realized fee-token disposals are non-taxable "
                        "utility costs under CIRS art. 10(1)(k) and are filtered out. The removed lots "
                        "are surfaced as Crypto Supplementary review rows (per-row DEBUG audit trail "
                        "preserved in the file log). "
                        "Source: Implementation decision (PT-C-034 OGR routing; DP-015/PT-C-036 fee "
                        "filtering)."
                    ),
                ),
                (
                    "Data Sources",
                    "",
                    (
                        "Interactive Brokers (dividends, capital gains), "
                        "Koinly (crypto capital gains, rewards, loans), config.ini (exchange rates). "
                        "IB CSV format and Koinly export settings documented in README.md. "
                        "Source: Project documentation."
                    ),
                ),
                (
                    "Cashback Treatment",
                    "DP-009",
                    (
                        "Cashbacks and fee refunds are recorded at fair market value at receipt as cost basis "
                        "for future FIFO, NOT as zero-cost deposits. "
                        "Zero-cost would overstate gains on later disposal. "
                        "Conservative approach: records true acquisition cost. "
                        "Koinly setting: 'Treat cashbacks and fee refunds as zero-cost deposits?' = OFF. "
                        "Source: Implementation decision (conservative)."
                    ),
                ),
                (
                    "Review Flag Reasons",
                    "PT-C-030",
                    (
                        "Review flags carry specific, actionable reasons via the review_reason field. "
                        "When review_required=True, the Excel column shows 'YES: <reason>' "
                        "instead of a bare boolean. This enables efficient manual review without "
                        "re-examining source data. Source: Implementation decision (UX optimization)."
                    ),
                ),
            ],
        ),
    ]

    # PT-C-028 run-count suffix: append the per-run materiality-filter counts to
    # the "Materiality Threshold" item when ``decision_counts`` is supplied
    # (crypto runs). IB-only runs pass None and the static text renders as-is.
    if decision_counts is not None:
        materiality_suffix = (
            f" [This run: filtered {decision_counts.sub_1_eur_filtered} entries, "
            f"{decision_counts.sub_1_eur_retained} retained.]"
        )
        for section_idx, (_section_title, items) in enumerate(methodology_items):
            for item_idx, (label, _rule_ids, description) in enumerate(items):
                if label == "Materiality Threshold":
                    methodology_items[section_idx][1][item_idx] = (
                        label,
                        _rule_ids,
                        description + materiality_suffix,
                    )

    # Dedup run-count suffix: append the per-run OGR-vs-CG / fee lot dedup
    # removal counts to the "OGR-vs-CG and Fee Lot Dedup" item when
    # ``decision_counts`` is supplied (crypto runs). IB-only runs pass None and
    # the static text renders as-is (backward compat).
    if decision_counts is not None:
        dedup_suffix = (
            f" [This run: derivatives removed {decision_counts.derivatives_dedup_removed}; "
            f"fee removed {decision_counts.fee_dedup_removed}.]"
        )
        for section_idx, (_section_title, items) in enumerate(methodology_items):
            for item_idx, (label, _rule_ids, description) in enumerate(items):
                if label == "OGR-vs-CG and Fee Lot Dedup":
                    methodology_items[section_idx][1][item_idx] = (
                        label,
                        _rule_ids,
                        description + dedup_suffix,
                    )

    # Render grouped methodology sections
    for section_title, items in methodology_items:
        # Blank row before each section (except first)
        worksheet.cell(row_no, 1, "")
        row_no += 1

        # Section header in bold
        worksheet.cell(row_no, 1, section_title)
        worksheet.cell(row_no, 1).font = openpyxl.styles.Font(bold=True, size=11)
        row_no += 1

        # Section items
        for label, rule_ids, description in items:
            worksheet.cell(row_no, 1, label)
            worksheet.cell(row_no, 1).font = openpyxl.styles.Font(bold=True)
            # Include rule_ids in the description for testability
            full_description = f"{description} Rule IDs: {rule_ids}." if rule_ids else description
            worksheet.cell(row_no, 2, full_description)
            row_no += 1

    auto_column_width(worksheet)
