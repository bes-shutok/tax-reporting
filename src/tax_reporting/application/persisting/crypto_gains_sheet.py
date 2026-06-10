"""Crypto capital gains sheet writer for the Excel tax report."""

from __future__ import annotations

from decimal import Decimal
from typing import TYPE_CHECKING

import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet

if TYPE_CHECKING:
    from ..crypto_reporting import CryptoCapitalGainEntry, CryptoTaxReport

from .excel_utils import apply_multi_date_row_fill, apply_review_row_fill, auto_column_width, safe_cell_value

_CAPITAL_GAINS_NUM_COLS = 20

# Fill colors for OGR validation
_RED_FILL = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")
_YELLOW_FILL = PatternFill(start_color="FFFFFF00", end_color="FFFFFF00", fill_type="solid")


def _render_capital_gain_row(
    worksheet: Worksheet,
    row_no: int,
    entry: CryptoCapitalGainEntry,
    threshold: Decimal,
) -> None:
    """Write a single capital gain entry row and apply conditional formatting."""
    worksheet.cell(row_no, 1, entry.disposal_date)
    worksheet.cell(row_no, 2, entry.acquisition_date)
    worksheet.cell(row_no, 3, entry.asset)
    worksheet.cell(row_no, 4, entry.amount)
    worksheet.cell(row_no, 5, entry.cost_eur)
    worksheet.cell(row_no, 6, entry.proceeds_eur)
    worksheet.cell(row_no, 7, entry.gain_loss_eur)
    worksheet.cell(row_no, 8, entry.holding_period)
    worksheet.cell(row_no, 9, entry.wallet)
    worksheet.cell(row_no, 10, entry.platform)
    worksheet.cell(row_no, 11, entry.chain)
    worksheet.cell(row_no, 12, entry.operator_origin.operator_entity)
    worksheet.cell(row_no, 13, entry.operator_origin.operator_country)
    worksheet.cell(row_no, 14, entry.annex_hint)
    review_display = f"YES: {entry.review_reason or '(no reason propagated)'}" if entry.review_required else "NO"
    worksheet.cell(row_no, 15, review_display)
    worksheet.cell(row_no, 16, safe_cell_value(entry.notes))
    worksheet.cell(row_no, 17, entry.token_swap_history or "")

    # OGR validation columns (18, 19, 20)
    ogr_val = entry.ogr_validation
    if ogr_val:
        worksheet.cell(row_no, 18, ogr_val.ogr_gain_loss)
        worksheet.cell(row_no, 19, ogr_val.magnitude_diff_percent)
        ogr_review_display = (
            f"YES: {ogr_val.review_reason}" if ogr_val.review_required else "NO"
        )
        worksheet.cell(row_no, 20, ogr_review_display)
    else:
        # Leave blank when no OGR validation
        worksheet.cell(row_no, 18, None)
        worksheet.cell(row_no, 19, None)
        worksheet.cell(row_no, 20, None)

    # Apply conditional formatting
    _apply_conditional_formatting(worksheet, row_no, entry, threshold)


def _apply_conditional_formatting(
    worksheet: Worksheet,
    row_no: int,
    entry: CryptoCapitalGainEntry,
    threshold: Decimal,
) -> None:
    """Apply conditional formatting to a capital gain row based on OGR validation and other rules.

    Priority order:
    1. RED fill for OGR direction override (highest priority)
    2. YELLOW fill for OGR magnitude differences
    3. RED fill for entry review required or zero-cost above threshold
    4. BLUE fill for multi-acquisition dates
    5. No fill (default)
    """
    # Check OGR validation conditions first (highest priority)
    if entry.ogr_validation and entry.ogr_validation.review_required:
        if entry.ogr_validation.review_reason and "OGR direction override" in entry.ogr_validation.review_reason:
            # RED fill for direction conflict
            for col in range(1, _CAPITAL_GAINS_NUM_COLS + 1):
                worksheet.cell(row_no, col).fill = _RED_FILL
            return
        else:
            # YELLOW fill for magnitude differences
            for col in range(1, _CAPITAL_GAINS_NUM_COLS + 1):
                worksheet.cell(row_no, col).fill = _YELLOW_FILL
            return

    # Existing rules for entry-level review and zero-cost
    needs_fill = entry.review_required or (entry.cost_eur == 0 and abs(entry.gain_loss_eur) >= threshold)
    if needs_fill:
        apply_review_row_fill(worksheet, row_no, 1, _CAPITAL_GAINS_NUM_COLS)
    elif entry.multi_acquisition_dates:
        apply_multi_date_row_fill(worksheet, row_no, 1, _CAPITAL_GAINS_NUM_COLS)


def write_crypto_gains_sheet(workbook: openpyxl.Workbook, crypto_tax_report: CryptoTaxReport) -> None:  # noqa: PLR0915
    """Create and populate a 'Crypto Gains' worksheet with capital gains entries and statistics.

    Writes:
    - Title and tax year metadata
    - PDF summary (if present)
    - Section 1: Capital gain entries in 17 columns
    - Section 1b: Capital gain statistics by holding period

    Args:
        workbook: The Excel workbook to add the sheet to.
        crypto_tax_report: The crypto tax report data.
    """
    worksheet = workbook.create_sheet("Crypto Gains")
    worksheet.cell(1, 1, "CRYPTO TAX REPORT - PORTUGAL").font = Font(bold=True)
    worksheet.cell(2, 1, "Tax year")
    worksheet.cell(2, 2, crypto_tax_report.tax_year)
    if crypto_tax_report.pdf_summary:
        worksheet.cell(3, 1, "PDF period")
        worksheet.cell(3, 2, crypto_tax_report.pdf_summary.period or "N/A")
        worksheet.cell(3, 3, "PDF timezone")
        worksheet.cell(3, 4, crypto_tax_report.pdf_summary.timezone or "N/A")
        worksheet.cell(3, 5, "PDF extracted tokens")
        worksheet.cell(3, 6, crypto_tax_report.pdf_summary.extracted_tokens)

    row_no = 5 if crypto_tax_report.pdf_summary else 4
    worksheet.cell(row_no, 1, "1. CAPITAL GAINS").font = Font(bold=True)
    row_no += 1

    capital_headers = [
        "Disposal date",
        "Acquisition date",
        "Asset",
        "Quantity",
        "Acquisition Cost (EUR)",
        "Disposal Proceeds (EUR)",
        "Gain/Loss (EUR)",
        "Holding period",
        "Wallet",
        "Platform",
        "Disposal chain",
        "Operator entity",
        "Operator country",
        "Annex hint",
        "Review flag",
        "Notes",
        "Token origin",
        "OGR Gain/Loss (EUR)",
        "OGR Diff (%)",
        "OGR Review",
    ]
    for idx, header in enumerate(capital_headers, start=1):
        worksheet.cell(row_no, idx, header)
    row_no += 1

    threshold = crypto_tax_report.zero_basis_review_threshold
    for entry in crypto_tax_report.capital_entries:
        _render_capital_gain_row(worksheet, row_no, entry, threshold)
        row_no += 1

    # 1b. CAPITAL GAINS STATISTICS
    row_no += 1
    worksheet.cell(row_no, 1, "1b. CAPITAL GAINS STATISTICS").font = Font(bold=True)
    row_no += 1

    stats_headers = [
        "Holding Period",
        "Count",
        "Acquisition Cost Total (EUR)",
        "Disposal Proceeds Total (EUR)",
        "Gain/Loss Total (EUR)",
    ]
    for idx, header in enumerate(stats_headers, start=1):
        header_cell = worksheet.cell(row_no, idx, header)
        header_cell.font = Font(bold=True)
    row_no += 1

    stats = crypto_tax_report.capital_gain_stats
    period_rows = [
        ("Short-term", stats.short_term),
        ("Long-term", stats.long_term),
        ("Mixed", stats.mixed),
        ("Unknown", stats.unknown),
        ("Grand Total", stats.grand_total),
    ]
    for label, period_stats in period_rows:
        worksheet.cell(row_no, 1, label)
        worksheet.cell(row_no, 2, period_stats.count)
        worksheet.cell(row_no, 3, period_stats.cost_total_eur)
        worksheet.cell(row_no, 4, period_stats.proceeds_total_eur)
        worksheet.cell(row_no, 5, period_stats.gain_loss_total_eur)
        row_no += 1

    auto_column_width(worksheet)
