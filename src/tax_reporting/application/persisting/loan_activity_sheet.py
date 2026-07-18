"""Loan Activity sheet writer for the Excel tax report."""

from __future__ import annotations

from typing import TYPE_CHECKING

import openpyxl
from openpyxl.styles import PatternFill

if TYPE_CHECKING:
    from ..crypto_reporting import CryptoTaxReport

from ...domain.constants import LOAN_STATUS_NO_EUR_PRICE, LOAN_STATUS_OVERPAID_VERIFY
from .excel_utils import auto_column_width, safe_cell_value

# Module-level fills (Invariant 6): the fill-color mapping matches entry.balance_status
# against sentinel constants. _LIGHT_RED_FILL mirrors the prior function-scoped red_fill;
# _NO_EUR_PRICE_FILL mirrors _YELLOW_FILL in crypto_gains_sheet.py. The two live side by
# side so the symmetry between the two highlight colors is visible.
_LIGHT_RED_FILL = PatternFill(start_color="FFFFCCCC", end_color="FFFFCCCC", fill_type="solid")
_NO_EUR_PRICE_FILL = PatternFill(start_color="FFFFFF00", end_color="FFFFFF00", fill_type="solid")


def write_loan_activity_sheet(workbook: openpyxl.Workbook, crypto_tax_report: CryptoTaxReport) -> None:
    """Create and populate a 'Loan Activity' worksheet with per-asset loan summaries.

    Shows all loan receipts and repayments found in the Koinly transaction history
    regardless of filing year; the tab reflects all-history balances so the reviewer
    can verify outstanding and over-repaid positions across years.

    Each row carries one of five loan-status sentinels (see ``domain/constants.py``
    and the classifier in ``loan_activity.py``). Fill mapping (by exact sentinel
    equality, Invariant 6): ``LOAN_STATUS_NO_EUR_PRICE`` -> yellow;
    ``LOAN_STATUS_OVERPAID_VERIFY`` -> light-red; the other three -> no fill. The
    varying overshoot percentage or repayment-only note is rendered in the sibling
    "Balance Detail" column (column 10), never interpolated into ``balance_status``.

    Args:
        workbook: The Excel workbook to add the sheet to.
        crypto_tax_report: The full crypto tax report; only its loan_activity field
            is used. Each LoanActivityEntry supplies per-asset received/repaid counts,
            amounts, EUR values, a balance status sentinel, and an optional
            balance_detail string.
    """
    _header_row = 3
    _data_start_row = 4

    worksheet = workbook.create_sheet("Loan Activity")

    # Row 1: visible note so the reviewer understands balances are all-history and how
    # to read the five status sentinels. This is intentional: cross-year loan repayments
    # must be visible so the reviewer can verify the outstanding balance. The five
    # sentinels and their fills: Settled / Open loan / Likely in-asset interest (neutral),
    # Cannot classify: no EUR price data (yellow), Overpaid (light-red). The Overpaid
    # sentinel's full value written to column 9 is "Overpaid (cross-year loan? verify)";
    # "Overpaid (verify)" below is the legend abbreviation for that sentinel.
    note_cell = worksheet.cell(
        1,
        1,
        "Note: balances shown across ALL years in Transaction History, not filtered to the filing year. "
        "Statuses: Settled / Open loan / Likely in-asset interest (neutral fill), "
        "Cannot classify: no EUR price data (yellow), "
        'Overpaid (verify) = "Overpaid (cross-year loan? verify)" (light-red fill).',
    )
    note_cell.font = openpyxl.styles.Font(italic=True)

    headers = [
        "Asset",
        "Received Count",
        "Received Amount",
        "Received Value (EUR)",
        "Repaid Count",
        "Repaid Amount",
        "Repaid Value (EUR)",
        "Balance Amount",
        "Balance Status",
        "Balance Detail",
    ]
    for idx, header in enumerate(headers, start=1):
        worksheet.cell(_header_row, idx, header)

    for row_offset, entry in enumerate(crypto_tax_report.loan_activity, start=_data_start_row):
        worksheet.cell(row_offset, 1, safe_cell_value(entry.asset))
        worksheet.cell(row_offset, 2, entry.received_count)
        worksheet.cell(row_offset, 3, entry.received_amount)
        worksheet.cell(row_offset, 4, entry.received_value_eur)
        worksheet.cell(row_offset, 5, entry.repaid_count)
        worksheet.cell(row_offset, 6, entry.repaid_amount)
        worksheet.cell(row_offset, 7, entry.repaid_value_eur)
        worksheet.cell(row_offset, 8, entry.balance_amount)
        worksheet.cell(row_offset, 9, entry.balance_status)
        # Column 10: sibling "Balance Detail" (overshoot % or repayment-only note).
        # Written only when non-None via safe_cell_value (matches module convention and
        # closes the CSV-injection vector if a future classifier edit interpolates a
        # CSV-derived field).
        if entry.balance_detail is not None:
            worksheet.cell(row_offset, 10, safe_cell_value(entry.balance_detail))

        # Fill mapping matches the sentinel constant by exact equality (Invariant 6),
        # NOT the rendered string. NO_EUR_PRICE -> yellow; OVERPAID_VERIFY (large
        # overshoot OR repayment-only asset) -> light-red; all others -> no fill.
        if entry.balance_status == LOAN_STATUS_NO_EUR_PRICE:
            fill: PatternFill | None = _NO_EUR_PRICE_FILL
        elif entry.balance_status == LOAN_STATUS_OVERPAID_VERIFY:
            fill = _LIGHT_RED_FILL
        else:
            fill = None
        if fill is not None:
            for col_idx in range(1, 11):
                worksheet.cell(row_offset, col_idx).fill = fill

    section_start = _data_start_row + len(crypto_tax_report.loan_activity) + 2
    scope_label_cell = worksheet.cell(section_start, 1, "FIFO Rebuild Scope")
    scope_label_cell.font = openpyxl.styles.Font(bold=True)
    note_text = (
        "Assets rebuilt from Transaction History instead of Koinly CG export "
        "(loan-affected under CIRS art. 10(20))"
    )
    worksheet.cell(section_start, 2, note_text).font = openpyxl.styles.Font(italic=True)
    if crypto_tax_report.fifo_rebuild_assets:
        for row_offset, asset in enumerate(sorted(crypto_tax_report.fifo_rebuild_assets), start=section_start + 1):
            worksheet.cell(row_offset, 1, asset)
            worksheet.cell(row_offset, 2, "Rebuilt from Transaction History")
    else:
        worksheet.cell(section_start + 1, 1, "None").font = openpyxl.styles.Font(italic=True)
        worksheet.cell(section_start + 1, 2, "FIFO rebuild not active or no loan-affected assets discovered")

    auto_column_width(worksheet)
