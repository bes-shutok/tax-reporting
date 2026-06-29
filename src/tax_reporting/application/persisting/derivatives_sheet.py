"""Derivatives P&L Excel sheet for financial derivatives (CIRS art. 10(1)(e))."""

from __future__ import annotations

import logging
from decimal import Decimal
from typing import TYPE_CHECKING

import openpyxl
from openpyxl.styles import Font, PatternFill

if TYPE_CHECKING:
    from ...infrastructure.config import TaxJurisdictionConfig
    from ..crypto_reporting import CryptoTaxReport

from .excel_utils import auto_column_width, safe_cell_value

_SHEET_NAME = "Derivatives P&L"
_HEADER_TITLE = "DERIVATIVES P&L (Financial Derivatives: CIRS art. 10(1)(e))"
_COLUMN_HEADERS = [
    "Date",
    "Asset",
    "Platform",
    "Event Type",
    "P&L (EUR)",
    "Operator entity",
    "Operator country",
    "Event count",
    "Notes",
    "Review",
    "Annex",
    "Código",
]
_EMPTY_STATE_MESSAGE = "No derivatives activity for this jurisdiction"
_LOSS_FOOTNOTE = "Losses are deductible against other Category G gains; carry-forward 5 years per PT-C-016"
_TOTAL_LABEL = "Total"
_NUM_COLUMNS = len(_COLUMN_HEADERS)

# Red fill for review rows (matches crypto_gains_sheet convention)
_REVIEW_FILL = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")


def write_derivatives_sheet(
    workbook: openpyxl.Workbook,
    crypto_tax_report: CryptoTaxReport,
    jurisdiction: TaxJurisdictionConfig,
) -> None:
    """Create and populate the 'Derivatives P&L' worksheet.

    Renders a single-tab P&L summary for derivatives realizations taxed under
    CIRS art. 10(1)(e). Each aggregated entry maps to one row; the totals row
    shows the net P&L; a loss-deductibility footnote appears whenever at least
    one entry is a loss.

    The annex routing (column 11) and AT operation code (column 12) are written
    per row from each entry's ``annex_hint`` and ``operation_code`` (blank when
    residency routing is disabled). When residency routing is enabled and any
    entry has a blank annex_hint, a warning is emitted so a failed/blank route is
    surfaced loudly rather than rendered silently (development_lessons.md #77/#118).

    The sheet is always rendered, even when derivatives_entries is empty (per
    development_lessons.md #93): the headers and an explicit empty-state row are
    written so the reviewer can see the category was considered.

    Args:
        workbook: The Excel workbook to add the sheet to.
        crypto_tax_report: The crypto tax report; only derivatives_entries is used.
        jurisdiction: The active tax jurisdiction; gates the blank-annex warning
            via ``route_derivatives_by_counterparty_residency``.
    """
    worksheet = workbook.create_sheet(_SHEET_NAME)

    title_cell = worksheet.cell(1, 1, _HEADER_TITLE)
    title_cell.font = Font(bold=True)

    entries = crypto_tax_report.derivatives_entries

    # Defensive guard: when residency routing is enabled, _derivatives_route always
    # resolves a non-blank annex for the single DerivativesPnLEntry construction
    # site, so this is not an observed production condition today. It catches a
    # future second construction site that forgets to route. Describe the
    # observation rather than diagnosing a cause the renderer cannot establish.
    if entries and jurisdiction.route_derivatives_by_counterparty_residency:
        blank_annex_rows = [
            e for e in entries if e.annex_hint == ""
        ]
        if blank_annex_rows:
            logging.getLogger(__name__).warning(
                "Derivatives P&L: %d of %d entries have a blank Annex (annex_hint) under "
                "the active jurisdiction. Affected rows: %s",
                len(blank_annex_rows),
                len(entries),
                sorted({(e.date, e.asset, e.platform) for e in blank_annex_rows}),
            )

    header_row = 3
    for idx, header in enumerate(_COLUMN_HEADERS, start=1):
        header_cell = worksheet.cell(header_row, idx, header)
        header_cell.font = Font(bold=True)

    current_row = header_row + 1

    if not entries:
        empty_cell = worksheet.cell(current_row, 1, _EMPTY_STATE_MESSAGE)
        empty_cell.font = Font(italic=True)
        auto_column_width(worksheet)
        return

    for entry in entries:
        worksheet.cell(current_row, 1, safe_cell_value(entry.date))
        worksheet.cell(current_row, 2, safe_cell_value(entry.asset))
        worksheet.cell(current_row, 3, safe_cell_value(entry.platform))
        worksheet.cell(current_row, 4, entry.event_type.value)
        worksheet.cell(current_row, 5, entry.pnl_eur)
        worksheet.cell(current_row, 6, safe_cell_value(entry.operator_entity))
        worksheet.cell(current_row, 7, safe_cell_value(entry.operator_country))
        worksheet.cell(current_row, 8, entry.event_count)
        worksheet.cell(current_row, 9, safe_cell_value(entry.notes))
        review_display = (
            f"YES: {entry.review_reason}" if entry.review_required and entry.review_reason else
            "YES: (no reason propagated)" if entry.review_required else "NO"
        )
        worksheet.cell(current_row, 10, review_display)
        worksheet.cell(current_row, 11, safe_cell_value(entry.annex_hint))
        worksheet.cell(current_row, 12, safe_cell_value(entry.operation_code))

        if entry.review_required:
            for col_idx in range(1, _NUM_COLUMNS + 1):
                worksheet.cell(current_row, col_idx).fill = _REVIEW_FILL

        current_row += 1

    current_row += 1
    total_label_cell = worksheet.cell(current_row, 1, _TOTAL_LABEL)
    total_label_cell.font = Font(bold=True)
    net_pnl = _sum_pnl_eur(entries)
    total_cell = worksheet.cell(current_row, 5, net_pnl)
    total_cell.font = Font(bold=True)

    if any(entry.pnl_eur < 0 for entry in entries):
        current_row += 1
        footnote_cell = worksheet.cell(current_row, 1, _LOSS_FOOTNOTE)
        footnote_cell.font = Font(italic=True)

    auto_column_width(worksheet)


def _sum_pnl_eur(entries: list) -> Decimal:  # type: ignore[type-arg]
    """Sum the pnl_eur field across entries.

    Args:
        entries: Iterable of DerivativesPnLEntry objects.

    Returns:
        Decimal total of pnl_eur across all entries.
    """
    return sum((entry.pnl_eur for entry in entries), start=Decimal("0"))
