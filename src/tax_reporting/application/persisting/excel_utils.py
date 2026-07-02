"""Shared Excel utilities: auto-width calculation, review row highlighting, and safe file removal."""

from __future__ import annotations

from pathlib import Path
from typing import TYPE_CHECKING

if TYPE_CHECKING:
    from os import PathLike

from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from ...infrastructure.logging_config import create_module_logger
from ...infrastructure.text_sanitize import strip_control_chars

# Column width bounds for auto_column_width
MAX_CELL_WIDTH = 50  # Maximum character width to measure per cell (caps outliers)
MIN_DATA_WIDTH = 12  # Minimum width for formula-only or empty columns
HEADER_THRESHOLD = 10  # Max length for a cell to be considered a header/label vs actual data

# Shared fill for rows that require manual review (review_required=True or placeholder buys)
REVIEW_ROW_FILL = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")

# Fill for rows where aggregated capital gain consumed lots from multiple acquisition dates
MULTI_DATE_ROW_FILL = PatternFill(start_color="FFCCFFFF", end_color="FFCCFFFF", fill_type="solid")


def auto_column_width(worksheet: Worksheet) -> None:
    """Set column widths to fit the widest non-formula cell value plus padding.

    Iterates every column in the worksheet. For each column, measures the
    character length of all non-formula, non-None cell values and sets the
    column width to the maximum found plus a 2-character padding. Columns
    that contain only formulas or are entirely empty receive MIN_DATA_WIDTH.
    Individual cell contributions are capped at MAX_CELL_WIDTH to prevent
    outliers (e.g., long explanatory notes) from blowing out column widths.

    Formula-heavy columns (where formulas outnumber non-formula cells)
    receive MIN_DATA_WIDTH to ensure rendered values have adequate space.

    Args:
        worksheet: The openpyxl worksheet to resize.
    """
    logger = create_module_logger(__name__)
    logger.debug("Auto-adjusting column widths")
    for column_cells in worksheet.columns:
        # Measure each non-formula cell, capped at MAX_CELL_WIDTH
        lengths = [
            min(len(str(cell.value)), MAX_CELL_WIDTH)
            for cell in column_cells
            if cell.value is not None and cell.data_type != "f"
        ]
        # Count formula cells to detect formula-heavy columns
        formula_count = sum(
            1 for cell in column_cells if cell.value is not None and cell.data_type == "f"
        )

        # Determine if this is a formula-heavy column.
        # A column is formula-heavy if formulas are present AND the column appears to
        # contain only headers/labels (all measured cells shorter than HEADER_THRESHOLD)
        # rather than actual data content. This catches typical headers like "Amount" (6)
        # while allowing data like "ABCDEFGHIJK" (11+) to be measured normally.
        is_formula_heavy = (
            formula_count > 0
            and lengths
            and all(cell_length < HEADER_THRESHOLD for cell_length in lengths)
        )

        # Calculate width with MIN_DATA_WIDTH floor for formula-heavy or empty columns
        # Formula-heavy columns use measured width (including headers) with MIN_DATA_WIDTH as floor
        if not lengths:
            width = MIN_DATA_WIDTH
        elif is_formula_heavy:
            width = max((*lengths, MIN_DATA_WIDTH))
        else:
            width = max((*lengths, 2)) + 2

        first_cell = column_cells[0]
        column_idx = None
        try:
            column_idx = first_cell.column
            if column_idx is not None:
                column_letter = get_column_letter(column_idx)
                worksheet.column_dimensions[column_letter].width = width
        except (AttributeError, TypeError) as e:
            logger.warning("Failed to set column width for column %s: %s", column_idx, e)


def apply_review_row_fill(worksheet: Worksheet, row_no: int, start_col: int, end_col: int) -> None:
    """Apply the shared review-row red fill to a range of cells in a row.

    Use this wherever a row is flagged for manual review (review_required=True,
    placeholder buys, missing data, etc.) to keep the visual convention consistent
    across all sheets.

    Args:
        worksheet: The worksheet containing the row.
        row_no: 1-based row index to fill.
        start_col: 1-based index of the first column to fill (inclusive).
        end_col: 1-based index of the last column to fill (inclusive).
    """
    for col_idx in range(start_col, end_col + 1):
        worksheet.cell(row_no, col_idx).fill = REVIEW_ROW_FILL  # type: ignore[assignment]


def apply_multi_date_row_fill(worksheet: Worksheet, row_no: int, start_col: int, end_col: int) -> None:
    """Apply the multi-date acquisition blue fill to a range of cells in a row.

    Use this for rows where the aggregated capital gain consumed lots from multiple
    acquisition dates, to visually distinguish them from single-date rows.
    This presentation enhancement supports PT-C-027 aggregation behavior by making
    multi-lot sales visually distinct.

    Args:
        worksheet: The worksheet containing the row.
        row_no: 1-based row index to fill.
        start_col: 1-based index of the first column to fill (inclusive).
        end_col: 1-based index of the last column to fill (inclusive).
    """
    for col_idx in range(start_col, end_col + 1):
        worksheet.cell(row_no, col_idx).fill = MULTI_DATE_ROW_FILL  # type: ignore[assignment]


def safe_cell_value(value: str) -> str:
    """Prevent Excel formula injection and strip control characters from cell values.

    Neutralizes formula prefixes (=, +, -, @) by prepending a space, and removes
    control/newline characters that would corrupt cell rendering. User-controlled
    strings (platform names, wallet labels, Koinly-derived values) should be wrapped
    with this function before writing to worksheet cells.

    Args:
        value: The raw string to write to a cell.

    Returns:
        The sanitized value safe for Excel cell writing. Empty strings are safe
        and returned as-is (no space prefix needed).
    """
    cleaned = strip_control_chars(value)
    # Empty strings (after stripping control chars) are safe - no prefix needed
    if cleaned[:1] in ("=", "+", "-", "@"):
        return f" {cleaned}"
    return cleaned


def safe_remove_file(path: str | PathLike[str]) -> None:
    """Safely remove a file if it exists, logging any errors.

    Args:
        path: File path to remove.
    """
    logger = create_module_logger(__name__)
    try:
        p = Path(path)
        if p.exists():
            p.unlink()
            logger.debug("Removed existing file: %s", p.name)
    except OSError as e:
        logger.warning("Failed to remove file %s: %s", path, e)
