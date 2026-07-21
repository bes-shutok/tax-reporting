"""Crypto supplementary sheet writer for the Excel tax report.

This sheet contains audit and support data for crypto transactions, including:
- Income code reference
- Taxable-now reward detail (per-row trace data)
- Deferred-by-law reward detail
- Rewards classification reconciliation
- Review required entries (for zero-value popular tokens)

Aggregated taxable-now rewards are reported on the Reporting worksheet under
OTHER CAPITAL INVESTMENT INCOME, not in this sheet. This sheet provides
classification support detail and traceability for auditability only.
"""

from __future__ import annotations

import openpyxl
from openpyxl.styles import Font

from ...infrastructure.config import TaxJurisdictionConfig
from ..crypto_reporting import (
    ZERO,
    CryptoReviewEntry,
    CryptoRewardIncomeEntry,
    CryptoTaxReport,
    RewardTaxClassification,
)
from .excel_utils import apply_review_row_fill, auto_column_width, safe_cell_value
from .tax_constants import _INCOME_CODE_DESCRIPTIONS

_REWARD_ROW_NUM_COLS = 11
_REVIEW_ROW_NUM_COLS = 5

_REWARD_DETAIL_HEADERS = [
    "Date",
    "Asset",
    "Value (EUR)",
    "Income type",
    "Wallet",
    "Platform",
    "Reward chain",
    "Country",
    "Foreign tax (EUR)",
    "Review flag",
    "Description",
]

# Suppressed-rewards / dust-summary table column layout. Mirrors the
# ``_REWARD_DETAIL_HEADERS`` shape (bold header row, one data row per group) so
# the summary blocks render as proper tables instead of clumped single lines.
# Deferred side sums the native-unit ``amount`` (the deferred skip is on crypto
# rewards whose EUR value is zero by definition); taxable-now sums ``value_eur``
# (the taxable-now dust partition keeps EUR-denominated rewards).
_SUPPRESSED_REWARDS_HEADERS = [
    "Asset",
    "Wallet",
    "Rows",
    "Summed amount",
    "Category",
]
_DUST_SUMMARY_HEADERS = [
    "Asset",
    "Wallet",
    "Rows",
    "Summed Value (EUR)",
    "Category",
]


def _write_reward_detail_rows(
    worksheet: openpyxl.worksheet.worksheet.Worksheet,
    row_no: int,
    entries: list[CryptoRewardIncomeEntry],
    empty_label: str,
) -> int:
    """Write a block of reward detail rows and return the next free row number.

    Renders the shared 11-column detail layout used for both taxable-now and
    deferred-by-law reward sections. When ``entries`` is empty, writes
    ``empty_label`` instead.
    """
    for idx, header in enumerate(_REWARD_DETAIL_HEADERS, start=1):
        header_cell = worksheet.cell(row_no, idx, header)
        header_cell.font = Font(bold=True)
    row_no += 1

    if not entries:
        worksheet.cell(row_no, 1, empty_label)
        return row_no + 1

    for entry in entries:
        worksheet.cell(row_no, 1, entry.date)
        worksheet.cell(row_no, 2, safe_cell_value(entry.asset))
        worksheet.cell(row_no, 3, float(entry.value_eur))
        worksheet.cell(row_no, 4, safe_cell_value(entry.source_type))
        worksheet.cell(row_no, 5, safe_cell_value(entry.wallet))
        worksheet.cell(row_no, 6, safe_cell_value(entry.platform))
        worksheet.cell(row_no, 7, safe_cell_value(entry.chain))
        worksheet.cell(row_no, 8, safe_cell_value(entry.operator_origin.operator_country))
        worksheet.cell(row_no, 9, float(entry.foreign_tax_eur))
        review_cell = f"YES: {safe_cell_value(entry.review_reason)}" if entry.review_required else "NO"
        worksheet.cell(row_no, 10, review_cell)
        worksheet.cell(row_no, 11, safe_cell_value(entry.description))
        if entry.review_required:
            apply_review_row_fill(worksheet, row_no, 1, _REWARD_ROW_NUM_COLS)
        row_no += 1

    return row_no


def _write_review_rows(
    worksheet: openpyxl.worksheet.worksheet.Worksheet,
    row_no: int,
    review_entries: list[CryptoReviewEntry],
) -> int:
    """Write the REVIEW REQUIRED detail rows and return the next free row number.

    Renders the 5-column review layout. Suspicious entries get a red bold asset
    cell. When ``review_entries`` is empty, writes "No review items".
    """
    review_headers = ["Source", "Date", "Asset", "Platform", "Review reason"]
    for idx, header in enumerate(review_headers, start=1):
        header_cell = worksheet.cell(row_no, idx, header)
        header_cell.font = Font(bold=True)
    row_no += 1

    if not review_entries:
        worksheet.cell(row_no, 1, "No review items")
        return row_no + 1

    for entry in review_entries:
        source_label = {
            "capital_gains": "Capital Gains",
            "transaction_history": "Transaction History",
        }.get(entry.source_section, "Income")
        worksheet.cell(row_no, 1, source_label)
        worksheet.cell(row_no, 2, entry.date)
        asset_cell = worksheet.cell(row_no, 3, safe_cell_value(entry.asset))
        worksheet.cell(row_no, 4, safe_cell_value(entry.platform))
        worksheet.cell(row_no, 5, entry.review_reason)
        if entry.is_suspicious:
            asset_cell.font = Font(color="FF0000", bold=True)
        row_no += 1

    return row_no


def _priced_assets_in_export(reward_entries: list[CryptoRewardIncomeEntry]) -> frozenset[str]:
    """Return the set of assets with at least one priced row in the export.

    The shared priced-asset discriminator (Invariant 3, AGENTS.md rule 30):
    "dust" vs "unpriced" splits on whether the asset has any ``value_eur > 0``
    row anywhere in ``reward_entries``. Used by BOTH ``_partition_taxable_now``
    (CRG-021, taxable-now side) and ``_partition_skipped_rewards`` (CRG-022,
    deferred side) so the two sibling aggregators cannot drift apart. A direct
    unit test guards this helper.
    """
    return frozenset(e.asset for e in reward_entries if e.value_eur > 0)


def _partition_taxable_now(
    taxable_now_entries: list[CryptoRewardIncomeEntry],
    reward_entries: list[CryptoRewardIncomeEntry],
) -> tuple[list[CryptoRewardIncomeEntry], list[CryptoRewardIncomeEntry]]:
    """Split taxable-now entries into (real_rows, dust_rows).

    Dust = zero-value rows on assets that have at least one priced row elsewhere
    in the export (Koinly 2-decimal rounding artifact). Genuinely-unpriced assets
    (every row zero) keep per-row YES; they stay in real_rows.
    See CRG-021.

    ``taxable_now_entries`` is passed explicitly (not rebuilt from reward_entries)
    so the caller's local stays the single source of truth for line 236's
    taxable_now_total_eur (r2 Blocker #2). The discriminator guard is the direct
    unit test `TestPartitionTaxableNow` (Task 3), not a count-check on this list.

    Uses the shared ``_priced_assets_in_export`` discriminator (Invariant 3) so
    taxable-now dust and deferred dust (CRG-022) cannot silently desynchronize.
    """
    priced_assets_in_export = _priced_assets_in_export(reward_entries)
    dust_rows = [e for e in taxable_now_entries if e.value_eur == 0 and e.asset in priced_assets_in_export]
    real_rows = [e for e in taxable_now_entries if not (e.value_eur == 0 and e.asset in priced_assets_in_export)]
    return real_rows, dust_rows


def _partition_skipped_rewards(
    skipped_zero_value_deferred_rewards: list[CryptoRewardIncomeEntry],
    reward_entries: list[CryptoRewardIncomeEntry],
) -> tuple[list[CryptoRewardIncomeEntry], list[CryptoRewardIncomeEntry]]:
    """Split skipped deferred rewards into (dust_rows, unpriced_rows).

    Sibling to ``_partition_taxable_now`` (CRG-022). Both call the shared
    ``_priced_assets_in_export`` discriminator (Invariant 3, AGENTS.md rule 30)
    so the priced/unpriced split is byte-identical across the two sides.

    - ``dust_rows``: zero-value rows on an asset that has at least one
      ``value_eur > 0`` row elsewhere in ``reward_entries`` (Koinly 2-decimal
      rounding artifact).
    - ``unpriced_rows``: zero-value rows on an asset with NO priced row anywhere
      (Koinly has no price feed for these assets).

    Returns ``(dust_rows, unpriced_rows)`` so the caller can pass each partition
    into ``_write_suppressed_deferred_rewards_block`` once and reuse the counts
    for the Section 4 reconciliation (Invariant 5: compute-once-reuse).
    """
    priced_assets_in_export = _priced_assets_in_export(reward_entries)
    dust_rows = [
        e for e in skipped_zero_value_deferred_rewards if e.asset in priced_assets_in_export
    ]
    unpriced_rows = [
        e for e in skipped_zero_value_deferred_rewards if e.asset not in priced_assets_in_export
    ]
    return dust_rows, unpriced_rows


def _write_suppressed_rewards_bucket_table(
    worksheet: openpyxl.worksheet.worksheet.Worksheet,
    row_no: int,
    rows: list[CryptoRewardIncomeEntry],
    category: str,
) -> int:
    """Render one bucket (dust or unpriced) as a column table.

    Columns: Asset | Wallet | Rows | Summed amount | Category. Groups by
    ``(asset, wallet)`` and sorts ascending. ``Summed amount`` is the native-unit
    ``entry.amount`` sum formatted with ``:.8f`` (8 dp matches Koinly's native
    precision; ``parse_koinly_decimal`` returns ``Decimal(text)`` verbatim and
    Koinly's Amount column has variable scale 2/6/8 dp). ``Category`` is the
    short discriminator (``"dust"`` or ``"unpriced"``); the sub-header above
    carries the bucket label only (the prior verbose reason was removed when the
    column-table restructure deleted the single-line format).
    """
    for idx, header in enumerate(_SUPPRESSED_REWARDS_HEADERS, start=1):
        worksheet.cell(row_no, idx, header).font = Font(bold=True)
    row_no += 1
    grouped: dict[tuple[str, str], list[CryptoRewardIncomeEntry]] = {}
    for entry in rows:
        grouped.setdefault((entry.asset, entry.wallet), []).append(entry)
    for (asset, wallet), group in sorted(grouped.items()):
        amount_summed = sum((entry.amount for entry in group), start=ZERO)
        worksheet.cell(row_no, 1, safe_cell_value(asset))
        worksheet.cell(row_no, 2, safe_cell_value(wallet))
        worksheet.cell(row_no, 3, len(group))
        worksheet.cell(row_no, 4, f"{amount_summed:.8f}")
        worksheet.cell(row_no, 5, category)
        row_no += 1
    return row_no


def _write_suppressed_deferred_rewards_block(
    worksheet: openpyxl.worksheet.worksheet.Worksheet,
    row_no: int,
    dust_rows: list[CryptoRewardIncomeEntry],
    unpriced_rows: list[CryptoRewardIncomeEntry],
) -> int:
    """Write the "Suppressed zero-value deferred rewards" block.

    Outer header + two sub-headers (Deferred dust / Deferred unpriced), each
    followed by its own column table (Asset | Wallet | Rows | Summed amount |
    Category) sorted per-`(asset, wallet)`.

    Signature takes PRE-PARTITIONED rows (Invariant 5: compute-once-reuse - the
    partition happens once at the caller ``write_crypto_supplementary_sheet``,
    never inside this helper). The two partitions are disjoint by ``(asset,
    wallet)`` because the discriminator is asset-level (a priced asset cannot
    also be unpriced), so each bucket's sort is independent and a given
    ``(asset, wallet)`` key appears in at most one bucket.

    Render sequence:
      1. Blank spacer row (matches the section-boundary spacer convention).
      2. Outer bold header: "Suppressed zero-value deferred rewards".
      3. Sub-header "Deferred dust (priced-asset rounding)" + column table
         (only when ``dust_rows`` is non-empty).
      4. Sub-header "Deferred unpriced (no Koinly price feed)" + column table
         (only when ``unpriced_rows`` is non-empty).

    ``safe_cell_value`` wraps all user-facing string interpolations (asset,
    wallet) to prevent Excel formula injection and ``None`` literals.

    History: the predecessor plan collapsed two blocks into one per its r1
    review (aesthetic call). User feedback on the rendered sheet iterated twice:
    first the merged clause was hard to scan, so the restructure split it into
    two sub-headers; then the clumped single-line format was hard to read
    column-by-column, so each bucket became a proper column table mirroring the
    sheet's other table sections.
    """
    # 1. Blank spacer row before the outer header.
    row_no += 1

    # 2. Outer bold header.
    worksheet.cell(row_no, 1, "Suppressed zero-value deferred rewards").font = Font(bold=True)
    row_no += 1

    # 3. Dust bucket: render only when non-empty (conditional sub-header).
    if dust_rows:
        worksheet.cell(row_no, 1, "Deferred dust (priced-asset rounding)").font = Font(bold=True)
        row_no += 1
        row_no = _write_suppressed_rewards_bucket_table(worksheet, row_no, dust_rows, "dust")

    # 4. Unpriced bucket: render only when non-empty (conditional sub-header).
    if unpriced_rows:
        worksheet.cell(row_no, 1, "Deferred unpriced (no Koinly price feed)").font = Font(bold=True)
        row_no += 1
        row_no = _write_suppressed_rewards_bucket_table(worksheet, row_no, unpriced_rows, "unpriced")
    return row_no


def _write_dust_summary_block(
    worksheet: openpyxl.worksheet.worksheet.Worksheet,
    row_no: int,
    dust_rows: list[CryptoRewardIncomeEntry],
) -> int:
    """Write the taxable-now "Dust summary:" block.

    Outer header + sub-header, followed by a column table
    (Asset | Wallet | Rows | Summed Value (EUR) | Category) sorted per-`(asset,
    wallet)`. Taxable-now sums ``value_eur`` (EUR), not native amount - the
    taxable-now dust partition keeps EUR-denominated rewards, so the column
    header is "Summed Value (EUR)" (asymmetric with the deferred-side
    ``_write_suppressed_rewards_bucket_table``, which sums native-unit amount).

    Render sequence:
      1. Blank spacer row (matches the section-boundary spacer convention).
      2. Outer bold header: "Dust summary:" (kept - many tests + docs pin it).
      3. Sub-header "Taxable-now dust (priced-asset rounding)" + column table.
         Taxable-now has only one bucket - unpriced taxable-now rows keep
         per-row YES in the detail table per CRG-021, so there is no second
         sub-header here.
    """
    # 1. Blank spacer row before the outer header.
    row_no += 1

    # 2. Outer bold header.
    worksheet.cell(row_no, 1, "Dust summary:").font = Font(bold=True)
    row_no += 1

    # 3. Sub-header + column table.
    worksheet.cell(row_no, 1, "Taxable-now dust (priced-asset rounding)").font = Font(bold=True)
    row_no += 1
    for idx, header in enumerate(_DUST_SUMMARY_HEADERS, start=1):
        worksheet.cell(row_no, idx, header).font = Font(bold=True)
    row_no += 1
    grouped: dict[tuple[str, str], list[CryptoRewardIncomeEntry]] = {}
    for entry in dust_rows:
        grouped.setdefault((entry.asset, entry.wallet), []).append(entry)
    for (asset, wallet), group in sorted(grouped.items()):
        summed = sum((entry.value_eur for entry in group), start=ZERO)
        worksheet.cell(row_no, 1, safe_cell_value(asset))
        worksheet.cell(row_no, 2, safe_cell_value(wallet))
        worksheet.cell(row_no, 3, len(group))
        worksheet.cell(row_no, 4, float(summed))
        worksheet.cell(row_no, 5, "dust")
        row_no += 1
    return row_no


def write_crypto_supplementary_sheet(  # noqa: PLR0915
    workbook: openpyxl.Workbook,
    crypto_tax_report: CryptoTaxReport,
    tax_jurisdiction: TaxJurisdictionConfig,
) -> None:
    """Create and populate a 'Crypto Supplementary' worksheet with audit data.

    Writes:
    - Section 1: Income Codes reference (rendered when income-code classification is enabled)
    - Section 2: Taxable-now support detail (per-row trace data)
    - Section 3: Deferred-by-law support detail
    - Section 4: Rewards classification reconciliation
    - Section 5: Review required (zero-value popular tokens, suspicious entries)

    This sheet is for auditability only. Aggregated taxable-now rewards are
    reported on the Reporting worksheet under OTHER CAPITAL INVESTMENT INCOME
    per SRG-008.

    Args:
        workbook: The Excel workbook to add the sheet to.
        crypto_tax_report: The crypto tax report data.
        tax_jurisdiction: The reporting jurisdiction. The Income Codes reference
            section is rendered only when income-code classification is enabled
            (``classify_rewards_with_income_codes``); otherwise the entire
            reference section is omitted.
    """
    worksheet = workbook.create_sheet("Crypto Supplementary")

    taxable_now_entries = [
        e for e in crypto_tax_report.reward_entries if e.tax_classification == RewardTaxClassification.TAXABLE_NOW
    ]
    deferred_entries = [
        e for e in crypto_tax_report.reward_entries if e.tax_classification == RewardTaxClassification.DEFERRED_BY_LAW
    ]

    # Partition taxable-now rows into (real_rows, dust_rows). The local
    # ``taxable_now_entries`` stays the single source of truth for line 236's
    # taxable_now_total_eur (CRG-021 / r2 Blocker #2): the helper returns new
    # lists rather than rebuilding the local, so the total is byte-for-byte
    # identical to the pre-partition render (Invariant 1).
    real_rows, dust_rows = _partition_taxable_now(taxable_now_entries, crypto_tax_report.reward_entries)

    # Partition skipped deferred rewards ONCE (Invariant 5: compute-once-reuse).
    # The resulting (dust, unpriced) partition feeds BOTH the Section 3
    # suppressed-rewards block AND (in Task 5) the Section 4 reconciliation
    # split; never re-partition inside those consumers (CRG-022 / r2 finding #5).
    deferred_dust_rows, deferred_unpriced_rows = _partition_skipped_rewards(
        crypto_tax_report.skipped_zero_value_deferred_rewards,
        crypto_tax_report.reward_entries,
    )

    row_no = 1
    # Sections are numbered dynamically so the list still reads 1, 2, 3, ...
    # when the "INCOME CODES REFERENCE" section is omitted (classification
    # disabled; a hardcoded "1."-prefixed list that starts at "2." implies a
    # missing predecessor).
    section_no = 0

    # 1. INCOME CODES REFERENCE (rendered when income-code classification is enabled)
    if tax_jurisdiction.classify_rewards_with_income_codes:
        section_no += 1
        worksheet.cell(row_no, 1, f"{section_no}. INCOME CODES REFERENCE").font = Font(bold=True)
        row_no += 1

        reference_note = worksheet.cell(
            row_no,
            1,
            "Portuguese Tabela V income codes used for crypto reward income classification. "
            "See https://www.gov.pt (search 'Tabela V' in IRS withholding tax tables) for official source.",
        )
        reference_note.font = Font(italic=True, size=9)
        row_no += 1

        worksheet.cell(row_no, 1, "Country").font = Font(bold=True)
        worksheet.cell(row_no, 2, "Income Code").font = Font(bold=True)
        worksheet.cell(row_no, 3, "Description").font = Font(bold=True)
        row_no += 1

        for code, description in sorted(_INCOME_CODE_DESCRIPTIONS.items()):
            worksheet.cell(row_no, 1, tax_jurisdiction.country.upper())
            worksheet.cell(row_no, 2, code)
            worksheet.cell(row_no, 3, description)
            row_no += 1

    # 2. TAXABLE-NOW SUPPORT DETAIL
    row_no += 1
    section_no += 1
    worksheet.cell(row_no, 1, f"{section_no}. TAXABLE-NOW - SUPPORT DETAIL").font = Font(bold=True)
    row_no += 1

    taxable_note = worksheet.cell(
        row_no,
        1,
        "Individual taxable-now reward rows. Aggregated totals are reported on the Reporting worksheet "
        "under OTHER CAPITAL INVESTMENT INCOME. Use this section to trace each aggregated line back to "
        "its source Koinly rows for audit purposes.",
    )
    taxable_note.font = Font(italic=True, size=9)
    row_no += 1

    # Three-state empty label for Section 2. ``_write_reward_detail_rows`` only
    # renders ``empty_label`` when entries is empty (guard at :66-68), so passing
    # "" in the mixed/all-real case is safe (real_rows is non-empty there).
    if not real_rows and dust_rows:
        taxable_empty_label = "All taxable-now rows classified as dust - see summary below"
    elif not real_rows and not dust_rows:
        taxable_empty_label = "No taxable-now rewards"
    else:
        taxable_empty_label = ""  # mixed or all-real: headers render, no empty label
    row_no = _write_reward_detail_rows(worksheet, row_no, real_rows, taxable_empty_label)

    # Per-(asset, wallet) dust summary block renders below the detail table when
    # any taxable-now row was routed to dust (CRG-021).
    if dust_rows:
        row_no = _write_dust_summary_block(worksheet, row_no, dust_rows)

    # 3. DEFERRED BY LAW SUPPORT DETAIL
    row_no += 1
    section_no += 1
    worksheet.cell(row_no, 1, f"{section_no}. DEFERRED BY LAW - SUPPORT DETAIL").font = Font(bold=True)
    row_no += 1

    deferred_note = worksheet.cell(
        row_no,
        1,
        "These rewards are crypto-denominated and taxation is deferred until disposal per CIRS art. 5(11). "
        "They are shown here for auditability but not included in immediate filing rows.",
    )
    deferred_note.font = Font(italic=True, size=9)
    row_no += 1

    row_no = _write_reward_detail_rows(worksheet, row_no, deferred_entries, "No deferred rewards")

    # Single "Suppressed zero-value deferred rewards" summary block reading from
    # ``skipped_zero_value_deferred_rewards`` (CRG-022). The pre-partitioned
    # (dust, unpriced) rows are computed once above and reused (Invariant 5).
    if crypto_tax_report.skipped_zero_value_deferred_rewards:
        row_no = _write_suppressed_deferred_rewards_block(
            worksheet, row_no, deferred_dust_rows, deferred_unpriced_rows
        )

    # 4. REWARDS CLASSIFICATION RECONCILIATION
    row_no += 1
    section_no += 1
    worksheet.cell(row_no, 1, f"{section_no}. REWARDS CLASSIFICATION RECONCILIATION").font = Font(bold=True)
    row_no += 1

    taxable_now_total_eur = sum((e.value_eur for e in taxable_now_entries), start=ZERO)
    deferred_total_eur = sum((e.value_eur for e in deferred_entries), start=ZERO)

    reconciliation_rewards_rows = [
        ("Total reward rows (raw)", len(crypto_tax_report.reward_entries)),
        ("Taxable-now detail rows", len(real_rows)),
        ("Taxable-now dust rows (suppressed from detail)", len(dust_rows)),
        # Three-line deferred split (CRG-022). REUSES the partition computed once
        # at the ``_partition_skipped_rewards`` call above (Invariant 5:
        # compute-once-reuse - never re-partition for the reconciliation).
        # ``deferred_dust_rows`` and
        # ``deferred_unpriced_rows`` are the same lists passed into
        # ``_write_suppressed_deferred_rewards_block``, so the Section 4 counts and
        # the rendered Section 3 block cannot drift apart. The old single
        # ``("Deferred-by-law rows (taxation deferred)", ...)`` line is REMOVED.
        # The cross-sheet audit count lives on the Crypto Reconciliation sheet,
        # not here (r1 review finding #2: a 4th ``("Skipped zero-value deferred
        # rewards (audit)", ...)`` line is tautological on this sheet - dust +
        # unpriced == total by construction).
        ("Deferred detail rows", len(deferred_entries)),
        ("Deferred dust rows (suppressed from detail)", len(deferred_dust_rows)),
        ("Deferred unpriced rows (suppressed from detail)", len(deferred_unpriced_rows)),
        ("Taxable-now total value (EUR)", float(taxable_now_total_eur)),
        ("Deferred total value (EUR)", float(deferred_total_eur)),
    ]

    for key, value in reconciliation_rewards_rows:
        worksheet.cell(row_no, 1, key)
        worksheet.cell(row_no, 2, value)
        row_no += 1

    # 5. REVIEW REQUIRED
    row_no += 1
    section_no += 1
    worksheet.cell(row_no, 1, f"{section_no}. REVIEW REQUIRED").font = Font(bold=True)
    row_no += 1

    review_note = worksheet.cell(
        row_no,
        1,
        "Entries requiring manual review. These are excluded from the main report to avoid clutter. "
        "Review each item and verify the underlying Koinly data before including in your tax filing.",
    )
    review_note.font = Font(italic=True, size=9)
    row_no += 1

    row_no = _write_review_rows(worksheet, row_no, crypto_tax_report.review_entries)

    auto_column_width(worksheet)
