"""IB reporting sheet writer for the Excel tax report."""

from __future__ import annotations

from typing import TYPE_CHECKING

import openpyxl
from openpyxl.comments import Comment
from openpyxl.styles import Font, PatternFill

from ...domain.collections import (
    CapitalGainLinesPerCompany,
    DividendIncomePerCompany,
)
from ...domain.constants import (
    EXCEL_COLUMN_OFFSET,
    EXCEL_NUMBER_FORMAT,
    PLACEHOLDER_YEAR,
    ZERO_QUANTITY,
)
from ...domain.exceptions import ReportGenerationError
from ...infrastructure.config import Config, ConversionRate
from ...infrastructure.logging_config import create_module_logger
from .excel_utils import REVIEW_ROW_FILL, auto_column_width, safe_cell_value
from .tax_constants import get_income_code_description

if TYPE_CHECKING:
    from ..crypto_reporting import AggregatedRewardIncomeEntry


def write_ib_reporting_sheet(  # noqa: PLR0912, PLR0915
    worksheet: openpyxl.worksheet.worksheet.Worksheet,
    config: Config,
    capital_gain_lines_per_company: CapitalGainLinesPerCompany,
    dividend_income_per_company: DividendIncomePerCompany | None = None,
    other_capital_income_entries: list[AggregatedRewardIncomeEntry] | None = None,
) -> None:
    """Write the IB Reporting sheet with capital gains and dividend income.

    Args:
        worksheet: The openpyxl worksheet to populate (must already be named).
        config: Application configuration with exchange rates.
        capital_gain_lines_per_company: Calculated capital gains grouped by company.
        dividend_income_per_company: Dividend income data grouped by company (optional).
        other_capital_income_entries: Aggregated taxable fiat reward entries for the
            OTHER CAPITAL INVESTMENT INCOME subsection (optional).
    """
    logger = create_module_logger(__name__)

    # Write CAPITAL GAINS section title at the top if there are capital gains
    line_number = 1
    if capital_gain_lines_per_company:
        section_title_cell = worksheet.cell(line_number, 1, "CAPITAL GAINS")
        section_title_cell.font = Font(bold=True)  # type: ignore[assignment]
        line_number += 1  # Leave one blank row after section title
    else:
        # No capital gains, start from row 1
        line_number = 0

    first_header = [
        "Country of Source",
        "SALE",
        "",
        "",
        "",
        "PURCHASE",
        "",
        "",
        "",
        "WITHOLDING TAX",
        "",
        "Expenses incurred with obtaining the capital gains",
        "",
        "Symbol",
        "Currency",
        "Sale amount",
        "Buy amount",
        "Expenses amount",
    ]
    second_header = [
        "",
        "Day ",
        "Month ",
        "Year",
        "Amount",
        "Day ",
        "Month ",
        "Year",
        "Amount",
        "Country",
        "Amount",
        "",
        "",
        "",
        "",
        "",
        "",
        "",
    ]

    last_column: int = max(len(first_header), len(second_header))
    exchange_rates: dict[str, str] = create_currency_table(worksheet, last_column + 2, line_number + 1, config)
    logger.debug("Created currency exchange table with %s rates", len(config.rates) + 1)

    for i in range(len(first_header)):
        _ = worksheet.cell(line_number + 1, i + 1, first_header[i])
        _ = worksheet.cell(line_number + 2, i + 1, second_header[i])

    # Merge header cells for grouped columns. Each precondition guards the merge
    # span against the fixed header array structure; a mismatch means the header
    # layout drifted from the merge ranges, which would silently corrupt the sheet.
    header_row_1 = line_number + 1
    if first_header[1] != "SALE":
        raise ReportGenerationError(
            f"Expected SALE header at column 2, found {first_header[1]!r}; header layout drifted from merge spans"
        )
    if first_header[4] != "":
        raise ReportGenerationError(
            f"Expected empty SALE sub-header at column 5, found {first_header[4]!r}; merge span mismatch"
        )
    # Merge SALE header across 4 columns (cols 2-5: SALE + 3 empty sub-headers)
    worksheet.merge_cells(start_row=header_row_1, start_column=2, end_row=header_row_1, end_column=5)
    if first_header[5] != "PURCHASE":
        raise ReportGenerationError(
            f"Expected PURCHASE header at column 6, found {first_header[5]!r}; header layout drifted from merge spans"
        )
    if first_header[8] != "":
        raise ReportGenerationError(
            f"Expected empty PURCHASE sub-header at column 9, found {first_header[8]!r}; merge span mismatch"
        )
    # Merge PURCHASE header across 4 columns (cols 6-9: PURCHASE + 3 empty sub-headers)
    worksheet.merge_cells(start_row=header_row_1, start_column=6, end_row=header_row_1, end_column=9)
    if first_header[9] != "WITHOLDING TAX":
        raise ReportGenerationError(
            f"Expected WITHOLDING TAX header at column 10, found {first_header[9]!r}; "
            "header layout drifted from merge spans"
        )
    if first_header[10] != "":
        raise ReportGenerationError(
            f"Expected empty WITHOLDING TAX sub-header at column 11, found {first_header[10]!r}; merge span mismatch"
        )
    # Merge WITHOLDING TAX header across 2 columns (cols 10-11: WITHOLDING TAX + 1 empty sub-header)
    worksheet.merge_cells(start_row=header_row_1, start_column=10, end_row=header_row_1, end_column=11)

    start_column = 2  # Start at column 2 (first SALE sub-column), country of source is column 1
    capital_gains_start_row = line_number + 3  # Save for country of source pass
    line_number += 3  # Move past title row, blank row, and two header rows
    processed_lines = ZERO_QUANTITY

    for currency_company, capital_gain_lines in capital_gain_lines_per_company.items():
        currency = currency_company.currency
        company = currency_company.company
        logger.debug("Processing capital gain lines for %s (%s)", company.ticker, currency.currency)

        for line in capital_gain_lines:
            if currency != line.get_currency():
                raise ReportGenerationError(f"Currency mismatch in line: {currency} != {line.get_currency()}")
            processed_lines += 1
            idx = start_column

            _ = worksheet.cell(line_number, start_column, line.get_sell_date().day)
            idx += 1
            _ = worksheet.cell(line_number, idx, line.get_sell_date().get_month_name())
            idx += 1
            _ = worksheet.cell(line_number, idx, line.get_sell_date().year)
            idx += 1
            _ = worksheet.cell(
                line_number,
                idx,
                "=" + exchange_rates[currency.currency] + "*(" + line.get_sell_amount() + ")",
            )

            idx += 1
            _ = worksheet.cell(line_number, idx, line.get_buy_date().day)
            idx += 1
            _ = worksheet.cell(line_number, idx, line.get_buy_date().get_month_name())
            idx += 1
            _ = worksheet.cell(line_number, idx, line.get_buy_date().year)
            idx += 1
            _ = worksheet.cell(
                line_number,
                idx,
                "=" + exchange_rates[currency.currency] + "*(" + line.get_buy_amount() + ")",
            )

            idx += EXCEL_COLUMN_OFFSET

            expense_cell = worksheet.cell(
                line_number,
                idx,
                "=" + exchange_rates[currency.currency] + "*(" + line.get_expense_amount() + ")",
            )
            expense_cell.number_format = EXCEL_NUMBER_FORMAT  # type: ignore[assignment]
            idx += 2

            _ = worksheet.cell(line_number, idx, company.ticker)
            idx += 1
            _ = worksheet.cell(line_number, idx, currency.currency)
            idx += 1

            sell_amount_cell = worksheet.cell(line_number, idx, "=" + line.get_sell_amount())
            sell_amount_cell.number_format = EXCEL_NUMBER_FORMAT  # type: ignore[assignment]
            idx += 1
            buy_amount_cell = worksheet.cell(line_number, idx, "=" + line.get_buy_amount())
            buy_amount_cell.number_format = EXCEL_NUMBER_FORMAT  # type: ignore[assignment]
            idx += 1
            expense_amount_cell = worksheet.cell(line_number, idx, "=" + line.get_expense_amount())
            expense_amount_cell.number_format = EXCEL_NUMBER_FORMAT  # type: ignore[assignment]

            if line.get_buy_date().year == PLACEHOLDER_YEAR:
                for col_idx in range(start_column, idx + 1):
                    cell = worksheet.cell(line_number, col_idx)
                    cell.fill = REVIEW_ROW_FILL  # type: ignore[assignment]

            line_number += 1

    logger.debug("Processed %s capital gain lines", processed_lines)

    line_number = capital_gains_start_row
    for currency_company, capital_gain_lines in capital_gain_lines_per_company.items():
        company = currency_company.company
        for _ in capital_gain_lines:
            _ = worksheet.cell(line_number, 1, company.country_of_issuance)
            _ = worksheet.cell(line_number, 10, company.country_of_issuance)
            line_number += 1

    _write_capital_investment_income_section(
        worksheet, exchange_rates, dividend_income_per_company, other_capital_income_entries
    )

    auto_column_width(worksheet)


def _write_capital_investment_income_section(
    worksheet: openpyxl.worksheet.worksheet.Worksheet,
    exchange_rates: dict[str, str],
    dividend_income_per_company: DividendIncomePerCompany | None,
    other_capital_income_entries: list[AggregatedRewardIncomeEntry] | None,
) -> None:
    """Write the CAPITAL INVESTMENT INCOME section with SHARE DIVIDENDS and OTHER subsections."""
    logger = create_module_logger(__name__)

    has_dividends = bool(dividend_income_per_company)
    has_other = bool(other_capital_income_entries)
    if not has_dividends and not has_other:
        return

    logger.info(
        "Adding CAPITAL INVESTMENT INCOME section (dividends=%s, other=%s)",
        has_dividends,
        len(other_capital_income_entries) if other_capital_income_entries else 0,
    )

    line_number = worksheet.max_row + 2

    section_title_cell = worksheet.cell(line_number, 1, "CAPITAL INVESTMENT INCOME")
    section_title_cell.font = Font(bold=True)  # type: ignore[assignment]
    line_number += 2

    income_headers = [
        "Type of capital income\n(choose one)",
        "Country of source",
        "ISIN",
        "Gross amount",
        "Withholding tax at source",
        "Withholding tax in Portugal\n(if any)",
        "",
        "Symbol",
        "Currency",
        "Original gross amount",
        "Original tax amount",
        "Net amount",
        "Source system",
        "Raw row count",
    ]

    for i, header in enumerate(income_headers):
        _ = worksheet.cell(line_number, i + 1, header)
    line_number += 1

    if has_dividends:
        subsection_cell = worksheet.cell(line_number, 1, "SHARE DIVIDENDS")
        subsection_cell.font = Font(bold=True)  # type: ignore[assignment]
        line_number += 1

        for symbol, dividend_data in dividend_income_per_company.items():
            _ = worksheet.cell(line_number, 1, "Share dividends")

            if dividend_data.isin == "MISSING_ISIN_REQUIRES_ATTENTION":
                country_cell = worksheet.cell(line_number, 2, "⚠️ MISSING DATA")
                country_cell.fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")  # type: ignore[assignment]

                isin_cell = worksheet.cell(line_number, 3, f"⚠️ {symbol}")
                isin_cell.fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")  # type: ignore[assignment]

                isin_cell.comment = Comment(  # type: ignore[assignment]
                    f"Security information missing for {symbol}. Please verify this symbol in your IB account.",
                    "Shares Reporting",
                )
            else:
                _ = worksheet.cell(line_number, 2, dividend_data.country)
                _ = worksheet.cell(line_number, 3, dividend_data.isin)

            gross_amount_cell = worksheet.cell(
                line_number,
                4,
                "=" + exchange_rates[dividend_data.currency.currency] + "*(" + str(dividend_data.gross_amount) + ")",
            )
            gross_amount_cell.number_format = EXCEL_NUMBER_FORMAT  # type: ignore[assignment]

            tax_amount_cell = worksheet.cell(
                line_number,
                5,
                "=" + exchange_rates[dividend_data.currency.currency] + "*(" + str(dividend_data.total_taxes) + ")",
            )
            tax_amount_cell.number_format = EXCEL_NUMBER_FORMAT  # type: ignore[assignment]

            _ = worksheet.cell(line_number, 6, "")

            _ = worksheet.cell(line_number, 8, symbol)
            _ = worksheet.cell(line_number, 9, dividend_data.currency.currency)

            original_gross_cell = worksheet.cell(line_number, 10, str(dividend_data.gross_amount))
            original_gross_cell.number_format = EXCEL_NUMBER_FORMAT

            original_tax_cell = worksheet.cell(line_number, 11, str(dividend_data.total_taxes))
            original_tax_cell.number_format = EXCEL_NUMBER_FORMAT

            net_amount = dividend_data.gross_amount - dividend_data.total_taxes
            net_amount_cell = worksheet.cell(line_number, 12, str(net_amount))
            net_amount_cell.number_format = EXCEL_NUMBER_FORMAT

            logger.debug(
                "Added dividend income row for %s: %s gross, %s tax, %s net (%s)",
                symbol,
                dividend_data.gross_amount,
                dividend_data.total_taxes,
                net_amount,
                dividend_data.currency.currency,
            )
            line_number += 1

    if other_capital_income_entries:
        if has_dividends:
            line_number += 1  # Add blank row separator
        other_subsection_cell = worksheet.cell(line_number, 1, "OTHER CAPITAL INVESTMENT INCOME")
        other_subsection_cell.font = Font(bold=True)  # type: ignore[assignment]
        line_number += 1

        line_number = _write_other_capital_income_subsection(
            worksheet, line_number, other_capital_income_entries
        )


def _write_other_capital_income_subsection(
    worksheet: openpyxl.worksheet.worksheet.Worksheet,
    line_number: int,
    other_capital_income_entries: list[AggregatedRewardIncomeEntry],
) -> int:
    """Write the OTHER CAPITAL INVESTMENT INCOME subsection rows.

    Args:
        worksheet: The openpyxl worksheet to populate.
        line_number: The starting row number (1-based).
        other_capital_income_entries: Aggregated taxable fiat reward entries.

    Returns:
        The updated line_number after writing all rows.
    """
    logger = create_module_logger(__name__)

    for entry in other_capital_income_entries:
        source_detail = safe_cell_value(", ".join(sorted(entry.chains)))
        income_type_label = safe_cell_value(get_income_code_description(entry.income_code))
        net_eur = entry.gross_income_eur - entry.foreign_tax_eur

        columns = [
            (1, income_type_label, None),
            (2, safe_cell_value(entry.source_country), None),
            (3, "", None),
            (4, str(entry.gross_income_eur), EXCEL_NUMBER_FORMAT),
            (5, str(entry.foreign_tax_eur), EXCEL_NUMBER_FORMAT),
            (6, "", None),
            (7, "", None),
            (8, source_detail, None),
            (9, "EUR", None),
            (10, str(entry.gross_income_eur), EXCEL_NUMBER_FORMAT),
            (11, str(entry.foreign_tax_eur), EXCEL_NUMBER_FORMAT),
            (12, str(net_eur), EXCEL_NUMBER_FORMAT),
            (13, "Koinly", None),
            (14, entry.raw_row_count, None),
        ]

        for col_idx, value, number_format in columns:
            cell = worksheet.cell(line_number, col_idx, value)
            if number_format:
                cell.number_format = number_format  # type: ignore[assignment]

        logger.debug(
            "Added other capital income row: code=%s country=%s gross=%s tax=%s net=%s",
            entry.income_code,
            entry.source_country,
            entry.gross_income_eur,
            entry.foreign_tax_eur,
            net_eur,
        )
        line_number += 1

    return line_number


def create_currency_table(
    worksheet: openpyxl.worksheet.worksheet.Worksheet,
    column_no: int,
    row_no: int,
    config: Config,
) -> dict[str, str]:
    """Create a currency configuration table in the excel worksheet.

    Args:
        worksheet: The target excel worksheet.
        column_no: The starting column number (1-based).
        row_no: The starting row number (1-based).
        config: The application configuration object.

    Returns:
        A dictionary mapping currency codes to their cell coordinate strings.
    """
    logger = create_module_logger(__name__)
    currency_header = ["Base/target", "Rate"]
    rates: list[ConversionRate] = config.rates

    logger.debug("Creating currency table starting at column %s, row %s", column_no, row_no)

    _ = worksheet.cell(row_no, column_no, "Currency exchange rate")
    row_no += 1
    for i in range(len(currency_header)):
        _ = worksheet.cell(row_no, column_no + i, currency_header[i])
    row_no += 1

    coordinates: dict[str, str] = {}
    for j in range(len(rates)):
        _ = worksheet.cell(row_no + j, column_no, rates[j].base + "/" + rates[j].calculated)
        cell = worksheet.cell(row_no + j, column_no + 1, str(rates[j].rate))
        coordinates[rates[j].calculated] = cell.coordinate
        logger.debug("Added currency rate %s/%s = %s", rates[j].base, rates[j].calculated, rates[j].rate)

    _ = worksheet.cell(row_no + len(rates), column_no, config.base + "/" + config.base)
    cell = worksheet.cell(row_no + len(rates), column_no + 1, "1")
    coordinates[config.base] = cell.coordinate
    logger.debug("Added base currency %s/%s = 1", config.base, config.base)

    logger.debug("Created currency table with %s exchange rates", len(coordinates))
    return coordinates
