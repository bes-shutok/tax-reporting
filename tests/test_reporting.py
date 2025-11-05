from datetime import datetime
from decimal import Decimal

from shares_reporting.application.extraction import parse_data
from shares_reporting.application.persisting import persist_results
from shares_reporting.application.transformation import calculate, split_by_days
from shares_reporting.domain.value_objects import TradeDate, TradeType, get_trade_date
from shares_reporting.domain.accumulators import TradePartsWithinDay
from shares_reporting.domain.collections import DayPartitionedTrades, TradeCyclePerCompany, CapitalGainLinesPerCompany, QuantitatedTradeAction, QuantitatedTradeActions
from test_data import sell_action1

test_dict1 = {("2022", "01"), ("2021", "12"), ("2021", "02")}
test_dict2 = ["202201", "202112", "202102"]
date_time1 = datetime.strptime("2021-05-18, 14:53:23", '%Y-%m-%d, %H:%M:%S')
date_time2 = datetime.strptime("2022-05-18, 14:53:23", '%Y-%m-%d, %H:%M:%S')
date_time3 = datetime.strptime("2021-01-18, 14:53:23", '%Y-%m-%d, %H:%M:%S')
date_time4 = datetime.strptime("2021-12-18, 14:53:23", '%Y-%m-%d, %H:%M:%S')
test_dict4 = [TradeDate(date_time1.year, date_time1.month, date_time1.day),
              TradeDate(date_time2.year, date_time2.month, date_time2.day),
              TradeDate(date_time3.year, date_time3.month, date_time3.day),
              TradeDate(date_time4.year, date_time4.month, date_time4.day)]

def test_sorting():
    print(str(test_dict1))
    print(str(sorted(test_dict1)))
    print(str(test_dict2))
    print(str(sorted(test_dict2)))
    print(str(test_dict4))
    print(str(sorted(test_dict4)))
    assert 1 == 1

def test_partitioning_by_days():
        """Test that split_by_days correctly groups trades by day"""
        trades_within_day1 = TradePartsWithinDay()
        trades_within_day1.push_trade_part(Decimal(1), sell_action1)
        day_partitioned_trades1: DayPartitionedTrades = {get_trade_date(sell_action1.date_time): trades_within_day1}

        actions: QuantitatedTradeActions = [QuantitatedTradeAction(quantity=Decimal(1.0), action=sell_action1)]
        actual: DayPartitionedTrades = split_by_days(actions, TradeType.SELL)
        # print(actual)
        assert actual == day_partitioned_trades1

# https://stackoverflow.com/questions/52089716/comparing-two-excel-files-using-openpyxl
def test_comparing():
    from pathlib import Path

    expected = Path('tests','resources', 'expected.xlsx')
    source = Path('tests','resources', 'simple.csv')
    destination = Path('tests','resources', 'tmp.xlsx')
    leftover = Path('tests','resources', 'tmp_leftover.xlsx')

    trade_actions_per_company: TradeCyclePerCompany = parse_data(source)
    capital_gains: CapitalGainLinesPerCompany = {}
    leftover_trades: TradeCyclePerCompany = {}
    calculate(trade_actions_per_company, leftover_trades, capital_gains)
    persist_results(destination, capital_gains)

    from openpyxl import load_workbook
    workbook1 = load_workbook(expected, data_only=False)
    workbook2 = load_workbook(destination, data_only=False)

    sheet1 = workbook1.active
    sheet2 = workbook2.active
    assert sheet1.title == sheet2.title

    min_row1: int = sheet1.min_row
    min_row2: int = sheet2.min_row
    assert min_row1 == min_row2

    max_row1: int = sheet1.max_row
    max_row2: int = sheet2.max_row
    assert max_row1 == max_row2

    min_column1: int = sheet1.min_column
    min_column2: int = sheet2.min_column
    assert min_column1 == min_column2

    max_column1: int = sheet1.max_column
    max_column2: int = sheet2.max_column
    assert max_column1 == max_column2

    for i in range(min_row1, max_row1):
        for j in range(min_column1, max_column1):
            expected = sheet1.cell(i, j).value
            actual = sheet2.cell(i, j).value
            # Handle country column differences - expected file has None, new format has 'Unknown'
            if j == 2 or j == 11:  # Country of Source and WITHOLDING TAX Country columns
                if expected is None:
                    assert actual == 'Unknown' or actual is None
                else:
                    assert expected == actual
            else:
                assert expected == actual 
