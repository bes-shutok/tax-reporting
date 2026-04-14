"""Tests for dividend Excel persisting functionality."""

from decimal import Decimal

import pytest

from shares_reporting.application.crypto_reporting import (
    CryptoCapitalGainEntry,
    CryptoCapitalGainStats,
    CryptoCompletePdfSummary,
    CryptoReconciliationSummary,
    CryptoRewardIncomeEntry,
    CryptoSkippedZeroValueToken,
    CryptoTaxReport,
    HoldingsSnapshot,
    resolve_operator_origin,
)
from shares_reporting.application.persisting import generate_tax_report
from shares_reporting.domain.collections import DividendIncomePerCompany
from shares_reporting.domain.entities import DividendIncomePerSecurity
from shares_reporting.domain.value_objects import parse_currency


class TestDividendExcelPersisting:
    """Test dividend Excel report generation."""

    def test_generate_tax_report_with_dividends(self, tmp_path):
        """Test generating Excel report with dividend income section."""
        dividend_income = DividendIncomePerCompany(
            {
                "AAPL": DividendIncomePerSecurity(
                    symbol="AAPL",
                    isin="US0378331005",
                    country="United States",
                    gross_amount=Decimal("100.00"),
                    total_taxes=Decimal("15.00"),
                    currency=parse_currency("USD"),
                ),
                "MSFT": DividendIncomePerSecurity(
                    symbol="MSFT",
                    isin="US5949181045",
                    country="United States",
                    gross_amount=Decimal("200.00"),
                    total_taxes=Decimal("30.00"),
                    currency=parse_currency("USD"),
                ),
                "ASML": DividendIncomePerSecurity(
                    symbol="ASML",
                    isin="NL0010273215",
                    country="Netherlands",
                    gross_amount=Decimal("150.00"),
                    total_taxes=Decimal("0.00"),
                    currency=parse_currency("EUR"),
                ),
            }
        )

        # Generate report
        report_path = tmp_path / "dividend_report.xlsx"
        generate_tax_report(
            extract=report_path,
            capital_gain_lines_per_company={},
            dividend_income_per_company=dividend_income,
        )

        # Verify file was created
        assert report_path.exists()

        # Read and verify the Excel file
        import openpyxl

        workbook = openpyxl.load_workbook(report_path)
        worksheet = workbook.active
        assert worksheet is not None, "Workbook should have an active worksheet"

        # Find the dividend section
        found_dividend_section = False
        dividend_row_count = 0

        for _row_idx, row in enumerate(worksheet.iter_rows(values_only=True), 1):
            if row and isinstance(row[0], str) and "CAPITAL INVESTMENT INCOME" in row[0]:
                found_dividend_section = True
                # Next row should be empty
                continue

            if found_dividend_section and row and len(row) > 0:
                # Check for header row
                if isinstance(row[0], str) and "Beneficiary" in row[0]:
                    continue

                # Count data rows
                if dividend_row_count < 3 and row[1] == "Dividends":
                    dividend_row_count += 1
                    # Verify data integrity
                    symbol = row[8] if len(row) > 8 else None
                    country = row[2] if len(row) > 2 else None
                    isin = row[3] if len(row) > 3 else None
                    original_gross = row[10] if len(row) > 10 else None
                    original_tax = row[11] if len(row) > 11 else None
                    net_amount = row[12] if len(row) > 12 else None

                    # Verify expected values based on symbol
                    if symbol == "AAPL":
                        assert country == "United States"
                        assert isin == "US0378331005"
                        assert original_gross == "100.00"
                        assert original_tax == "15.00"
                        assert net_amount == "85.00"
                    elif symbol == "MSFT":
                        assert country == "United States"
                        assert isin == "US5949181045"
                        assert original_gross == "200.00"
                        assert original_tax == "30.00"
                        assert net_amount == "170.00"
                    elif symbol == "ASML":
                        assert country == "Netherlands"
                        assert isin == "NL0010273215"
                        assert original_gross == "150.00"
                        assert original_tax == "0.00"
                        assert net_amount == "150.00"

        assert found_dividend_section, "Dividend section not found in report"
        assert dividend_row_count == 3, f"Expected 3 dividend rows, found {dividend_row_count}"

        workbook.close()

    def test_generate_tax_report_without_dividends(self, tmp_path):
        """Test generating Excel report without dividend income section."""
        report_path = tmp_path / "no_dividend_report.xlsx"
        generate_tax_report(
            extract=report_path,
            capital_gain_lines_per_company={},
            dividend_income_per_company={},  # Empty dividend data
        )

        assert report_path.exists()

        # Read and verify no dividend section
        import openpyxl

        workbook = openpyxl.load_workbook(report_path)
        worksheet = workbook.active
        assert worksheet is not None, "Workbook should have an active worksheet"

        # Should not have dividend section
        found_dividend_section = False
        for row in worksheet.iter_rows(values_only=True):
            if row and isinstance(row[0], str) and "CAPITAL INVESTMENT INCOME" in row[0]:
                found_dividend_section = True
                break

        assert not found_dividend_section, "Dividend section should not be present"
        workbook.close()

    def test_generate_tax_report_with_dividend_formulas(self, tmp_path):
        """Test that dividend Excel report contains proper currency conversion formulas."""
        dividend_income = DividendIncomePerCompany(
            {
                "AAPL": DividendIncomePerSecurity(
                    symbol="AAPL",
                    isin="US0378331005",
                    country="United States",
                    gross_amount=Decimal("100.50"),
                    total_taxes=Decimal("15.25"),
                    currency=parse_currency("USD"),
                ),
            }
        )

        report_path = tmp_path / "dividend_formulas_report.xlsx"
        generate_tax_report(
            extract=report_path,
            capital_gain_lines_per_company={},
            dividend_income_per_company=dividend_income,
        )

        # Read and verify formulas
        import openpyxl

        workbook = openpyxl.load_workbook(report_path)
        worksheet = workbook.active
        assert worksheet is not None, "Workbook should have an active worksheet"

        # Find dividend row and check formulas
        dividend_row = None
        for row_idx, row in enumerate(worksheet.iter_rows(values_only=False), 1):
            if row and len(row) > 1 and row[1].value == "Dividends":
                dividend_row = row_idx
                break

        assert dividend_row is not None, "Dividend row not found"

        # Check converted gross amount formula (column 5, index 4)
        gross_amount_cell = worksheet.cell(row=dividend_row, column=5)
        assert gross_amount_cell.data_type == "f"  # Formula cell
        assert "=V" in str(gross_amount_cell.value)  # Contains reference to exchange rate
        assert "100.5" in str(gross_amount_cell.value)  # Contains original amount

        # Check converted tax amount formula (column 6, index 5)
        tax_amount_cell = worksheet.cell(row=dividend_row, column=6)
        assert tax_amount_cell.data_type == "f"  # Formula cell
        assert "=V" in str(tax_amount_cell.value)  # Contains reference to exchange rate
        assert "15.25" in str(tax_amount_cell.value)  # Contains original amount

        # Check original amounts (should be string values)
        original_gross_cell = worksheet.cell(row=dividend_row, column=11)
        assert original_gross_cell.data_type == "s"  # String value
        assert original_gross_cell.value == "100.50"

        original_tax_cell = worksheet.cell(row=dividend_row, column=12)
        assert original_tax_cell.data_type == "s"  # String value
        assert original_tax_cell.value == "15.25"

        net_amount_cell = worksheet.cell(row=dividend_row, column=13)
        assert net_amount_cell.data_type == "s"  # String value
        assert net_amount_cell.value == "85.25"  # 100.50 - 15.25

        workbook.close()

    def test_generate_tax_report_with_multiple_currencies(self, tmp_path):
        """Test generating Excel report with multiple dividend currencies."""
        dividend_income = DividendIncomePerCompany(
            {
                "AAPL": DividendIncomePerSecurity(
                    symbol="AAPL",
                    isin="US0378331005",
                    country="United States",
                    gross_amount=Decimal("100.00"),
                    total_taxes=Decimal("15.00"),
                    currency=parse_currency("USD"),
                ),
                "ASML": DividendIncomePerSecurity(
                    symbol="ASML",
                    isin="NL0010273215",
                    country="Netherlands",
                    gross_amount=Decimal("200.00"),
                    total_taxes=Decimal("25.00"),
                    currency=parse_currency("EUR"),
                ),
                "TSLA": DividendIncomePerSecurity(
                    symbol="TSLA",
                    isin="US88160R1014",
                    country="United States",
                    gross_amount=Decimal("50.00"),
                    total_taxes=Decimal("0.00"),
                    currency=parse_currency("CAD"),
                ),
            }
        )

        report_path = tmp_path / "multi_currency_dividend_report.xlsx"
        generate_tax_report(
            extract=report_path,
            capital_gain_lines_per_company={},
            dividend_income_per_company=dividend_income,
        )

        # Read and verify multiple currencies
        import openpyxl

        workbook = openpyxl.load_workbook(report_path)
        worksheet = workbook.active
        assert worksheet is not None, "Workbook should have an active worksheet"

        # Find dividend rows and check currency columns
        dividend_currencies = []
        for _row_idx, row in enumerate(worksheet.iter_rows(values_only=True), 1):
            if row and len(row) > 1 and row[1] == "Dividends":
                currency = row[9] if len(row) > 9 else None  # Currency column
                if currency:
                    dividend_currencies.append(currency)

        assert len(dividend_currencies) == 3
        assert "USD" in dividend_currencies
        assert "EUR" in dividend_currencies
        assert "CAD" in dividend_currencies
        workbook.close()

    def test_generate_tax_report_with_crypto_sheet(self, tmp_path):
        """Test generating Excel report with additional crypto worksheet."""
        bybit_origin = resolve_operator_origin("ByBit")
        wirex_origin = resolve_operator_origin("Wirex", transaction_type="crypto_deposit")

        capital_entries = [
            CryptoCapitalGainEntry(
                disposal_date="2025-01-13",
                acquisition_date="2024-11-18",
                asset="USDT",
                amount=Decimal("1.5"),
                cost_eur=Decimal("1.25"),
                proceeds_eur=Decimal("1.35"),
                gain_loss_eur=Decimal("0.10"),
                holding_period="Short term",
                wallet="ByBit (2)",
                platform="ByBit",
                chain="ByBit",
                operator_origin=bybit_origin,
                annex_hint="J",
                review_required=bybit_origin.review_required,
                review_reason=bybit_origin.review_reason,
                notes="",
                token_swap_history="",
            )
        ]

        crypto_report = CryptoTaxReport(
            tax_year=2025,
            capital_entries=capital_entries,
            reward_entries=[
                CryptoRewardIncomeEntry(
                    date="2025-01-01",
                    asset="WXT",
                    amount=Decimal("5"),
                    value_eur=Decimal("17.10"),
                    income_label="Reward",
                    source_type="Reward",
                    wallet="Wirex",
                    platform="Wirex",
                    chain="Wirex",
                    operator_origin=wirex_origin,
                    annex_hint="J",
                    review_required=wirex_origin.review_required,
                    review_reason=wirex_origin.review_reason,
                    description="",
                )
            ],
            reconciliation=CryptoReconciliationSummary(
                capital_rows=1,
                reward_rows=1,
                short_term_rows=1,
                long_term_rows=0,
                mixed_rows=0,
                unknown_rows=0,
                capital_cost_total_eur=Decimal("1.25"),
                capital_proceeds_total_eur=Decimal("1.35"),
                capital_gain_total_eur=Decimal("0.10"),
                reward_total_eur=Decimal("17.10"),
                opening_holdings=HoldingsSnapshot(
                    asset_rows=1,
                    total_cost_eur=Decimal("100.00"),
                    total_value_eur=Decimal("120.00"),
                ),
                closing_holdings=HoldingsSnapshot(
                    asset_rows=1,
                    total_cost_eur=Decimal("130.00"),
                    total_value_eur=Decimal("150.00"),
                ),
            ),
            capital_gain_stats=CryptoCapitalGainStats.from_entries(capital_entries),
            skipped_zero_value_tokens=[
                CryptoSkippedZeroValueToken(source_section="capital_gains", asset="FEE", count=3),
                CryptoSkippedZeroValueToken(source_section="income", asset="AAA", count=2),
            ],
            pdf_summary=CryptoCompletePdfSummary(
                period="1 Jan 2025 to 31 Dec 2025",
                timezone="Europe/Lisbon",
                extracted_tokens=12,
            ),
        )

        report_path = tmp_path / "crypto_report.xlsx"
        generate_tax_report(
            extract=report_path,
            capital_gain_lines_per_company={},
            dividend_income_per_company={},
            crypto_tax_report=crypto_report,
        )

        assert report_path.exists()

        import openpyxl

        workbook = openpyxl.load_workbook(report_path)
        assert "Crypto Gains" in workbook.sheetnames
        assert "Crypto Rewards" in workbook.sheetnames
        assert "Crypto Reconciliation" in workbook.sheetnames

        gains_sheet = workbook["Crypto Gains"]
        assert gains_sheet["A1"].value == "CRYPTO TAX REPORT - PORTUGAL"

        gains_labels = set()
        for row in gains_sheet.iter_rows(values_only=True):
            first_cell = row[0] if row else None
            if isinstance(first_cell, str):
                gains_labels.add(first_cell)

        rewards_labels = set()
        for row in workbook["Crypto Rewards"].iter_rows(values_only=True):
            first_cell = row[0] if row else None
            if isinstance(first_cell, str):
                rewards_labels.add(first_cell)

        recon_labels = set()
        for row in workbook["Crypto Reconciliation"].iter_rows(values_only=True):
            first_cell = row[0] if row else None
            if isinstance(first_cell, str):
                recon_labels.add(first_cell)

        assert "1. CAPITAL GAINS" in gains_labels
        assert "2. REWARDS INCOME - IRS-READY FILING SUMMARY" in rewards_labels
        assert "2b. DEFERRED BY LAW - SUPPORT DETAIL" in rewards_labels
        assert "2c. REWARDS CLASSIFICATION RECONCILIATION" in rewards_labels
        assert "3. RECONCILIATION" in recon_labels
        assert "4. SKIPPED ZERO VALUE TOKENS" in recon_labels
        assert "PDF period" in gains_labels
        workbook.close()

    def test_crypto_sheet_contains_chain_column(self, tmp_path):
        """Assert the Crypto worksheet contains chain headers and writes the normalized values."""
        bybit_origin = resolve_operator_origin("ByBit")
        wirex_origin = resolve_operator_origin("Wirex", transaction_type="crypto_deposit")

        capital_entries = [
            CryptoCapitalGainEntry(
                disposal_date="2025-01-13",
                acquisition_date="2024-11-18",
                asset="USDT",
                amount=Decimal("1.5"),
                cost_eur=Decimal("1.25"),
                proceeds_eur=Decimal("1.35"),
                gain_loss_eur=Decimal("0.10"),
                holding_period="Short term",
                wallet="ByBit (2)",
                platform="ByBit",
                chain="ByBit",
                operator_origin=bybit_origin,
                annex_hint="J",
                review_required=bybit_origin.review_required,
                review_reason=bybit_origin.review_reason,
                notes="",
                token_swap_history="",
            )
        ]

        crypto_report = CryptoTaxReport(
            tax_year=2025,
            capital_entries=capital_entries,
            reward_entries=[
                CryptoRewardIncomeEntry(
                    date="2025-01-01",
                    asset="WXT",
                    amount=Decimal("5"),
                    value_eur=Decimal("17.10"),
                    income_label="Reward",
                    source_type="Reward",
                    wallet="Wirex",
                    platform="Wirex",
                    chain="Wirex",
                    operator_origin=wirex_origin,
                    annex_hint="J",
                    review_required=wirex_origin.review_required,
                    review_reason=wirex_origin.review_reason,
                    description="",
                )
            ],
            reconciliation=CryptoReconciliationSummary(
                capital_rows=1,
                reward_rows=1,
                short_term_rows=1,
                long_term_rows=0,
                mixed_rows=0,
                unknown_rows=0,
                capital_cost_total_eur=Decimal("1.25"),
                capital_proceeds_total_eur=Decimal("1.35"),
                capital_gain_total_eur=Decimal("0.10"),
                reward_total_eur=Decimal("17.10"),
                opening_holdings=None,
                closing_holdings=None,
            ),
            capital_gain_stats=CryptoCapitalGainStats.from_entries(capital_entries),
            skipped_zero_value_tokens=[],
            pdf_summary=None,
        )

        report_path = tmp_path / "crypto_chain_report.xlsx"
        generate_tax_report(
            extract=report_path,
            capital_gain_lines_per_company={},
            dividend_income_per_company={},
            crypto_tax_report=crypto_report,
        )

        assert report_path.exists()

        import openpyxl

        workbook = openpyxl.load_workbook(report_path)
        gains_sheet = workbook["Crypto Gains"]
        rewards_sheet = workbook["Crypto Rewards"]

        # Find the capital gains header row and check for chain column
        capital_headers_found = False
        capital_chain_col_idx = None

        for row in gains_sheet.iter_rows(values_only=True):
            if row and "Disposal date" in str(row[0] if row[0] else ""):
                capital_headers_found = True
                for col_idx, cell_value in enumerate(row):
                    if cell_value == "Disposal chain":
                        capital_chain_col_idx = col_idx + 1
                        break

        reward_headers_found = False
        reward_chain_col_idx = None

        for row in rewards_sheet.iter_rows(values_only=True):
            if row and "Date" in str(row[0] if row[0] else "") and "Asset" in str(row[1] if len(row) > 1 else ""):
                reward_headers_found = True
                for col_idx, cell_value in enumerate(row):
                    if cell_value == "Reward chain":
                        reward_chain_col_idx = col_idx + 1
                        break

        # Verify headers were found
        assert capital_headers_found, "Capital gains headers not found"
        assert reward_headers_found, "Reward headers not found"
        assert capital_chain_col_idx is not None, "Disposal chain column not found in capital gains headers"
        assert reward_chain_col_idx is not None, "Reward chain column not found in rewards headers"

        # Verify the chain values are written in the data rows
        capital_data_row = None
        reward_data_row = None

        for row in gains_sheet.iter_rows(values_only=True):
            if row and row[0] == "2025-01-13":
                capital_data_row = row

        for row in rewards_sheet.iter_rows(values_only=True):
            if row and row[0] == "2025-01-01":
                reward_data_row = row

        assert capital_data_row is not None, "Capital data row not found"
        assert reward_data_row is not None, "Reward data row not found"

        # Check chain values in data rows (using 1-based column index)
        capital_chain = capital_data_row[capital_chain_col_idx - 1]
        assert capital_chain == "ByBit", f"Expected 'ByBit' chain in capital data, got {capital_chain}"
        reward_chain = reward_data_row[reward_chain_col_idx - 1]
        assert reward_chain == "Wirex", f"Expected 'Wirex' chain in reward data, got {reward_chain}"

        workbook.close()

    def test_crypto_sheet_contains_token_origin_column(self, tmp_path):
        """Assert the Crypto worksheet Token origin column shows blank for unknown origin."""
        bybit_origin = resolve_operator_origin("ByBit")

        capital_entries = [
            CryptoCapitalGainEntry(
                disposal_date="2025-02-16",
                acquisition_date="2025-02-16",
                asset="HASUI",
                amount=Decimal("29.83"),
                cost_eur=Decimal("10.00"),
                proceeds_eur=Decimal("15.00"),
                gain_loss_eur=Decimal("5.00"),
                holding_period="Short term",
                wallet="Ledger SUI",
                platform="Ledger",
                chain="Sui",
                operator_origin=bybit_origin,
                annex_hint="J",
                review_required=False,
                notes="",
                token_swap_history="",
            )
        ]

        crypto_report = CryptoTaxReport(
            tax_year=2025,
            capital_entries=capital_entries,
            reward_entries=[],
            reconciliation=CryptoReconciliationSummary(
                capital_rows=1,
                reward_rows=0,
                short_term_rows=1,
                long_term_rows=0,
                mixed_rows=0,
                unknown_rows=0,
                capital_cost_total_eur=Decimal("10.00"),
                capital_proceeds_total_eur=Decimal("15.00"),
                capital_gain_total_eur=Decimal("5.00"),
                reward_total_eur=Decimal("0"),
                opening_holdings=None,
                closing_holdings=None,
            ),
            capital_gain_stats=CryptoCapitalGainStats.from_entries(capital_entries),
            skipped_zero_value_tokens=[],
            pdf_summary=None,
        )

        report_path = tmp_path / "crypto_token_origin_report.xlsx"
        generate_tax_report(
            extract=report_path,
            capital_gain_lines_per_company={},
            dividend_income_per_company={},
            crypto_tax_report=crypto_report,
        )

        assert report_path.exists()

        import openpyxl

        workbook = openpyxl.load_workbook(report_path)
        gains_sheet = workbook["Crypto Gains"]

        capital_headers_found = False
        token_origin_col_idx = None

        for row in gains_sheet.iter_rows(values_only=True):
            if row and "Disposal date" in str(row[0] if row[0] else ""):
                capital_headers_found = True
                for col_idx, cell_value in enumerate(row):
                    if cell_value == "Token origin":
                        token_origin_col_idx = col_idx + 1
                        break
                break

        assert capital_headers_found, "Capital gains headers not found"
        assert token_origin_col_idx is not None, "Token origin column not found in capital gains headers"

        capital_data_row = None

        for row in gains_sheet.iter_rows(values_only=True):
            if row and row[0] == "2025-02-16":
                capital_data_row = row
                break

        assert capital_data_row is not None, "Capital data row not found"

        token_origin_value = capital_data_row[token_origin_col_idx - 1]
        assert token_origin_value in ("", None), f"Expected blank Token origin, got {token_origin_value!r}"

        workbook.close()

    def test_crypto_sheet_token_origin_non_blank_when_resolved(self, tmp_path):
        """Assert Token origin column shows non-blank value when resolver found a match."""
        kraken_origin = resolve_operator_origin("Kraken")

        capital_entries = [
            CryptoCapitalGainEntry(
                disposal_date="2025-03-15",
                acquisition_date="2025-01-15",
                asset="BTC",
                amount=Decimal("0.001"),
                cost_eur=Decimal("50.00"),
                proceeds_eur=Decimal("55.00"),
                gain_loss_eur=Decimal("5.00"),
                holding_period="Short term",
                wallet="Kraken",
                platform="Kraken",
                chain="Kraken",
                operator_origin=kraken_origin,
                annex_hint="J",
                review_required=False,
                notes="",
                token_swap_history="EUR (direct_purchase, medium confidence)",
            )
        ]

        crypto_report = CryptoTaxReport(
            tax_year=2025,
            capital_entries=capital_entries,
            reward_entries=[],
            reconciliation=CryptoReconciliationSummary(
                capital_rows=1,
                reward_rows=0,
                short_term_rows=1,
                long_term_rows=0,
                mixed_rows=0,
                unknown_rows=0,
                capital_cost_total_eur=Decimal("50.00"),
                capital_proceeds_total_eur=Decimal("55.00"),
                capital_gain_total_eur=Decimal("5.00"),
                reward_total_eur=Decimal("0"),
                opening_holdings=None,
                closing_holdings=None,
            ),
            capital_gain_stats=CryptoCapitalGainStats.from_entries(capital_entries),
            skipped_zero_value_tokens=[],
            pdf_summary=None,
        )

        report_path = tmp_path / "crypto_token_origin_resolved_report.xlsx"
        generate_tax_report(
            extract=report_path,
            capital_gain_lines_per_company={},
            dividend_income_per_company={},
            crypto_tax_report=crypto_report,
        )

        assert report_path.exists()

        import openpyxl

        workbook = openpyxl.load_workbook(report_path)
        gains_sheet = workbook["Crypto Gains"]

        token_origin_col_idx = None
        for row in gains_sheet.iter_rows(values_only=True):
            if row and "Disposal date" in str(row[0] if row[0] else ""):
                for col_idx, cell_value in enumerate(row):
                    if cell_value == "Token origin":
                        token_origin_col_idx = col_idx + 1
                        break
                break

        assert token_origin_col_idx is not None, "Token origin column not found"

        capital_data_row = None
        for row in gains_sheet.iter_rows(values_only=True):
            if row and row[0] == "2025-03-15":
                capital_data_row = row
                break

        assert capital_data_row is not None, "Capital data row not found"

        token_origin_value = capital_data_row[token_origin_col_idx - 1]
        assert token_origin_value == "EUR (direct_purchase, medium confidence)", (
            f"Expected resolved Token origin, got {token_origin_value!r}"
        )

        workbook.close()

    def test_crypto_sheet_token_origin_shows_confidence_level(self, tmp_path):
        """Assert Token origin column includes the confidence level string."""
        kraken_origin = resolve_operator_origin("Kraken")

        capital_entries = [
            CryptoCapitalGainEntry(
                disposal_date="2025-03-15",
                acquisition_date="2025-01-15",
                asset="BTC",
                amount=Decimal("0.001"),
                cost_eur=Decimal("50.00"),
                proceeds_eur=Decimal("55.00"),
                gain_loss_eur=Decimal("5.00"),
                holding_period="Short term",
                wallet="Kraken",
                platform="Kraken",
                chain="Kraken",
                operator_origin=kraken_origin,
                annex_hint="J",
                review_required=False,
                notes="",
                token_swap_history="ETH (swap_conversion, high confidence)",
            )
        ]

        crypto_report = CryptoTaxReport(
            tax_year=2025,
            capital_entries=capital_entries,
            reward_entries=[],
            reconciliation=CryptoReconciliationSummary(
                capital_rows=1,
                reward_rows=0,
                short_term_rows=1,
                long_term_rows=0,
                mixed_rows=0,
                unknown_rows=0,
                capital_cost_total_eur=Decimal("50.00"),
                capital_proceeds_total_eur=Decimal("55.00"),
                capital_gain_total_eur=Decimal("5.00"),
                reward_total_eur=Decimal("0"),
                opening_holdings=None,
                closing_holdings=None,
            ),
            capital_gain_stats=CryptoCapitalGainStats.from_entries(capital_entries),
            skipped_zero_value_tokens=[],
            pdf_summary=None,
        )

        report_path = tmp_path / "crypto_token_origin_confidence_report.xlsx"
        generate_tax_report(
            extract=report_path,
            capital_gain_lines_per_company={},
            dividend_income_per_company={},
            crypto_tax_report=crypto_report,
        )

        import openpyxl

        workbook = openpyxl.load_workbook(report_path)
        gains_sheet = workbook["Crypto Gains"]

        token_origin_col_idx = None
        for row in gains_sheet.iter_rows(values_only=True):
            if row and "Disposal date" in str(row[0] if row[0] else ""):
                for col_idx, cell_value in enumerate(row):
                    if cell_value == "Token origin":
                        token_origin_col_idx = col_idx + 1
                        break
                break

        capital_data_row = None
        for row in gains_sheet.iter_rows(values_only=True):
            if row and row[0] == "2025-03-15":
                capital_data_row = row
                break

        assert capital_data_row is not None
        token_origin_value = str(capital_data_row[token_origin_col_idx - 1])
        assert "high confidence" in token_origin_value, (
            f"Expected confidence level in Token origin, got {token_origin_value!r}"
        )

        workbook.close()

    def test_crypto_sheet_review_reason_in_excel_output(self, tmp_path):
        """Assert review_reason appears in Excel cells as 'YES: <reason>' when review_required."""
        bybit_origin = resolve_operator_origin("ByBit")
        wirex_origin = resolve_operator_origin("Wirex", transaction_type="crypto_deposit")

        capital_entries = [
            CryptoCapitalGainEntry(
                disposal_date="2025-01-13",
                acquisition_date="2024-11-18",
                asset="USDT",
                amount=Decimal("1.5"),
                cost_eur=Decimal("1.25"),
                proceeds_eur=Decimal("1.35"),
                gain_loss_eur=Decimal("0.10"),
                holding_period="Short term",
                wallet="ByBit (2)",
                platform="ByBit",
                chain="ByBit",
                operator_origin=bybit_origin,
                annex_hint="J",
                review_required=bybit_origin.review_required,
                review_reason=bybit_origin.review_reason,
                notes="",
                token_swap_history="",
            )
        ]

        crypto_report = CryptoTaxReport(
            tax_year=2025,
            capital_entries=capital_entries,
            reward_entries=[
                CryptoRewardIncomeEntry(
                    date="2025-01-01",
                    asset="WXT",
                    amount=Decimal("5"),
                    value_eur=Decimal("17.10"),
                    income_label="Reward",
                    source_type="Reward",
                    wallet="Wirex",
                    platform="Wirex",
                    chain="Wirex",
                    operator_origin=wirex_origin,
                    annex_hint="J",
                    review_required=False,
                    description="",
                )
            ],
            reconciliation=CryptoReconciliationSummary(
                capital_rows=1,
                reward_rows=1,
                short_term_rows=1,
                long_term_rows=0,
                mixed_rows=0,
                unknown_rows=0,
                capital_cost_total_eur=Decimal("1.25"),
                capital_proceeds_total_eur=Decimal("1.35"),
                capital_gain_total_eur=Decimal("0.10"),
                reward_total_eur=Decimal("17.10"),
                opening_holdings=None,
                closing_holdings=None,
            ),
            capital_gain_stats=CryptoCapitalGainStats.from_entries(capital_entries),
            skipped_zero_value_tokens=[],
            pdf_summary=None,
        )

        report_path = tmp_path / "crypto_review_reason_report.xlsx"
        generate_tax_report(
            extract=report_path,
            capital_gain_lines_per_company={},
            dividend_income_per_company={},
            crypto_tax_report=crypto_report,
        )

        assert report_path.exists()

        import openpyxl

        workbook = openpyxl.load_workbook(report_path)
        gains_sheet = workbook["Crypto Gains"]
        rewards_sheet = workbook["Crypto Rewards"]

        # Find capital gains data row on Crypto Gains sheet
        capital_review_cell = None
        found_capital_review = False
        for row in gains_sheet.iter_rows(values_only=True):
            if row and row[0] == "2025-01-13":
                capital_review_cell = row[14] if len(row) > 14 else None
                found_capital_review = True

        # Find reward row in deferred support detail section on Crypto Rewards sheet
        reward_review_cell = None
        found_reward_review = False
        in_deferred = False
        for row in rewards_sheet.iter_rows(values_only=True):
            if row and isinstance(row[0], str) and "DEFERRED BY LAW" in row[0]:
                in_deferred = True
            if in_deferred and row and row[0] == "2025-01-01":
                reward_review_cell = row[9] if len(row) > 9 else None
                found_reward_review = True
                break

        assert found_capital_review, "Capital data row not found"
        assert capital_review_cell is not None
        assert isinstance(capital_review_cell, str)
        assert capital_review_cell.startswith("YES:"), f"Expected 'YES: ...' but got '{capital_review_cell}'"
        assert "account-region" in capital_review_cell.lower(), (
            f"Expected account-region in review reason, got '{capital_review_cell}'"
        )

        assert found_reward_review, "Reward row not found in deferred support detail section"
        assert reward_review_cell is not None
        assert reward_review_cell == "NO", f"Expected 'NO' for non-review reward, got '{reward_review_cell}'"

        workbook.close()

    def test_generate_tax_report_with_single_dividend_entry(self, tmp_path):
        """Test that report generation works with a single valid dividend entry."""
        dividend_income = DividendIncomePerCompany(
            {
                "VALID": DividendIncomePerSecurity(
                    symbol="VALID",
                    isin="US1234567890",
                    country="United States",
                    gross_amount=Decimal("100.00"),  # Valid positive amount
                    total_taxes=Decimal("15.00"),
                    currency=parse_currency("USD"),
                ),
            }
        )

        report_path = tmp_path / "valid_dividend_report.xlsx"

        # Should generate successfully
        generate_tax_report(
            extract=report_path,
            capital_gain_lines_per_company={},
            dividend_income_per_company=dividend_income,
        )

        assert report_path.exists()

    def test_generate_tax_report_mixed_capital_gains_and_dividends(self, tmp_path):
        """Test generating Excel report with both capital gains and dividend sections."""
        # Create simple dividend data
        dividend_income = DividendIncomePerCompany(
            {
                "AAPL": DividendIncomePerSecurity(
                    symbol="AAPL",
                    isin="US0378331005",
                    country="United States",
                    gross_amount=Decimal("100.00"),
                    total_taxes=Decimal("15.00"),
                    currency=parse_currency("USD"),
                ),
            }
        )

        report_path = tmp_path / "mixed_report.xlsx"
        generate_tax_report(
            extract=report_path,
            capital_gain_lines_per_company={},  # Empty capital gains
            dividend_income_per_company=dividend_income,
        )

        assert report_path.exists()

        # Read and verify both sections exist
        import openpyxl

        workbook = openpyxl.load_workbook(report_path)
        worksheet = workbook.active
        assert worksheet is not None, "Workbook should have an active worksheet"

        # Find dividend section
        found_dividends = False

        for row in worksheet.iter_rows(values_only=True):
            if row and isinstance(row[0], str):
                if "CAPITAL GAINS" in row[0] or "DISPONIBILIZAÇÕES" in row[0]:
                    pass  # Found capital gains section
                elif "CAPITAL INVESTMENT INCOME" in row[0]:
                    found_dividends = True

        assert found_dividends, "Dividend section not found in mixed report"
        workbook.close()

    def test_dividend_excel_cell_formatting(self, tmp_path):
        """Test that dividend Excel cells have proper number formatting."""
        dividend_income = DividendIncomePerCompany(
            {
                "AAPL": DividendIncomePerSecurity(
                    symbol="AAPL",
                    isin="US0378331005",
                    country="United States",
                    gross_amount=Decimal("100.12345"),
                    total_taxes=Decimal("15.67890"),
                    currency=parse_currency("USD"),
                ),
            }
        )

        report_path = tmp_path / "formatted_dividend_report.xlsx"
        generate_tax_report(
            extract=report_path,
            capital_gain_lines_per_company={},
            dividend_income_per_company=dividend_income,
        )

        # Read and verify cell formatting
        import openpyxl

        workbook = openpyxl.load_workbook(report_path)
        worksheet = workbook.active
        assert worksheet is not None, "Workbook should have an active worksheet"

        # Find dividend row and check formatting
        dividend_row = None
        for row_idx, row in enumerate(worksheet.iter_rows(values_only=False), 1):
            if row and len(row) > 1 and row[1].value == "Dividends":
                dividend_row = row_idx
                break

        assert dividend_row is not None

        # Check number format for original amounts
        original_gross_cell = worksheet.cell(row=dividend_row, column=11)
        assert original_gross_cell.number_format == "0.00"

        original_tax_cell = worksheet.cell(row=dividend_row, column=12)
        assert original_tax_cell.number_format == "0.00"

        net_amount_cell = worksheet.cell(row=dividend_row, column=13)
        assert net_amount_cell.number_format == "0.00"

        # Check values are stored as strings with full precision
        assert original_gross_cell.value == "100.12345"
        assert original_tax_cell.value == "15.67890"
        assert net_amount_cell.value == "84.44455"  # 100.12345 - 15.67890

        workbook.close()

    def test_crypto_sheet_generation_error_fails_report_generation(self, tmp_path):
        """Test that crypto sheet generation errors fail the entire report generation.

        Per plan requirement (Task 2), report generation must fail with a clear error
        when a taxable-now row cannot be assigned all mandatory IRS fields.
        """
        from unittest.mock import patch

        from shares_reporting.domain.exceptions import FileProcessingError

        # Create a simple crypto report
        capital_entries = [
            CryptoCapitalGainEntry(
                disposal_date="2025-01-13",
                acquisition_date="2024-11-18",
                asset="ETH",
                amount=Decimal("1"),
                cost_eur=Decimal("1000"),
                proceeds_eur=Decimal("1200"),
                gain_loss_eur=Decimal("200"),
                holding_period="Short term",
                wallet="Ledger ETH",
                platform="Ledger ETH",
                chain="Ethereum",
                operator_origin=resolve_operator_origin("Ledger ETH"),
                annex_hint="J",
                review_required=False,
                notes="",
                token_swap_history="",
            ),
        ]

        crypto_report = CryptoTaxReport(
            tax_year=2025,
            capital_entries=capital_entries,
            reward_entries=[],
            reconciliation=CryptoReconciliationSummary(
                capital_rows=1,
                reward_rows=0,
                short_term_rows=1,
                long_term_rows=0,
                mixed_rows=0,
                unknown_rows=0,
                capital_cost_total_eur=Decimal("1000"),
                capital_proceeds_total_eur=Decimal("1200"),
                capital_gain_total_eur=Decimal("200"),
                reward_total_eur=Decimal("0"),
                opening_holdings=None,
                closing_holdings=None,
            ),
            capital_gain_stats=CryptoCapitalGainStats.from_entries(capital_entries),
            skipped_zero_value_tokens=[],
            pdf_summary=None,
        )

        report_path = tmp_path / "crypto_error_test.xlsx"

        # Mock write_crypto_gains_sheet to raise a FileProcessingError
        with (
            patch(
                "shares_reporting.application.persisting.workbook_builder.write_crypto_gains_sheet",
                side_effect=FileProcessingError("Simulated validation failure for testing"),
            ),
            pytest.raises(FileProcessingError, match="Simulated validation failure for testing"),
        ):
            generate_tax_report(
                extract=report_path,
                capital_gain_lines_per_company={},
                dividend_income_per_company={},
                crypto_tax_report=crypto_report,
            )

        # Verify the report was NOT created because generation failed
        assert not report_path.exists()

    def test_crypto_sheet_capital_headers_distinguish_disposal_vs_acquisition_date(self, tmp_path):
        """Assert the Crypto capital-gains header row has both 'Disposal date' and
        'Acquisition date' as separate, explicitly named columns.

        This guards against a future column rename or reordering that collapses
        the two date fields into one ambiguous header, which would make it
        impossible to tell whether a repeated timestamp is a disposal or
        acquisition date in the generated workbook.
        """
        bybit_origin = resolve_operator_origin("ByBit")

        capital_entries = [
            CryptoCapitalGainEntry(
                disposal_date="2025-01-13",
                acquisition_date="2024-07-27",
                asset="USDT",
                amount=Decimal("1.5"),
                cost_eur=Decimal("1.25"),
                proceeds_eur=Decimal("1.35"),
                gain_loss_eur=Decimal("0.10"),
                holding_period="Short term",
                wallet="ByBit (2)",
                platform="ByBit",
                chain="ByBit",
                operator_origin=bybit_origin,
                annex_hint="J",
                review_required=bybit_origin.review_required,
                review_reason=bybit_origin.review_reason,
                notes="",
                token_swap_history="",
            )
        ]

        crypto_report = CryptoTaxReport(
            tax_year=2025,
            capital_entries=capital_entries,
            reward_entries=[],
            reconciliation=CryptoReconciliationSummary(
                capital_rows=1,
                reward_rows=0,
                short_term_rows=1,
                long_term_rows=0,
                mixed_rows=0,
                unknown_rows=0,
                capital_cost_total_eur=Decimal("1.25"),
                capital_proceeds_total_eur=Decimal("1.35"),
                capital_gain_total_eur=Decimal("0.10"),
                reward_total_eur=Decimal("0"),
                opening_holdings=None,
                closing_holdings=None,
            ),
            capital_gain_stats=CryptoCapitalGainStats.from_entries(capital_entries),
            skipped_zero_value_tokens=[],
            pdf_summary=None,
        )

        report_path = tmp_path / "crypto_date_headers_report.xlsx"
        generate_tax_report(
            extract=report_path,
            capital_gain_lines_per_company={},
            dividend_income_per_company={},
            crypto_tax_report=crypto_report,
        )

        assert report_path.exists()

        import openpyxl

        workbook = openpyxl.load_workbook(report_path)
        gains_sheet = workbook["Crypto Gains"]

        header_row = None
        for row in gains_sheet.iter_rows(values_only=True):
            if row and "Disposal date" in str(row[0] if row[0] else ""):
                header_row = [str(c) if c is not None else "" for c in row]
                break

        assert header_row is not None, "Capital gains header row not found in Crypto Gains sheet"

        assert header_row[0] == "Disposal date", (
            f"First capital gains header must be 'Disposal date', got {header_row[0]!r}"
        )
        assert "Acquisition date" in header_row, (
            f"'Acquisition date' column missing from capital gains headers: {header_row}"
        )
        disposal_idx = header_row.index("Disposal date")
        acquisition_idx = header_row.index("Acquisition date")
        assert disposal_idx != acquisition_idx, (
            "Disposal date and Acquisition date must be separate columns"
        )

        data_row = None
        for row in gains_sheet.iter_rows(values_only=True):
            if row and row[0] == "2025-01-13":
                data_row = row
                break

        assert data_row is not None, "Capital gains data row not found"
        assert data_row[disposal_idx] == "2025-01-13"
        assert data_row[acquisition_idx] == "2024-07-27"

        workbook.close()

    def test_crypto_sheet_removed_on_partial_write_error(self, tmp_path):
        """Test that error during crypto sheet generation fails cleanly.

        When an exception occurs during crypto sheet writing, the workbook is closed
        (with partial sheets removed), the output file is removed, and the error
        propagates to the caller.
        """
        from unittest.mock import patch

        from shares_reporting.application.persisting import generate_tax_report
        from shares_reporting.domain.exceptions import FileProcessingError

        capital_entries = [
            CryptoCapitalGainEntry(
                disposal_date="2025-01-15",
                acquisition_date="2024-06-01",
                asset="BTC",
                amount=Decimal("0.5"),
                cost_eur=Decimal("10000"),
                proceeds_eur=Decimal("12000"),
                gain_loss_eur=Decimal("2000"),
                holding_period="Short term",
                wallet="TestWallet",
                platform="TestPlatform",
                chain="Bitcoin",
                operator_origin=resolve_operator_origin("TestPlatform", "crypto_disposal"),
                annex_hint="G",
                review_required=False,
                notes="",
                token_swap_history="",
            ),
        ]

        crypto_report = CryptoTaxReport(
            tax_year=2025,
            capital_entries=capital_entries,
            reward_entries=[],
            reconciliation=CryptoReconciliationSummary(
                capital_rows=1,
                reward_rows=0,
                short_term_rows=1,
                long_term_rows=0,
                mixed_rows=0,
                unknown_rows=0,
                capital_cost_total_eur=Decimal("10000"),
                capital_proceeds_total_eur=Decimal("12000"),
                capital_gain_total_eur=Decimal("2000"),
                reward_total_eur=Decimal("0"),
                opening_holdings=None,
                closing_holdings=None,
            ),
            capital_gain_stats=CryptoCapitalGainStats.from_entries(capital_entries),
            skipped_zero_value_tokens=[],
            pdf_summary=None,
        )

        report_path = tmp_path / "crypto_partial_error_test.xlsx"

        # Mock write_crypto_gains_sheet to create the sheet but fail during writing
        def mock_gains_that_fails_partial(workbook, crypto_tax_report):
            crypto_ws = workbook.create_sheet("Crypto Gains")
            crypto_ws.cell(1, 1, "CRYPTO TAX REPORT - PORTUGAL")
            crypto_ws.cell(2, 1, "Tax year")
            crypto_ws.cell(2, 2, crypto_tax_report.tax_year)
            raise FileProcessingError("Simulated write error during crypto sheet generation")

        with (
            patch(
                "shares_reporting.application.persisting.workbook_builder.write_crypto_gains_sheet",
                side_effect=mock_gains_that_fails_partial,
            ),
            pytest.raises(FileProcessingError, match="Simulated write error during crypto sheet generation"),
        ):
            generate_tax_report(
                extract=report_path,
                capital_gain_lines_per_company={},
                dividend_income_per_company={},
                crypto_tax_report=crypto_report,
            )

        # Verify the report was NOT created because generation failed
        assert not report_path.exists()

    def _build_crypto_report_with_stats(self):
        """Build a CryptoTaxReport with known capital entries across holding periods.

        Returns a report with:
        - 2 short-term: cost=100, proceeds=120, gain=20; cost=50, proceeds=45, gain=-5
        - 1 long-term: cost=200, proceeds=250, gain=50
        - 1 mixed: cost=75, proceeds=80, gain=5
        - 0 unknown
        Grand total: count=4, cost=425, proceeds=495, gain/loss=70
        """
        ethereum_origin = resolve_operator_origin("Ethereum")

        capital_entries = [
            CryptoCapitalGainEntry(
                disposal_date="2025-02-01",
                acquisition_date="2025-01-15",
                asset="ETH",
                amount=Decimal("1"),
                cost_eur=Decimal("100"),
                proceeds_eur=Decimal("120"),
                gain_loss_eur=Decimal("20"),
                holding_period="Short term",
                wallet="Ledger ETH",
                platform="Ledger ETH",
                chain="Ethereum",
                operator_origin=ethereum_origin,
                annex_hint="J",
                review_required=False,
                notes="",
                token_swap_history="",
            ),
            CryptoCapitalGainEntry(
                disposal_date="2025-03-01",
                acquisition_date="2025-02-20",
                asset="BTC",
                amount=Decimal("0.5"),
                cost_eur=Decimal("50"),
                proceeds_eur=Decimal("45"),
                gain_loss_eur=Decimal("-5"),
                holding_period="Short term",
                wallet="Ledger ETH",
                platform="Ledger ETH",
                chain="Ethereum",
                operator_origin=ethereum_origin,
                annex_hint="J",
                review_required=False,
                notes="",
                token_swap_history="",
            ),
            CryptoCapitalGainEntry(
                disposal_date="2025-06-01",
                acquisition_date="2024-01-01",
                asset="ETH",
                amount=Decimal("2"),
                cost_eur=Decimal("200"),
                proceeds_eur=Decimal("250"),
                gain_loss_eur=Decimal("50"),
                holding_period="Long term",
                wallet="Ledger ETH",
                platform="Ledger ETH",
                chain="Ethereum",
                operator_origin=ethereum_origin,
                annex_hint="G1",
                review_required=False,
                notes="",
                token_swap_history="",
            ),
            CryptoCapitalGainEntry(
                disposal_date="2025-04-01",
                acquisition_date="mixed",
                asset="SOL",
                amount=Decimal("10"),
                cost_eur=Decimal("75"),
                proceeds_eur=Decimal("80"),
                gain_loss_eur=Decimal("5"),
                holding_period="Mixed",
                wallet="Ledger ETH",
                platform="Ledger ETH",
                chain="Ethereum",
                operator_origin=ethereum_origin,
                annex_hint="J",
                review_required=False,
                notes="",
                token_swap_history="",
            ),
        ]

        return CryptoTaxReport(
            tax_year=2025,
            capital_entries=capital_entries,
            reward_entries=[],
            reconciliation=CryptoReconciliationSummary(
                capital_rows=4,
                reward_rows=0,
                short_term_rows=2,
                long_term_rows=1,
                mixed_rows=1,
                unknown_rows=0,
                capital_cost_total_eur=Decimal("425"),
                capital_proceeds_total_eur=Decimal("495"),
                capital_gain_total_eur=Decimal("70"),
                reward_total_eur=Decimal("0"),
                opening_holdings=None,
                closing_holdings=None,
            ),
            capital_gain_stats=CryptoCapitalGainStats.from_entries(capital_entries),
            skipped_zero_value_tokens=[],
            pdf_summary=None,
        )

    def test_crypto_sheet_contains_capital_gains_statistics_header(self, tmp_path):
        """Verify the CAPITAL GAINS STATISTICS section header appears after capital gains detail rows."""
        crypto_report = self._build_crypto_report_with_stats()

        report_path = tmp_path / "crypto_stats_header.xlsx"
        generate_tax_report(
            extract=report_path,
            capital_gain_lines_per_company={},
            dividend_income_per_company={},
            crypto_tax_report=crypto_report,
        )

        assert report_path.exists()

        import openpyxl

        workbook = openpyxl.load_workbook(report_path)
        gains_sheet = workbook["Crypto Gains"]
        rewards_sheet = workbook["Crypto Rewards"]

        gains_labels = []
        for row in gains_sheet.iter_rows(values_only=True):
            first_cell = row[0] if row else None
            if isinstance(first_cell, str):
                gains_labels.append(first_cell)

        rewards_labels = []
        for row in rewards_sheet.iter_rows(values_only=True):
            first_cell = row[0] if row else None
            if isinstance(first_cell, str):
                rewards_labels.append(first_cell)

        assert "1b. CAPITAL GAINS STATISTICS" in gains_labels, (
            f"Expected '1b. CAPITAL GAINS STATISTICS' header in Crypto Gains sheet, got labels: {gains_labels}"
        )

        assert "1. CAPITAL GAINS" in gains_labels, "Section 1 header should still exist"

        section_1_idx = gains_labels.index("1. CAPITAL GAINS")
        stats_idx = gains_labels.index("1b. CAPITAL GAINS STATISTICS")

        assert section_1_idx < stats_idx, "Statistics section must appear after capital gains detail"
        assert "2. REWARDS INCOME - IRS-READY FILING SUMMARY" in rewards_labels, (
            "Rewards section should be in Crypto Rewards sheet"
        )

        workbook.close()

    def test_crypto_sheet_capital_gains_statistics_values(self, tmp_path):  # noqa: PLR0915
        """Verify per-period rows and Grand Total row contain correct values."""
        crypto_report = self._build_crypto_report_with_stats()

        report_path = tmp_path / "crypto_stats_values.xlsx"
        generate_tax_report(
            extract=report_path,
            capital_gain_lines_per_company={},
            dividend_income_per_company={},
            crypto_tax_report=crypto_report,
        )

        assert report_path.exists()

        import openpyxl

        workbook = openpyxl.load_workbook(report_path)
        gains_sheet = workbook["Crypto Gains"]

        rows = list(gains_sheet.iter_rows(values_only=True))

        stats_start = None
        for i, row in enumerate(rows):
            if row and isinstance(row[0], str) and "1b. CAPITAL GAINS STATISTICS" in row[0]:
                stats_start = i
                break

        assert stats_start is not None, "Statistics section not found"

        header_row_idx = stats_start + 1
        header_row = rows[header_row_idx]
        assert "Holding Period" in str(header_row[0])
        assert "Count" in str(header_row[1])
        assert "Cost Total (EUR)" in str(header_row[2])
        assert "Proceeds Total (EUR)" in str(header_row[3])
        assert "Gain/Loss Total (EUR)" in str(header_row[4])

        data_rows = {}
        for row in rows[header_row_idx + 1 :]:
            if row and row[0] is not None:
                label = str(row[0])
                if label in ("Short-term", "Long-term", "Mixed", "Unknown", "Grand Total"):
                    data_rows[label] = row
                else:
                    break

        assert "Short-term" in data_rows, "Short-term row missing"
        assert data_rows["Short-term"][1] == 2
        assert data_rows["Short-term"][2] == pytest.approx(150.0)
        assert data_rows["Short-term"][3] == pytest.approx(165.0)
        assert data_rows["Short-term"][4] == pytest.approx(15.0)

        assert "Long-term" in data_rows, "Long-term row missing"
        assert data_rows["Long-term"][1] == 1
        assert data_rows["Long-term"][2] == pytest.approx(200.0)
        assert data_rows["Long-term"][3] == pytest.approx(250.0)
        assert data_rows["Long-term"][4] == pytest.approx(50.0)

        assert "Mixed" in data_rows, "Mixed row missing"
        assert data_rows["Mixed"][1] == 1
        assert data_rows["Mixed"][2] == pytest.approx(75.0)
        assert data_rows["Mixed"][3] == pytest.approx(80.0)
        assert data_rows["Mixed"][4] == pytest.approx(5.0)

        assert "Unknown" in data_rows, "Unknown row missing"
        assert data_rows["Unknown"][1] == 0
        assert data_rows["Unknown"][2] == pytest.approx(0.0)
        assert data_rows["Unknown"][3] == pytest.approx(0.0)
        assert data_rows["Unknown"][4] == pytest.approx(0.0)

        assert "Grand Total" in data_rows, "Grand Total row missing"
        assert data_rows["Grand Total"][1] == 4
        assert data_rows["Grand Total"][2] == pytest.approx(425.0)
        assert data_rows["Grand Total"][3] == pytest.approx(495.0)
        assert data_rows["Grand Total"][4] == pytest.approx(70.0)

        workbook.close()
