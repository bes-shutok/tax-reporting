"""Tests for workbook builder orchestrator."""

from __future__ import annotations

from decimal import Decimal
from unittest.mock import patch

import pytest

from tax_reporting.application.crypto_reporting import CryptoTaxReport
from tax_reporting.domain.exceptions import FileProcessingError


def _make_crypto_tax_report_with_invalid_country():
    """Build a CryptoTaxReport whose taxable rewards will fail country validation."""
    from tax_reporting.application.crypto_reporting import (
        CryptoCapitalGainStats,
        CryptoReconciliationSummary,
        CryptoRewardIncomeEntry,
        HoldingsSnapshot,
        OperatorOrigin,
        RewardTaxClassification,
    )

    operator = OperatorOrigin(
        platform="TestOp",
        service_scope="global",
        operator_entity="TestOp Ltd",
        operator_country="XX_INVALID",
        source_url="https://example.com",
        source_checked_on="2025-01-01",
        confidence="high",
        review_required=False,
        platform_review_required=False,
    )
    reward = CryptoRewardIncomeEntry(
        date="2025-01-15",
        asset="EUR",
        amount=Decimal("10"),
        value_eur=Decimal("10"),
        income_label="Lending interest",
        source_type="fiat_deposit",
        wallet="TestWallet",
        platform="TestOp",
        chain="TestChain",
        operator_origin=operator,
        annex_hint="E",
        review_required=False,
        description="Test reward",
        tax_classification=RewardTaxClassification.TAXABLE_NOW,
        foreign_tax_eur=Decimal("0"),
    )
    return CryptoTaxReport(
        tax_year=2025,
        capital_entries=[],
        reward_entries=[reward],
        reconciliation=CryptoReconciliationSummary(
            capital_rows=0,
            reward_rows=1,
            short_term_rows=0,
            long_term_rows=0,
            mixed_rows=0,
            unknown_rows=0,
            capital_cost_total_eur=Decimal("0"),
            capital_proceeds_total_eur=Decimal("0"),
            capital_gain_total_eur=Decimal("0"),
            reward_total_eur=Decimal("10"),
            opening_holdings=HoldingsSnapshot(asset_rows=0, total_cost_eur=Decimal("0"), total_value_eur=Decimal("0")),
            closing_holdings=HoldingsSnapshot(asset_rows=0, total_cost_eur=Decimal("0"), total_value_eur=Decimal("0")),
        ),
        capital_gain_stats=CryptoCapitalGainStats.from_entries([]),
    )


@pytest.mark.unit
class TestGenerateTaxReportErrorPaths:
    """Error-path coverage for generate_tax_report."""

    def test_invalid_taxable_country_propagates_file_processing_error(self, tmp_path):
        from tax_reporting.application.persisting.workbook_builder import generate_tax_report

        output = tmp_path / "report.xlsx"
        report = _make_crypto_tax_report_with_invalid_country()

        with pytest.raises(FileProcessingError, match="valid Tabela X country"):
            generate_tax_report(str(output), {}, crypto_tax_report=report)

    def test_invalid_taxable_country_cleans_up_output_file(self, tmp_path):
        output = tmp_path / "report.xlsx"
        report = _make_crypto_tax_report_with_invalid_country()

        from tax_reporting.application.persisting.workbook_builder import generate_tax_report

        with pytest.raises(FileProcessingError):
            generate_tax_report(str(output), {}, crypto_tax_report=report)

        assert not output.exists(), "No stale output file should remain after aggregate error"

    def test_no_crypto_report_returns_false(self, tmp_path):
        from tax_reporting.application.persisting.workbook_builder import generate_tax_report

        output = tmp_path / "report.xlsx"
        result = generate_tax_report(str(output), {})
        assert result is False


def _make_valid_crypto_tax_report() -> CryptoTaxReport:
    """Build a valid CryptoTaxReport for happy path testing."""
    from tax_reporting.application.crypto_reporting import (
        CryptoCapitalGainEntry,
        CryptoCapitalGainStats,
        CryptoReconciliationSummary,
        CryptoRewardIncomeEntry,
        HoldingsSnapshot,
        OperatorOrigin,
        RewardTaxClassification,
    )

    operator = OperatorOrigin(
        platform="Kraken",
        service_scope="crypto",
        operator_entity="Kraken",
        operator_country="PT",
        source_url="https://example.com",
        source_checked_on="2025-01-01",
        confidence="high",
        review_required=False,
        platform_review_required=False,
    )

    capital_entry = CryptoCapitalGainEntry(
        disposal_date="2025-01-15",
        acquisition_date="2024-01-15",
        asset="BTC",
        amount=Decimal("1"),
        cost_eur=Decimal("1000"),
        proceeds_eur=Decimal("1200"),
        gain_loss_eur=Decimal("200"),
        holding_period="Long term",
        wallet="Kraken",
        platform="Kraken",
        chain="BTC",
        operator_origin=operator,
        annex_hint="J",
        review_required=False,
        notes="",
        review_reason=None,
        token_swap_history="",
    )

    reward_entry = CryptoRewardIncomeEntry(
        date="2025-01-15",
        asset="EUR",
        amount=Decimal("10"),
        value_eur=Decimal("10"),
        income_label="Lending interest",
        source_type="fiat_deposit",
        wallet="Wirex",
        platform="Wirex",
        chain="Ethereum",
        operator_origin=operator,
        annex_hint="E",
        review_required=False,
        description="Test reward",
        tax_classification=RewardTaxClassification.TAXABLE_NOW,
        foreign_tax_eur=Decimal("0"),
    )

    return CryptoTaxReport(
        tax_year=2025,
        capital_entries=[capital_entry],
        reward_entries=[reward_entry],
        reconciliation=CryptoReconciliationSummary(
            capital_rows=1,
            reward_rows=1,
            short_term_rows=0,
            long_term_rows=1,
            mixed_rows=0,
            unknown_rows=0,
            capital_cost_total_eur=Decimal("1000"),
            capital_proceeds_total_eur=Decimal("1200"),
            capital_gain_total_eur=Decimal("200"),
            reward_total_eur=Decimal("10"),
            opening_holdings=HoldingsSnapshot(asset_rows=0, total_cost_eur=Decimal("0"), total_value_eur=Decimal("0")),
            closing_holdings=HoldingsSnapshot(asset_rows=0, total_cost_eur=Decimal("0"), total_value_eur=Decimal("0")),
        ),
        capital_gain_stats=CryptoCapitalGainStats.from_entries([capital_entry]),
    )


@pytest.mark.unit
class TestGenerateTaxReportHappyPath:
    """Happy path coverage for generate_tax_report."""

    def test_generates_all_expected_sheets_with_crypto_data(self, tmp_path):
        from tax_reporting.application.persisting.workbook_builder import generate_tax_report

        output = tmp_path / "report.xlsx"
        report = _make_valid_crypto_tax_report()

        result = generate_tax_report(str(output), {}, crypto_tax_report=report)

        assert result is True
        assert output.exists()

        # Verify all expected sheets exist (Derivatives P&L is gated on
        # separate_derivatives_reporting, which is True in tests/config.ini
        # and the 2025 decision points TOML for PT)
        from openpyxl import load_workbook

        wb = load_workbook(output)
        expected_sheets = {
            "Reporting",
            "Crypto Gains",
            "Derivatives P&L",
            "Crypto Supplementary",
            "Crypto Reconciliation",
            "Loan Activity",
            "Assumptions & Methodology",
        }
        actual_sheets = set(wb.sheetnames)
        assert actual_sheets == expected_sheets, f"Expected sheets {expected_sheets}, got {actual_sheets}"
        wb.close()

    def test_generates_only_reporting_sheet_without_crypto_data(self, tmp_path):
        from tax_reporting.application.persisting.workbook_builder import generate_tax_report

        output = tmp_path / "report.xlsx"
        result = generate_tax_report(str(output), {})

        assert result is False
        assert output.exists()

        # Verify only Reporting sheet exists
        from openpyxl import load_workbook

        wb = load_workbook(output)
        actual_sheets = set(wb.sheetnames)
        assert actual_sheets == {"Reporting"}
        wb.close()

    def test_returns_true_when_crypto_tax_report_provided(self, tmp_path):
        from tax_reporting.application.persisting.workbook_builder import generate_tax_report

        output = tmp_path / "report.xlsx"
        report = _make_valid_crypto_tax_report()

        result = generate_tax_report(str(output), {}, crypto_tax_report=report)

        assert result is True

    def test_creates_output_file_on_success(self, tmp_path):
        from tax_reporting.application.persisting.workbook_builder import generate_tax_report

        output = tmp_path / "report.xlsx"
        report = _make_valid_crypto_tax_report()

        generate_tax_report(str(output), {}, crypto_tax_report=report)

        assert output.exists()
        assert output.stat().st_size > 0


def _build_test_config(*, separate_derivatives_reporting: bool):
    """Build a Config with the requested separate_derivatives_reporting flag."""
    from tax_reporting.domain.jurisdiction import TaxJurisdictionConfig
    from tax_reporting.infrastructure.config import Config
    from tax_reporting.infrastructure.validation import SecurityConfig

    tax_jurisdiction = TaxJurisdictionConfig(
        country="PT",
        fiscal_year=2025,
        exclude_loan_repayment_gains=True,
        zero_basis_review_threshold=Decimal("50"),
        futures_derivatives_taxable=True,
        use_other_gains_report=True,
        separate_derivatives_reporting=separate_derivatives_reporting,
    )
    return Config(
        base="EUR",
        rates=[],
        tax_jurisdiction=tax_jurisdiction,
        security=SecurityConfig(),
    )


@pytest.mark.unit
class TestDerivativesSheetRegistration:
    """Verify the Derivatives P&L sheet is registered only when the flag is on."""

    def test_tab_registered_when_flag_on(self, tmp_path):
        from openpyxl import load_workbook

        from tax_reporting.application.persisting.workbook_builder import generate_tax_report

        output = tmp_path / "report.xlsx"
        report = _make_valid_crypto_tax_report()

        with patch(
            "tax_reporting.application.persisting.workbook_builder.load_configuration_from_file",
            return_value=_build_test_config(separate_derivatives_reporting=True),
        ):
            generate_tax_report(str(output), {}, crypto_tax_report=report)

        wb = load_workbook(output)
        assert "Derivatives P&L" in wb.sheetnames, "Derivatives P&L sheet must exist when flag is on"
        wb.close()

    def test_tab_skipped_when_flag_off(self, tmp_path):
        from openpyxl import load_workbook

        from tax_reporting.application.persisting.workbook_builder import generate_tax_report

        output = tmp_path / "report.xlsx"
        report = _make_valid_crypto_tax_report()

        with patch(
            "tax_reporting.application.persisting.workbook_builder.load_configuration_from_file",
            return_value=_build_test_config(separate_derivatives_reporting=False),
        ):
            generate_tax_report(str(output), {}, crypto_tax_report=report)

        wb = load_workbook(output)
        assert "Derivatives P&L" not in wb.sheetnames, "Derivatives P&L sheet must NOT exist when flag is off"
        wb.close()

