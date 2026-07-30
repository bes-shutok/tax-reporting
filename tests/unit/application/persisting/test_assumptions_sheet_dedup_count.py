"""Tests for the OGR-vs-CG / fee lot dedup run-count suffix on the
Assumptions & Methodology sheet.

Task 7 of the relocate-crypto-warnings plan: a new methodology item documents the
OGR-vs-CG (derivatives) and fee lot dedup decisions and appends a per-run suffix
``[This run: derivatives removed N; fee removed M.]`` when ``CryptoDecisionCounts``
is threaded into the writer. IB-only runs (decision_counts=None) keep the static
description (backward compat).
"""

from __future__ import annotations

import openpyxl
import pytest

from tax_reporting.application.crypto.entities import CryptoDecisionCounts
from tax_reporting.application.persisting.assumptions_sheet import (
    write_assumptions_and_methodology_sheet,
)


@pytest.mark.unit
class TestAssumptionsSheetDedupCount:
    """The OGR-vs-CG lot dedup methodology item carries the per-run dedup suffix."""

    def test_dedup_lines_carry_run_counts(self) -> None:
        """Given CryptoDecisionCounts threaded into the writer, the dedup
        methodology item renders the per-run dedup-count suffix."""
        wb = openpyxl.Workbook()
        write_assumptions_and_methodology_sheet(
            wb,
            decision_counts=CryptoDecisionCounts(
                derivatives_dedup_removed=123,
                fee_dedup_removed=763,
            ),
        )
        ws = wb["Assumptions & Methodology"]

        for row_idx in range(1, 200):
            label = ws.cell(row_idx, 1).value
            if label and "dedup" in str(label).lower():
                description = ws.cell(row_idx, 2).value
                assert (
                    "[This run: derivatives removed 123; fee removed 763.]"
                    in description
                ), f"suffix missing; description was: {description!r}"
                break
        else:
            pytest.fail("dedup methodology item not found")

    def test_dedup_lines_omit_suffix_when_counts_absent(self) -> None:
        """Given decision_counts=None (IB-only run), the dedup item renders its
        static description with NO suffix (backward compat)."""
        wb = openpyxl.Workbook()
        write_assumptions_and_methodology_sheet(wb, decision_counts=None)
        ws = wb["Assumptions & Methodology"]

        for row_idx in range(1, 200):
            label = ws.cell(row_idx, 1).value
            if label and "dedup" in str(label).lower():
                description = ws.cell(row_idx, 2).value
                assert "[This run:" not in description
                # The static description still documents the methodology.
                assert description
                break
        else:
            pytest.fail("dedup methodology item not found")
