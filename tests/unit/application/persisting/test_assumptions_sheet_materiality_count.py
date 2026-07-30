"""Tests for the PT-C-028 run-count suffix on the Assumptions & Methodology sheet.

Task 2 of the relocate-crypto-warnings plan: the W10 sub-1-EUR materiality filter
demotes its aggregate WARNING to INFO and surfaces the per-run dropped/retained
counts as a suffix on the PT-C-028 "Materiality Threshold" methodology item.
"""

from __future__ import annotations

import pytest

import openpyxl

from tax_reporting.application.crypto.entities import CryptoDecisionCounts
from tax_reporting.application.persisting.assumptions_sheet import (
    write_assumptions_and_methodology_sheet,
)


@pytest.mark.unit
class TestAssumptionsSheetMaterialityCount:
    """PT-C-028 Materiality Threshold cell carries the per-run count suffix."""

    def test_ptc028_line_carries_run_count(self) -> None:
        """Given CryptoDecisionCounts threaded into the writer, the PT-C-028
        Materiality Threshold cell text contains the run-count suffix."""
        wb = openpyxl.Workbook()
        write_assumptions_and_methodology_sheet(
            wb,
            decision_counts=CryptoDecisionCounts(
                sub_1_eur_filtered=173,
                sub_1_eur_retained=8,
            ),
        )
        ws = wb["Assumptions & Methodology"]

        for row_idx in range(1, 200):
            if ws.cell(row_idx, 1).value == "Materiality Threshold":
                description = ws.cell(row_idx, 2).value
                assert "[This run: filtered 173 entries, 8 retained.]" in description
                break
        else:
            pytest.fail("Materiality Threshold row not found")

    def test_ptc028_line_omits_suffix_when_counts_absent(self) -> None:
        """Given decision_counts=None (IB-only run), the PT-C-028 line renders
        the static description with NO suffix (backward compat)."""
        wb = openpyxl.Workbook()
        write_assumptions_and_methodology_sheet(wb, decision_counts=None)
        ws = wb["Assumptions & Methodology"]

        for row_idx in range(1, 200):
            if ws.cell(row_idx, 1).value == "Materiality Threshold":
                description = ws.cell(row_idx, 2).value
                assert "[This run:" not in description
                assert "PT-C-028" in description
                break
        else:
            pytest.fail("Materiality Threshold row not found")
