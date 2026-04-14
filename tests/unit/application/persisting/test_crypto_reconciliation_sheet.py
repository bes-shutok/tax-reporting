"""Tests for the crypto reconciliation sheet writer."""

from decimal import Decimal

import openpyxl
import pytest

from shares_reporting.application.crypto_reporting import (
    CapitalGainPeriodStats,
    CryptoCapitalGainStats,
    CryptoReconciliationSummary,
    CryptoSkippedZeroValueToken,
    CryptoTaxReport,
    HoldingsSnapshot,
)
from shares_reporting.application.persisting.crypto_reconciliation_sheet import (
    write_crypto_reconciliation_sheet,
)


def _make_crypto_tax_report(
    reconciliation: CryptoReconciliationSummary | None = None,
    skipped_tokens: list[CryptoSkippedZeroValueToken] | None = None,
) -> CryptoTaxReport:
    empty_stats = CapitalGainPeriodStats(
        count=0, cost_total_eur=Decimal("0"), proceeds_total_eur=Decimal("0"), gain_loss_total_eur=Decimal("0")
    )
    capital_gain_stats = CryptoCapitalGainStats(
        short_term=empty_stats, long_term=empty_stats, mixed=empty_stats, unknown=empty_stats, grand_total=empty_stats
    )
    recon = reconciliation or CryptoReconciliationSummary(
        capital_rows=5,
        reward_rows=3,
        short_term_rows=2,
        long_term_rows=1,
        mixed_rows=1,
        unknown_rows=1,
        capital_cost_total_eur=Decimal("10000"),
        capital_proceeds_total_eur=Decimal("12000"),
        capital_gain_total_eur=Decimal("2000"),
        reward_total_eur=Decimal("500"),
        opening_holdings=None,
        closing_holdings=None,
    )
    return CryptoTaxReport(
        tax_year=2025,
        capital_entries=[],
        reward_entries=[],
        reconciliation=recon,
        capital_gain_stats=capital_gain_stats,
        skipped_zero_value_tokens=skipped_tokens or [],
        pdf_summary=None,
    )


@pytest.mark.unit
class TestCryptoReconciliationSheetName:
    """Tests that the sheet is created with the correct name."""

    def test_sheet_named_crypto_reconciliation(self):
        wb = openpyxl.Workbook()
        report = _make_crypto_tax_report()
        write_crypto_reconciliation_sheet(wb, report)
        assert "Crypto Reconciliation" in wb.sheetnames


@pytest.mark.unit
class TestCryptoReconciliationSection3:
    """Tests for section 3: Reconciliation key-value rows."""

    def test_section_3_title_written(self):
        wb = openpyxl.Workbook()
        report = _make_crypto_tax_report()
        write_crypto_reconciliation_sheet(wb, report)
        ws = wb["Crypto Reconciliation"]
        found = False
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=1):
            if row[0].value == "3. RECONCILIATION":
                found = True
                break
        assert found

    def test_section_3_title_is_bold(self):
        wb = openpyxl.Workbook()
        report = _make_crypto_tax_report()
        write_crypto_reconciliation_sheet(wb, report)
        ws = wb["Crypto Reconciliation"]
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=1):
            if row[0].value == "3. RECONCILIATION":
                assert row[0].font.bold is True
                break

    def test_reconciliation_key_value_pairs(self):
        wb = openpyxl.Workbook()
        report = _make_crypto_tax_report()
        write_crypto_reconciliation_sheet(wb, report)
        ws = wb["Crypto Reconciliation"]
        section_start = None
        for r in range(1, ws.max_row + 1):
            if ws.cell(r, 1).value == "3. RECONCILIATION":
                section_start = r
                break
        assert section_start is not None
        data_start = section_start + 1
        keys = {}
        for r in range(data_start, data_start + 10):
            key = ws.cell(r, 1).value
            value = ws.cell(r, 2).value
            if key:
                keys[key] = value
        assert keys["Capital sale events (aggregated)"] == 5
        assert keys["Reward rows"] == 3
        assert keys["Short term rows"] == 2
        assert keys["Long term rows"] == 1
        assert keys["Mixed holding period rows"] == 1
        assert keys["Unknown holding period rows"] == 1
        assert keys["Capital cost total (EUR)"] == float(Decimal("10000"))
        assert keys["Capital proceeds total (EUR)"] == float(Decimal("12000"))
        assert keys["Capital gain total (EUR)"] == float(Decimal("2000"))
        assert keys["Rewards total (EUR)"] == float(Decimal("500"))

    def test_opening_holdings_written_when_present(self):
        holdings = HoldingsSnapshot(asset_rows=3, total_cost_eur=Decimal("5000"), total_value_eur=Decimal("6000"))
        recon = CryptoReconciliationSummary(
            capital_rows=1,
            reward_rows=0,
            short_term_rows=1,
            long_term_rows=0,
            mixed_rows=0,
            unknown_rows=0,
            capital_cost_total_eur=Decimal("5000"),
            capital_proceeds_total_eur=Decimal("6000"),
            capital_gain_total_eur=Decimal("1000"),
            reward_total_eur=Decimal("0"),
            opening_holdings=holdings,
            closing_holdings=None,
        )
        report = _make_crypto_tax_report(reconciliation=recon)
        wb = openpyxl.Workbook()
        write_crypto_reconciliation_sheet(wb, report)
        ws = wb["Crypto Reconciliation"]
        keys = {}
        for r in range(1, ws.max_row + 1):
            key = ws.cell(r, 1).value
            value = ws.cell(r, 2).value
            if key:
                keys[key] = value
        assert keys["Opening holdings rows"] == 3
        assert keys["Opening holdings cost (EUR)"] == float(Decimal("5000"))
        assert keys["Opening holdings value (EUR)"] == float(Decimal("6000"))

    def test_closing_holdings_written_when_present(self):
        holdings = HoldingsSnapshot(asset_rows=2, total_cost_eur=Decimal("3000"), total_value_eur=Decimal("4000"))
        recon = CryptoReconciliationSummary(
            capital_rows=1,
            reward_rows=0,
            short_term_rows=1,
            long_term_rows=0,
            mixed_rows=0,
            unknown_rows=0,
            capital_cost_total_eur=Decimal("3000"),
            capital_proceeds_total_eur=Decimal("4000"),
            capital_gain_total_eur=Decimal("1000"),
            reward_total_eur=Decimal("0"),
            opening_holdings=None,
            closing_holdings=holdings,
        )
        report = _make_crypto_tax_report(reconciliation=recon)
        wb = openpyxl.Workbook()
        write_crypto_reconciliation_sheet(wb, report)
        ws = wb["Crypto Reconciliation"]
        keys = {}
        for r in range(1, ws.max_row + 1):
            key = ws.cell(r, 1).value
            value = ws.cell(r, 2).value
            if key:
                keys[key] = value
        assert keys["Closing holdings rows"] == 2
        assert keys["Closing holdings cost (EUR)"] == float(Decimal("3000"))
        assert keys["Closing holdings value (EUR)"] == float(Decimal("4000"))

    def test_no_holdings_rows_when_absent(self):
        wb = openpyxl.Workbook()
        report = _make_crypto_tax_report()
        write_crypto_reconciliation_sheet(wb, report)
        ws = wb["Crypto Reconciliation"]
        keys = {}
        for r in range(1, ws.max_row + 1):
            key = ws.cell(r, 1).value
            if key:
                keys[str(key)] = ws.cell(r, 2).value
        assert "Opening holdings rows" not in keys
        assert "Closing holdings rows" not in keys

    def test_both_holdings_present(self):
        opening = HoldingsSnapshot(asset_rows=3, total_cost_eur=Decimal("5000"), total_value_eur=Decimal("6000"))
        closing = HoldingsSnapshot(asset_rows=2, total_cost_eur=Decimal("3000"), total_value_eur=Decimal("4000"))
        recon = CryptoReconciliationSummary(
            capital_rows=1,
            reward_rows=0,
            short_term_rows=1,
            long_term_rows=0,
            mixed_rows=0,
            unknown_rows=0,
            capital_cost_total_eur=Decimal("5000"),
            capital_proceeds_total_eur=Decimal("6000"),
            capital_gain_total_eur=Decimal("1000"),
            reward_total_eur=Decimal("0"),
            opening_holdings=opening,
            closing_holdings=closing,
        )
        report = _make_crypto_tax_report(reconciliation=recon)
        wb = openpyxl.Workbook()
        write_crypto_reconciliation_sheet(wb, report)
        ws = wb["Crypto Reconciliation"]
        keys = {}
        for r in range(1, ws.max_row + 1):
            key = ws.cell(r, 1).value
            value = ws.cell(r, 2).value
            if key:
                keys[key] = value
        assert keys["Opening holdings rows"] == 3
        assert keys["Closing holdings rows"] == 2


@pytest.mark.unit
class TestCryptoReconciliationSection4:
    """Tests for section 4: Skipped zero value tokens."""

    def test_section_4_title_written(self):
        wb = openpyxl.Workbook()
        report = _make_crypto_tax_report()
        write_crypto_reconciliation_sheet(wb, report)
        ws = wb["Crypto Reconciliation"]
        found = False
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=1):
            if row[0].value == "4. SKIPPED ZERO VALUE TOKENS":
                found = True
                break
        assert found

    def test_section_4_title_is_bold(self):
        wb = openpyxl.Workbook()
        report = _make_crypto_tax_report()
        write_crypto_reconciliation_sheet(wb, report)
        ws = wb["Crypto Reconciliation"]
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=1):
            if row[0].value == "4. SKIPPED ZERO VALUE TOKENS":
                assert row[0].font.bold is True
                break

    def test_skipped_tokens_headers(self):
        skipped = [CryptoSkippedZeroValueToken(source_section="Capital Gains", asset="DUST", count=2)]
        report = _make_crypto_tax_report(skipped_tokens=skipped)
        wb = openpyxl.Workbook()
        write_crypto_reconciliation_sheet(wb, report)
        ws = wb["Crypto Reconciliation"]
        section_start = None
        for r in range(1, ws.max_row + 1):
            if ws.cell(r, 1).value == "4. SKIPPED ZERO VALUE TOKENS":
                section_start = r
                break
        assert section_start is not None
        header_row = section_start + 1
        assert ws.cell(header_row, 1).value == "Source section"
        assert ws.cell(header_row, 2).value == "Asset"
        assert ws.cell(header_row, 3).value == "Skipped rows"

    def test_skipped_tokens_data_rows(self):
        skipped = [
            CryptoSkippedZeroValueToken(source_section="Capital Gains", asset="DUST", count=2),
            CryptoSkippedZeroValueToken(source_section="Rewards", asset="PEPE", count=5),
        ]
        report = _make_crypto_tax_report(skipped_tokens=skipped)
        wb = openpyxl.Workbook()
        write_crypto_reconciliation_sheet(wb, report)
        ws = wb["Crypto Reconciliation"]
        header_row = None
        for r in range(1, ws.max_row + 1):
            if ws.cell(r, 1).value == "Source section" and ws.cell(r, 2).value == "Asset":
                header_row = r
                break
        assert header_row is not None
        assert ws.cell(header_row + 1, 1).value == "Capital Gains"
        assert ws.cell(header_row + 1, 2).value == "DUST"
        assert ws.cell(header_row + 1, 3).value == 2
        assert ws.cell(header_row + 2, 1).value == "Rewards"
        assert ws.cell(header_row + 2, 2).value == "PEPE"
        assert ws.cell(header_row + 2, 3).value == 5

    def test_no_skipped_tokens_shows_none(self):
        report = _make_crypto_tax_report(skipped_tokens=[])
        wb = openpyxl.Workbook()
        write_crypto_reconciliation_sheet(wb, report)
        ws = wb["Crypto Reconciliation"]
        header_row = None
        for r in range(1, ws.max_row + 1):
            if ws.cell(r, 1).value == "Source section" and ws.cell(r, 2).value == "Asset":
                header_row = r
                break
        assert header_row is not None
        data_row = header_row + 1
        assert ws.cell(data_row, 1).value == "none"
        assert ws.cell(data_row, 2).value == ""
        assert ws.cell(data_row, 3).value == 0


@pytest.mark.unit
class TestCryptoReconciliationSheetAutoWidth:
    """Tests that auto_column_width is called."""

    def test_auto_width_adjusts_columns(self):
        wb = openpyxl.Workbook()
        report = _make_crypto_tax_report()
        write_crypto_reconciliation_sheet(wb, report)
        ws = wb["Crypto Reconciliation"]
        assert ws.column_dimensions["A"].width > 0
