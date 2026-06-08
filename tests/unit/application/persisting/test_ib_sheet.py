"""Tests for IB reporting sheet writer and currency table."""

from decimal import Decimal

import openpyxl
import pytest

from tax_reporting.application.crypto_reporting import AggregatedRewardIncomeEntry
from tax_reporting.application.persisting.ib_sheet import (
    create_currency_table,
    write_ib_reporting_sheet,
)
from tax_reporting.application.persisting.tax_constants import get_income_code_description
from tax_reporting.domain.collections import CapitalGainLinesPerCompany, DividendIncomePerCompany
from tax_reporting.domain.constants import EXCEL_HEADER_ROW_1, EXCEL_HEADER_ROW_2, EXCEL_START_ROW
from tax_reporting.domain.entities import (
    CapitalGainLine,
    CurrencyCompany,
    DividendIncomePerSecurity,
)
from tax_reporting.domain.value_objects import Company, Currency, TradeDate
from tax_reporting.infrastructure.config import Config, ConversionRate, TaxJurisdictionConfig


def _make_capital_gain_line(  # noqa: PLR0913
    sell_date: TradeDate | None = None,
    buy_date: TradeDate | None = None,
    currency: Currency | None = None,
    sell_quantities: list[Decimal] | None = None,
    buy_quantities: list[Decimal] | None = None,
    sell_prices: list[Decimal] | None = None,
    buy_prices: list[Decimal] | None = None,
    sell_fees: list[Decimal] | None = None,
    buy_fees: list[Decimal] | None = None,
) -> CapitalGainLine:
    """Helper to create a CapitalGainLine with minimal setup."""
    from tax_reporting.domain.entities import TradeAction

    _sell_date = sell_date or TradeDate(2025, 6, 15)
    _buy_date = buy_date or TradeDate(2024, 3, 10)
    _currency = currency or Currency("USD")
    sq = sell_quantities or [Decimal("10")]
    bq = buy_quantities or [Decimal("10")]
    sp = sell_prices or [Decimal("100")]
    bp = buy_prices or [Decimal("80")]
    sf = sell_fees or [Decimal("1")]
    bf = buy_fees or [Decimal("1")]

    sell_trades = []
    for p, f in zip(sp, sf, strict=True):
        sell_trades.append(
            TradeAction(
                company=Company("AAPL"),
                date_time="2025-06-15, 10:00:00",
                currency=_currency,
                quantity="-10",
                price=str(p),
                fee=str(f),
            )
        )

    buy_trades = []
    for p, f in zip(bp, bf, strict=True):
        buy_trades.append(
            TradeAction(
                company=Company("AAPL"),
                date_time="2024-03-10, 10:00:00",
                currency=_currency,
                quantity="10",
                price=str(p),
                fee=str(f),
            )
        )

    return CapitalGainLine(
        ticker="AAPL",
        currency=_currency,
        sell_date=_sell_date,
        sell_quantities=sq,
        sell_trades=sell_trades,
        buy_date=_buy_date,
        buy_quantities=bq,
        buy_trades=buy_trades,
    )


def _make_config() -> Config:
    """Create a test config with USD/EUR rate."""
    return Config(
        base="EUR",
        rates=[ConversionRate(base="EUR", calculated="USD", rate=Decimal("1.10"))],
        tax_jurisdiction=TaxJurisdictionConfig(
            country="PT", fiscal_year=2025, exclude_loan_repayment_gains=False,
            zero_basis_review_threshold=Decimal("50"),
        ),
    )


def _make_aggregated_reward_entry(  # noqa: PLR0913
    income_code: str = "402",
    source_country: str = "GB",
    gross_income_eur: Decimal | None = None,
    foreign_tax_eur: Decimal | None = None,
    raw_row_count: int = 50,
    chains: tuple[str, ...] | None = None,
    description: str = "Wirex",
) -> AggregatedRewardIncomeEntry:
    """Helper to create an AggregatedRewardIncomeEntry for testing."""
    return AggregatedRewardIncomeEntry(
        income_code=income_code,
        source_country=source_country,
        gross_income_eur=gross_income_eur if gross_income_eur is not None else Decimal("1195.79"),
        foreign_tax_eur=foreign_tax_eur if foreign_tax_eur is not None else Decimal("0"),
        raw_row_count=raw_row_count,
        chains=chains if chains is not None else ("Wirex",),
        description=description,
    )


@pytest.mark.unit
class TestCreateCurrencyTable:
    """Tests for create_currency_table."""

    def test_writes_title_in_first_cell(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        config = _make_config()
        create_currency_table(ws, column_no=1, row_no=1, config=config)
        assert ws.cell(1, 1).value == "Currency exchange rate"

    def test_writes_headers_in_row_below_title(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        config = _make_config()
        create_currency_table(ws, column_no=1, row_no=1, config=config)
        assert ws.cell(2, 1).value == "Base/target"
        assert ws.cell(2, 2).value == "Rate"

    def test_returns_coordinate_map_with_base_currency(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        config = _make_config()
        result = create_currency_table(ws, column_no=1, row_no=1, config=config)
        assert "USD" in result
        assert "EUR" in result

    def test_writes_rate_value_in_cell(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        config = _make_config()
        create_currency_table(ws, column_no=1, row_no=1, config=config)
        rate_cell = ws.cell(3, 2)
        assert rate_cell.value == "1.10"

    def test_writes_base_self_rate_as_one(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        config = _make_config()
        create_currency_table(ws, column_no=1, row_no=1, config=config)
        base_cell = ws.cell(4, 2)
        assert base_cell.value == "1"

    def test_offsets_correctly_with_nonzero_start(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        config = _make_config()
        result = create_currency_table(ws, column_no=5, row_no=10, config=config)
        assert ws.cell(10, 5).value == "Currency exchange rate"
        assert ws.cell(11, 5).value == "Base/target"
        assert ws.cell(11, 6).value == "Rate"
        assert "USD" in result

    def test_both_headers_in_separate_columns(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        config = _make_config()
        create_currency_table(ws, column_no=3, row_no=1, config=config)
        assert ws.cell(2, 3).value == "Base/target"
        assert ws.cell(2, 4).value == "Rate"


@pytest.mark.unit
class TestWriteIbReportingSheetHeaders:
    """Tests that write_ib_reporting_sheet writes correct header rows."""

    def test_writes_first_header_row(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Reporting"
        config = _make_config()
        lines: CapitalGainLinesPerCompany = {}
        write_ib_reporting_sheet(ws, config, lines)
        # When there are no capital gains, no section title is written, headers at rows 1-2
        assert ws.cell(1, 1).value == "Country of Source"
        assert ws.cell(1, 2).value == "SALE"

    def test_writes_second_header_row(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Reporting"
        config = _make_config()
        lines: CapitalGainLinesPerCompany = {}
        write_ib_reporting_sheet(ws, config, lines)
        # When there are no capital gains, no section title is written, headers at rows 1-2
        assert ws.cell(2, 2).value == "Day "


@pytest.mark.unit
class TestWriteIbReportingSheetCapitalGains:
    """Tests that capital gain data rows are written correctly."""

    def test_writes_sell_day_for_capital_gain_line(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Reporting"
        config = _make_config()
        line = _make_capital_gain_line(sell_date=TradeDate(2025, 6, 15))
        cc = CurrencyCompany(Currency("USD"), Company("AAPL", country_of_issuance="US"))
        lines: CapitalGainLinesPerCompany = {cc: [line]}
        write_ib_reporting_sheet(ws, config, lines)
        assert ws.cell(5, 2).value == 15

    def test_writes_sell_month_name(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Reporting"
        config = _make_config()
        line = _make_capital_gain_line(sell_date=TradeDate(2025, 6, 15))
        cc = CurrencyCompany(Currency("USD"), Company("AAPL", country_of_issuance="US"))
        lines: CapitalGainLinesPerCompany = {cc: [line]}
        write_ib_reporting_sheet(ws, config, lines)
        assert ws.cell(5, 3).value == "June"

    def test_writes_sell_year(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Reporting"
        config = _make_config()
        line = _make_capital_gain_line(sell_date=TradeDate(2025, 6, 15))
        cc = CurrencyCompany(Currency("USD"), Company("AAPL", country_of_issuance="US"))
        lines: CapitalGainLinesPerCompany = {cc: [line]}
        write_ib_reporting_sheet(ws, config, lines)
        assert ws.cell(5, 4).value == 2025

    def test_sell_amount_is_formula(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Reporting"
        config = _make_config()
        line = _make_capital_gain_line()
        cc = CurrencyCompany(Currency("USD"), Company("AAPL", country_of_issuance="US"))
        lines: CapitalGainLinesPerCompany = {cc: [line]}
        write_ib_reporting_sheet(ws, config, lines)
        sell_amount_cell = ws.cell(5, 5)
        assert sell_amount_cell.data_type == "f"
        assert sell_amount_cell.value.startswith("=")

    def test_buy_day_for_capital_gain_line(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Reporting"
        config = _make_config()
        line = _make_capital_gain_line(buy_date=TradeDate(2024, 3, 10))
        cc = CurrencyCompany(Currency("USD"), Company("AAPL", country_of_issuance="US"))
        lines: CapitalGainLinesPerCompany = {cc: [line]}
        write_ib_reporting_sheet(ws, config, lines)
        assert ws.cell(5, 6).value == 10

    def test_buy_amount_is_formula(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Reporting"
        config = _make_config()
        line = _make_capital_gain_line()
        cc = CurrencyCompany(Currency("USD"), Company("AAPL", country_of_issuance="US"))
        lines: CapitalGainLinesPerCompany = {cc: [line]}
        write_ib_reporting_sheet(ws, config, lines)
        buy_amount_cell = ws.cell(5, 9)
        assert buy_amount_cell.data_type == "f"

    def test_expense_cell_is_formula(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Reporting"
        config = _make_config()
        line = _make_capital_gain_line()
        cc = CurrencyCompany(Currency("USD"), Company("AAPL", country_of_issuance="US"))
        lines: CapitalGainLinesPerCompany = {cc: [line]}
        write_ib_reporting_sheet(ws, config, lines)
        expense_cell = ws.cell(5, 12)
        assert expense_cell.data_type == "f"
        assert expense_cell.number_format == "0.00"

    def test_country_of_source_populated(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Reporting"
        config = _make_config()
        line = _make_capital_gain_line()
        cc = CurrencyCompany(Currency("USD"), Company("AAPL", country_of_issuance="US"))
        lines: CapitalGainLinesPerCompany = {cc: [line]}
        write_ib_reporting_sheet(ws, config, lines)
        assert ws.cell(5, 1).value == "US"
        assert ws.cell(5, 10).value == "US"

    def test_symbol_and_currency_written(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Reporting"
        config = _make_config()
        line = _make_capital_gain_line()
        cc = CurrencyCompany(Currency("USD"), Company("AAPL", country_of_issuance="US"))
        lines: CapitalGainLinesPerCompany = {cc: [line]}
        write_ib_reporting_sheet(ws, config, lines)
        assert ws.cell(5, 14).value == "AAPL"
        assert ws.cell(5, 15).value == "USD"

    def test_multiple_lines_write_on_separate_rows(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Reporting"
        config = _make_config()
        line1 = _make_capital_gain_line(sell_date=TradeDate(2025, 1, 5))
        line2 = _make_capital_gain_line(sell_date=TradeDate(2025, 2, 10))
        cc = CurrencyCompany(Currency("USD"), Company("AAPL", country_of_issuance="US"))
        lines: CapitalGainLinesPerCompany = {cc: [line1, line2]}
        write_ib_reporting_sheet(ws, config, lines)
        assert ws.cell(5, 2).value == 5
        assert ws.cell(6, 2).value == 10

    def test_placeholder_buy_row_has_red_fill(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Reporting"
        config = _make_config()
        line = _make_capital_gain_line(buy_date=TradeDate(1000, 1, 1))
        cc = CurrencyCompany(Currency("USD"), Company("AAPL", country_of_issuance="US"))
        lines: CapitalGainLinesPerCompany = {cc: [line]}
        write_ib_reporting_sheet(ws, config, lines)
        cell = ws.cell(5, 2)
        assert cell.fill.start_color.rgb == "FFFF0000"

    def test_section_title_at_row_1(self):
        """Verify CAPITAL GAINS section title is at row 1 with bold font."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Reporting"
        config = _make_config()
        line = _make_capital_gain_line()
        cc = CurrencyCompany(Currency("USD"), Company("AAPL", country_of_issuance="US"))
        lines: CapitalGainLinesPerCompany = {cc: [line]}
        write_ib_reporting_sheet(ws, config, lines)

        assert ws.cell(1, 1).value == "CAPITAL GAINS"
        assert ws.cell(1, 1).font.bold == True

    def test_single_blank_row_after_title(self):
        """Verify there is exactly one blank row after the section title."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Reporting"
        config = _make_config()
        line = _make_capital_gain_line()
        cc = CurrencyCompany(Currency("USD"), Company("AAPL", country_of_issuance="US"))
        lines: CapitalGainLinesPerCompany = {cc: [line]}
        write_ib_reporting_sheet(ws, config, lines)

        # Row 2 should be empty (the blank row after title)
        # Check all 19 columns (header has 18, plus column 19) to catch accidental writes beyond expected structure
        for col in range(1, 20):
            assert ws.cell(2, col).value is None

    def test_first_header_row_structure(self):
        """Verify first header row has correct structure at row 3."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Reporting"
        config = _make_config()
        line = _make_capital_gain_line()
        cc = CurrencyCompany(Currency("USD"), Company("AAPL", country_of_issuance="US"))
        lines: CapitalGainLinesPerCompany = {cc: [line]}
        write_ib_reporting_sheet(ws, config, lines)

        assert ws.cell(3, 1).value == "Country of Source"
        assert ws.cell(3, 2).value == "SALE"
        assert ws.cell(3, 6).value == "PURCHASE"

    def test_second_header_row_structure(self):
        """Verify second header row has correct sub-headers at row 4."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Reporting"
        config = _make_config()
        line = _make_capital_gain_line()
        cc = CurrencyCompany(Currency("USD"), Company("AAPL", country_of_issuance="US"))
        lines: CapitalGainLinesPerCompany = {cc: [line]}
        write_ib_reporting_sheet(ws, config, lines)

        assert ws.cell(4, 2).value == "Day "
        assert ws.cell(4, 3).value == "Month "
        assert ws.cell(4, 4).value == "Year"
        assert ws.cell(4, 5).value == "Amount"

    def test_sale_header_merged_across_4_columns(self):
        """Verify SALE header is merged across columns 2-5."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Reporting"
        config = _make_config()
        line = _make_capital_gain_line()
        cc = CurrencyCompany(Currency("USD"), Company("AAPL", country_of_issuance="US"))
        lines: CapitalGainLinesPerCompany = {cc: [line]}
        write_ib_reporting_sheet(ws, config, lines)

        # Check that the range (2-5, 3) is merged
        # Note: String matching on merged_cells.ranges is used because openpyxl does not provide
        # a direct range coordinate comparison API. The format is stable across versions.
        merge_ranges = [str(r) for r in ws.merged_cells.ranges]
        assert any("B3:E3" in r or r.startswith("B3") and r.endswith("E3") for r in merge_ranges)

    def test_purchase_header_merged_across_4_columns(self):
        """Verify PURCHASE header is merged across columns 6-9."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Reporting"
        config = _make_config()
        line = _make_capital_gain_line()
        cc = CurrencyCompany(Currency("USD"), Company("AAPL", country_of_issuance="US"))
        lines: CapitalGainLinesPerCompany = {cc: [line]}
        write_ib_reporting_sheet(ws, config, lines)

        # Check that the range (6-9, 3) is merged
        # Note: String matching on merged_cells.ranges is used because openpyxl does not provide
        # a direct range coordinate comparison API. The format is stable across versions.
        merge_ranges = [str(r) for r in ws.merged_cells.ranges]
        assert any("F3:I3" in r or r.startswith("F3") and r.endswith("I3") for r in merge_ranges)

    def test_data_starts_at_row_5(self):
        """Verify first data row is at row 5 with sell_day value."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Reporting"
        config = _make_config()
        line = _make_capital_gain_line(sell_date=TradeDate(2025, 6, 15))
        cc = CurrencyCompany(Currency("USD"), Company("AAPL", country_of_issuance="US"))
        lines: CapitalGainLinesPerCompany = {cc: [line]}
        write_ib_reporting_sheet(ws, config, lines)

        # Data should start at row 5
        assert ws.cell(5, 2).value == 15

    def test_country_and_sell_day_at_different_columns(self):
        """Regression test: verify Country of Source (col 1) and sell_day (col 2) are in different columns."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Reporting"
        config = _make_config()
        line = _make_capital_gain_line(sell_date=TradeDate(2025, 6, 15))
        cc = CurrencyCompany(Currency("USD"), Company("AAPL", country_of_issuance="US"))
        lines: CapitalGainLinesPerCompany = {cc: [line]}
        write_ib_reporting_sheet(ws, config, lines)

        # Country at col 1, sell_day at col 2 -- must be different columns
        assert ws.cell(5, 1).value == "US"
        assert ws.cell(5, 2).value == 15


@pytest.mark.unit
class TestWriteIbReportingSheetDividends:
    """Tests that dividend section is written correctly."""

    def test_section_title_written_when_dividends_provided(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Reporting"
        config = _make_config()
        lines: CapitalGainLinesPerCompany = {}
        dividends: DividendIncomePerCompany = {
            "AAPL": DividendIncomePerSecurity(
                symbol="AAPL",
                isin="US0378331005",
                country="US",
                gross_amount=Decimal("100"),
                total_taxes=Decimal("15"),
                currency=Currency("USD"),
            )
        }
        write_ib_reporting_sheet(ws, config, lines, dividends)
        found = False
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=1):
            if row[0].value and "CAPITAL INVESTMENT INCOME" in str(row[0].value):
                found = True
                break
        assert found

    def test_dividend_data_row_values(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Reporting"
        config = _make_config()
        lines: CapitalGainLinesPerCompany = {}
        dividends: DividendIncomePerCompany = {
            "AAPL": DividendIncomePerSecurity(
                symbol="AAPL",
                isin="US0378331005",
                country="US",
                gross_amount=Decimal("100"),
                total_taxes=Decimal("15"),
                currency=Currency("USD"),
            )
        }
        write_ib_reporting_sheet(ws, config, lines, dividends)
        div_type_cell = None
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=12):
            if row[0].value == "Share dividends":
                div_type_cell = row
                break
        assert div_type_cell is not None
        assert div_type_cell[1].value == "US"
        assert div_type_cell[2].value == "US0378331005"
        assert div_type_cell[7].value == "AAPL"
        assert div_type_cell[8].value == "USD"

    def test_dividend_gross_amount_is_formula(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Reporting"
        config = _make_config()
        lines: CapitalGainLinesPerCompany = {}
        dividends: DividendIncomePerCompany = {
            "AAPL": DividendIncomePerSecurity(
                symbol="AAPL",
                isin="US0378331005",
                country="US",
                gross_amount=Decimal("100"),
                total_taxes=Decimal("15"),
                currency=Currency("USD"),
            )
        }
        write_ib_reporting_sheet(ws, config, lines, dividends)
        gross_cell = None
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=12):
            if row[0].value == "Share dividends":
                gross_cell = row[3]
                break
        assert gross_cell is not None
        assert gross_cell.data_type == "f"

    def test_missing_isin_shows_warning(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Reporting"
        config = _make_config()
        lines: CapitalGainLinesPerCompany = {}
        dividends: DividendIncomePerCompany = {
            "UNKNOWN": DividendIncomePerSecurity(
                symbol="UNKNOWN",
                isin="MISSING_ISIN_REQUIRES_ATTENTION",
                country="UNKNOWN_COUNTRY",
                gross_amount=Decimal("50"),
                total_taxes=Decimal("0"),
                currency=Currency("USD"),
            )
        }
        write_ib_reporting_sheet(ws, config, lines, dividends)
        warning_cell = None
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=12):
            if row[0].value == "Share dividends":
                warning_cell = row[1]
                break
        assert warning_cell is not None
        assert "MISSING DATA" in str(warning_cell.value)

    def test_no_dividend_section_when_none_provided(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Reporting"
        config = _make_config()
        lines: CapitalGainLinesPerCompany = {}
        write_ib_reporting_sheet(ws, config, lines)
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=1):
            assert row[0].value is None or "CAPITAL INVESTMENT INCOME" not in str(row[0].value)


@pytest.mark.unit
class TestWriteIbReportingSheetCurrencyTable:
    """Tests that currency table is embedded in the IB sheet."""

    def test_currency_table_present_in_sheet(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Reporting"
        config = _make_config()
        lines: CapitalGainLinesPerCompany = {}
        write_ib_reporting_sheet(ws, config, lines)
        found = False
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=20, max_col=20):
            if row[0].value == "Currency exchange rate":
                found = True
                break
        assert found

    def test_returns_none(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Reporting"
        config = _make_config()
        lines: CapitalGainLinesPerCompany = {}
        result = write_ib_reporting_sheet(ws, config, lines)
        assert result is None


@pytest.mark.unit
class TestWriteIbReportingSheetAutoWidth:
    """Tests that auto_column_width is called and respects bounds."""

    def test_auto_width_adjusts_columns(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Reporting"
        config = _make_config()
        line = _make_capital_gain_line()
        cc = CurrencyCompany(Currency("USD"), Company("AAPL", country_of_issuance="US"))
        lines: CapitalGainLinesPerCompany = {cc: [line]}
        write_ib_reporting_sheet(ws, config, lines)
        assert ws.column_dimensions["A"].width > 0

    def test_formula_heavy_columns_get_reasonable_widths(self):
        """Verify formula-heavy columns preserve header width with MIN_DATA_WIDTH as floor."""
        from tax_reporting.application.persisting.excel_utils import MIN_DATA_WIDTH

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Reporting"
        config = _make_config()
        # Add multiple lines to trigger formula-heavy detection (formulas > non-formula cells)
        line1 = _make_capital_gain_line()
        line2 = _make_capital_gain_line(sell_date=TradeDate(2025, 7, 20))
        line3 = _make_capital_gain_line(sell_date=TradeDate(2025, 8, 10))
        cc = CurrencyCompany(Currency("USD"), Company("AAPL", country_of_issuance="US"))
        lines: CapitalGainLinesPerCompany = {cc: [line1, line2, line3]}
        write_ib_reporting_sheet(ws, config, lines)

        # Expected widths for headers: L=45 (longest), P=11, Q=10, R=15, E=6, I=6
        # All formula-heavy columns should use measured header length with MIN_DATA_WIDTH as floor
        expected_widths = {
            "E": 6,    # "Amount" (sell)
            "I": 6,    # "Amount" (buy)
            "L": 45,   # "Expenses incurred with obtaining the capital gains"
            "P": 11,   # "Sale amount"
            "Q": 10,   # "Buy amount"
            "R": 15,   # "Expenses amount"
        }

        for col_letter, expected_header_len in expected_widths.items():
            width = ws.column_dimensions[col_letter].width
            assert width is not None, f"Column {col_letter} should have a width set"
            # Formula-heavy columns get max(header_length, MIN_DATA_WIDTH), not MIN_DATA_WIDTH alone
            expected_min = max(expected_header_len, MIN_DATA_WIDTH)
            assert width >= expected_min, (
                f"Column {col_letter} width {width} is too small; "
                f"expected at least {expected_min} (header={expected_header_len}, MIN_DATA_WIDTH={MIN_DATA_WIDTH})"
            )

    def test_empty_column_gets_min_data_width(self):
        """Verify a column with only formulas/empty cells gets MIN_DATA_WIDTH."""
        from tax_reporting.application.persisting.excel_utils import MIN_DATA_WIDTH, auto_column_width

        # Create a worksheet with one column that has only formulas
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Test"

        # Column A: only a formula
        ws["A1"] = "=SUM(B1:B10)"

        auto_column_width(ws)

        # Column A should get MIN_DATA_WIDTH since it has no data cells
        width = ws.column_dimensions["A"].width
        msg = f"Column with only formulas should get MIN_DATA_WIDTH ({MIN_DATA_WIDTH}), got {width}"
        assert width == MIN_DATA_WIDTH, msg


@pytest.mark.unit
class TestWriteIbReportingSheetCapitalInvestmentIncome:
    """Tests for mixed capital investment income rendering (dividends + fiat rewards)."""

    @staticmethod
    def _find_row_with_text(ws, text_fragment, column=1):
        for row in range(1, ws.max_row + 1):
            val = ws.cell(row, column).value
            if val is not None and text_fragment in str(val):
                return row
        return None

    def test_share_dividends_use_explicit_share_dividend_label(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Reporting"
        config = _make_config()
        lines: CapitalGainLinesPerCompany = {}
        dividends: DividendIncomePerCompany = {
            "AAPL": DividendIncomePerSecurity(
                symbol="AAPL",
                isin="US0378331005",
                country="US",
                gross_amount=Decimal("100"),
                total_taxes=Decimal("15"),
                currency=Currency("USD"),
            )
        }
        write_ib_reporting_sheet(ws, config, lines, dividends)

        section_row = self._find_row_with_text(ws, "CAPITAL INVESTMENT INCOME")
        assert section_row is not None, "CAPITAL INVESTMENT INCOME section title should exist"

        subsection_row = self._find_row_with_text(ws, "SHARE DIVIDENDS")
        assert subsection_row is not None, "SHARE DIVIDENDS subsection label should exist"

        div_row = self._find_row_with_text(ws, "Share dividends", column=1)
        assert div_row is not None, "Dividend row should use 'Share dividends' as income type"

    def test_taxable_fiat_reward_aggregate_is_written_under_capital_investment_income(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Reporting"
        config = _make_config()
        lines: CapitalGainLinesPerCompany = {}
        entry = _make_aggregated_reward_entry()
        write_ib_reporting_sheet(ws, config, lines, None, other_capital_income_entries=[entry])

        section_row = self._find_row_with_text(ws, "CAPITAL INVESTMENT INCOME")
        assert section_row is not None

        other_row = self._find_row_with_text(ws, "OTHER CAPITAL INVESTMENT INCOME")
        assert other_row is not None

        data_row = self._find_row_with_text(ws, "Crypto interest", column=1)
        assert data_row is not None

        assert ws.cell(data_row, 2).value == "GB"
        assert ws.cell(data_row, 3).value in (None, "")
        assert float(ws.cell(data_row, 4).value) == pytest.approx(1195.79)
        assert float(ws.cell(data_row, 5).value) == pytest.approx(0)
        assert ws.cell(data_row, 6).value in (None, "")
        assert ws.cell(data_row, 7).value in (None, "")
        assert ws.cell(data_row, 8).value == "Wirex"
        assert ws.cell(data_row, 9).value == "EUR"
        assert float(ws.cell(data_row, 10).value) == pytest.approx(1195.79)
        assert float(ws.cell(data_row, 11).value) == pytest.approx(0)
        assert float(ws.cell(data_row, 12).value) == pytest.approx(1195.79)
        assert ws.cell(data_row, 13).value == "Koinly"
        assert ws.cell(data_row, 14).value == 50

    def test_non_wirex_taxable_fiat_reward_uses_derived_source_fields(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Reporting"
        config = _make_config()
        lines: CapitalGainLinesPerCompany = {}
        entry = _make_aggregated_reward_entry(
            source_country="IE",
            chains=("Kraken",),
            description="Kraken",
        )
        write_ib_reporting_sheet(ws, config, lines, None, other_capital_income_entries=[entry])

        data_row = self._find_row_with_text(ws, "Crypto interest", column=1)
        assert data_row is not None
        assert ws.cell(data_row, 2).value == "IE"
        assert ws.cell(data_row, 8).value == "Kraken"

    def test_foreign_tax_is_summed_and_net_is_gross_minus_tax(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Reporting"
        config = _make_config()
        lines: CapitalGainLinesPerCompany = {}
        entry = _make_aggregated_reward_entry(
            gross_income_eur=Decimal("100"),
            foreign_tax_eur=Decimal("15"),
            raw_row_count=2,
        )
        write_ib_reporting_sheet(ws, config, lines, None, other_capital_income_entries=[entry])

        data_row = self._find_row_with_text(ws, "Crypto interest", column=1)
        assert data_row is not None
        assert float(ws.cell(data_row, 4).value) == pytest.approx(100)
        assert float(ws.cell(data_row, 5).value) == pytest.approx(15)
        assert float(ws.cell(data_row, 12).value) == pytest.approx(85)
        assert ws.cell(data_row, 14).value == 2

    def test_empty_other_capital_income_list_does_not_write_other_income_subsection(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Reporting"
        config = _make_config()
        lines: CapitalGainLinesPerCompany = {}
        write_ib_reporting_sheet(ws, config, lines, None, other_capital_income_entries=[])

        other_row = self._find_row_with_text(ws, "OTHER CAPITAL INVESTMENT INCOME")
        assert other_row is None, "OTHER CAPITAL INVESTMENT INCOME should not appear with empty list"

    def test_no_capital_investment_section_when_no_dividends_and_no_taxable_rewards(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Reporting"
        config = _make_config()
        lines: CapitalGainLinesPerCompany = {}
        write_ib_reporting_sheet(ws, config, lines)

        section_row = self._find_row_with_text(ws, "CAPITAL INVESTMENT INCOME")
        assert section_row is None, "CAPITAL INVESTMENT INCOME section should not appear with no income"

    def test_koinly_source_strings_are_safe_for_excel_cells(self):
        # Single-character prefixes
        for prefix in ("=", "+", "-", "@"):
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Reporting"
            config = _make_config()
            lines: CapitalGainLinesPerCompany = {}
            entry = _make_aggregated_reward_entry(
                chains=(f"{prefix}Chain",),
                description=f"{prefix}Chain test",
            )
            write_ib_reporting_sheet(ws, config, lines, None, other_capital_income_entries=[entry])

            data_row = self._find_row_with_text(ws, "Crypto interest", column=1)
            assert data_row is not None, f"Data row for prefix '{prefix}'"

            cell = ws.cell(data_row, 8)
            assert not str(cell.value).startswith(prefix), (
                f"Source detail should not start with '{prefix}', got: {cell.value}"
            )

        # Multi-character prefixes
        for prefix in ("==", "++", "--", "@@"):
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Reporting"
            config = _make_config()
            lines: CapitalGainLinesPerCompany = {}
            entry = _make_aggregated_reward_entry(
                chains=(f"{prefix}Chain",),
                description=f"{prefix}Chain test",
            )
            write_ib_reporting_sheet(ws, config, lines, None, other_capital_income_entries=[entry])

            data_row = self._find_row_with_text(ws, "Crypto interest", column=1)
            assert data_row is not None, f"Data row for prefix '{prefix}'"

            cell = ws.cell(data_row, 8)
            assert not str(cell.value).startswith(prefix), (
                f"Source detail should not start with '{prefix}', got: {cell.value}"
            )

        # Control characters with prefix
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Reporting"
        config = _make_config()
        lines: CapitalGainLinesPerCompany = {}
        entry = _make_aggregated_reward_entry(
            chains=("=Chain\nValue",),
            description="=\nChain test",
        )
        write_ib_reporting_sheet(ws, config, lines, None, other_capital_income_entries=[entry])

        data_row = self._find_row_with_text(ws, "Crypto interest", column=1)
        assert data_row is not None, "Data row for control char test"

        cell = ws.cell(data_row, 8)
        # Control chars should be stripped and prefix neutralized
        assert "\n" not in str(cell.value), f"Newline should be stripped, got: {cell.value}"
        assert not str(cell.value).startswith("="), f"Prefix should be neutralized, got: {cell.value}"

    def test_zero_gross_income_renders_as_zero(self):
        """Zero gross_income_eur should render as zero value in Excel."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Reporting"
        config = _make_config()
        lines: CapitalGainLinesPerCompany = {}
        entry = _make_aggregated_reward_entry(gross_income_eur=Decimal("0"))
        write_ib_reporting_sheet(ws, config, lines, None, other_capital_income_entries=[entry])

        data_row = self._find_row_with_text(ws, "Crypto interest", column=1)
        assert data_row is not None
        # Zero value should render as 0, not None or empty
        assert float(ws.cell(data_row, 4).value) == pytest.approx(0)
        assert float(ws.cell(data_row, 12).value) == pytest.approx(0)

    def test_foreign_tax_exceeding_gross_income_renders_negative_net(self):
        """Foreign tax exceeding gross income should render negative net value."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Reporting"
        config = _make_config()
        lines: CapitalGainLinesPerCompany = {}
        entry = _make_aggregated_reward_entry(
            gross_income_eur=Decimal("100"),
            foreign_tax_eur=Decimal("150"),  # Tax exceeds gross
        )
        write_ib_reporting_sheet(ws, config, lines, None, other_capital_income_entries=[entry])

        data_row = self._find_row_with_text(ws, "Crypto interest", column=1)
        assert data_row is not None
        # Net value should be negative (gross - tax = 100 - 150 = -50)
        assert float(ws.cell(data_row, 4).value) == pytest.approx(100)
        assert float(ws.cell(data_row, 5).value) == pytest.approx(150)
        assert float(ws.cell(data_row, 12).value) == pytest.approx(-50)

    def test_empty_chains_tuple_renders_gracefully(self):
        """Empty chains tuple should render without error."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Reporting"
        config = _make_config()
        lines: CapitalGainLinesPerCompany = {}
        entry = _make_aggregated_reward_entry(chains=())
        write_ib_reporting_sheet(ws, config, lines, None, other_capital_income_entries=[entry])

        data_row = self._find_row_with_text(ws, "Crypto interest", column=1)
        assert data_row is not None
        # Empty chains should render as empty string in source detail column (column 8)
        source_detail_value = ws.cell(data_row, 8).value
        assert source_detail_value in ("", None)

    def test_empty_chains_tuple_renders_gracefully(self):
        """Empty chains tuple should render without error."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Reporting"
        config = _make_config()
        lines: CapitalGainLinesPerCompany = {}
        entry = _make_aggregated_reward_entry(chains=())
        write_ib_reporting_sheet(ws, config, lines, None, other_capital_income_entries=[entry])

        data_row = self._find_row_with_text(ws, "Crypto interest", column=1)
        assert data_row is not None
        # Empty chains should not cause rendering error


@pytest.mark.unit
class TestGetIncomeCodeDescription:
    """Tests for get_income_code_description function."""

    def test_returns_known_code_description(self):
        """Known income codes return their descriptions."""
        assert get_income_code_description("401") == "Crypto capital income (staking, rewards, airdrops)"
        assert get_income_code_description("402") == "Crypto interest (lending, deposit interest)"
        assert get_income_code_description("403") == "Mining income"
        assert get_income_code_description("404") == "Fork income"
        assert get_income_code_description("405") == "Crypto dividends"

    def test_returns_fallback_for_unknown_code(self):
        """Unknown income codes return a fallback string with the code."""
        unknown_code = "999"
        result = get_income_code_description(unknown_code)
        assert result == f"Income code {unknown_code}"

    def test_fallback_includes_original_code(self):
        """Fallback preserves the original code value."""
        code = "ABC"
        result = get_income_code_description(code)
        assert code in result
        assert "Income code" in result


