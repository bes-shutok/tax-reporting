"""Tests for the crypto supplementary sheet writer."""

import re
from decimal import Decimal

import openpyxl
import pytest

from tax_reporting.application.crypto_reporting import (
    AggregatedRewardIncomeEntry,
    CapitalGainPeriodStats,
    CryptoCapitalGainStats,
    CryptoReconciliationSummary,
    CryptoReviewEntry,
    CryptoRewardIncomeEntry,
    CryptoTaxReport,
    RewardTaxClassification,
)
from tax_reporting.application.persisting.crypto_supplementary_sheet import write_crypto_supplementary_sheet
from tax_reporting.application.persisting.tax_constants import _INCOME_CODE_DESCRIPTIONS

# make_operator_origin is a module-level helper from conftest, not a pytest fixture
from tests.conftest import build_koinly_jurisdiction, make_operator_origin


def _make_reward_entry(
    classification: RewardTaxClassification = RewardTaxClassification.TAXABLE_NOW,
    **overrides: object,
) -> CryptoRewardIncomeEntry:
    defaults = {
        "date": "2025-03-15",
        "asset": "ETH",
        "amount": Decimal("0.5"),
        "value_eur": Decimal("1500"),
        "income_label": "Staking Reward",
        "source_type": "staking",
        "wallet": "Kraken",
        "platform": "Kraken",
        "chain": "Ethereum",
        "operator_origin": make_operator_origin(),
        "annex_hint": "J",
        "review_required": False,
        "description": "Staking reward payout",
        "tax_classification": classification,
        "foreign_tax_eur": Decimal("0"),
    }
    defaults.update(overrides)
    return CryptoRewardIncomeEntry(**defaults)  # type: ignore[arg-type]


def _make_aggregated_reward(**overrides: object) -> AggregatedRewardIncomeEntry:
    defaults = {
        "income_code": "",
        "source_country": "US",
        "gross_income_eur": Decimal("1500"),
        "foreign_tax_eur": Decimal("0"),
        "raw_row_count": 1,
        "chains": ("Ethereum",),
        "description": "Staking income",
    }
    defaults.update(overrides)
    return AggregatedRewardIncomeEntry(**defaults)  # type: ignore[arg-type]


def _make_crypto_tax_report(
    reward_entries: list[CryptoRewardIncomeEntry] | None = None,
    review_entries: list[CryptoReviewEntry] | None = None,
    skipped_deferred_rewards: list[CryptoRewardIncomeEntry] | None = None,
) -> CryptoTaxReport:
    entries = reward_entries if reward_entries is not None else [_make_reward_entry()]
    reconciliation = CryptoReconciliationSummary(
        capital_rows=0,
        reward_rows=len(entries),
        short_term_rows=0,
        long_term_rows=0,
        mixed_rows=0,
        unknown_rows=0,
        capital_cost_total_eur=Decimal("0"),
        capital_proceeds_total_eur=Decimal("0"),
        capital_gain_total_eur=Decimal("0"),
        reward_total_eur=sum((e.value_eur for e in entries), start=Decimal("0")),
        opening_holdings=None,
        closing_holdings=None,
    )
    empty_stats = CapitalGainPeriodStats(
        count=0, cost_total_eur=Decimal("0"), proceeds_total_eur=Decimal("0"), gain_loss_total_eur=Decimal("0")
    )
    capital_gain_stats = CryptoCapitalGainStats(
        short_term=empty_stats, long_term=empty_stats, mixed=empty_stats, unknown=empty_stats, grand_total=empty_stats
    )
    return CryptoTaxReport(
        tax_year=2025,
        capital_entries=[],
        reward_entries=entries,
        reconciliation=reconciliation,
        capital_gain_stats=capital_gain_stats,
        pdf_summary=None,
        review_entries=review_entries if review_entries is not None else [],
        skipped_zero_value_deferred_rewards=(
            skipped_deferred_rewards if skipped_deferred_rewards is not None else []
        ),
    )


@pytest.mark.unit
class TestCryptoSupplementarySheetName:
    """Tests that the sheet is created with the correct name."""

    def test_sheet_named_crypto_supplementary(self):
        wb = openpyxl.Workbook()
        report = _make_crypto_tax_report()
        write_crypto_supplementary_sheet(wb, report, build_koinly_jurisdiction())
        assert "Crypto Supplementary" in wb.sheetnames


@pytest.mark.unit
class TestCryptoSupplementarySheetIncomeCodes:
    """Tests for section 1: Income Codes reference."""

    def test_section_1_title_written(self):
        wb = openpyxl.Workbook()
        report = _make_crypto_tax_report()
        write_crypto_supplementary_sheet(wb, report, build_koinly_jurisdiction())
        ws = wb["Crypto Supplementary"]
        found = False
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=1):
            if row[0].value == "1. INCOME CODES REFERENCE":
                found = True
                break
        assert found

    def test_section_1_title_is_bold(self):
        wb = openpyxl.Workbook()
        report = _make_crypto_tax_report()
        write_crypto_supplementary_sheet(wb, report, build_koinly_jurisdiction())
        ws = wb["Crypto Supplementary"]
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=1):
            if row[0].value == "1. INCOME CODES REFERENCE":
                assert row[0].font.bold is True
                break

    def test_reference_note_is_italic(self):
        wb = openpyxl.Workbook()
        report = _make_crypto_tax_report()
        write_crypto_supplementary_sheet(wb, report, build_koinly_jurisdiction())
        ws = wb["Crypto Supplementary"]
        found = False
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=1):
            if row[0].value and "Tabela V" in str(row[0].value):
                assert row[0].font.italic is True
                found = True
                break
        assert found

    def test_income_codes_headers_written(self):
        wb = openpyxl.Workbook()
        report = _make_crypto_tax_report()
        write_crypto_supplementary_sheet(wb, report, build_koinly_jurisdiction())
        ws = wb["Crypto Supplementary"]
        header_row = None
        for r in range(1, ws.max_row + 1):
            if ws.cell(r, 1).value == "Country":
                header_row = r
                break
        assert header_row is not None
        assert ws.cell(header_row, 1).value == "Country"
        assert ws.cell(header_row, 2).value == "Income Code"
        assert ws.cell(header_row, 3).value == "Description"

    def test_income_codes_headers_are_bold(self):
        wb = openpyxl.Workbook()
        report = _make_crypto_tax_report()
        write_crypto_supplementary_sheet(wb, report, build_koinly_jurisdiction())
        ws = wb["Crypto Supplementary"]
        header_row = None
        for r in range(1, ws.max_row + 1):
            if ws.cell(r, 1).value == "Country":
                header_row = r
                break
        assert header_row is not None
        for idx in range(1, 4):
            assert ws.cell(header_row, idx).font.bold is True

    def test_all_income_codes_listed(self):
        wb = openpyxl.Workbook()
        report = _make_crypto_tax_report()
        write_crypto_supplementary_sheet(wb, report, build_koinly_jurisdiction())
        ws = wb["Crypto Supplementary"]
        header_row = None
        for r in range(1, ws.max_row + 1):
            if ws.cell(r, 1).value == "Country":
                header_row = r
                break
        assert header_row is not None

        codes_found = []
        for r in range(header_row + 1, ws.max_row + 1):
            first_cell = ws.cell(r, 1).value
            # Stop when we hit the next section
            if first_cell and isinstance(first_cell, str) and first_cell.startswith("2."):
                break
            code = ws.cell(r, 2).value
            if code and isinstance(code, str):
                codes_found.append(code)

        # Only the official Tabela V crypto code (E25) is rendered under PT.
        # The set is sourced from the consolidated owner, not a hardcoded literal,
        # so the assertion tracks the owner without duplicating its contents.
        assert set(codes_found) == set(_INCOME_CODE_DESCRIPTIONS.keys())

    def test_income_codes_sorted_alphabetically(self):
        wb = openpyxl.Workbook()
        report = _make_crypto_tax_report()
        write_crypto_supplementary_sheet(wb, report, build_koinly_jurisdiction())
        ws = wb["Crypto Supplementary"]
        header_row = None
        for r in range(1, ws.max_row + 1):
            if ws.cell(r, 1).value == "Country":
                header_row = r
                break
        assert header_row is not None

        codes = []
        for r in range(header_row + 1, ws.max_row + 1):
            first_cell = ws.cell(r, 1).value
            # Stop when we hit the next section
            if first_cell and isinstance(first_cell, str) and first_cell.startswith("2."):
                break
            code = ws.cell(r, 2).value
            if code and isinstance(code, str):
                codes.append(code)

        # Codes are rendered via sorted(_INCOME_CODE_DESCRIPTIONS.items())
        assert codes == sorted(codes)

    def test_income_code_table_renders_official_codes_under_pt(self):
        """Under PT the reference table lists the official E25 code with its
        official description, the Country column is sourced from the
        jurisdiction (not a hardcoded literal), and the code matches the
        consolidated owner."""
        wb = openpyxl.Workbook()
        report = _make_crypto_tax_report()
        write_crypto_supplementary_sheet(wb, report, build_koinly_jurisdiction())
        ws = wb["Crypto Supplementary"]
        header_row = None
        for r in range(1, ws.max_row + 1):
            if ws.cell(r, 1).value == "Country":
                header_row = r
                break
        assert header_row is not None

        # Single official code under PT
        assert ws.cell(header_row + 1, 1).value == "PT"
        assert ws.cell(header_row + 1, 2).value == "E25"
        assert ws.cell(header_row + 1, 3).value == _INCOME_CODE_DESCRIPTIONS["E25"]
        # No second code row before section 2
        assert ws.cell(header_row + 2, 1).value != "PT"

    def test_income_code_reference_omitted_when_classification_off(self):
        """When classify_rewards_with_income_codes is False the entire
        '1. INCOME CODES REFERENCE' section is omitted (structural absence,
        not field-blanked)."""
        wb = openpyxl.Workbook()
        report = _make_crypto_tax_report()
        write_crypto_supplementary_sheet(
            wb, report, build_koinly_jurisdiction(country="DE", classify_rewards_with_income_codes=False)
        )
        ws = wb["Crypto Supplementary"]

        for r in range(1, ws.max_row + 1):
            val = ws.cell(r, 1).value
            assert val != "1. INCOME CODES REFERENCE"
            assert val != "Country"
        # The first visible section is renumbered to "1." so the list does not
        # start at "2." (which would imply a missing predecessor).
        found_renumbered_first = False
        for r in range(1, ws.max_row + 1):
            if ws.cell(r, 1).value == "1. TAXABLE-NOW - SUPPORT DETAIL":
                found_renumbered_first = True
                break
        assert found_renumbered_first


@pytest.mark.unit
class TestCryptoSupplementarySheetTaxableNowDetail:
    """Tests for section 2: Taxable-now support detail."""

    DETAIL_HEADERS = [
        "Date",
        "Asset",
        "Value (EUR)",
        "Income type",
        "Wallet",
        "Platform",
        "Reward chain",
        "Country",
        "Foreign tax (EUR)",
        "Review flag",
        "Description",
    ]

    def test_section_2_title_written(self):
        wb = openpyxl.Workbook()
        report = _make_crypto_tax_report()
        write_crypto_supplementary_sheet(wb, report, build_koinly_jurisdiction())
        ws = wb["Crypto Supplementary"]
        found = False
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=1):
            if row[0].value == "2. TAXABLE-NOW - SUPPORT DETAIL":
                found = True
                break
        assert found

    def test_section_2_title_is_bold(self):
        wb = openpyxl.Workbook()
        report = _make_crypto_tax_report()
        write_crypto_supplementary_sheet(wb, report, build_koinly_jurisdiction())
        ws = wb["Crypto Supplementary"]
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=1):
            if row[0].value == "2. TAXABLE-NOW - SUPPORT DETAIL":
                assert row[0].font.bold is True
                break

    def test_taxable_note_mentions_reporting_worksheet(self):
        wb = openpyxl.Workbook()
        report = _make_crypto_tax_report()
        write_crypto_supplementary_sheet(wb, report, build_koinly_jurisdiction())
        ws = wb["Crypto Supplementary"]
        found = False
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=1):
            val = row[0].value
            if val and "Reporting worksheet" in str(val) and "OTHER CAPITAL INVESTMENT INCOME" in str(val):
                found = True
                break
        assert found, "Expected note to mention Reporting worksheet and OTHER CAPITAL INVESTMENT INCOME"

    def test_detail_headers_written(self):
        wb = openpyxl.Workbook()
        report = _make_crypto_tax_report(reward_entries=[_make_reward_entry()])
        write_crypto_supplementary_sheet(wb, report, build_koinly_jurisdiction())
        ws = wb["Crypto Supplementary"]
        section_row = None
        for r in range(1, ws.max_row + 1):
            if ws.cell(r, 1).value == "2. TAXABLE-NOW - SUPPORT DETAIL":
                section_row = r
                break
        assert section_row is not None
        header_row = None
        for r in range(section_row, ws.max_row + 1):
            if ws.cell(r, 1).value == "Date":
                header_row = r
                break
        assert header_row is not None
        for idx, expected in enumerate(self.DETAIL_HEADERS, start=1):
            assert ws.cell(header_row, idx).value == expected

    def test_taxable_now_entry_values(self):
        entry = _make_reward_entry(
            date="2025-03-15",
            asset="ETH",
            value_eur=Decimal("1500"),
            source_type="staking",
            wallet="Kraken",
            platform="Kraken",
            chain="Ethereum",
        )
        report = _make_crypto_tax_report(reward_entries=[entry])
        wb = openpyxl.Workbook()
        write_crypto_supplementary_sheet(wb, report, build_koinly_jurisdiction())
        ws = wb["Crypto Supplementary"]
        header_row = None
        for r in range(1, ws.max_row + 1):
            if ws.cell(r, 1).value == "Date" and ws.cell(r, 4).value == "Income type":
                header_row = r
                break
        assert header_row is not None
        data_row = header_row + 1
        assert ws.cell(data_row, 1).value == "2025-03-15"
        assert ws.cell(data_row, 2).value == "ETH"
        assert ws.cell(data_row, 3).value == float(Decimal("1500"))
        assert ws.cell(data_row, 4).value == "staking"
        assert ws.cell(data_row, 5).value == "Kraken"
        assert ws.cell(data_row, 6).value == "Kraken"
        assert ws.cell(data_row, 7).value == "Ethereum"
        assert ws.cell(data_row, 8).value == "US"
        assert ws.cell(data_row, 9).value == float(Decimal("0"))
        assert ws.cell(data_row, 10).value == "NO"
        assert ws.cell(data_row, 11).value == "Staking reward payout"

    def test_taxable_now_review_flag_yes_with_reason(self):
        entry = _make_reward_entry(review_required=True, review_reason="Missing cost basis")
        report = _make_crypto_tax_report(reward_entries=[entry])
        wb = openpyxl.Workbook()
        write_crypto_supplementary_sheet(wb, report, build_koinly_jurisdiction())
        ws = wb["Crypto Supplementary"]
        header_row = None
        for r in range(1, ws.max_row + 1):
            if ws.cell(r, 1).value == "Date" and ws.cell(r, 4).value == "Income type":
                header_row = r
                break
        data_row = header_row + 1
        assert ws.cell(data_row, 10).value == "YES: Missing cost basis"

    def test_no_taxable_now_entries_shows_note(self):
        deferred = _make_reward_entry(classification=RewardTaxClassification.DEFERRED_BY_LAW)
        report = _make_crypto_tax_report(reward_entries=[deferred])
        wb = openpyxl.Workbook()
        write_crypto_supplementary_sheet(wb, report, build_koinly_jurisdiction())
        ws = wb["Crypto Supplementary"]
        found = False
        section_start = None
        for r in range(1, ws.max_row + 1):
            if ws.cell(r, 1).value == "2. TAXABLE-NOW - SUPPORT DETAIL":
                section_start = r
                break
        assert section_start is not None
        for r in range(section_start, ws.max_row + 1):
            val = ws.cell(r, 1).value
            if val and "No taxable-now rewards" in str(val):
                found = True
                break
        assert found


@pytest.mark.unit
class TestCryptoSupplementarySheetDeferredDetail:
    """Tests for section 3: Deferred by law support detail."""

    def test_section_3_title_written(self):
        wb = openpyxl.Workbook()
        report = _make_crypto_tax_report()
        write_crypto_supplementary_sheet(wb, report, build_koinly_jurisdiction())
        ws = wb["Crypto Supplementary"]
        found = False
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=1):
            if row[0].value == "3. DEFERRED BY LAW - SUPPORT DETAIL":
                found = True
                break
        assert found

    def test_section_3_title_is_bold(self):
        wb = openpyxl.Workbook()
        report = _make_crypto_tax_report()
        write_crypto_supplementary_sheet(wb, report, build_koinly_jurisdiction())
        ws = wb["Crypto Supplementary"]
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=1):
            if row[0].value == "3. DEFERRED BY LAW - SUPPORT DETAIL":
                assert row[0].font.bold is True
                break

    def test_deferred_note_is_italic(self):
        wb = openpyxl.Workbook()
        report = _make_crypto_tax_report()
        write_crypto_supplementary_sheet(wb, report, build_koinly_jurisdiction())
        ws = wb["Crypto Supplementary"]
        found = False
        section_start = None
        for r in range(1, ws.max_row + 1):
            if ws.cell(r, 1).value == "3. DEFERRED BY LAW - SUPPORT DETAIL":
                section_start = r
                break
        assert section_start is not None
        for r in range(section_start, ws.max_row + 1):
            val = ws.cell(r, 1).value
            if val and "deferred until disposal" in str(val):
                assert ws.cell(r, 1).font.italic is True
                found = True
                break
        assert found

    def test_deferred_entry_values(self):
        deferred = _make_reward_entry(
            classification=RewardTaxClassification.DEFERRED_BY_LAW,
            date="2025-04-01",
            asset="BTC",
            value_eur=Decimal("500"),
            source_type="mining",
            wallet="Ledger",
            platform="Ledger",
            chain="Bitcoin",
            description="Mining payout",
        )
        report = _make_crypto_tax_report(reward_entries=[deferred])
        wb = openpyxl.Workbook()
        write_crypto_supplementary_sheet(wb, report, build_koinly_jurisdiction())
        ws = wb["Crypto Supplementary"]
        section_start = None
        for r in range(1, ws.max_row + 1):
            if ws.cell(r, 1).value == "3. DEFERRED BY LAW - SUPPORT DETAIL":
                section_start = r
                break
        assert section_start is not None
        header_row = None
        for r in range(section_start, ws.max_row + 1):
            if ws.cell(r, 1).value == "Date" and ws.cell(r, 2).value == "Asset":
                header_row = r
                break
        assert header_row is not None
        data_row = header_row + 1
        assert ws.cell(data_row, 1).value == "2025-04-01"
        assert ws.cell(data_row, 2).value == "BTC"
        assert ws.cell(data_row, 3).value == float(Decimal("500"))
        assert ws.cell(data_row, 4).value == "mining"
        assert ws.cell(data_row, 10).value == "NO"
        assert ws.cell(data_row, 11).value == "Mining payout"

    def test_no_deferred_entries_shows_note(self):
        taxable = _make_reward_entry(classification=RewardTaxClassification.TAXABLE_NOW)
        report = _make_crypto_tax_report(reward_entries=[taxable])
        wb = openpyxl.Workbook()
        write_crypto_supplementary_sheet(wb, report, build_koinly_jurisdiction())
        ws = wb["Crypto Supplementary"]
        section_start = None
        for r in range(1, ws.max_row + 1):
            if ws.cell(r, 1).value == "3. DEFERRED BY LAW - SUPPORT DETAIL":
                section_start = r
                break
        assert section_start is not None
        found = False
        for r in range(section_start, ws.max_row + 1):
            val = ws.cell(r, 1).value
            if val and "No deferred rewards" in str(val):
                found = True
                break
        assert found


@pytest.mark.unit
class TestCryptoSupplementarySheetClassificationReconciliation:
    """Tests for section 4: Rewards classification reconciliation."""

    def test_section_4_title_written(self):
        wb = openpyxl.Workbook()
        report = _make_crypto_tax_report()
        write_crypto_supplementary_sheet(wb, report, build_koinly_jurisdiction())
        ws = wb["Crypto Supplementary"]
        found = False
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=1):
            if row[0].value == "4. REWARDS CLASSIFICATION RECONCILIATION":
                found = True
                break
        assert found

    def test_section_4_title_is_bold(self):
        wb = openpyxl.Workbook()
        report = _make_crypto_tax_report()
        write_crypto_supplementary_sheet(wb, report, build_koinly_jurisdiction())
        ws = wb["Crypto Supplementary"]
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=1):
            if row[0].value == "4. REWARDS CLASSIFICATION RECONCILIATION":
                assert row[0].font.bold is True
                break

    def test_section4_splits_detail_and_dust_counts(self):
        """Section 4 reconciliation splits the taxable-now count into detail + dust.

        Given a CryptoTaxReport with 2 detail taxable-now rows (BTC priced) and
        3 dust taxable-now rows (BTC zero-value, priced elsewhere in the export),
        Section 4 must render the new split lines
        ``("Taxable-now detail rows", 2)`` and
        ``("Taxable-now dust rows (suppressed from detail)", 3)`` and must NOT
        render the old ``("Taxable-now rows (immediately taxable)", 5)`` line.

        RED for Task 5: the Section 4 reconciliation still emits the old single
        line; Task 6 flips it GREEN by replacing the single reconciliation row
        with the two split rows.
        """
        # Two priced BTC taxable-now rows -> detail (real_rows).
        btc_real_1 = _make_reward_entry(
            asset="BTC", value_eur=Decimal("0.50"), wallet="Demo Spot", platform="Demo Spot", chain="Bitcoin"
        )
        btc_real_2 = _make_reward_entry(
            asset="BTC", value_eur=Decimal("0.30"), wallet="Wirex", platform="Wirex", chain="Bitcoin"
        )
        # Three zero-value BTC taxable-now rows -> dust (BTC has priced rows above).
        btc_dust_1 = _make_reward_entry(
            asset="BTC", value_eur=Decimal("0"), wallet="Demo Spot", platform="Demo Spot", chain="Bitcoin"
        )
        btc_dust_2 = _make_reward_entry(
            asset="BTC", value_eur=Decimal("0"), wallet="Wirex", platform="Wirex", chain="Bitcoin"
        )
        btc_dust_3 = _make_reward_entry(
            asset="BTC", value_eur=Decimal("0"), wallet="Demo Spot", platform="Demo Spot", chain="Bitcoin"
        )
        report = _make_crypto_tax_report(
            reward_entries=[btc_real_1, btc_real_2, btc_dust_1, btc_dust_2, btc_dust_3]
        )
        wb = openpyxl.Workbook()
        write_crypto_supplementary_sheet(wb, report, build_koinly_jurisdiction())
        ws = wb["Crypto Supplementary"]
        keys = _section4_reconciliation_keys(ws)
        assert keys["Taxable-now detail rows"] == 2, (
            f"expected new 'Taxable-now detail rows' split line with value 2, got keys={keys}"
        )
        assert keys["Taxable-now dust rows (suppressed from detail)"] == 3, (
            f"expected new 'Taxable-now dust rows (suppressed from detail)' split line with value 3, "
            f"got keys={keys}"
        )
        # The old single-line reconciliation row must be absent after the split.
        assert "Taxable-now rows (immediately taxable)" not in keys, (
            f"old 'Taxable-now rows (immediately taxable)' line still present after split; keys={keys}"
        )

    def test_section4_splits_deferred_detail_dust_unpriced(self):
        """Section 4 reconciliation splits the deferred count into detail + dust + unpriced.

        Given a CryptoTaxReport with 2 non-zero deferred rows, 3 dust deferred rows
        (in ``skipped_zero_value_deferred_rewards``, priced assets) and 4 unpriced
        deferred rows (in ``skipped_zero_value_deferred_rewards``, unpriced assets),
        Section 4 must render the new three-line split
        ``("Deferred detail rows", 2)``,
        ``("Deferred dust rows (suppressed from detail)", 3)`` and
        ``("Deferred unpriced rows (suppressed from detail)", 4)`` and must NOT
        render the old ``("Deferred-by-law rows (taxation deferred)", 9)`` line NOR
        a ``("Skipped zero-value deferred rewards (audit)", ...)`` line (r1 review
        finding #2: the 4th audit line is tautological and dropped on the
        Supplementary sheet - it lives on the Reconciliation sheet instead, where
        it makes the cross-sheet count auditable).

        RED for Task 5: the Section 4 reconciliation still emits the old single
        deferred line; Task 5 flips it GREEN by replacing the single
        reconciliation row with the three split rows, reusing the partition
        computed once in Task 4 (Invariant 5: compute-once-reuse).
        """
        # Two priced WBERA + WBTC deferred rows -> deferred detail (deferred_entries).
        wbera_real = _make_reward_entry(
            classification=RewardTaxClassification.DEFERRED_BY_LAW,
            asset="WBERA",
            value_eur=Decimal("1.00"),
            wallet="Wirex",
            platform="Wirex",
            chain="Berachain",
            amount=Decimal("1"),
        )
        wbtc_real = _make_reward_entry(
            classification=RewardTaxClassification.DEFERRED_BY_LAW,
            asset="WBTC",
            value_eur=Decimal("2.00"),
            wallet="Wirex",
            platform="Wirex",
            chain="Bitcoin",
            amount=Decimal("1"),
        )
        # Three zero-value WBERA/WBTC deferred rows -> dust (priced assets, in
        # skipped_zero_value_deferred_rewards; WBERA/WBTC have priced rows above).
        wbera_dust_1 = _make_reward_entry(
            classification=RewardTaxClassification.DEFERRED_BY_LAW,
            asset="WBERA",
            value_eur=Decimal("0"),
            wallet="Wirex",
            platform="Wirex",
            chain="Berachain",
            amount=Decimal("1"),
        )
        wbera_dust_2 = _make_reward_entry(
            classification=RewardTaxClassification.DEFERRED_BY_LAW,
            asset="WBERA",
            value_eur=Decimal("0"),
            wallet="Kraken",
            platform="Kraken",
            chain="Berachain",
            amount=Decimal("1"),
        )
        wbtc_dust_1 = _make_reward_entry(
            classification=RewardTaxClassification.DEFERRED_BY_LAW,
            asset="WBTC",
            value_eur=Decimal("0"),
            wallet="Wirex",
            platform="Wirex",
            chain="Bitcoin",
            amount=Decimal("1"),
        )
        # Four zero-value OSBGT/BNB deferred rows -> unpriced (no priced row
        # anywhere; in skipped_zero_value_deferred_rewards).
        osbgt_unpriced_1 = _make_reward_entry(
            classification=RewardTaxClassification.DEFERRED_BY_LAW,
            asset="OSBGT",
            value_eur=Decimal("0"),
            wallet="Wirex",
            platform="Wirex",
            chain="Berachain",
            amount=Decimal("1"),
        )
        osbgt_unpriced_2 = _make_reward_entry(
            classification=RewardTaxClassification.DEFERRED_BY_LAW,
            asset="OSBGT",
            value_eur=Decimal("0"),
            wallet="Kraken",
            platform="Kraken",
            chain="Berachain",
            amount=Decimal("1"),
        )
        bnb_unpriced_1 = _make_reward_entry(
            classification=RewardTaxClassification.DEFERRED_BY_LAW,
            asset="BNB",
            value_eur=Decimal("0"),
            wallet="Kraken",
            platform="Kraken",
            chain="Binance",
            amount=Decimal("1"),
        )
        bnb_unpriced_2 = _make_reward_entry(
            classification=RewardTaxClassification.DEFERRED_BY_LAW,
            asset="BNB",
            value_eur=Decimal("0"),
            wallet="Wirex",
            platform="Wirex",
            chain="Binance",
            amount=Decimal("1"),
        )
        report = _make_crypto_tax_report(
            reward_entries=[wbera_real, wbtc_real],
            skipped_deferred_rewards=[
                wbera_dust_1,
                wbera_dust_2,
                wbtc_dust_1,
                osbgt_unpriced_1,
                osbgt_unpriced_2,
                bnb_unpriced_1,
                bnb_unpriced_2,
            ],
        )
        wb = openpyxl.Workbook()
        write_crypto_supplementary_sheet(wb, report, build_koinly_jurisdiction())
        ws = wb["Crypto Supplementary"]
        keys = _section4_reconciliation_keys(ws)
        assert keys["Deferred detail rows"] == 2, (
            f"expected new 'Deferred detail rows' split line with value 2, got keys={keys}"
        )
        assert keys["Deferred dust rows (suppressed from detail)"] == 3, (
            f"expected new 'Deferred dust rows (suppressed from detail)' split line with value 3, "
            f"got keys={keys}"
        )
        assert keys["Deferred unpriced rows (suppressed from detail)"] == 4, (
            f"expected new 'Deferred unpriced rows (suppressed from detail)' split line with value 4, "
            f"got keys={keys}"
        )
        # The old single-line deferred reconciliation row must be absent after
        # the split.
        assert "Deferred-by-law rows (taxation deferred)" not in keys, (
            f"old 'Deferred-by-law rows (taxation deferred)' line still present after split; keys={keys}"
        )
        # r1 review finding #2: the audit line is tautological on the Supplementary
        # sheet (dust + unpriced == total by construction) and lives only on the
        # Reconciliation sheet.
        assert "Skipped zero-value deferred rewards (audit)" not in keys, (
            f"tautological audit line present on Supplementary sheet; keys={keys}"
        )

    def test_reconciliation_key_value_pairs(self):
        taxable = _make_reward_entry(classification=RewardTaxClassification.TAXABLE_NOW, value_eur=Decimal("100"))
        deferred = _make_reward_entry(
            classification=RewardTaxClassification.DEFERRED_BY_LAW, value_eur=Decimal("200"), asset="BTC"
        )
        report = _make_crypto_tax_report(reward_entries=[taxable, deferred])
        wb = openpyxl.Workbook()
        write_crypto_supplementary_sheet(wb, report, build_koinly_jurisdiction())
        ws = wb["Crypto Supplementary"]
        keys = _section4_reconciliation_keys(ws)
        # Unchanged reconciliation lines (preserved intent): the raw count and
        # the EUR totals are still rendered.
        assert keys["Total reward rows (raw)"] == 2
        assert keys["Taxable-now total value (EUR)"] == float(Decimal("100"))
        assert keys["Deferred total value (EUR)"] == float(Decimal("200"))
        # New split lines replace the old single taxable-now count. The priced
        # ETH taxable-now row stays in detail (real_rows), and there are no
        # zero-value rows on a priced asset, so dust_rows is empty.
        assert keys["Taxable-now detail rows"] == 1, (
            f"expected new 'Taxable-now detail rows' split line with value 1, got keys={keys}"
        )
        assert keys["Taxable-now dust rows (suppressed from detail)"] == 0, (
            f"expected new 'Taxable-now dust rows (suppressed from detail)' split line with value 0, "
            f"got keys={keys}"
        )
        # New three-line deferred split replaces the old single deferred count.
        # The priced BTC deferred row stays in detail (deferred_entries), and
        # skipped_zero_value_deferred_rewards is empty so dust + unpriced are 0.
        assert keys["Deferred detail rows"] == 1, (
            f"expected new 'Deferred detail rows' split line with value 1, got keys={keys}"
        )
        assert keys["Deferred dust rows (suppressed from detail)"] == 0, (
            f"expected new 'Deferred dust rows (suppressed from detail)' split line with value 0, "
            f"got keys={keys}"
        )
        assert keys["Deferred unpriced rows (suppressed from detail)"] == 0, (
            f"expected new 'Deferred unpriced rows (suppressed from detail)' split line with value 0, "
            f"got keys={keys}"
        )
        # The old single-line reconciliation rows must be absent after the split.
        assert "Taxable-now rows (immediately taxable)" not in keys, (
            f"old 'Taxable-now rows (immediately taxable)' line still present after split; keys={keys}"
        )
        assert "Deferred-by-law rows (taxation deferred)" not in keys, (
            f"old 'Deferred-by-law rows (taxation deferred)' line still present after split; keys={keys}"
        )

    def test_reconciliation_empty_rewards(self):
        report = _make_crypto_tax_report(reward_entries=[])
        wb = openpyxl.Workbook()
        write_crypto_supplementary_sheet(wb, report, build_koinly_jurisdiction())
        ws = wb["Crypto Supplementary"]
        keys = _section4_reconciliation_keys(ws)
        # Empty-path regression guard: both new split lines render with 0,
        # replacing the old single ("Taxable-now rows (immediately taxable)", 0).
        assert keys["Taxable-now detail rows"] == 0, (
            f"expected 'Taxable-now detail rows' == 0 on empty path, got keys={keys}"
        )
        assert keys["Taxable-now dust rows (suppressed from detail)"] == 0, (
            f"expected 'Taxable-now dust rows (suppressed from detail)' == 0 on empty path, got keys={keys}"
        )
        assert "Taxable-now rows (immediately taxable)" not in keys, (
            f"old 'Taxable-now rows (immediately taxable)' line still present on empty path; keys={keys}"
        )


@pytest.mark.unit
class TestCryptoSupplementarySheetAutoWidth:
    """Tests that auto_column_width is called."""

    def test_auto_width_adjusts_columns(self):
        wb = openpyxl.Workbook()
        report = _make_crypto_tax_report()
        write_crypto_supplementary_sheet(wb, report, build_koinly_jurisdiction())
        ws = wb["Crypto Supplementary"]
        assert ws.column_dimensions["A"].width > 0

    def test_column_widths_respect_max_cell_width_cap(self):
        """Verify no column exceeds MAX_CELL_WIDTH + 2 even with long notes."""
        from tax_reporting.application.persisting.excel_utils import MAX_CELL_WIDTH

        wb = openpyxl.Workbook()
        report = _make_crypto_tax_report()
        write_crypto_supplementary_sheet(wb, report, build_koinly_jurisdiction())
        ws = wb["Crypto Supplementary"]

        # All column widths should be capped at MAX_CELL_WIDTH + 2
        max_allowed = MAX_CELL_WIDTH + 2
        for col_letter, col_dim in ws.column_dimensions.items():
            if col_dim.width is not None:
                assert (
                    col_dim.width <= max_allowed
                ), f"Column {col_letter} width {col_dim.width} exceeds cap {max_allowed}"

    def test_all_columns_have_reasonable_widths(self):
        """Verify no column is collapsed: all have width >= MIN_DATA_WIDTH floor."""

        wb = openpyxl.Workbook()
        report = _make_crypto_tax_report()
        write_crypto_supplementary_sheet(wb, report, build_koinly_jurisdiction())
        ws = wb["Crypto Supplementary"]

        for col_idx in range(1, 12):  # Columns A through K
            col_letter = openpyxl.utils.get_column_letter(col_idx)
            width = ws.column_dimensions[col_letter].width
            assert width is not None, f"Column {col_letter} should have a width set"
            assert width >= 4, f"Column {col_letter} width {width} is too small"


@pytest.mark.unit
class TestCryptoSupplementarySheetReviewRequired:
    """Tests for Section 5: REVIEW REQUIRED."""

    def test_review_required_section_title(self):
        """Test that the Review Required section title is rendered correctly."""
        review_entries = [
            CryptoReviewEntry(
                source_section="capital_gains",
                date="2025-01-15",
                asset="BTC",
                platform="Kraken",
                review_reason="Zero cost basis",
            ),
        ]
        wb = openpyxl.Workbook()
        report = _make_crypto_tax_report(review_entries=review_entries)
        write_crypto_supplementary_sheet(wb, report, build_koinly_jurisdiction())
        ws = wb["Crypto Supplementary"]

        # Find the section title
        section_row = None
        for r in range(1, ws.max_row + 1):
            if ws.cell(r, 1).value == "5. REVIEW REQUIRED":
                section_row = r
                break
        assert section_row is not None
        assert ws.cell(section_row, 1).font.bold is True

    def test_review_required_headers(self):
        """Test that Review Required section headers are rendered correctly."""
        review_entries = [
            CryptoReviewEntry(
                source_section="income",
                date="2025-01-15",
                asset="ETH",
                platform="ByBit",
                review_reason="Missing price data",
            ),
        ]
        wb = openpyxl.Workbook()
        report = _make_crypto_tax_report(review_entries=review_entries)
        write_crypto_supplementary_sheet(wb, report, build_koinly_jurisdiction())
        ws = wb["Crypto Supplementary"]

        # Find the header row (after section title and note)
        header_row = None
        for r in range(1, ws.max_row + 1):
            if ws.cell(r, 1).value == "Source":
                header_row = r
                break
        assert header_row is not None
        assert ws.cell(header_row, 1).value == "Source"
        assert ws.cell(header_row, 2).value == "Date"
        assert ws.cell(header_row, 3).value == "Asset"
        assert ws.cell(header_row, 4).value == "Platform"
        assert ws.cell(header_row, 5).value == "Review reason"
        # Verify headers are bold
        for col in range(1, 6):
            assert ws.cell(header_row, col).font.bold is True

    def test_review_required_data_rows(self):
        """Test that review entry data rows are rendered correctly."""
        review_entries = [
            CryptoReviewEntry(
                source_section="capital_gains",
                date="2025-01-15",
                asset="BTC",
                platform="Kraken",
                review_reason="Zero cost basis",
            ),
            CryptoReviewEntry(
                source_section="income",
                date="2025-01-20",
                asset="ETH",
                platform="ByBit",
                review_reason="Missing price data",
            ),
        ]
        wb = openpyxl.Workbook()
        report = _make_crypto_tax_report(review_entries=review_entries)
        write_crypto_supplementary_sheet(wb, report, build_koinly_jurisdiction())
        ws = wb["Crypto Supplementary"]

        # Find the data start row (after headers)
        header_row = None
        for r in range(1, ws.max_row + 1):
            if ws.cell(r, 1).value == "Source":
                header_row = r
                break
        assert header_row is not None

        # First data row
        data_row = header_row + 1
        assert ws.cell(data_row, 1).value == "Capital Gains"
        assert ws.cell(data_row, 2).value == "2025-01-15"
        assert ws.cell(data_row, 3).value == "BTC"
        assert ws.cell(data_row, 4).value == "Kraken"
        assert ws.cell(data_row, 5).value == "Zero cost basis"

        # Second data row
        data_row += 1
        assert ws.cell(data_row, 1).value == "Income"
        assert ws.cell(data_row, 2).value == "2025-01-20"
        assert ws.cell(data_row, 3).value == "ETH"
        assert ws.cell(data_row, 4).value == "ByBit"
        assert ws.cell(data_row, 5).value == "Missing price data"

    def test_review_required_suspicious_flag_formatting(self):
        """Test that suspicious assets are highlighted with red font."""
        review_entries = [
            CryptoReviewEntry(
                source_section="income",
                date="2025-01-15",
                asset="BTC",
                platform="Kraken",
                review_reason="Zero cost basis",
                is_suspicious=False,
            ),
            CryptoReviewEntry(
                source_section="capital_gains",
                date="2025-01-20",
                asset="РUB",
                platform="ByBit",
                review_reason="Non-Latin characters",
                is_suspicious=True,
            ),
        ]
        wb = openpyxl.Workbook()
        report = _make_crypto_tax_report(review_entries=review_entries)
        write_crypto_supplementary_sheet(wb, report, build_koinly_jurisdiction())
        ws = wb["Crypto Supplementary"]

        # Find the data start row
        header_row = None
        for r in range(1, ws.max_row + 1):
            if ws.cell(r, 1).value == "Source":
                header_row = r
                break
        assert header_row is not None

        # First row - not suspicious, should have default formatting
        non_suspicious_row = header_row + 1
        asset_cell = ws.cell(non_suspicious_row, 3)
        assert asset_cell.value == "BTC"
        assert asset_cell.font.bold is False
        # Default color (None or theme color) - not the suspicious red
        assert asset_cell.font.color.rgb not in ("FFFF0000", "00FF0000")

        # Second row - suspicious, should have red bold formatting
        suspicious_row = header_row + 2
        asset_cell = ws.cell(suspicious_row, 3)
        assert asset_cell.value == "РUB"
        assert asset_cell.font.bold is True
        assert asset_cell.font.color.rgb in ("FFFF0000", "00FF0000")  # Red text (alpha may vary)

    def test_review_required_no_items_shows_message(self):
        """Test that 'No review items' message is shown when there are no review entries."""
        wb = openpyxl.Workbook()
        report = _make_crypto_tax_report(review_entries=[])
        write_crypto_supplementary_sheet(wb, report, build_koinly_jurisdiction())
        ws = wb["Crypto Supplementary"]

        # Find the header row
        header_row = None
        for r in range(1, ws.max_row + 1):
            if ws.cell(r, 1).value == "Source":
                header_row = r
                break
        assert header_row is not None

        # Data row should show "No review items"
        data_row = header_row + 1
        assert ws.cell(data_row, 1).value == "No review items"


def _section2_bounds(ws) -> tuple[int, int]:
    """Return the (start_row, end_row_exclusive) for Section 2.

    ``end_row_exclusive`` is the row of the next section title (3. DEFERRED
    BY LAW) or ``ws.max_row + 1`` if not found, so callers can iterate
    Section 2 with ``range(start, end)``.
    """
    start = None
    for r in range(1, ws.max_row + 1):
        if ws.cell(r, 1).value == "2. TAXABLE-NOW - SUPPORT DETAIL":
            start = r
            break
    assert start is not None, "Section 2 title not rendered"
    end = ws.max_row + 1
    for r in range(start + 1, ws.max_row + 1):
        val = ws.cell(r, 1).value
        if isinstance(val, str) and val.startswith("3."):
            end = r
            break
    return start, end


def _section2_strings(ws) -> list[str]:
    """Collect all column-A string cell values rendered inside Section 2."""
    start, end = _section2_bounds(ws)
    out: list[str] = []
    for r in range(start, end):
        val = ws.cell(r, 1).value
        if isinstance(val, str):
            out.append(val)
    return out


def _section2_reward_detail_rows(ws) -> list[dict[int, object]]:
    """Return Section 2 detail data rows (skipping the header row, labels, and the dust block).

    A row is treated as a detail row iff column 2 (Asset) is populated AND
    column 4 (Income type) is populated AND column 1 is not the header row's
    ``"Date"`` and not a label/dust line. Column 4 (``source_type``) is the
    cleanest discriminator: it is empty on title/note/label/dust rows and on
    the header row's empty trailing cells, but populated on every real detail
    row produced by ``_write_reward_detail_rows``.
    """
    start, end = _section2_bounds(ws)
    rows: list[dict[int, object]] = []
    skip_prefixes = (
        "Dust summary:",
        "Taxable-now dust",
        "No taxable-now",
        "All taxable-now",
    )
    for r in range(start, end):
        c1 = ws.cell(r, 1).value
        c2 = ws.cell(r, 2).value
        c4 = ws.cell(r, 4).value
        if c2 in (None, ""):
            continue
        if c4 in (None, ""):
            continue
        if isinstance(c1, str) and (c1 == "Date" or c1.startswith(skip_prefixes)):
            continue
        row: dict[int, object] = {}
        for col in range(1, 12):
            row[col] = ws.cell(r, col).value
        rows.append(row)
    return rows


def _section3_bounds(ws) -> tuple[int, int]:
    """Return the (start_row, end_row_exclusive) for Section 3.

    ``end_row_exclusive`` is the row of the next section title (4. REWARDS
    CLASSIFICATION RECONCILIATION) or ``ws.max_row + 1`` if not found, so
    callers can iterate Section 3 with ``range(start, end)``.
    """
    start = None
    for r in range(1, ws.max_row + 1):
        if ws.cell(r, 1).value == "3. DEFERRED BY LAW - SUPPORT DETAIL":
            start = r
            break
    assert start is not None, "Section 3 title not rendered"
    end = ws.max_row + 1
    for r in range(start + 1, ws.max_row + 1):
        val = ws.cell(r, 1).value
        if isinstance(val, str) and val.startswith("4."):
            end = r
            break
    return start, end


def _section3_strings(ws) -> list[str]:
    """Collect all column-A string cell values rendered inside Section 3."""
    start, end = _section3_bounds(ws)
    out: list[str] = []
    for r in range(start, end):
        val = ws.cell(r, 1).value
        if isinstance(val, str):
            out.append(val)
    return out


def _section3_reward_detail_rows(ws) -> list[dict[int, object]]:
    """Return Section 3 detail data rows (skipping header, labels, and the suppressed block).

    Mirrors :func:`_section2_reward_detail_rows` for the deferred section. A row
    is a detail row iff column 2 (Asset) and column 4 (Income type) are both
    populated AND column 1 is not the header ``"Date"`` and not a label /
    suppressed-block line.
    """
    start, end = _section3_bounds(ws)
    rows: list[dict[int, object]] = []
    skip_prefixes = (
        "Suppressed zero-value deferred rewards",
        "Deferred dust",
        "Deferred unpriced",
        "No deferred rewards",
        "All deferred",
    )
    for r in range(start, end):
        c1 = ws.cell(r, 1).value
        c2 = ws.cell(r, 2).value
        c4 = ws.cell(r, 4).value
        if c2 in (None, ""):
            continue
        if c4 in (None, ""):
            continue
        if isinstance(c1, str) and (c1 == "Date" or c1.startswith(skip_prefixes)):
            continue
        row: dict[int, object] = {}
        for col in range(1, 12):
            row[col] = ws.cell(r, col).value
        rows.append(row)
    return rows


def _all_sheet_strings(ws) -> list[str]:
    """Collect every column-A string cell value rendered anywhere on the sheet.

    Used to scan for the single-block "Suppressed zero-value deferred rewards"
    header without having to predict which numbered section it falls under
    (it renders AFTER the deferred detail table, so it lives inside Section 3).
    """
    out: list[str] = []
    for r in range(1, ws.max_row + 1):
        val = ws.cell(r, 1).value
        if isinstance(val, str):
            out.append(val)
    return out


def _find_table_rows_after(ws, anchor_text: str, num_cols: int) -> list[dict[int, object]]:
    """Return the data rows of the column table whose header row is the first
    row after the cell containing ``anchor_text`` in column A.

    ``anchor_text`` is the sub-header line that immediately precedes the table's
    header row (e.g. ``"Taxable-now dust (priced-asset rounding)"`` or
    ``"Deferred dust (priced-asset rounding)"``). The next row is the table
    header (Asset | Wallet | Rows | ...), and subsequent rows are data rows
    until a blank cell or a new header/sub-header appears in column A.

    Returns each data row as ``{col_idx_1based: value}`` for ``num_cols`` cols.
    """
    # Locate the anchor row (sub-header).
    anchor_row = None
    for r in range(1, ws.max_row + 1):
        if ws.cell(r, 1).value == anchor_text:
            anchor_row = r
            break
    assert anchor_row is not None, f"anchor {anchor_text!r} not found"
    # Header row is the next row.
    header_row = anchor_row + 1
    rows: list[dict[int, object]] = []
    for r in range(header_row + 1, ws.max_row + 1):
        c1 = ws.cell(r, 1).value
        # End the table at a blank cell, a new section header, a new sub-header,
        # or a new outer header - whichever comes first.
        if c1 in (None, ""):
            break
        if isinstance(c1, str) and (
            c1.startswith(("Suppressed zero-value deferred rewards", "Deferred dust", "Deferred unpriced"))
            or c1.startswith(("Dust summary:", "Taxable-now dust"))
            or c1.startswith(("1.", "2.", "3.", "4.", "5."))
        ):
            break
        row: dict[int, object] = {}
        for col in range(1, num_cols + 1):
            row[col] = ws.cell(r, col).value
        rows.append(row)
    return rows


def _section4_reconciliation_keys(ws) -> dict[str, object]:
    """Return the Section 4 key-value pairs as a dict keyed by column-A label.

    Iterates from the ``4. REWARDS CLASSIFICATION RECONCILIATION`` title row
    through the next section title (or end of sheet), collecting every row
    whose column-A cell is a non-empty string into ``{key: col_b_value}``.
    Robust to the post-Task-6 split (the title is followed by N key/value rows
    rather than a fixed count), so callers can assert on individual keys
    without pinning row offsets.
    """
    start = None
    for r in range(1, ws.max_row + 1):
        if ws.cell(r, 1).value == "4. REWARDS CLASSIFICATION RECONCILIATION":
            start = r
            break
    assert start is not None, "Section 4 title not rendered"
    keys: dict[str, object] = {}
    for r in range(start + 1, ws.max_row + 1):
        key = ws.cell(r, 1).value
        if isinstance(key, str) and re.match(r"^\d+\.\s", key):
            # Hit the next numbered section title (e.g. "5. REVIEW REQUIRED").
            break
        if isinstance(key, str) and key:
            keys[key] = ws.cell(r, 2).value
    return keys


@pytest.mark.unit
class TestCryptoSupplementarySheetDustSummary:
    """Worksheet-level tests for the Section 2 dust partition (Part 7).

    RED for Task 3: ``write_crypto_supplementary_sheet`` does not yet partition
    taxable-now rows into (real, dust), does not render a "Dust summary:" block,
    and uses the old two-state empty label. These tests fail via assertion
    against the unchanged render; Task 4 (GREEN) adds the partition + helpers.
    """

    def test_btc_zero_collapses_to_dust(self):
        """A zero-value BTC row whose asset has a priced row collapses to dust."""
        btc_priced = _make_reward_entry(
            asset="BTC", value_eur=Decimal("0.50"), wallet="Demo Spot", platform="Demo Spot", chain="Bitcoin"
        )
        btc_zero = _make_reward_entry(
            asset="BTC", value_eur=Decimal("0"), wallet="Demo Spot", platform="Demo Spot", chain="Bitcoin"
        )
        report = _make_crypto_tax_report(reward_entries=[btc_priced, btc_zero])
        wb = openpyxl.Workbook()
        write_crypto_supplementary_sheet(wb, report, build_koinly_jurisdiction())
        ws = wb["Crypto Supplementary"]

        # The zero-value BTC row must NOT appear in the Section 2 detail table.
        for row in _section2_reward_detail_rows(ws):
            assert not (
                str(row.get(2, "")) == "BTC" and float(row.get(3, -1)) == 0.0
            ), "zero-value BTC row leaked into Section 2 detail table"

        # The dust summary block now renders as a column table under the
        # "Taxable-now dust (...)" sub-header. Asset=BTC, Wallet=Demo Spot,
        # Rows=1, Summed Value (EUR)=0.0, Category=dust.
        section_strs = _section2_strings(ws)
        assert any(s == "Dust summary:" for s in section_strs), "Dust summary header missing"
        table = _find_table_rows_after(ws, _TAXABLE_NOW_DUST_SUBHEADER, num_cols=5)
        btc_rows = [r for r in table if r.get(1) == "BTC"]
        assert len(btc_rows) == 1, (
            f"expected one BTC dust row in taxable-now table, got {btc_rows}; full table={table}"
        )
        assert btc_rows[0].get(2) == "Demo Spot", f"wallet wrong: {btc_rows[0]}"
        assert btc_rows[0].get(3) == 1, f"rows count wrong: {btc_rows[0]}"
        assert float(btc_rows[0].get(4)) == 0.0, f"summed value wrong: {btc_rows[0]}"
        assert btc_rows[0].get(5) == "dust", f"category wrong: {btc_rows[0]}"

    def test_osbgt_zero_keeps_per_row_yes(self):
        """OSBGT (no priced row anywhere) keeps per-row YES (the r9 headline fix).

        Fails under a popular-token discriminator because OSBGT *is* popular;
        the has-any-priced-row discriminator keeps it in detail.
        """
        osbgt = _make_reward_entry(
            asset="OSBGT",
            value_eur=Decimal("0"),
            wallet="Demo Spot",
            platform="Demo Spot",
            chain="Berachain",
            review_required=True,
            review_reason="Unpriced wrapper token - manual pricing required",
        )
        report = _make_crypto_tax_report(reward_entries=[osbgt])
        wb = openpyxl.Workbook()
        write_crypto_supplementary_sheet(wb, report, build_koinly_jurisdiction())
        ws = wb["Crypto Supplementary"]

        detail = _section2_reward_detail_rows(ws)
        osbgt_rows = [r for r in detail if str(r.get(2, "")) == "OSBGT"]
        assert osbgt_rows, "OSBGT row should appear in Section 2 detail table"
        review_flag = str(osbgt_rows[0].get(10, ""))
        assert review_flag.startswith("YES:"), (
            f"OSBGT review flag must start 'YES:' (r9 headline fix); got {review_flag!r}"
        )

        # No dust summary block should be rendered (nothing routed to dust).
        section_strs = _section2_strings(ws)
        assert not any(s == "Dust summary:" for s in section_strs), (
            "Dust summary rendered when no row routed to dust"
        )

    def test_bgt_zero_with_priced_row_collapses_to_dust(self):
        """BGT (popular AND priced) zero row collapses to dust.

        Fails if the discriminator is popular-set-only (which would keep BGT
        in detail because BGT is in the popular set).
        """
        bgt_priced = _make_reward_entry(
            asset="BGT", value_eur=Decimal("10"), wallet="Demo Spot", platform="Demo Spot", chain="Berachain"
        )
        bgt_zero = _make_reward_entry(
            asset="BGT", value_eur=Decimal("0"), wallet="Demo Spot", platform="Demo Spot", chain="Berachain"
        )
        report = _make_crypto_tax_report(reward_entries=[bgt_priced, bgt_zero])
        wb = openpyxl.Workbook()
        write_crypto_supplementary_sheet(wb, report, build_koinly_jurisdiction())
        ws = wb["Crypto Supplementary"]

        # The zero-value BGT row must NOT appear in Section 2 detail table.
        for row in _section2_reward_detail_rows(ws):
            assert not (
                str(row.get(2, "")) == "BGT" and float(row.get(3, -1)) == 0.0
            ), "zero-value BGT row leaked into Section 2 detail table (popular-set discriminator?)"

        section_strs = _section2_strings(ws)
        assert any(s == "Dust summary:" for s in section_strs), "Dust summary header missing"
        table = _find_table_rows_after(ws, _TAXABLE_NOW_DUST_SUBHEADER, num_cols=5)
        bgt_rows = [r for r in table if r.get(1) == "BGT"]
        assert len(bgt_rows) == 1, (
            f"expected one BGT dust row in taxable-now table, got {bgt_rows}; full table={table}"
        )
        assert bgt_rows[0].get(3) == 1, f"rows count wrong: {bgt_rows[0]}"
        assert float(bgt_rows[0].get(4)) == 0.0, f"summed value wrong: {bgt_rows[0]}"
        assert bgt_rows[0].get(5) == "dust", f"category wrong: {bgt_rows[0]}"

    def test_mixed_dust_and_detail_both_render(self):
        """A mix: priced BTC detail row, zero BTC dust row, zero OSBGT YES detail row."""
        btc_priced = _make_reward_entry(
            asset="BTC", value_eur=Decimal("0.50"), wallet="Demo Spot", platform="Demo Spot", chain="Bitcoin"
        )
        btc_zero = _make_reward_entry(
            asset="BTC", value_eur=Decimal("0"), wallet="Demo Spot", platform="Demo Spot", chain="Bitcoin"
        )
        osbgt_zero = _make_reward_entry(
            asset="OSBGT",
            value_eur=Decimal("0"),
            wallet="Demo Spot",
            platform="Demo Spot",
            chain="Berachain",
            review_required=True,
            review_reason="Unpriced wrapper token - manual pricing required",
        )
        report = _make_crypto_tax_report(reward_entries=[btc_priced, btc_zero, osbgt_zero])
        wb = openpyxl.Workbook()
        write_crypto_supplementary_sheet(wb, report, build_koinly_jurisdiction())
        ws = wb["Crypto Supplementary"]

        detail_assets = [str(r.get(2, "")) for r in _section2_reward_detail_rows(ws)]
        assert "BTC" in detail_assets, "priced BTC reward should appear in detail table"
        assert "OSBGT" in detail_assets, "zero-value OSBGT should appear in detail table"
        # Only the priced BTC row should appear (not the zero-value BTC).
        btc_rows_in_detail = [r for r in _section2_reward_detail_rows(ws) if str(r.get(2, "")) == "BTC"]
        assert all(float(r.get(3, -1)) != 0.0 for r in btc_rows_in_detail), (
            "zero-value BTC row leaked into detail in mixed scenario"
        )

        section_strs = _section2_strings(ws)
        assert any(s == "Dust summary:" for s in section_strs), "Dust summary header missing"
        table = _find_table_rows_after(ws, _TAXABLE_NOW_DUST_SUBHEADER, num_cols=5)
        btc_dust_rows = [r for r in table if r.get(1) == "BTC"]
        assert len(btc_dust_rows) == 1, (
            f"expected one BTC dust row in taxable-now table, got {btc_dust_rows}; full table={table}"
        )
        assert float(btc_dust_rows[0].get(4)) == 0.0, f"summed value wrong: {btc_dust_rows[0]}"

    def test_all_dust_empty_label(self):
        """All taxable-now rows are dust: detail table shows the all-dust label and a dust block renders."""
        # Two BTC wallets each with a priced row, plus a zero-value row per asset,
        # so every taxable-now row is dust. The priced rows are classified
        # DEFERRED_BY_LAW so they stay in reward_entries as discriminator evidence
        # (proving BTC/ETH are priced assets) without entering taxable_now_entries
        # (which would make them non-dust real_rows and defeat the all-dust case).
        btc_priced = _make_reward_entry(
            asset="BTC",
            value_eur=Decimal("1"),
            wallet="Demo Spot",
            platform="Demo Spot",
            chain="Bitcoin",
            classification=RewardTaxClassification.DEFERRED_BY_LAW,
        )
        btc_zero = _make_reward_entry(
            asset="BTC", value_eur=Decimal("0"), wallet="Demo Spot", platform="Demo Spot", chain="Bitcoin"
        )
        eth_priced = _make_reward_entry(
            asset="ETH",
            value_eur=Decimal("2"),
            wallet="Demo Spot",
            platform="Demo Spot",
            chain="Ethereum",
            classification=RewardTaxClassification.DEFERRED_BY_LAW,
        )
        eth_zero = _make_reward_entry(
            asset="ETH", value_eur=Decimal("0"), wallet="Demo Spot", platform="Demo Spot", chain="Ethereum"
        )
        report = _make_crypto_tax_report(reward_entries=[btc_priced, btc_zero, eth_priced, eth_zero])
        wb = openpyxl.Workbook()
        write_crypto_supplementary_sheet(wb, report, build_koinly_jurisdiction())
        ws = wb["Crypto Supplementary"]

        section_strs = _section2_strings(ws)
        assert any(
            s == "All taxable-now rows classified as dust - see summary below" for s in section_strs
        ), f"all-dust empty label missing in {section_strs}"
        assert any(s == "Dust summary:" for s in section_strs), "dust summary header missing in all-dust case"

    def test_no_rewards_empty_label_unchanged(self):
        """Regression guard: with no taxable-now entries, Section 2 still shows 'No taxable-now rewards'."""
        deferred = _make_reward_entry(classification=RewardTaxClassification.DEFERRED_BY_LAW)
        report = _make_crypto_tax_report(reward_entries=[deferred])
        wb = openpyxl.Workbook()
        write_crypto_supplementary_sheet(wb, report, build_koinly_jurisdiction())
        ws = wb["Crypto Supplementary"]

        section_strs = _section2_strings(ws)
        assert any(s == "No taxable-now rewards" for s in section_strs), (
            f"'No taxable-now rewards' label missing in {section_strs}"
        )
        # And no dust summary block should render.
        assert not any(s == "Dust summary:" for s in section_strs), (
            "dust summary rendered when there are no taxable-now entries"
        )

    def test_dust_line_empty_wallet_renders_explicitly(self):
        """Empty-string wallet (reachable production degradation) renders as
        Wallet="" (empty cell value, not a 'None' literal)."""
        btc_priced = _make_reward_entry(
            asset="BTC", value_eur=Decimal("0.50"), wallet="Demo Spot", platform="Demo Spot", chain="Bitcoin"
        )
        btc_zero = _make_reward_entry(
            asset="BTC", value_eur=Decimal("0"), wallet="", platform="Demo Spot", chain="Bitcoin"
        )
        report = _make_crypto_tax_report(reward_entries=[btc_priced, btc_zero])
        wb = openpyxl.Workbook()
        write_crypto_supplementary_sheet(wb, report, build_koinly_jurisdiction())
        ws = wb["Crypto Supplementary"]

        table = _find_table_rows_after(ws, _TAXABLE_NOW_DUST_SUBHEADER, num_cols=5)
        btc_rows = [r for r in table if r.get(1) == "BTC"]
        assert len(btc_rows) == 1, f"expected one BTC dust row, got {btc_rows}; table={table}"
        # Wallet column renders empty-string, NOT "None".
        wallet_val = btc_rows[0].get(2)
        assert wallet_val in ("", None), (
            f"expected empty wallet (None or '') for empty-string input, got {wallet_val!r}"
        )
        assert wallet_val != "None", f"'None' literal leaked into Wallet column: {btc_rows[0]}"

    def test_dust_sorted_by_asset_wallet(self):
        """Dust rows are sorted by (asset, wallet) ascending across multiple keys."""
        btc_priced = _make_reward_entry(
            asset="BTC", value_eur=Decimal("0.50"), wallet="Demo Spot", platform="Demo Spot", chain="Bitcoin"
        )
        btc_zero_demo = _make_reward_entry(
            asset="BTC", value_eur=Decimal("0"), wallet="Demo Spot", platform="Demo Spot", chain="Bitcoin"
        )
        btc_zero_wirex = _make_reward_entry(
            asset="BTC", value_eur=Decimal("0"), wallet="Wirex", platform="Wirex", chain="Bitcoin"
        )
        eth_priced = _make_reward_entry(
            asset="ETH", value_eur=Decimal("0.30"), wallet="Demo Spot", platform="Demo Spot", chain="Ethereum"
        )
        eth_zero_demo = _make_reward_entry(
            asset="ETH", value_eur=Decimal("0"), wallet="Demo Spot", platform="Demo Spot", chain="Ethereum"
        )
        report = _make_crypto_tax_report(
            reward_entries=[btc_priced, btc_zero_demo, btc_zero_wirex, eth_priced, eth_zero_demo]
        )
        wb = openpyxl.Workbook()
        write_crypto_supplementary_sheet(wb, report, build_koinly_jurisdiction())
        ws = wb["Crypto Supplementary"]

        table = _find_table_rows_after(ws, _TAXABLE_NOW_DUST_SUBHEADER, num_cols=5)

        # Each (asset, wallet) group appears in the expected ascending order.
        # The expected order is (BTC, Demo Spot), (BTC, Wirex), (ETH, Demo Spot).
        expected_order = [
            ("BTC", "Demo Spot"),
            ("BTC", "Wirex"),
            ("ETH", "Demo Spot"),
        ]
        assert len(table) == len(expected_order), (
            f"expected {len(expected_order)} dust rows, got {table}"
        )
        for row, (asset, wallet) in zip(table, expected_order, strict=True):
            assert row.get(1) == asset, f"expected asset {asset!r}, got {row}"
            assert row.get(2) == wallet, f"expected wallet {wallet!r}, got {row}"

    def test_taxable_now_block_has_outer_header_and_subheader(self):
        """The taxable-now Dust summary block now has an outer 'Dust summary:'
        header followed by a 'Taxable-now dust (...)' sub-header, then the
        dust lines. Mirrors the deferred-side shape for consistency (user
        answer #3).
        """
        btc_priced = _make_reward_entry(
            asset="BTC", value_eur=Decimal("0.50"), wallet="Demo Spot", platform="Demo Spot", chain="Bitcoin"
        )
        btc_zero = _make_reward_entry(
            asset="BTC", value_eur=Decimal("0"), wallet="Demo Spot", platform="Demo Spot", chain="Bitcoin"
        )
        report = _make_crypto_tax_report(reward_entries=[btc_priced, btc_zero])
        wb = openpyxl.Workbook()
        write_crypto_supplementary_sheet(wb, report, build_koinly_jurisdiction())
        ws = wb["Crypto Supplementary"]

        section_strs = _section2_strings(ws)
        try:
            outer_idx = section_strs.index("Dust summary:")
        except ValueError as exc:
            raise AssertionError(
                f"'Dust summary:' outer header missing from Section 2: {section_strs}"
            ) from exc
        body = section_strs[outer_idx + 1 :]
        assert _TAXABLE_NOW_DUST_SUBHEADER in body, (
            f"taxable-now dust sub-header {_TAXABLE_NOW_DUST_SUBHEADER!r} missing after "
            f"'Dust summary:' outer header; body={body}"
        )
        # The dust body row renders under the sub-header as a column table.
        table = _find_table_rows_after(ws, _TAXABLE_NOW_DUST_SUBHEADER, num_cols=5)
        btc_rows = [r for r in table if r.get(1) == "BTC"]
        assert len(btc_rows) == 1, (
            f"BTC dust row missing under {_TAXABLE_NOW_DUST_SUBHEADER!r}; table={table}"
        )

    def test_taxable_now_outer_header_preceded_by_blank_spacer_row(self):
        """The taxable-now 'Dust summary:' outer header is preceded by a blank
        spacer row (matches the section-header spacer convention). Fails RED
        against unchanged production (today the header is flush against the
        taxable-now detail table).
        """
        btc_priced = _make_reward_entry(
            asset="BTC", value_eur=Decimal("0.50"), wallet="Demo Spot", platform="Demo Spot", chain="Bitcoin"
        )
        btc_zero = _make_reward_entry(
            asset="BTC", value_eur=Decimal("0"), wallet="Demo Spot", platform="Demo Spot", chain="Bitcoin"
        )
        report = _make_crypto_tax_report(reward_entries=[btc_priced, btc_zero])
        wb = openpyxl.Workbook()
        write_crypto_supplementary_sheet(wb, report, build_koinly_jurisdiction())
        ws = wb["Crypto Supplementary"]

        header_row = None
        for r in range(1, ws.max_row + 1):
            if ws.cell(r, 1).value == "Dust summary:":
                header_row = r
                break
        assert header_row is not None, "'Dust summary:' header not rendered"
        above = ws.cell(header_row - 1, 1).value
        assert above in (None, ""), (
            f"expected blank spacer row above 'Dust summary:' header (row {header_row}); "
            f"got row {header_row - 1} = {above!r}"
        )


@pytest.mark.unit
class TestRewardDetailRowsContract:
    """Pins the ``_write_reward_detail_rows`` empty-label contract (r1 Advisory #11).

    The three-state empty-label logic in Task 4 depends on the helper rendering
    ``empty_label`` ONLY when ``entries`` is empty (guard at
    ``crypto_supplementary_sheet.py:66-68``). This test pins that contract so a
    future refactor that inverts the guard fails loudly rather than silently
    rendering a stray empty-label row in the mixed / all-real case.
    """

    def test_empty_label_only_rendered_when_entries_empty(self):
        from tax_reporting.application.persisting.crypto_supplementary_sheet import _write_reward_detail_rows

        entry = _make_reward_entry(asset="BTC", value_eur=Decimal("1"))
        wb = openpyxl.Workbook()
        ws = wb.create_sheet("probe")
        # Place the call at row 1; with non-empty entries and empty_label="" the
        # empty label must NOT render.
        next_row = _write_reward_detail_rows(ws, 1, [entry], empty_label="")
        for r in range(1, next_row):
            assert ws.cell(r, 1).value != "", (
                f"empty_label leaked into row {r} when entries is non-empty "
                f"(cell value={ws.cell(r, 1).value!r})"
            )


@pytest.mark.unit
class TestPartitionTaxableNow:
    """Direct unit test of the load-bearing r9 discriminator at the helper boundary.

    The helper ``_partition_taxable_now`` does not exist yet (Task 4 adds it).
    Per AGENTS.md rule 113 a committed RED test that is itself the deliverable
    must fail via ``pytest.fail(<message>)`` naming the resolving task, never an
    unhandled exception. We use a guarded import.
    """

    def test_priced_asset_zero_routes_to_dust(self):
        try:
            from tax_reporting.application.persisting.crypto_supplementary_sheet import _partition_taxable_now
        except ImportError:
            pytest.fail("Task 4 must add _partition_taxable_now helper")

        btc_zero = _make_reward_entry(
            asset="BTC", value_eur=Decimal("0"), wallet="Demo Spot", platform="Demo Spot", chain="Bitcoin"
        )
        btc_priced = _make_reward_entry(
            asset="BTC", value_eur=Decimal("0.50"), wallet="Demo Spot", platform="Demo Spot", chain="Bitcoin"
        )
        taxable_now_entries = [btc_zero]
        reward_entries = [btc_zero, btc_priced]
        real_rows, dust_rows = _partition_taxable_now(taxable_now_entries, reward_entries)
        assert real_rows == []
        assert dust_rows == [btc_zero]

    def test_unpriced_asset_zero_stays_in_real_rows(self):
        try:
            from tax_reporting.application.persisting.crypto_supplementary_sheet import _partition_taxable_now
        except ImportError:
            pytest.fail("Task 4 must add _partition_taxable_now helper")

        osbgt = _make_reward_entry(
            asset="OSBGT",
            value_eur=Decimal("0"),
            wallet="Demo Spot",
            platform="Demo Spot",
            chain="Berachain",
            review_required=True,
            review_reason="Unpriced wrapper token - manual pricing required",
        )
        taxable_now_entries = [osbgt]
        reward_entries = [osbgt]  # no non-zero OSBGT row anywhere in the export
        real_rows, dust_rows = _partition_taxable_now(taxable_now_entries, reward_entries)
        assert real_rows == [osbgt]
        assert dust_rows == []


@pytest.mark.unit
class TestPricedAssetsInExport:
    """Direct unit test of the shared priced-asset discriminator (Invariant 3).

    ``_priced_assets_in_export`` is the ONE helper both ``_partition_taxable_now``
    (CRG-021) and ``_partition_skipped_rewards`` (CRG-022) call, so taxable-now
    dust and deferred dust cannot silently desynchronize. The guarded import
    keeps this test RED against the pre-Task-4 codebase.
    """

    def test_helper_exists(self):
        try:
            from tax_reporting.application.persisting.crypto_supplementary_sheet import (  # noqa: F401
                _priced_assets_in_export,
            )
        except ImportError:
            pytest.fail("Task 4 must add _priced_assets_in_export helper")

    def test_empty_reward_entries_returns_empty_set(self):
        from tax_reporting.application.persisting.crypto_supplementary_sheet import _priced_assets_in_export

        assert _priced_assets_in_export([]) == frozenset()

    def test_all_zero_rows_returns_empty_set(self):
        from tax_reporting.application.persisting.crypto_supplementary_sheet import _priced_assets_in_export

        btc_zero = _make_reward_entry(asset="BTC", value_eur=Decimal("0"))
        eth_zero = _make_reward_entry(asset="ETH", value_eur=Decimal("0"))
        assert _priced_assets_in_export([btc_zero, eth_zero]) == frozenset()

    def test_all_priced_rows_returns_every_asset(self):
        from tax_reporting.application.persisting.crypto_supplementary_sheet import _priced_assets_in_export

        btc = _make_reward_entry(asset="BTC", value_eur=Decimal("1"))
        eth = _make_reward_entry(asset="ETH", value_eur=Decimal("2"))
        assert _priced_assets_in_export([btc, eth]) == frozenset({"BTC", "ETH"})

    def test_mixed_priced_and_zero_returns_only_priced_assets(self):
        from tax_reporting.application.persisting.crypto_supplementary_sheet import _priced_assets_in_export

        btc_priced = _make_reward_entry(asset="BTC", value_eur=Decimal("1"))
        btc_zero = _make_reward_entry(asset="BTC", value_eur=Decimal("0"))
        osbgt_zero = _make_reward_entry(asset="OSBGT", value_eur=Decimal("0"))
        # Only BTC has a priced row; the zero rows on BTC and OSBGT do not add
        # OSBGT to the set.
        result = _priced_assets_in_export([btc_priced, btc_zero, osbgt_zero])
        assert result == frozenset({"BTC"})

    def test_returns_frozenset_type(self):
        from tax_reporting.application.persisting.crypto_supplementary_sheet import _priced_assets_in_export

        result = _priced_assets_in_export([_make_reward_entry(asset="BTC", value_eur=Decimal("1"))])
        assert isinstance(result, frozenset)


@pytest.mark.unit
class TestPartitionSkippedRewards:
    """Direct unit test of the deferred-side partition (CRG-022).

    Mirrors ``TestPartitionTaxableNow`` from Part 7. ``_partition_skipped_rewards``
    is the deferred-side sibling that uses the shared ``_priced_assets_in_export``
    discriminator (Invariant 3, AGENTS.md rule 30). Returns ``(dust_rows,
    unpriced_rows)`` so the caller can reuse the partition once for both the
    Section 3 block and the Section 4 reconciliation (Invariant 5).
    """

    def test_helper_exists(self):
        try:
            from tax_reporting.application.persisting.crypto_supplementary_sheet import (  # noqa: F401
                _partition_skipped_rewards,
            )
        except ImportError:
            pytest.fail("Task 4 must add _partition_skipped_rewards helper")

    def test_empty_skipped_list_returns_two_empty_lists(self):
        from tax_reporting.application.persisting.crypto_supplementary_sheet import _partition_skipped_rewards

        dust_rows, unpriced_rows = _partition_skipped_rewards([], [])
        assert dust_rows == []
        assert unpriced_rows == []

    def test_priced_asset_zero_routes_to_dust(self):
        from tax_reporting.application.persisting.crypto_supplementary_sheet import _partition_skipped_rewards

        btc_priced = _make_reward_entry(
            asset="BTC", value_eur=Decimal("0.50"), wallet="Demo Spot", platform="Demo Spot", chain="Bitcoin"
        )
        btc_zero = _make_reward_entry(
            asset="BTC", value_eur=Decimal("0"), wallet="Demo Spot", platform="Demo Spot", chain="Bitcoin"
        )
        # The priced BTC row is in reward_entries; the zero BTC row is in the
        # skipped list.
        reward_entries = [btc_priced]
        skipped = [btc_zero]
        dust_rows, unpriced_rows = _partition_skipped_rewards(skipped, reward_entries)
        assert dust_rows == [btc_zero]
        assert unpriced_rows == []

    def test_unpriced_asset_zero_routes_to_unpriced(self):
        from tax_reporting.application.persisting.crypto_supplementary_sheet import _partition_skipped_rewards

        osbgt_zero = _make_reward_entry(
            asset="OSBGT",
            value_eur=Decimal("0"),
            wallet="Demo Spot",
            platform="Demo Spot",
            chain="Berachain",
        )
        # No OSBGT row anywhere in reward_entries -> OSBGT is unpriced.
        dust_rows, unpriced_rows = _partition_skipped_rewards([osbgt_zero], [])
        assert dust_rows == []
        assert unpriced_rows == [osbgt_zero]

    def test_mixed_dust_and_unpriced_route_correctly(self):
        from tax_reporting.application.persisting.crypto_supplementary_sheet import _partition_skipped_rewards

        btc_priced = _make_reward_entry(
            asset="BTC", value_eur=Decimal("1.00"), wallet="Demo Spot", platform="Demo Spot", chain="Bitcoin"
        )
        btc_zero = _make_reward_entry(
            asset="BTC", value_eur=Decimal("0"), wallet="Demo Spot", platform="Demo Spot", chain="Bitcoin"
        )
        osbgt_zero = _make_reward_entry(
            asset="OSBGT",
            value_eur=Decimal("0"),
            wallet="Demo Spot",
            platform="Demo Spot",
            chain="Berachain",
        )
        skipped = [btc_zero, osbgt_zero]
        reward_entries = [btc_priced]
        dust_rows, unpriced_rows = _partition_skipped_rewards(skipped, reward_entries)
        assert dust_rows == [btc_zero]
        assert unpriced_rows == [osbgt_zero]

    def test_discriminator_uses_shared_helper(self):
        """The two sibling partitions must call the same discriminator helper
        (AGENTS.md rule 30). The contract: a priced-asset zero row routed to
        taxable-now dust must also be routed to deferred dust when it appears in
        the skipped list. If the helpers desynchronize, this test fails.
        """
        from tax_reporting.application.persisting.crypto_supplementary_sheet import (
            _partition_skipped_rewards,
            _partition_taxable_now,
        )

        btc_priced = _make_reward_entry(
            asset="BTC", value_eur=Decimal("1.00"), wallet="Demo Spot", platform="Demo Spot", chain="Bitcoin"
        )
        btc_zero = _make_reward_entry(
            asset="BTC", value_eur=Decimal("0"), wallet="Demo Spot", platform="Demo Spot", chain="Bitcoin"
        )
        # Same asset, same export -> both sides agree BTC is "priced".
        real_rows, taxable_dust = _partition_taxable_now([btc_zero], [btc_priced, btc_zero])
        deferred_dust, deferred_unpriced = _partition_skipped_rewards([btc_zero], [btc_priced, btc_zero])
        assert taxable_dust == [btc_zero]
        assert deferred_dust == [btc_zero]
        assert deferred_unpriced == []
        assert real_rows == []


@pytest.mark.unit
class TestWriteDustSummaryBlock:
    """Direct unit test of the extracted dust-block render helper.

    The helper ``_write_dust_summary_block`` does not exist yet (Task 4 adds it).
    Per AGENTS.md rule 113 the RED test fails via ``pytest.fail`` naming the
    resolving task through a guarded import.
    """

    def test_groups_by_asset_wallet_and_sorts(self):
        try:
            from tax_reporting.application.persisting.crypto_supplementary_sheet import _write_dust_summary_block
        except ImportError:
            pytest.fail("Task 4 must add _write_dust_summary_block helper")

        btc_demo_1 = _make_reward_entry(
            asset="BTC", value_eur=Decimal("0"), wallet="Demo Spot", platform="Demo Spot", chain="Bitcoin"
        )
        btc_demo_2 = _make_reward_entry(
            asset="BTC", value_eur=Decimal("0"), wallet="Demo Spot", platform="Demo Spot", chain="Bitcoin"
        )
        btc_wirex = _make_reward_entry(
            asset="BTC", value_eur=Decimal("0"), wallet="Wirex", platform="Wirex", chain="Bitcoin"
        )
        eth_demo = _make_reward_entry(
            asset="ETH", value_eur=Decimal("0"), wallet="Demo Spot", platform="Demo Spot", chain="Ethereum"
        )
        # Deliberately out of (asset, wallet) order to exercise the sort.
        dust_rows = [eth_demo, btc_wirex, btc_demo_1, btc_demo_2]

        wb = openpyxl.Workbook()
        ws = wb.create_sheet("dust")
        start_row = 1
        next_row = _write_dust_summary_block(ws, start_row, dust_rows)

        # The helper writes: spacer row (row 1), outer "Dust summary:" header
        # (row 2), sub-header "Taxable-now dust (...)" (row 3), table header
        # row (row 4), then data rows.
        assert ws.cell(2, 1).value == "Dust summary:", (
            f"outer header wrong: {ws.cell(2, 1).value!r}"
        )
        assert ws.cell(3, 1).value == _TAXABLE_NOW_DUST_SUBHEADER, (
            f"sub-header wrong: {ws.cell(3, 1).value!r}"
        )
        # Data rows start at row 5 (after the table header at row 4).
        data_rows: list[tuple[str, str, int, float]] = []
        for r in range(5, next_row):
            asset = ws.cell(r, 1).value
            wallet = ws.cell(r, 2).value
            count = ws.cell(r, 3).value
            summed = ws.cell(r, 4).value
            data_rows.append((asset, wallet, count, summed))
        assert len(data_rows) == 3, f"expected 3 data rows, got {data_rows!r}"

        # (BTC, Demo Spot) group sums two zero rows.
        assert data_rows[0] == ("BTC", "Demo Spot", 2, 0.0), (
            f"first row wrong: {data_rows[0]!r}"
        )
        # (BTC, Wirex) and (ETH, Demo Spot) each sum one zero row.
        assert data_rows[1] == ("BTC", "Wirex", 1, 0.0), (
            f"second row wrong: {data_rows[1]!r}"
        )
        assert data_rows[2] == ("ETH", "Demo Spot", 1, 0.0), (
            f"third row wrong: {data_rows[2]!r}"
        )


# Outer header text the renderer must emit before the suppressed-rewards body
# lines. The exact literal pins the outer-header design.
_SUPPRESSED_HEADER = "Suppressed zero-value deferred rewards"

# Sub-header text the renderer emits between the outer header and each bucket's
# body lines. Pinned so a maintainer renaming a sub-header sees the change here.
_DUST_SUBHEADER = "Deferred dust (priced-asset rounding)"
_UNPRICED_SUBHEADER = "Deferred unpriced (no Koinly price feed)"

# Symmetric taxable-now sub-header (CRG-021). Taxable-now has only one bucket
# (unpriced taxable-now rows keep per-row YES in detail), so one sub-header.
_TAXABLE_NOW_DUST_SUBHEADER = "Taxable-now dust (priced-asset rounding)"


@pytest.mark.unit
class TestCryptoSupplementarySheetDeferredSkip:
    """Worksheet-level tests for the Section 3 suppressed-rewards block (CRG-022).

    RED for Task 4: ``write_crypto_supplementary_sheet`` does not yet render the
    single "Suppressed zero-value deferred rewards" block from
    ``crypto_tax_report.skipped_zero_value_deferred_rewards``. The
    ``skipped_zero_value_deferred_rewards`` report field itself already exists
    (Task 2), so these tests fail via natural assertion failure (header /
    line regex not found in the rendered sheet), not via ``pytest.fail``.
    """

    def test_dust_row_renders_with_dust_reason(self):
        """A zero-value WBERA deferred row whose asset has a priced row renders
        as a "dust" line in the suppressed-rewards block.

        Discriminator: WBERA has a non-zero row in ``reward_entries``, so the
        zero row is priced-asset rounding dust. The block must NOT render it
        with the "unpriced" / "no Koinly price feed" reason.
        """
        wbera_priced = _make_reward_entry(
            classification=RewardTaxClassification.DEFERRED_BY_LAW,
            asset="WBERA",
            value_eur=Decimal("1.00"),
            wallet="Wirex",
            platform="Wirex",
            chain="Berachain",
            amount=Decimal("2.0"),
        )
        wbera_zero = _make_reward_entry(
            classification=RewardTaxClassification.DEFERRED_BY_LAW,
            asset="WBERA",
            value_eur=Decimal("0"),
            wallet="Wirex",
            platform="Wirex",
            chain="Berachain",
            amount=Decimal("1.5"),
        )
        report = _make_crypto_tax_report(
            reward_entries=[wbera_priced],
            skipped_deferred_rewards=[wbera_zero],
        )
        wb = openpyxl.Workbook()
        write_crypto_supplementary_sheet(wb, report, build_koinly_jurisdiction())
        ws = wb["Crypto Supplementary"]

        # The header renders exactly once.
        all_strs = _all_sheet_strings(ws)
        assert all_strs.count(_SUPPRESSED_HEADER) == 1, (
            f"expected exactly one {_SUPPRESSED_HEADER!r} header, got {all_strs}"
        )

        # The dust sub-header renders, and the WBERA row appears under it as a
        # column-table row with Category="dust" (NOT "unpriced").
        table = _find_table_rows_after(ws, _DUST_SUBHEADER, num_cols=5)
        wbera_rows = [r for r in table if r.get(1) == "WBERA"]
        assert len(wbera_rows) == 1, (
            f"expected one WBERA row in dust table, got {wbera_rows}; full table={table}"
        )
        row = wbera_rows[0]
        assert row.get(2) == "Wirex", f"wallet wrong: {row}"
        assert row.get(3) == 1, f"rows count wrong: {row}"
        # Summed amount = 1.50000000 (Decimal "1.5" formatted :.8f).
        assert row.get(4) == "1.50000000", f"summed amount wrong: {row}"
        assert row.get(5) == "dust", f"category must be 'dust', got {row}"

        # The zero-value WBERA row must NOT appear in the Section 3 detail table.
        detail = _section3_reward_detail_rows(ws)
        assert not any(
            str(r.get(2, "")) == "WBERA" and float(r.get(3, -1)) == 0.0 for r in detail
        ), "zero-value WBERA row leaked into Section 3 detail table"

    def test_unpriced_row_renders_with_unpriced_reason(self):
        """A zero-value OSBGT deferred row with NO priced row anywhere renders
        as an "unpriced" line. OSBGT is in ``popular_crypto_tokens.json`` so it
        survives the parse-time ``is_known`` gate and reaches the deferred-skip
        path. The block must NOT render the "Re-export" reason.
        """
        osbgt_zero = _make_reward_entry(
            classification=RewardTaxClassification.DEFERRED_BY_LAW,
            asset="OSBGT",
            value_eur=Decimal("0"),
            wallet="Wirex",
            platform="Wirex",
            chain="Berachain",
            amount=Decimal("3.0"),
        )
        # No OSBGT row anywhere in reward_entries -> OSBGT is unpriced.
        report = _make_crypto_tax_report(
            reward_entries=[],
            skipped_deferred_rewards=[osbgt_zero],
        )
        wb = openpyxl.Workbook()
        write_crypto_supplementary_sheet(wb, report, build_koinly_jurisdiction())
        ws = wb["Crypto Supplementary"]

        all_strs = _all_sheet_strings(ws)
        assert all_strs.count(_SUPPRESSED_HEADER) == 1, (
            f"expected exactly one {_SUPPRESSED_HEADER!r} header, got {all_strs}"
        )

        # The unpriced sub-header renders, and the OSBGT row appears under it
        # with Category="unpriced".
        table = _find_table_rows_after(ws, _UNPRICED_SUBHEADER, num_cols=5)
        osbgt_rows = [r for r in table if r.get(1) == "OSBGT"]
        assert len(osbgt_rows) == 1, (
            f"expected one OSBGT row in unpriced table, got {osbgt_rows}; full table={table}"
        )
        row = osbgt_rows[0]
        assert row.get(2) == "Wirex", f"wallet wrong: {row}"
        assert row.get(3) == 1, f"rows count wrong: {row}"
        assert row.get(4) == "3.00000000", f"summed amount wrong: {row}"
        assert row.get(5) == "unpriced", f"category must be 'unpriced', got {row}"

        # The OSBGT row must NOT appear in the Section 3 detail table.
        detail = _section3_reward_detail_rows(ws)
        assert not any(str(r.get(2, "")) == "OSBGT" for r in detail), (
            "OSBGT row leaked into Section 3 detail table"
        )

    def test_dust_and_unpriced_rows_render_under_separate_subheaders(self):
        """A WBERA dust row and an OSBGT unpriced row render under one outer
        header but TWO separate sub-headers. Reverses the predecessor plan's
        single-block design per user feedback: the merged clause was hard to
        read in a real export.

        Asserts:
          (a) exactly one outer header;
          (b) both sub-headers render exactly once;
          (c) the WBERA line appears AFTER the dust sub-header and BEFORE the
              unpriced sub-header;
          (d) the OSBGT line appears AFTER the unpriced sub-header.
        """
        wbera_priced = _make_reward_entry(
            classification=RewardTaxClassification.DEFERRED_BY_LAW,
            asset="WBERA",
            value_eur=Decimal("1.00"),
            wallet="Wirex",
            platform="Wirex",
            chain="Berachain",
            amount=Decimal("1.0"),
        )
        wbera_dust = _make_reward_entry(
            classification=RewardTaxClassification.DEFERRED_BY_LAW,
            asset="WBERA",
            value_eur=Decimal("0"),
            wallet="Wirex",
            platform="Wirex",
            chain="Berachain",
            amount=Decimal("1.0"),
        )
        osbgt_unpriced = _make_reward_entry(
            classification=RewardTaxClassification.DEFERRED_BY_LAW,
            asset="OSBGT",
            value_eur=Decimal("0"),
            wallet="Wirex",
            platform="Wirex",
            chain="Berachain",
            amount=Decimal("1.0"),
        )
        report = _make_crypto_tax_report(
            reward_entries=[wbera_priced],
            skipped_deferred_rewards=[wbera_dust, osbgt_unpriced],
        )
        wb = openpyxl.Workbook()
        write_crypto_supplementary_sheet(wb, report, build_koinly_jurisdiction())
        ws = wb["Crypto Supplementary"]

        all_strs = _all_sheet_strings(ws)
        # (a) Exactly one outer header in the whole sheet.
        assert all_strs.count(_SUPPRESSED_HEADER) == 1, (
            f"expected exactly one {_SUPPRESSED_HEADER!r} outer header, "
            f"got {all_strs.count(_SUPPRESSED_HEADER)} in {all_strs}"
        )
        # (b) Both sub-headers render exactly once.
        assert all_strs.count(_DUST_SUBHEADER) == 1, (
            f"expected exactly one {_DUST_SUBHEADER!r} sub-header, "
            f"got {all_strs.count(_DUST_SUBHEADER)} in {all_strs}"
        )
        assert all_strs.count(_UNPRICED_SUBHEADER) == 1, (
            f"expected exactly one {_UNPRICED_SUBHEADER!r} sub-header, "
            f"got {all_strs.count(_UNPRICED_SUBHEADER)} in {all_strs}"
        )
        # (c) + (d) Ordering: outer -> dust sub -> WBERA row -> unpriced sub -> OSBGT row.
        # The data rows now put Asset in column A (no trailing paren), so the
        # ordering check looks for "WBERA" / "OSBGT" as the cell value (the
        # column-table header row "Asset" also matches as a column-A string but
        # sorts before/after via the sub-header indices).
        outer_idx = all_strs.index(_SUPPRESSED_HEADER)
        dust_sub_idx = all_strs.index(_DUST_SUBHEADER)
        unpriced_sub_idx = all_strs.index(_UNPRICED_SUBHEADER)
        wbera_idx = all_strs.index("WBERA")
        osbgt_idx = all_strs.index("OSBGT")
        assert outer_idx < dust_sub_idx < wbera_idx < unpriced_sub_idx < osbgt_idx, (
            f"expected ordering outer < dust-sub < WBERA < unpriced-sub < OSBGT; "
            f"got outer={outer_idx}, dust-sub={dust_sub_idx}, WBERA={wbera_idx}, "
            f"unpriced-sub={unpriced_sub_idx}, OSBGT={osbgt_idx}"
        )

    def test_block_has_outer_header_and_two_subheaders(self):
        """The suppressed block has an outer header followed by two sub-headers
        in order: outer, dust sub, unpriced sub. Fails RED against unchanged
        production (today there are no sub-headers).
        """
        wbera_priced = _make_reward_entry(
            classification=RewardTaxClassification.DEFERRED_BY_LAW,
            asset="WBERA",
            value_eur=Decimal("1.00"),
            wallet="Wirex",
            platform="Wirex",
            chain="Berachain",
            amount=Decimal("1.0"),
        )
        wbera_dust = _make_reward_entry(
            classification=RewardTaxClassification.DEFERRED_BY_LAW,
            asset="WBERA",
            value_eur=Decimal("0"),
            wallet="Wirex",
            platform="Wirex",
            chain="Berachain",
            amount=Decimal("1.0"),
        )
        osbgt_unpriced = _make_reward_entry(
            classification=RewardTaxClassification.DEFERRED_BY_LAW,
            asset="OSBGT",
            value_eur=Decimal("0"),
            wallet="Wirex",
            platform="Wirex",
            chain="Berachain",
            amount=Decimal("1.0"),
        )
        report = _make_crypto_tax_report(
            reward_entries=[wbera_priced],
            skipped_deferred_rewards=[wbera_dust, osbgt_unpriced],
        )
        wb = openpyxl.Workbook()
        write_crypto_supplementary_sheet(wb, report, build_koinly_jurisdiction())
        ws = wb["Crypto Supplementary"]

        section_strs = _section3_strings(ws)
        try:
            outer_idx = section_strs.index(_SUPPRESSED_HEADER)
        except ValueError as exc:
            raise AssertionError(
                f"{_SUPPRESSED_HEADER!r} header missing from Section 3: {section_strs}"
            ) from exc
        body = section_strs[outer_idx + 1 :]
        assert _DUST_SUBHEADER in body, (
            f"dust sub-header {_DUST_SUBHEADER!r} missing after outer header; body={body}"
        )
        assert _UNPRICED_SUBHEADER in body, (
            f"unpriced sub-header {_UNPRICED_SUBHEADER!r} missing after outer header; body={body}"
        )
        # Sub-headers render in the canonical order (dust before unpriced).
        dust_idx = body.index(_DUST_SUBHEADER)
        unpriced_idx = body.index(_UNPRICED_SUBHEADER)
        assert dust_idx < unpriced_idx, (
            f"expected dust sub-header before unpriced sub-header; "
            f"got dust={dust_idx}, unpriced={unpriced_idx} in body={body}"
        )

    def test_outer_header_preceded_by_blank_spacer_row(self):
        """The outer header is preceded by a blank spacer row (matches the
        section-header spacer convention at every section boundary). Fails RED
        against unchanged production (today the header is flush against the
        deferred detail table).
        """
        wbera_priced = _make_reward_entry(
            classification=RewardTaxClassification.DEFERRED_BY_LAW,
            asset="WBERA",
            value_eur=Decimal("1.00"),
            wallet="Wirex",
            platform="Wirex",
            chain="Berachain",
            amount=Decimal("1.0"),
        )
        wbera_dust = _make_reward_entry(
            classification=RewardTaxClassification.DEFERRED_BY_LAW,
            asset="WBERA",
            value_eur=Decimal("0"),
            wallet="Wirex",
            platform="Wirex",
            chain="Berachain",
            amount=Decimal("1.0"),
        )
        report = _make_crypto_tax_report(
            reward_entries=[wbera_priced],
            skipped_deferred_rewards=[wbera_dust],
        )
        wb = openpyxl.Workbook()
        write_crypto_supplementary_sheet(wb, report, build_koinly_jurisdiction())
        ws = wb["Crypto Supplementary"]

        # Locate the outer header cell row.
        header_row = None
        for r in range(1, ws.max_row + 1):
            if ws.cell(r, 1).value == _SUPPRESSED_HEADER:
                header_row = r
                break
        assert header_row is not None, f"{_SUPPRESSED_HEADER!r} header not rendered"
        # Row immediately above MUST be blank (spacer row convention).
        above = ws.cell(header_row - 1, 1).value
        assert above in (None, ""), (
            f"expected blank spacer row above {_SUPPRESSED_HEADER!r} header "
            f"(row {header_row}); got row {header_row - 1} = {above!r}"
        )

    def test_dust_only_renders_dust_subheader_only(self):
        """A dust-only skipped list renders the outer header + dust sub-header
        but NOT the unpriced sub-header (conditional render guard).
        """
        wbera_priced = _make_reward_entry(
            classification=RewardTaxClassification.DEFERRED_BY_LAW,
            asset="WBERA",
            value_eur=Decimal("1.00"),
            wallet="Wirex",
            platform="Wirex",
            chain="Berachain",
            amount=Decimal("1.0"),
        )
        wbera_dust = _make_reward_entry(
            classification=RewardTaxClassification.DEFERRED_BY_LAW,
            asset="WBERA",
            value_eur=Decimal("0"),
            wallet="Wirex",
            platform="Wirex",
            chain="Berachain",
            amount=Decimal("1.0"),
        )
        report = _make_crypto_tax_report(
            reward_entries=[wbera_priced],
            skipped_deferred_rewards=[wbera_dust],
        )
        wb = openpyxl.Workbook()
        write_crypto_supplementary_sheet(wb, report, build_koinly_jurisdiction())
        ws = wb["Crypto Supplementary"]

        all_strs = _all_sheet_strings(ws)
        assert _SUPPRESSED_HEADER in all_strs, "outer header missing"
        assert _DUST_SUBHEADER in all_strs, "dust sub-header missing"
        assert _UNPRICED_SUBHEADER not in all_strs, (
            f"unpriced sub-header should NOT render when there are no unpriced rows; "
            f"sheet={all_strs}"
        )

    def test_unpriced_only_renders_unpriced_subheader_only(self):
        """An unpriced-only skipped list renders the outer header + unpriced
        sub-header but NOT the dust sub-header (symmetric conditional guard).
        """
        osbgt_unpriced = _make_reward_entry(
            classification=RewardTaxClassification.DEFERRED_BY_LAW,
            asset="OSBGT",
            value_eur=Decimal("0"),
            wallet="Wirex",
            platform="Wirex",
            chain="Berachain",
            amount=Decimal("1.0"),
        )
        report = _make_crypto_tax_report(
            reward_entries=[],
            skipped_deferred_rewards=[osbgt_unpriced],
        )
        wb = openpyxl.Workbook()
        write_crypto_supplementary_sheet(wb, report, build_koinly_jurisdiction())
        ws = wb["Crypto Supplementary"]

        all_strs = _all_sheet_strings(ws)
        assert _SUPPRESSED_HEADER in all_strs, "outer header missing"
        assert _UNPRICED_SUBHEADER in all_strs, "unpriced sub-header missing"
        assert _DUST_SUBHEADER not in all_strs, (
            f"dust sub-header should NOT render when there are no dust rows; "
            f"sheet={all_strs}"
        )


    def test_amount_sum_uniform_no_pattern_exclusion(self):
        """Two unpriced rows for the same (BGT, Wirex) group with amount=1000
        each sum to ``2000.00000000 BGT``. The literal ``2000.00000000`` pins
        the ``:.8f`` format spec and the absence of any pattern-based exclusion
        (r1 review DROPPED ``_UNPRICED_NON_DISPOSEABLE_PATTERNS``). BGT is in
        ``popular_crypto_tokens.json`` so it survives the parse-time ``is_known``
        gate and reaches the deferred-skip path.
        """
        bgt_zero_1 = _make_reward_entry(
            classification=RewardTaxClassification.DEFERRED_BY_LAW,
            asset="BGT",
            value_eur=Decimal("0"),
            wallet="Wirex",
            platform="Wirex",
            chain="Berachain",
            amount=Decimal("1000"),
        )
        bgt_zero_2 = _make_reward_entry(
            classification=RewardTaxClassification.DEFERRED_BY_LAW,
            asset="BGT",
            value_eur=Decimal("0"),
            wallet="Wirex",
            platform="Wirex",
            chain="Berachain",
            amount=Decimal("1000"),
        )
        # No BGT row anywhere in reward_entries -> BGT is unpriced.
        report = _make_crypto_tax_report(
            reward_entries=[],
            skipped_deferred_rewards=[bgt_zero_1, bgt_zero_2],
        )
        wb = openpyxl.Workbook()
        write_crypto_supplementary_sheet(wb, report, build_koinly_jurisdiction())
        ws = wb["Crypto Supplementary"]

        # BGT appears in the unpriced column table. The Summed amount column
        # pins the :.8f spec: str(Decimal("2000")) omits trailing zeros, so a
        # no-spec regression renders "2000" and fails this assertion. Uniform
        # sum (no pattern exclusion): r1 review DROPPED
        # _UNPRICED_NON_DISPOSEABLE_PATTERNS.
        table = _find_table_rows_after(ws, _UNPRICED_SUBHEADER, num_cols=5)
        bgt_rows = [r for r in table if r.get(1) == "BGT"]
        assert len(bgt_rows) == 1, (
            f"expected one BGT row in unpriced table, got {bgt_rows}; full table={table}"
        )
        assert bgt_rows[0].get(4) == "2000.00000000", (
            f"BGT summed amount did not pin :.8f spec: {bgt_rows[0]!r}"
        )

    def test_block_sorted_by_asset_wallet(self):
        """Within each sub-header, the body lines appear in (asset, wallet)
        ascending order. The two buckets (dust / unpriced) are now rendered
        as separate sorted groups under their own sub-headers, not merged
        (user feedback reversed the predecessor plan's single-block merge).
        """
        # Priced-asset rows make WBERA and WBTC "dust" (priced) assets; OSBGT
        # and BNB are unpriced (no priced row anywhere).
        wbera_priced = _make_reward_entry(
            classification=RewardTaxClassification.DEFERRED_BY_LAW,
            asset="WBERA",
            value_eur=Decimal("1.00"),
            wallet="Wirex",
            platform="Wirex",
            chain="Berachain",
            amount=Decimal("1"),
        )
        wbtc_priced = _make_reward_entry(
            classification=RewardTaxClassification.DEFERRED_BY_LAW,
            asset="WBTC",
            value_eur=Decimal("2.00"),
            wallet="Wirex",
            platform="Wirex",
            chain="Bitcoin",
            amount=Decimal("1"),
        )
        # Zero-value skipped rows (deliberately out of (asset, wallet) order).
        wbera_wirex_dust = _make_reward_entry(
            classification=RewardTaxClassification.DEFERRED_BY_LAW,
            asset="WBERA",
            value_eur=Decimal("0"),
            wallet="Wirex",
            platform="Wirex",
            chain="Berachain",
            amount=Decimal("1"),
        )
        wbera_kraken_dust = _make_reward_entry(
            classification=RewardTaxClassification.DEFERRED_BY_LAW,
            asset="WBERA",
            value_eur=Decimal("0"),
            wallet="Kraken",
            platform="Kraken",
            chain="Berachain",
            amount=Decimal("1"),
        )
        wbtc_wirex_dust = _make_reward_entry(
            classification=RewardTaxClassification.DEFERRED_BY_LAW,
            asset="WBTC",
            value_eur=Decimal("0"),
            wallet="Wirex",
            platform="Wirex",
            chain="Bitcoin",
            amount=Decimal("1"),
        )
        osbgt_wirex_unpriced = _make_reward_entry(
            classification=RewardTaxClassification.DEFERRED_BY_LAW,
            asset="OSBGT",
            value_eur=Decimal("0"),
            wallet="Wirex",
            platform="Wirex",
            chain="Berachain",
            amount=Decimal("1"),
        )
        bnb_kraken_unpriced = _make_reward_entry(
            classification=RewardTaxClassification.DEFERRED_BY_LAW,
            asset="BNB",
            value_eur=Decimal("0"),
            wallet="Kraken",
            platform="Kraken",
            chain="Binance",
            amount=Decimal("1"),
        )
        skipped = [
            wbera_wirex_dust,
            wbera_kraken_dust,
            wbtc_wirex_dust,
            osbgt_wirex_unpriced,
            bnb_kraken_unpriced,
        ]
        report = _make_crypto_tax_report(
            reward_entries=[wbera_priced, wbtc_priced],
            skipped_deferred_rewards=skipped,
        )
        wb = openpyxl.Workbook()
        write_crypto_supplementary_sheet(wb, report, build_koinly_jurisdiction())
        ws = wb["Crypto Supplementary"]

        # Dust sub-header appears first, followed by its column-table rows;
        # unpriced sub-header appears second. Both sub-headers must render.
        all_strs = _all_sheet_strings(ws)
        try:
            outer_idx = all_strs.index(_SUPPRESSED_HEADER)
        except ValueError as exc:
            raise AssertionError(
                f"{_SUPPRESSED_HEADER!r} header missing in {all_strs}"
            ) from exc
        try:
            dust_sub_idx = all_strs.index(_DUST_SUBHEADER)
        except ValueError as exc:
            raise AssertionError(
                f"dust sub-header {_DUST_SUBHEADER!r} missing in {all_strs}"
            ) from exc
        try:
            unpriced_sub_idx = all_strs.index(_UNPRICED_SUBHEADER)
        except ValueError as exc:
            raise AssertionError(
                f"unpriced sub-header {_UNPRICED_SUBHEADER!r} missing in {all_strs}"
            ) from exc
        assert outer_idx < dust_sub_idx < unpriced_sub_idx, (
            f"expected ordering outer < dust-sub < unpriced-sub; "
            f"got outer={outer_idx}, dust-sub={dust_sub_idx}, unpriced-sub={unpriced_sub_idx}"
        )

        # Dust column-table rows, sorted by (asset, wallet) ascending.
        dust_table = _find_table_rows_after(ws, _DUST_SUBHEADER, num_cols=5)
        assert len(dust_table) == 3, f"expected 3 dust rows, got {dust_table!r}"
        expected_dust_keys = [
            ("WBERA", "Kraken"),
            ("WBERA", "Wirex"),
            ("WBTC", "Wirex"),
        ]
        for row, (asset, wallet) in zip(dust_table, expected_dust_keys, strict=True):
            assert row.get(1) == asset, f"expected dust asset {asset!r}, got {row}"
            assert row.get(2) == wallet, f"expected dust wallet {wallet!r}, got {row}"

        # Unpriced column-table rows, sorted by (asset, wallet) ascending.
        unpriced_table = _find_table_rows_after(ws, _UNPRICED_SUBHEADER, num_cols=5)
        assert len(unpriced_table) == 2, f"expected 2 unpriced rows, got {unpriced_table!r}"
        expected_unpriced_keys = [
            ("BNB", "Kraken"),
            ("OSBGT", "Wirex"),
        ]
        for row, (asset, wallet) in zip(unpriced_table, expected_unpriced_keys, strict=True):
            assert row.get(1) == asset, f"expected unpriced asset {asset!r}, got {row}"
            assert row.get(2) == wallet, f"expected unpriced wallet {wallet!r}, got {row}"

    def test_empty_skipped_list_renders_no_block(self):
        """With an empty ``skipped_zero_value_deferred_rewards`` list the block
        header must NOT render (regression guard on the conditional render).
        """
        report = _make_crypto_tax_report(
            reward_entries=[_make_reward_entry(classification=RewardTaxClassification.DEFERRED_BY_LAW)],
            skipped_deferred_rewards=[],
        )
        wb = openpyxl.Workbook()
        write_crypto_supplementary_sheet(wb, report, build_koinly_jurisdiction())
        ws = wb["Crypto Supplementary"]

        all_strs = _all_sheet_strings(ws)
        assert _SUPPRESSED_HEADER not in all_strs, (
            f"{_SUPPRESSED_HEADER!r} rendered on empty skipped list: {all_strs}"
        )

    def test_safe_cell_value_wrappers_on_line(self):
        """A zero-value WBERA deferred row with an empty wallet renders as
        ``"WBERA (): ..."`` with no ``"None"`` literal (mirrors the Part 7
        Invariant 6 test; production contract is ``str``).
        """
        wbera_priced = _make_reward_entry(
            classification=RewardTaxClassification.DEFERRED_BY_LAW,
            asset="WBERA",
            value_eur=Decimal("1.00"),
            wallet="Wirex",
            platform="Wirex",
            chain="Berachain",
            amount=Decimal("1"),
        )
        wbera_zero = _make_reward_entry(
            classification=RewardTaxClassification.DEFERRED_BY_LAW,
            asset="WBERA",
            value_eur=Decimal("0"),
            wallet="",
            platform="Wirex",
            chain="Berachain",
            amount=Decimal("1.5"),
        )
        report = _make_crypto_tax_report(
            reward_entries=[wbera_priced],
            skipped_deferred_rewards=[wbera_zero],
        )
        wb = openpyxl.Workbook()
        write_crypto_supplementary_sheet(wb, report, build_koinly_jurisdiction())
        ws = wb["Crypto Supplementary"]

        # The WBERA row appears in the dust column table with an empty Wallet
        # cell (empty string, NOT 'None').
        table = _find_table_rows_after(ws, _DUST_SUBHEADER, num_cols=5)
        wbera_rows = [r for r in table if r.get(1) == "WBERA"]
        assert len(wbera_rows) == 1, (
            f"expected one WBERA row in dust table, got {wbera_rows}; full table={table}"
        )
        wallet_val = wbera_rows[0].get(2)
        assert wallet_val in ("", None), (
            f"expected empty wallet (None or '') for empty-string input, got {wallet_val!r}"
        )
        assert wallet_val != "None", f"'None' literal leaked into Wallet column: {wbera_rows[0]}"
