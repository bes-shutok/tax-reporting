"""Tests for the crypto capital gains sheet writer."""

from decimal import Decimal

import openpyxl
import pytest
from openpyxl.styles import PatternFill

from tax_reporting.application.crypto_reporting import (
    CapitalGainPeriodStats,
    CryptoCapitalGainEntry,
    CryptoCapitalGainStats,
    CryptoReconciliationSummary,
    CryptoRewardIncomeEntry,
    CryptoTaxReport,
)
from tax_reporting.application.persisting.crypto_gains_sheet import write_crypto_gains_sheet
from tax_reporting.domain.entities import OgrValidationResult
from tests.conftest import make_operator_origin


def _make_capital_entry(**overrides: object) -> CryptoCapitalGainEntry:
    defaults = {
        "disposal_date": "2025-06-15",
        "acquisition_date": "2025-01-10",
        "asset": "BTC",
        "amount": Decimal("0.5"),
        "cost_eur": Decimal("20000"),
        "proceeds_eur": Decimal("25000"),
        "gain_loss_eur": Decimal("5000"),
        "holding_period": "Short-term",
        "wallet": "Kraken",
        "platform": "Kraken",
        "chain": "Ethereum",
        "operator_origin": make_operator_origin(),
        "annex_hint": "J",
        "review_required": False,
        "notes": "",
    }
    defaults.update(overrides)
    return CryptoCapitalGainEntry(**defaults)  # type: ignore[arg-type]


def _make_stats() -> CryptoCapitalGainStats:
    short_term = CapitalGainPeriodStats(
        count=1,
        cost_total_eur=Decimal("20000"),
        proceeds_total_eur=Decimal("25000"),
        gain_loss_total_eur=Decimal("5000"),
    )
    empty = CapitalGainPeriodStats(
        count=0,
        cost_total_eur=Decimal("0"),
        proceeds_total_eur=Decimal("0"),
        gain_loss_total_eur=Decimal("0"),
    )
    grand_total = CapitalGainPeriodStats(
        count=1,
        cost_total_eur=Decimal("20000"),
        proceeds_total_eur=Decimal("25000"),
        gain_loss_total_eur=Decimal("5000"),
    )
    return CryptoCapitalGainStats(
        short_term=short_term, long_term=empty, mixed=empty, unknown=empty, grand_total=grand_total
    )


def _make_crypto_tax_report(  # noqa: PLR0913
    capital_entries: list[CryptoCapitalGainEntry] | None = None,
    stats: CryptoCapitalGainStats | None = None,
    reward_entries: list[CryptoRewardIncomeEntry] | None = None,
    pdf_summary: object = None,
    zero_basis_review_threshold: Decimal = Decimal("50"),
) -> CryptoTaxReport:

    entries = capital_entries if capital_entries is not None else [_make_capital_entry()]
    reconciliation = CryptoReconciliationSummary(
        capital_rows=len(entries),
        reward_rows=len(reward_entries) if reward_entries else 0,
        short_term_rows=sum(1 for e in entries if e.holding_period.lower().startswith("short")),
        long_term_rows=sum(1 for e in entries if e.holding_period.lower().startswith("long")),
        mixed_rows=sum(1 for e in entries if e.holding_period.lower() == "mixed"),
        unknown_rows=sum(1 for e in entries if e.holding_period.lower() == "unknown"),
        capital_cost_total_eur=sum((e.cost_eur for e in entries), start=Decimal("0")),
        capital_proceeds_total_eur=sum((e.proceeds_eur for e in entries), start=Decimal("0")),
        capital_gain_total_eur=sum((e.gain_loss_eur for e in entries), start=Decimal("0")),
        reward_total_eur=Decimal("0"),
        opening_holdings=None,
        closing_holdings=None,
    )
    return CryptoTaxReport(
        tax_year=2025,
        capital_entries=entries,
        reward_entries=reward_entries or [],
        reconciliation=reconciliation,
        capital_gain_stats=stats or _make_stats(),
        pdf_summary=pdf_summary,  # type: ignore[arg-type]
        zero_basis_review_threshold=zero_basis_review_threshold,
    )


@pytest.mark.unit
class TestCryptoGainsSheetName:
    """Tests that the sheet is created with the correct name."""

    def test_sheet_named_crypto_gains(self):
        wb = openpyxl.Workbook()
        report = _make_crypto_tax_report()
        write_crypto_gains_sheet(wb, report)
        assert "Crypto Gains" in wb.sheetnames


@pytest.mark.unit
class TestCryptoGainsSheetTitleAndMetadata:
    """Tests that the sheet writes title, tax year, and PDF summary."""

    def test_title_row_present(self):
        wb = openpyxl.Workbook()
        report = _make_crypto_tax_report()
        write_crypto_gains_sheet(wb, report)
        ws = wb["Crypto Gains"]
        assert ws.cell(1, 1).value == "CRYPTO TAX REPORT - PORTUGAL"

    def test_title_is_bold(self):
        wb = openpyxl.Workbook()
        report = _make_crypto_tax_report()
        write_crypto_gains_sheet(wb, report)
        ws = wb["Crypto Gains"]
        assert ws.cell(1, 1).font.bold is True

    def test_tax_year_written(self):
        wb = openpyxl.Workbook()
        report = _make_crypto_tax_report()
        write_crypto_gains_sheet(wb, report)
        ws = wb["Crypto Gains"]
        assert ws.cell(2, 1).value == "Tax year"
        assert ws.cell(2, 2).value == 2025

    def test_pdf_summary_written_when_present(self):
        from tax_reporting.application.crypto_reporting import CryptoCompletePdfSummary

        pdf = CryptoCompletePdfSummary(
            period="01 Jan 2025 to 31 Dec 2025", timezone="Europe/Lisbon", extracted_tokens=42
        )
        report = _make_crypto_tax_report(pdf_summary=pdf)
        wb = openpyxl.Workbook()
        write_crypto_gains_sheet(wb, report)
        ws = wb["Crypto Gains"]
        assert ws.cell(3, 1).value == "PDF period"
        assert ws.cell(3, 2).value == "01 Jan 2025 to 31 Dec 2025"
        assert ws.cell(3, 3).value == "PDF timezone"
        assert ws.cell(3, 4).value == "Europe/Lisbon"

    def test_pdf_summary_absent_no_pdf_row(self):
        report = _make_crypto_tax_report()
        wb = openpyxl.Workbook()
        write_crypto_gains_sheet(wb, report)
        ws = wb["Crypto Gains"]
        assert ws.cell(3, 1).value is None


@pytest.mark.unit
class TestCryptoGainsSheetCapitalEntries:
    """Tests that capital gain entries are written in 17 columns."""

    CAPITAL_HEADERS = [
        "Disposal date",
        "Acquisition date",
        "Asset",
        "Quantity",
        "Acquisition Cost (EUR)",
        "Disposal Proceeds (EUR)",
        "Gain/Loss (EUR)",
        "Holding period",
        "Wallet",
        "Platform",
        "Disposal chain",
        "Operator entity",
        "Operator country",
        "Annex hint",
        "Review flag",
        "Notes",
        "Token origin",
    ]

    def test_section_1_title_written(self):
        wb = openpyxl.Workbook()
        report = _make_crypto_tax_report()
        write_crypto_gains_sheet(wb, report)
        ws = wb["Crypto Gains"]
        found = False
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=1):
            if row[0].value == "1. CAPITAL GAINS":
                found = True
                break
        assert found

    def test_section_1_title_is_bold(self):
        wb = openpyxl.Workbook()
        report = _make_crypto_tax_report()
        write_crypto_gains_sheet(wb, report)
        ws = wb["Crypto Gains"]
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=1):
            if row[0].value == "1. CAPITAL GAINS":
                assert row[0].font.bold is True
                break

    def test_capital_headers_written(self):
        wb = openpyxl.Workbook()
        report = _make_crypto_tax_report()
        write_crypto_gains_sheet(wb, report)
        ws = wb["Crypto Gains"]
        header_row = None
        for r in range(1, ws.max_row + 1):
            if ws.cell(r, 1).value == "Disposal date":
                header_row = r
                break
        assert header_row is not None
        for idx, expected in enumerate(self.CAPITAL_HEADERS, start=1):
            assert ws.cell(header_row, idx).value == expected

    def test_capital_entry_values(self):
        entry = _make_capital_entry()
        report = _make_crypto_tax_report(capital_entries=[entry])
        wb = openpyxl.Workbook()
        write_crypto_gains_sheet(wb, report)
        ws = wb["Crypto Gains"]
        header_row = None
        for r in range(1, ws.max_row + 1):
            if ws.cell(r, 1).value == "Disposal date":
                header_row = r
                break
        assert header_row is not None
        data_row = header_row + 1
        assert ws.cell(data_row, 1).value == "2025-06-15"
        assert ws.cell(data_row, 2).value == "2025-01-10"
        assert ws.cell(data_row, 3).value == "BTC"
        assert ws.cell(data_row, 4).value == Decimal("0.5")
        assert ws.cell(data_row, 5).value == Decimal("20000")
        assert ws.cell(data_row, 6).value == Decimal("25000")
        assert ws.cell(data_row, 7).value == Decimal("5000")
        assert ws.cell(data_row, 8).value == "Short-term"
        assert ws.cell(data_row, 9).value == "Kraken"
        assert ws.cell(data_row, 10).value == "Kraken"
        assert ws.cell(data_row, 11).value == "Ethereum"
        assert ws.cell(data_row, 12).value == "Test Entity"
        assert ws.cell(data_row, 13).value == "US"
        assert ws.cell(data_row, 14).value == "J"
        assert ws.cell(data_row, 15).value == "NO"
        assert ws.cell(data_row, 16).value == ""
        assert ws.cell(data_row, 17).value == ""

    def test_capital_entry_review_flag_yes_with_reason(self):
        entry = _make_capital_entry(review_required=True, review_reason="Missing cost basis")
        report = _make_crypto_tax_report(capital_entries=[entry])
        wb = openpyxl.Workbook()
        write_crypto_gains_sheet(wb, report)
        ws = wb["Crypto Gains"]
        header_row = None
        for r in range(1, ws.max_row + 1):
            if ws.cell(r, 1).value == "Disposal date":
                header_row = r
                break
        data_row = header_row + 1
        assert ws.cell(data_row, 15).value == "YES: Missing cost basis"

    def test_token_origin_written(self):
        entry = _make_capital_entry(token_swap_history="ETH (swap_conversion, high confidence)")
        report = _make_crypto_tax_report(capital_entries=[entry])
        wb = openpyxl.Workbook()
        write_crypto_gains_sheet(wb, report)
        ws = wb["Crypto Gains"]
        header_row = None
        for r in range(1, ws.max_row + 1):
            if ws.cell(r, 1).value == "Disposal date":
                header_row = r
                break
        data_row = header_row + 1
        assert ws.cell(data_row, 17).value == "ETH (swap_conversion, high confidence)"

    def test_multiple_entries_on_separate_rows(self):
        entry1 = _make_capital_entry(asset="BTC", disposal_date="2025-01-15")
        entry2 = _make_capital_entry(asset="ETH", disposal_date="2025-02-20")
        report = _make_crypto_tax_report(capital_entries=[entry1, entry2])
        wb = openpyxl.Workbook()
        write_crypto_gains_sheet(wb, report)
        ws = wb["Crypto Gains"]
        header_row = None
        for r in range(1, ws.max_row + 1):
            if ws.cell(r, 1).value == "Disposal date":
                header_row = r
                break
        assert ws.cell(header_row + 1, 3).value == "BTC"
        assert ws.cell(header_row + 2, 3).value == "ETH"


@pytest.mark.unit
class TestCryptoGainsSheetStatistics:
    """Tests for the 1b. CAPITAL GAINS STATISTICS section."""

    STATS_HEADERS = [
        "Holding Period",
        "Count",
        "Acquisition Cost Total (EUR)",
        "Disposal Proceeds Total (EUR)",
        "Gain/Loss Total (EUR)",
    ]

    def test_statistics_section_title_written(self):
        wb = openpyxl.Workbook()
        report = _make_crypto_tax_report()
        write_crypto_gains_sheet(wb, report)
        ws = wb["Crypto Gains"]
        found = False
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=1):
            if row[0].value == "1b. CAPITAL GAINS STATISTICS":
                found = True
                break
        assert found

    def test_statistics_section_title_is_bold(self):
        wb = openpyxl.Workbook()
        report = _make_crypto_tax_report()
        write_crypto_gains_sheet(wb, report)
        ws = wb["Crypto Gains"]
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=1):
            if row[0].value == "1b. CAPITAL GAINS STATISTICS":
                assert row[0].font.bold is True
                break

    def test_statistics_headers_written(self):
        wb = openpyxl.Workbook()
        report = _make_crypto_tax_report()
        write_crypto_gains_sheet(wb, report)
        ws = wb["Crypto Gains"]
        stats_title_row = None
        for r in range(1, ws.max_row + 1):
            if ws.cell(r, 1).value == "1b. CAPITAL GAINS STATISTICS":
                stats_title_row = r
                break
        assert stats_title_row is not None
        header_row = stats_title_row + 1
        for idx, expected in enumerate(self.STATS_HEADERS, start=1):
            assert ws.cell(header_row, idx).value == expected

    def test_statistics_headers_are_bold(self):
        wb = openpyxl.Workbook()
        report = _make_crypto_tax_report()
        write_crypto_gains_sheet(wb, report)
        ws = wb["Crypto Gains"]
        stats_title_row = None
        for r in range(1, ws.max_row + 1):
            if ws.cell(r, 1).value == "1b. CAPITAL GAINS STATISTICS":
                stats_title_row = r
                break
        header_row = stats_title_row + 1
        for idx in range(1, len(self.STATS_HEADERS) + 1):
            assert ws.cell(header_row, idx).font.bold is True

    def test_statistics_period_rows(self):
        wb = openpyxl.Workbook()
        report = _make_crypto_tax_report()
        write_crypto_gains_sheet(wb, report)
        ws = wb["Crypto Gains"]
        stats_title_row = None
        for r in range(1, ws.max_row + 1):
            if ws.cell(r, 1).value == "1b. CAPITAL GAINS STATISTICS":
                stats_title_row = r
                break
        data_start = stats_title_row + 2
        assert ws.cell(data_start, 1).value == "Short-term"
        assert ws.cell(data_start, 2).value == 1
        assert ws.cell(data_start + 1, 1).value == "Long-term"
        assert ws.cell(data_start + 1, 2).value == 0
        assert ws.cell(data_start + 2, 1).value == "Mixed"
        assert ws.cell(data_start + 3, 1).value == "Unknown"
        assert ws.cell(data_start + 4, 1).value == "Grand Total"

    def test_statistics_values_match_report(self):
        wb = openpyxl.Workbook()
        report = _make_crypto_tax_report()
        write_crypto_gains_sheet(wb, report)
        ws = wb["Crypto Gains"]
        stats_title_row = None
        for r in range(1, ws.max_row + 1):
            if ws.cell(r, 1).value == "1b. CAPITAL GAINS STATISTICS":
                stats_title_row = r
                break
        data_start = stats_title_row + 2
        assert ws.cell(data_start, 3).value == Decimal("20000")
        assert ws.cell(data_start, 4).value == Decimal("25000")
        assert ws.cell(data_start, 5).value == Decimal("5000")


@pytest.mark.unit
class TestCryptoGainsSheetAutoWidth:
    """Tests that auto_column_width is called."""

    def test_auto_width_adjusts_columns(self):
        wb = openpyxl.Workbook()
        report = _make_crypto_tax_report()
        write_crypto_gains_sheet(wb, report)
        ws = wb["Crypto Gains"]
        assert ws.column_dimensions["A"].width > 0

    def test_column_widths_respect_max_cell_width_cap(self):
        """Verify no column exceeds MAX_CELL_WIDTH + 2 even with long token origins."""
        from tax_reporting.application.persisting.excel_utils import MAX_CELL_WIDTH

        wb = openpyxl.Workbook()
        report = _make_crypto_tax_report()
        write_crypto_gains_sheet(wb, report)
        ws = wb["Crypto Gains"]

        # All column widths should be capped at MAX_CELL_WIDTH + 2
        max_allowed = MAX_CELL_WIDTH + 2
        for col_letter, col_dim in ws.column_dimensions.items():
            if col_dim.width is not None:
                assert (
                    col_dim.width <= max_allowed
                ), f"Column {col_letter} width {col_dim.width} exceeds cap {max_allowed}"

    def test_all_columns_have_reasonable_widths(self):
        """Verify no column is collapsed: all have width >= MIN_DATA_WIDTH floor."""
        wb = openpyxl.Workbook()
        report = _make_crypto_tax_report()
        write_crypto_gains_sheet(wb, report)
        ws = wb["Crypto Gains"]

        for col_idx in range(1, 21):  # Columns A through T (20 columns with OGR validation)
            col_letter = openpyxl.utils.get_column_letter(col_idx)
            width = ws.column_dimensions[col_letter].width
            assert width is not None, f"Column {col_letter} should have a width set"
            assert width >= 4, f"Column {col_letter} width {width} is too small"


@pytest.mark.unit
class TestCryptoGainsSheetEmptyEntries:
    """Tests handling of empty capital entries list."""

    def test_empty_entries_writes_headers_no_data(self):
        report = _make_crypto_tax_report(capital_entries=[])
        wb = openpyxl.Workbook()
        write_crypto_gains_sheet(wb, report)
        ws = wb["Crypto Gains"]
        header_row = None
        for r in range(1, ws.max_row + 1):
            if ws.cell(r, 1).value == "Disposal date":
                header_row = r
                break
        assert header_row is not None
        next_value = ws.cell(header_row + 1, 1).value
        assert next_value is None or "1b. CAPITAL GAINS STATISTICS" in str(next_value)


_RED_FILL = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")
_YELLOW_FILL = PatternFill(start_color="FFFFFF00", end_color="FFFFFF00", fill_type="solid")
_NUM_CAPITAL_COLUMNS = 20  # Updated to include OGR validation columns


def _find_header_row(ws: openpyxl.worksheet.worksheet.Worksheet) -> int:
    for r in range(1, ws.max_row + 1):
        if ws.cell(r, 1).value == "Disposal date":
            return r
    raise AssertionError("Header row not found")


def _is_red_fill(cell: openpyxl.cell.cell.Cell) -> bool:
    fill = cell.fill
    return (
        fill.start_color.rgb == "FFFF0000"
        and fill.end_color.rgb == "FFFF0000"
        and fill.fill_type == "solid"
    )


@pytest.mark.unit
class TestCryptoGainsSheetZeroCostRedBackground:
    """Tests for red background rendering on zero-cost entries above threshold."""

    def test_render_zero_cost_entry_has_red_background(self):
        entry = _make_capital_entry(cost_eur=Decimal("0"), gain_loss_eur=Decimal("500"))
        report = _make_crypto_tax_report(capital_entries=[entry])
        wb = openpyxl.Workbook()
        write_crypto_gains_sheet(wb, report)
        ws = wb["Crypto Gains"]
        data_row = _find_header_row(ws) + 1
        for col in range(1, _NUM_CAPITAL_COLUMNS + 1):
            assert _is_red_fill(ws.cell(data_row, col)), f"Column {col} should have red fill"

    def test_render_zero_cost_below_threshold_no_red_background(self):
        entry = _make_capital_entry(cost_eur=Decimal("0"), gain_loss_eur=Decimal("0.5"))
        report = _make_crypto_tax_report(capital_entries=[entry])
        wb = openpyxl.Workbook()
        write_crypto_gains_sheet(wb, report)
        ws = wb["Crypto Gains"]
        data_row = _find_header_row(ws) + 1
        for col in range(1, _NUM_CAPITAL_COLUMNS + 1):
            assert not _is_red_fill(ws.cell(data_row, col)), f"Column {col} should NOT have red fill"

    def test_render_nonzero_cost_no_red_background(self):
        entry = _make_capital_entry(cost_eur=Decimal("100"), gain_loss_eur=Decimal("5000"))
        report = _make_crypto_tax_report(capital_entries=[entry])
        wb = openpyxl.Workbook()
        write_crypto_gains_sheet(wb, report)
        ws = wb["Crypto Gains"]
        data_row = _find_header_row(ws) + 1
        for col in range(1, _NUM_CAPITAL_COLUMNS + 1):
            assert not _is_red_fill(ws.cell(data_row, col)), f"Column {col} should NOT have red fill"

    def test_render_normal_row_no_red_background(self):
        entry = _make_capital_entry(cost_eur=Decimal("20000"), gain_loss_eur=Decimal("5000"))
        report = _make_crypto_tax_report(capital_entries=[entry])
        wb = openpyxl.Workbook()
        write_crypto_gains_sheet(wb, report)
        ws = wb["Crypto Gains"]
        data_row = _find_header_row(ws) + 1
        for col in range(1, _NUM_CAPITAL_COLUMNS + 1):
            assert not _is_red_fill(ws.cell(data_row, col)), f"Column {col} should NOT have red fill"

    def test_render_zero_cost_at_exact_threshold_has_red_background(self):
        """Gain/loss exactly equal to threshold must trigger red fill (>= boundary)."""
        entry = _make_capital_entry(cost_eur=Decimal("0"), gain_loss_eur=Decimal("50"))
        report = _make_crypto_tax_report(
            capital_entries=[entry], zero_basis_review_threshold=Decimal("50")
        )
        wb = openpyxl.Workbook()
        write_crypto_gains_sheet(wb, report)
        ws = wb["Crypto Gains"]
        data_row = _find_header_row(ws) + 1
        for col in range(1, _NUM_CAPITAL_COLUMNS + 1):
            assert _is_red_fill(ws.cell(data_row, col)), f"Column {col} should have red fill at threshold"

    def test_render_zero_cost_just_below_threshold_no_red_background(self):
        """Gain/loss just below threshold must NOT trigger red fill."""
        entry = _make_capital_entry(cost_eur=Decimal("0"), gain_loss_eur=Decimal("49.99"))
        report = _make_crypto_tax_report(
            capital_entries=[entry], zero_basis_review_threshold=Decimal("50")
        )
        wb = openpyxl.Workbook()
        write_crypto_gains_sheet(wb, report)
        ws = wb["Crypto Gains"]
        data_row = _find_header_row(ws) + 1
        for col in range(1, _NUM_CAPITAL_COLUMNS + 1):
            assert not _is_red_fill(ws.cell(data_row, col)), f"Column {col} should NOT have red fill below threshold"


_BLUE_FILL = PatternFill(start_color="FFCCFFFF", end_color="FFCCFFFF", fill_type="solid")


def _is_blue_fill(cell: openpyxl.cell.cell.Cell) -> bool:
    fill = cell.fill
    return (
        fill.start_color.rgb == "FFCCFFFF"
        and fill.end_color.rgb == "FFCCFFFF"
        and fill.fill_type == "solid"
    )


@pytest.mark.unit
class TestCryptoGainsSheetMultiDateBlueBackground:
    """Tests for blue background rendering on multi-acquisition-date entries."""

    def test_render_multi_date_entry_has_blue_background(self):
        """Entry with multi_acquisition_dates=True and review_required=False gets blue fill."""
        entry = _make_capital_entry(
            multi_acquisition_dates=True,
            notes="Acquired: 2024-04-13, 2024-04-19 (2 lots)",
        )
        report = _make_crypto_tax_report(capital_entries=[entry])
        wb = openpyxl.Workbook()
        write_crypto_gains_sheet(wb, report)
        ws = wb["Crypto Gains"]
        data_row = _find_header_row(ws) + 1
        for col in range(1, _NUM_CAPITAL_COLUMNS + 1):
            assert _is_blue_fill(ws.cell(data_row, col)), f"Column {col} should have blue fill"

    def test_render_review_required_takes_precedence_over_blue(self):
        """Entry with multi_acquisition_dates=True and review_required=True gets red fill, not blue."""
        entry = _make_capital_entry(
            multi_acquisition_dates=True,
            review_required=True,
            review_reason="Test reason",
            notes="",
        )
        report = _make_crypto_tax_report(capital_entries=[entry])
        wb = openpyxl.Workbook()
        write_crypto_gains_sheet(wb, report)
        ws = wb["Crypto Gains"]
        data_row = _find_header_row(ws) + 1
        for col in range(1, _NUM_CAPITAL_COLUMNS + 1):
            assert _is_red_fill(ws.cell(data_row, col)), f"Column {col} should have red fill (precedence)"
            assert not _is_blue_fill(ws.cell(data_row, col)), f"Column {col} should NOT have blue fill"

    def test_render_single_date_entry_no_background(self):
        """Entry with multi_acquisition_dates=False has no fill (default)."""
        entry = _make_capital_entry(multi_acquisition_dates=False)
        report = _make_crypto_tax_report(capital_entries=[entry])
        wb = openpyxl.Workbook()
        write_crypto_gains_sheet(wb, report)
        ws = wb["Crypto Gains"]
        data_row = _find_header_row(ws) + 1
        for col in range(1, _NUM_CAPITAL_COLUMNS + 1):
            cell = ws.cell(data_row, col)
            assert not _is_red_fill(cell), f"Column {col} should NOT have red fill"
            assert not _is_blue_fill(cell), f"Column {col} should NOT have blue fill"

    def test_render_zero_cost_red_takes_precedence_over_blue(self):
        """Zero-cost entry above threshold gets red fill even with multi_acquisition_dates=True."""
        entry = _make_capital_entry(
            cost_eur=Decimal("0"),
            gain_loss_eur=Decimal("100"),
            multi_acquisition_dates=True,
        )
        report = _make_crypto_tax_report(capital_entries=[entry])
        wb = openpyxl.Workbook()
        write_crypto_gains_sheet(wb, report)
        ws = wb["Crypto Gains"]
        data_row = _find_header_row(ws) + 1
        for col in range(1, _NUM_CAPITAL_COLUMNS + 1):
            assert _is_red_fill(ws.cell(data_row, col)), f"Column {col} should have red fill (zero-cost)"
            assert not _is_blue_fill(ws.cell(data_row, col)), f"Column {col} should NOT have blue fill"


@pytest.mark.unit
class TestCryptoGainsSheetLossValues:
    """Tests that negative gain/loss values are preserved in Excel output."""

    def test_loss_value_written_to_excel(self):
        """Entry with negative gain_loss_eur appears as negative value in Excel (not absolute)."""
        entry = _make_capital_entry(
            disposal_date="2025-01-13",
            asset="USDT",
            gain_loss_eur=Decimal("-138.73"),
        )
        report = _make_crypto_tax_report(capital_entries=[entry])
        wb = openpyxl.Workbook()
        write_crypto_gains_sheet(wb, report)
        ws = wb["Crypto Gains"]
        data_row = _find_header_row(ws) + 1
        # Column 7 is Gain/Loss (EUR)
        gain_loss_cell = ws.cell(data_row, 7)
        assert gain_loss_cell.value == Decimal("-138.73"), f"Expected -138.73, got {gain_loss_cell.value}"
        # Verify it's actually negative (not positive absolute value)
        assert gain_loss_cell.value < 0, "Loss value must be negative"

    def test_ogr_override_reflected_in_excel(self):
        """OGR-overridden loss shows correct negative value in Excel output."""
        # This test simulates the case where OGR override changes +22.71 EUR (CG)
        # to -138.73 EUR (OGR Type="Loss")
        entry = _make_capital_entry(
            disposal_date="2025-01-13",
            asset="USDT",
            platform="ByBit",
            gain_loss_eur=Decimal("-138.73"),  # OGR-overridden value
        )
        report = _make_crypto_tax_report(capital_entries=[entry])
        wb = openpyxl.Workbook()
        write_crypto_gains_sheet(wb, report)
        ws = wb["Crypto Gains"]
        data_row = _find_header_row(ws) + 1
        # Verify the full entry including the overridden loss value
        assert ws.cell(data_row, 1).value == "2025-01-13"  # Disposal date
        assert ws.cell(data_row, 3).value == "USDT"  # Asset
        assert ws.cell(data_row, 10).value == "ByBit"  # Platform
        # Column 7 is Gain/Loss (EUR) - must show the OGR loss value
        gain_loss_cell = ws.cell(data_row, 7)
        assert gain_loss_cell.value == Decimal("-138.73"), f"Expected -138.73 (OGR loss), got {gain_loss_cell.value}"
        assert gain_loss_cell.value < 0, "OGR loss value must be negative in Excel"


def _is_yellow_fill(cell: openpyxl.cell.cell.Cell) -> bool:
    fill = cell.fill
    return (
        fill.start_color.rgb == "FFFFFF00"
        and fill.end_color.rgb == "FFFFFF00"
        and fill.fill_type == "solid"
    )


@pytest.mark.unit
class TestCryptoGainsSheetOgrValidation:
    """Tests for OGR validation columns in the Crypto Gains sheet."""

    OGR_VALIDATION_HEADERS = [
        "OGR Gain/Loss (EUR)",
        "OGR Diff (%)",
        "OGR Review",
    ]

    def test_ogr_validation_headers_written(self):
        """Verify that OGR validation headers are present in the sheet."""
        wb = openpyxl.Workbook()
        report = _make_crypto_tax_report()
        write_crypto_gains_sheet(wb, report)
        ws = wb["Crypto Gains"]
        header_row = _find_header_row(ws)
        # Columns 18, 19, 20 should be the OGR validation headers
        assert ws.cell(header_row, 18).value == "OGR Gain/Loss (EUR)"
        assert ws.cell(header_row, 19).value == "OGR Diff (%)"
        assert ws.cell(header_row, 20).value == "OGR Review"

    def test_ogr_validation_columns_blank_when_ogr_validation_none(self):
        """When ogr_validation is None, OGR columns should be blank."""
        entry = _make_capital_entry(ogr_validation=None)
        report = _make_crypto_tax_report(capital_entries=[entry])
        wb = openpyxl.Workbook()
        write_crypto_gains_sheet(wb, report)
        ws = wb["Crypto Gains"]
        data_row = _find_header_row(ws) + 1
        assert ws.cell(data_row, 18).value is None
        assert ws.cell(data_row, 19).value is None
        assert ws.cell(data_row, 20).value is None

    def test_ogr_review_shows_no_when_review_required_false(self):
        """When ogr_validation.review_required=False, OGR Review shows NO."""
        ogr_val = OgrValidationResult(
            ogr_gain_loss=Decimal("100"),
            calculated_gain_loss=Decimal("100"),
            direction_conflict=False,
            magnitude_diff_percent=Decimal("0"),
            review_required=False,
            review_reason=None,
        )
        entry = _make_capital_entry(ogr_validation=ogr_val)
        report = _make_crypto_tax_report(capital_entries=[entry])
        wb = openpyxl.Workbook()
        write_crypto_gains_sheet(wb, report)
        ws = wb["Crypto Gains"]
        data_row = _find_header_row(ws) + 1
        assert ws.cell(data_row, 18).value == Decimal("100")  # OGR Gain/Loss
        assert ws.cell(data_row, 19).value == Decimal("0")  # OGR Diff (%)
        assert ws.cell(data_row, 20).value == "NO"  # OGR Review
        # No special fill
        assert not _is_red_fill(ws.cell(data_row, 18))
        assert not _is_yellow_fill(ws.cell(data_row, 18))

    def test_ogr_direction_override_shows_red_fill(self):
        """When review_reason contains 'OGR direction override', entire row gets RED fill."""
        ogr_val = OgrValidationResult(
            ogr_gain_loss=Decimal("-50"),
            calculated_gain_loss=Decimal("50"),
            direction_conflict=True,
            magnitude_diff_percent=Decimal("200"),
            review_required=True,
            review_reason="OGR direction override: OGR shows loss, CG shows gain",
        )
        entry = _make_capital_entry(ogr_validation=ogr_val)
        report = _make_crypto_tax_report(capital_entries=[entry])
        wb = openpyxl.Workbook()
        write_crypto_gains_sheet(wb, report)
        ws = wb["Crypto Gains"]
        data_row = _find_header_row(ws) + 1
        # Verify OGR values
        assert ws.cell(data_row, 18).value == Decimal("-50")  # OGR Gain/Loss
        assert ws.cell(data_row, 19).value == Decimal("200")  # OGR Diff (%)
        assert ws.cell(data_row, 20).value == "YES: OGR direction override: OGR shows loss, CG shows gain"
        # Verify RED fill on the entire row (all 20 columns)
        for col in range(1, _NUM_CAPITAL_COLUMNS + 1):
            assert _is_red_fill(ws.cell(data_row, col)), f"Column {col} should have RED fill for direction override"

    def test_ogr_magnitude_diff_shows_yellow_fill(self):
        """When magnitude differs, entire row gets YELLOW fill."""
        ogr_val = OgrValidationResult(
            ogr_gain_loss=Decimal("100"),
            calculated_gain_loss=Decimal("90"),
            direction_conflict=False,
            magnitude_diff_percent=Decimal("10"),
            review_required=True,
            review_reason="magnitude differs by 10%",
        )
        entry = _make_capital_entry(ogr_validation=ogr_val)
        report = _make_crypto_tax_report(capital_entries=[entry])
        wb = openpyxl.Workbook()
        write_crypto_gains_sheet(wb, report)
        ws = wb["Crypto Gains"]
        data_row = _find_header_row(ws) + 1
        # Verify OGR values
        assert ws.cell(data_row, 18).value == Decimal("100")  # OGR Gain/Loss
        assert ws.cell(data_row, 19).value == Decimal("10")  # OGR Diff (%)
        assert ws.cell(data_row, 20).value == "YES: magnitude differs by 10%"
        # Verify YELLOW fill on the entire row (all 20 columns)
        for col in range(1, _NUM_CAPITAL_COLUMNS + 1):
            assert _is_yellow_fill(ws.cell(data_row, col)), f"Column {col} should have YELLOW fill for magnitude diff"

    def test_ogr_gain_loss_value_displayed(self):
        """OGR Gain/Loss column shows the ogr_gain_loss value when present."""
        ogr_val = OgrValidationResult(
            ogr_gain_loss=Decimal("150.25"),
            calculated_gain_loss=Decimal("150.25"),
            direction_conflict=False,
            magnitude_diff_percent=Decimal("0"),
            review_required=False,
            review_reason=None,
        )
        entry = _make_capital_entry(ogr_validation=ogr_val)
        report = _make_crypto_tax_report(capital_entries=[entry])
        wb = openpyxl.Workbook()
        write_crypto_gains_sheet(wb, report)
        ws = wb["Crypto Gains"]
        data_row = _find_header_row(ws) + 1
        assert ws.cell(data_row, 18).value == Decimal("150.25")

    def test_ogr_diff_percent_displayed(self):
        """OGR Diff (%) column shows the magnitude_diff_percent value."""
        ogr_val = OgrValidationResult(
            ogr_gain_loss=Decimal("100"),
            calculated_gain_loss=Decimal("85"),
            direction_conflict=False,
            magnitude_diff_percent=Decimal("15"),
            review_required=True,
            review_reason="magnitude differs by 15%",
        )
        entry = _make_capital_entry(ogr_validation=ogr_val)
        report = _make_crypto_tax_report(capital_entries=[entry])
        wb = openpyxl.Workbook()
        write_crypto_gains_sheet(wb, report)
        ws = wb["Crypto Gains"]
        data_row = _find_header_row(ws) + 1
        assert ws.cell(data_row, 19).value == Decimal("15")
