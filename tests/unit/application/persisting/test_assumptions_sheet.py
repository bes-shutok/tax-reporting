"""Tests for the Assumptions & Methodology sheet writer."""

import re
from decimal import Decimal

import openpyxl
import pytest

from tax_reporting.application.crypto_reporting import (
    CapitalGainPeriodStats,
    CryptoCapitalGainEntry,
    CryptoCapitalGainStats,
    CryptoReconciliationSummary,
    CryptoRewardIncomeEntry,
    CryptoTaxReport,
    RewardTaxClassification,
)
from tax_reporting.application.persisting.assumptions_sheet import write_assumptions_and_methodology_sheet
from tests.conftest import make_operator_origin


def _make_origin(platform: str, assumption: str | None = None, platform_review_required: bool = False):
    """Create an OperatorOrigin for the given platform."""
    return make_operator_origin(
        platform=platform,
        operator_entity=f"{platform} Entity",
        operator_country="US",
        review_required=False,
        platform_assumption=assumption,
        platform_review_required=platform_review_required,
    )


def _make_capital_entry(
    platform: str = "Kraken",
    operator_assumption: str | None = None,
    platform_review_required: bool = False,
) -> CryptoCapitalGainEntry:
    """Create a capital gain entry with optional platform assumption/review flag."""
    origin = _make_origin(platform, operator_assumption, platform_review_required)
    return CryptoCapitalGainEntry(
        disposal_date="2025-06-15",
        acquisition_date="2025-01-10",
        asset="BTC",
        amount=Decimal("0.5"),
        cost_eur=Decimal("20000"),
        proceeds_eur=Decimal("25000"),
        gain_loss_eur=Decimal("5000"),
        holding_period="Short-term",
        wallet="kraken-wallet",
        platform=platform,
        chain="Ethereum",
        operator_origin=origin,
        annex_hint="J",
        review_required=False,
        notes="",
    )


def _make_reward_entry(
    platform: str = "Kraken",
    operator_assumption: str | None = None,
    platform_review_required: bool = False,
) -> CryptoRewardIncomeEntry:
    """Create a reward entry with optional platform assumption/review flag."""
    origin = _make_origin(platform, operator_assumption, platform_review_required)
    return CryptoRewardIncomeEntry(
        date="2025-01-15",
        asset="ETH",
        amount=Decimal("1"),
        value_eur=Decimal("100"),
        income_label="Staking",
        source_type="staking",
        wallet="kraken-wallet",
        platform=platform,
        chain="Ethereum",
        operator_origin=origin,
        annex_hint="J",
        foreign_tax_eur=Decimal("0"),
        tax_classification=RewardTaxClassification.TAXABLE_NOW,
        review_required=False,
        description="Staking reward",
    )


def _make_stats() -> CryptoCapitalGainStats:
    """Create empty stats for test reports."""
    empty = CapitalGainPeriodStats(
        count=0,
        cost_total_eur=Decimal("0"),
        proceeds_total_eur=Decimal("0"),
        gain_loss_total_eur=Decimal("0"),
    )
    return CryptoCapitalGainStats(
        short_term=empty,
        long_term=empty,
        mixed=empty,
        unknown=empty,
        grand_total=empty,
    )


def _make_crypto_tax_report(
    capital_entries: list[CryptoCapitalGainEntry] | None = None,
    reward_entries: list[CryptoRewardIncomeEntry] | None = None,
) -> CryptoTaxReport:
    """Create a test crypto tax report."""
    entries = capital_entries or []
    rewards = reward_entries or []
    reconciliation = CryptoReconciliationSummary(
        capital_rows=len(entries),
        reward_rows=len(rewards),
        short_term_rows=0,
        long_term_rows=0,
        mixed_rows=0,
        unknown_rows=0,
        capital_cost_total_eur=Decimal("0"),
        capital_proceeds_total_eur=Decimal("0"),
        capital_gain_total_eur=Decimal("0"),
        reward_total_eur=Decimal("0"),
        opening_holdings=None,
        closing_holdings=None,
    )
    return CryptoTaxReport(
        tax_year=2025,
        capital_entries=entries,
        reward_entries=rewards,
        reconciliation=reconciliation,
        capital_gain_stats=_make_stats(),
        pdf_summary=None,
        zero_basis_review_threshold=Decimal("50"),
    )


@pytest.mark.unit
class TestAssumptionsAndMethodologySheetName:
    """Tests that the sheet is created with the correct name."""

    def test_sheet_named_assumptions_and_methodology(self):
        wb = openpyxl.Workbook()
        write_assumptions_and_methodology_sheet(wb)
        assert "Assumptions & Methodology" in wb.sheetnames


@pytest.mark.unit
class TestAssumptionsAndMethodologySheetTitle:
    """Tests that the sheet writes the title correctly."""

    def test_title_row_present(self):
        wb = openpyxl.Workbook()
        write_assumptions_and_methodology_sheet(wb)
        ws = wb["Assumptions & Methodology"]
        assert ws.cell(1, 1).value == "Assumptions & Methodology"

    def test_title_is_bold(self):
        wb = openpyxl.Workbook()
        write_assumptions_and_methodology_sheet(wb)
        ws = wb["Assumptions & Methodology"]
        assert ws.cell(1, 1).font.bold is True


@pytest.mark.unit
class TestPlatformAssumptionsSection:
    """Tests that the Platform Assumptions section (section 1) works correctly."""

    def test_no_entries_shows_message(self):
        wb = openpyxl.Workbook()
        write_assumptions_and_methodology_sheet(wb)
        ws = wb["Assumptions & Methodology"]
        assert "No platform data found" in ws.cell(5, 1).value

    def test_all_platforms_appear_even_without_assumption(self):
        """Every platform from the data should appear, not just ones with assumption text."""
        wb = openpyxl.Workbook()
        entry = _make_capital_entry(platform="Kraken", operator_assumption=None)
        report = _make_crypto_tax_report(capital_entries=[entry])
        write_assumptions_and_methodology_sheet(wb, capital_entries=report.capital_entries)
        ws = wb["Assumptions & Methodology"]
        # Row 6 should contain Kraken (headers at row 5)
        assert ws.cell(6, 1).value == "Kraken"

    def test_platform_review_required_shows_yes(self):
        """Platforms with platform_review_required=True show YES in column 5."""
        wb = openpyxl.Workbook()
        entry = _make_capital_entry(
            platform="Bybit",
            operator_assumption="verify account region",
            platform_review_required=True,
        )
        report = _make_crypto_tax_report(capital_entries=[entry])
        write_assumptions_and_methodology_sheet(wb, capital_entries=report.capital_entries)
        ws = wb["Assumptions & Methodology"]
        # New layout: Platform | Operator Entity | Country | Confidence | Review Required | Assumption | Count
        assert ws.cell(6, 1).value == "Bybit"
        assert ws.cell(6, 5).value == "YES"

    def test_platform_no_review_shows_no(self):
        """Platforms without platform_review_required show NO in column 5."""
        wb = openpyxl.Workbook()
        entry = _make_capital_entry(platform="Kraken", platform_review_required=False)
        report = _make_crypto_tax_report(capital_entries=[entry])
        write_assumptions_and_methodology_sheet(wb, capital_entries=report.capital_entries)
        ws = wb["Assumptions & Methodology"]
        assert ws.cell(6, 5).value == "NO"


    def test_row_review_flag_does_not_trigger_platform_review(self):
        """Row-level review_required must not bleed into platform review column/fill."""
        wb = openpyxl.Workbook()
        origin = make_operator_origin(
            platform="Kraken",
            operator_entity="Kraken Entity",
            operator_country="US",
            review_required=True,
            review_reason="Origin row needs review",
            platform_review_required=False,
        )
        entry = CryptoCapitalGainEntry(
            disposal_date="2025-06-15",
            acquisition_date="2025-01-10",
            asset="BTC",
            amount=Decimal("0.5"),
            cost_eur=Decimal("20000"),
            proceeds_eur=Decimal("25000"),
            gain_loss_eur=Decimal("5000"),
            holding_period="Short-term",
            wallet="kraken-wallet",
            platform="Kraken",
            chain="Ethereum",
            operator_origin=origin,
            annex_hint="J",
            review_required=True,
            review_reason="Row needs manual review",
            notes="",
        )
        report = _make_crypto_tax_report(capital_entries=[entry])
        write_assumptions_and_methodology_sheet(wb, capital_entries=report.capital_entries)
        ws = wb["Assumptions & Methodology"]

        assert ws.cell(6, 5).value == "NO"
        assert ws.cell(6, 1).fill.fill_type is None

    def test_assumption_text_in_column_6(self):
        """Assumption text appears in column 6."""
        wb = openpyxl.Workbook()
        entry = _make_capital_entry(
            platform="Bybit",
            operator_assumption="Bybit uses account-region specific entities; verify your account region",
        )
        report = _make_crypto_tax_report(capital_entries=[entry])
        write_assumptions_and_methodology_sheet(wb, capital_entries=report.capital_entries)
        ws = wb["Assumptions & Methodology"]
        assert "account-region" in ws.cell(6, 6).value

    def test_transaction_count_in_column_7(self):
        """Transaction count appears in column 7."""
        wb = openpyxl.Workbook()
        bybit_entries = [_make_capital_entry(platform="Bybit") for _ in range(3)]
        report = _make_crypto_tax_report(capital_entries=bybit_entries)
        write_assumptions_and_methodology_sheet(wb, capital_entries=report.capital_entries)
        ws = wb["Assumptions & Methodology"]
        assert ws.cell(6, 7).value == 3

    def test_multiple_platforms_all_listed(self):
        """Multiple platforms from capital and reward entries all appear."""
        wb = openpyxl.Workbook()
        bybit_entry = _make_capital_entry(platform="Bybit", platform_review_required=True)
        mantle_reward = _make_reward_entry(platform="Mantle")
        kraken_entry = _make_capital_entry(platform="Kraken")
        report = _make_crypto_tax_report(
            capital_entries=[bybit_entry, kraken_entry],
            reward_entries=[mantle_reward],
        )
        write_assumptions_and_methodology_sheet(
            wb,
            capital_entries=report.capital_entries,
            reward_entries=report.reward_entries,
        )
        ws = wb["Assumptions & Methodology"]
        platforms = [ws.cell(r, 1).value for r in range(6, 10) if ws.cell(r, 1).value]
        assert "Bybit" in platforms
        assert "Mantle" in platforms
        assert "Kraken" in platforms

    def test_review_required_platforms_sorted_first(self):
        """Platforms with platform_review_required=True appear before non-review platforms."""
        wb = openpyxl.Workbook()
        kraken_entry = _make_capital_entry(platform="Kraken", platform_review_required=False)
        bybit_entry = _make_capital_entry(platform="Bybit", platform_review_required=True)
        report = _make_crypto_tax_report(capital_entries=[kraken_entry, bybit_entry])
        write_assumptions_and_methodology_sheet(wb, capital_entries=report.capital_entries)
        ws = wb["Assumptions & Methodology"]
        # Bybit (review_required) must appear before Kraken (no review)
        platforms = [ws.cell(r, 1).value for r in range(6, 9) if ws.cell(r, 1).value]
        assert platforms[0] == "Bybit"
        assert platforms[1] == "Kraken"


@pytest.mark.unit
class TestMethodologyAssumptionsSection:
    """Tests that the Methodology Assumptions section (section 2) is rendered correctly."""

    def test_methodology_section_header_present(self):
        """The 'Methodology Assumptions' header should be present after platform data."""
        wb = openpyxl.Workbook()
        entry = _make_capital_entry(platform="Kraken")
        report = _make_crypto_tax_report(capital_entries=[entry])
        write_assumptions_and_methodology_sheet(wb, capital_entries=report.capital_entries)
        ws = wb["Assumptions & Methodology"]

        # Find the methodology header (after platform data)
        methodology_header_found = False
        for row_idx in range(1, 50):
            cell_value = ws.cell(row_idx, 1).value
            if cell_value == "Methodology Assumptions":
                methodology_header_found = True
                assert ws.cell(row_idx, 1).font.bold is True
                break

        assert methodology_header_found, "Methodology Assumptions header not found"

    def test_methodology_section_has_all_item_labels(self):
        """All five methodology subsections should be present."""
        wb = openpyxl.Workbook()
        entry = _make_capital_entry(platform="Kraken")
        report = _make_crypto_tax_report(capital_entries=[entry])
        write_assumptions_and_methodology_sheet(wb, capital_entries=report.capital_entries)
        ws = wb["Assumptions & Methodology"]

        # Updated: 7 sections with new labels matching the grouped structure
        expected_labels = {
            "Aggregation Approach",
            "Cost Basis Method: FIFO",
            "365-Day Exemption",
            "Materiality Threshold",
            "Data Sources",
        }

        found_labels = set()
        for row_idx in range(1, 100):
            cell_value = ws.cell(row_idx, 1).value
            if cell_value in expected_labels:
                found_labels.add(cell_value)

        assert found_labels == expected_labels, f"Expected {expected_labels}, found {found_labels}"

    def test_aggregation_approach_mentions_pt_c_025(self):
        """Aggregation Approach should reference PT-C-025."""
        wb = openpyxl.Workbook()
        entry = _make_capital_entry(platform="Kraken")
        report = _make_crypto_tax_report(capital_entries=[entry])
        write_assumptions_and_methodology_sheet(wb, capital_entries=report.capital_entries)
        ws = wb["Assumptions & Methodology"]

        # Find the Aggregation Approach row
        for row_idx in range(1, 100):
            if ws.cell(row_idx, 1).value == "Aggregation Approach":
                description = ws.cell(row_idx, 2).value
                assert "PT-C-025" in description
                assert "Quadro 9.4" in description
                assert "blue fill" in description
                break
        else:
            pytest.fail("Aggregation Approach row not found")

    def test_fifo_methodology_mentions_cirs_art_43(self):
        """FIFO Methodology should reference CIRS art. 43."""
        wb = openpyxl.Workbook()
        entry = _make_capital_entry(platform="Kraken")
        report = _make_crypto_tax_report(capital_entries=[entry])
        write_assumptions_and_methodology_sheet(wb, capital_entries=report.capital_entries)
        ws = wb["Assumptions & Methodology"]

        # Find the FIFO Methodology row (updated label)
        for row_idx in range(1, 100):
            if ws.cell(row_idx, 1).value == "Cost Basis Method: FIFO":
                description = ws.cell(row_idx, 2).value
                assert "CIRS art. 43" in description
                assert "PT-C-008" in description
                break
        else:
            pytest.fail("Cost Basis Method: FIFO row not found")

    def test_holding_period_mentions_calendar_year_arithmetic(self):
        """Holding Period Classification should mention calendar-year arithmetic."""
        wb = openpyxl.Workbook()
        entry = _make_capital_entry(platform="Kraken")
        report = _make_crypto_tax_report(capital_entries=[entry])
        write_assumptions_and_methodology_sheet(wb, capital_entries=report.capital_entries)
        ws = wb["Assumptions & Methodology"]

        # Find the Holding Period Classification row (updated label)
        for row_idx in range(1, 100):
            if ws.cell(row_idx, 1).value == "365-Day Exemption":
                description = ws.cell(row_idx, 2).value
                assert "Calendar-year arithmetic" in description
                assert "2024-02-29" in description
                assert "PT-C-011" in description
                break
        else:
            pytest.fail("365-Day Exemption row not found")

    def test_materiality_threshold_mentions_sub_1_eur(self):
        """Materiality Threshold should mention |gain/loss| < 1 EUR."""
        wb = openpyxl.Workbook()
        entry = _make_capital_entry(platform="Kraken")
        report = _make_crypto_tax_report(capital_entries=[entry])
        write_assumptions_and_methodology_sheet(wb, capital_entries=report.capital_entries)
        ws = wb["Assumptions & Methodology"]

        # Find the Materiality Threshold row
        for row_idx in range(1, 100):
            if ws.cell(row_idx, 1).value == "Materiality Threshold":
                description = ws.cell(row_idx, 2).value
                assert "< 1 EUR" in description
                assert "PT-C-028" in description
                break
        else:
            pytest.fail("Materiality Threshold row not found")

    def test_data_sources_mentions_ib_and_koinly(self):
        """Data Sources should mention Interactive Brokers and Koinly."""
        wb = openpyxl.Workbook()
        entry = _make_capital_entry(platform="Kraken")
        report = _make_crypto_tax_report(capital_entries=[entry])
        write_assumptions_and_methodology_sheet(wb, capital_entries=report.capital_entries)
        ws = wb["Assumptions & Methodology"]

        # Find the Data Sources row
        for row_idx in range(1, 100):
            if ws.cell(row_idx, 1).value == "Data Sources":
                description = ws.cell(row_idx, 2).value
                assert "Interactive Brokers" in description
                assert "Koinly" in description
                assert "config.ini" in description
                break
        else:
            pytest.fail("Data Sources row not found")

    def test_methodology_renders_without_crypto_data(self):
        """Methodology section must render even when crypto data is empty (Design Invariant #3)."""
        wb = openpyxl.Workbook()
        # Call with no entries (empty lists)
        write_assumptions_and_methodology_sheet(wb, capital_entries=[], reward_entries=[])
        ws = wb["Assumptions & Methodology"]

        # Platform Assumptions section should show the "no data" message
        no_data_message_found = False
        for row_idx in range(1, 10):
            cell_value = ws.cell(row_idx, 1).value
            if "No platform data found" in str(cell_value):
                no_data_message_found = True
                break
        assert no_data_message_found, "Expected 'No platform data found' message when crypto data is empty"

        # Methodology section MUST still render with header and all content
        methodology_header_found = False
        for row_idx in range(1, 50):
            cell_value = ws.cell(row_idx, 1).value
            if cell_value == "Methodology Assumptions":
                methodology_header_found = True
                assert ws.cell(row_idx, 1).font.bold is True
                break

        assert methodology_header_found, "Methodology Assumptions header must render even without crypto data"

        # Representative methodology items must be present in the rendered output
        expected_labels = {
            "Aggregation Approach",
            "Cost Basis Method: FIFO",
            "365-Day Exemption",
            "Materiality Threshold",
            "Data Sources",
        }

        found_labels = set()
        for row_idx in range(1, 100):
            cell_value = ws.cell(row_idx, 1).value
            if cell_value in expected_labels:
                found_labels.add(cell_value)

        assert found_labels == expected_labels, (
            f"Expected all methodology labels {expected_labels}, found {found_labels}"
        )

    def test_methodology_sections_render(self):
        """Methodology section renders with grouped section headers in bold."""
        wb = openpyxl.Workbook()
        entry = _make_capital_entry(platform="Kraken")
        report = _make_crypto_tax_report(capital_entries=[entry])
        write_assumptions_and_methodology_sheet(wb, capital_entries=report.capital_entries)
        ws = wb["Assumptions & Methodology"]

        # Expected 7 section headers in bold
        expected_sections = {
            "Taxable Events",
            "Holding Period & Exemptions",
            "Capital Gains Calculation",
            "Losses",
            "Tax Rates",
            "Other Gains",
            "Implementation",
        }

        found_sections = set()
        for row_idx in range(1, 200):
            cell_value = ws.cell(row_idx, 1).value
            if cell_value in expected_sections:
                found_sections.add(cell_value)
                # Verify section header is bold
                assert ws.cell(row_idx, 1).font.bold is True, f"Section '{cell_value}' must be bold"
                # Verify blank row before section (except first section)
                if row_idx > 10:  # Skip title rows
                    prev_cell = ws.cell(row_idx - 1, 1).value
                    assert prev_cell is None or prev_cell == "", f"Section '{cell_value}' must have blank row before it"

        assert found_sections == expected_sections, f"Expected sections {expected_sections}, found {found_sections}"

    def test_methodology_items_have_legal_citations(self):
        """Each methodology item includes legal citation (CIRS, AT folheto, or PIV)."""
        wb = openpyxl.Workbook()
        entry = _make_capital_entry(platform="Kraken")
        report = _make_crypto_tax_report(capital_entries=[entry])
        write_assumptions_and_methodology_sheet(wb, capital_entries=report.capital_entries)
        ws = wb["Assumptions & Methodology"]

        # Find all methodology item labels by structural properties:
        # - Has a label in column 1 (non-empty)
        # - Has a description in column 2 (non-empty, not a section header)
        # - Exclude platform data headers
        # - Start scanning from methodology section (row 11+)
        methodology_labels = []
        platform_data_headers = {
            "Platform",
            "Operator Entity",
            "Country",
            "Confidence",
            "Review Required",
            "Assumption / Verification Note",
            "Transaction Count",
        }
        for row_idx in range(1, 200):
            cell_value = ws.cell(row_idx, 1).value
            description = ws.cell(row_idx, 2).value

            if row_idx < 11:
                continue
            # Methodology items: label + description present, not a platform header
            if cell_value and description and cell_value not in platform_data_headers:
                methodology_labels.append((cell_value, row_idx))

        # Should have at least 15 methodology items across all sections
        assert len(methodology_labels) >= 15, f"Expected at least 15 methodology items, found {len(methodology_labels)}"

        # Each description must include legal citation
        legal_citation_patterns = [
            "CIRS art.",
            "AT folheto",
            "Ofício Circulado",
            "AT PIV",
            "Lei n.",
            "Implementation decision",
            "Source:",
        ]

        for label, row_idx in methodology_labels:
            description = ws.cell(row_idx, 2).value
            assert description is not None, f"Item '{label}' at row {row_idx} must have description"

            has_legal_citation = any(pattern in description for pattern in legal_citation_patterns)
            assert has_legal_citation, (
                f"Item '{label}' description must include legal citation "
                "(CIRS, AT folheto, PIV, or Implementation decision)"
            )

    def test_all_decision_points_documented(self):
        """All decision points from decision_points/2025.md are documented in methodology.

        Verifies completeness: each DP-001 through DP-011 from the decision points
        document must have a corresponding methodology item with that DP reference.
        This test ensures the methodology section stays in sync with the canonical
        decision points list.
        """
        wb = openpyxl.Workbook()
        entry = _make_capital_entry(platform="Kraken")
        report = _make_crypto_tax_report(capital_entries=[entry])
        write_assumptions_and_methodology_sheet(wb, capital_entries=report.capital_entries)
        ws = wb["Assumptions & Methodology"]

        # Expected decision points from decision_points/2025.md
        expected_decision_points = {
            "DP-001",  # Loan Repayment Exclusion
            "DP-002",  # Crypto-to-Crypto Deferral
            "DP-003",  # Cost Basis Method: FIFO
            "DP-004",  # Per-Wallet FIFO
            "DP-005",  # Liquidity Provision
            "DP-006",  # Transfer Fees
            "DP-007",  # Fee Deductibility
            "DP-008",  # Other Gains Classification
            "DP-009",  # Cashback Treatment
            "DP-010",  # Futures/Derivatives Losses
            "DP-011",  # OGR Usage for Derivatives
            "DP-012",  # Separate Derivatives Reporting
        }

        # Collect all decision point references from methodology descriptions
        found_decision_points = set()
        for row_idx in range(1, 200):
            description = ws.cell(row_idx, 2).value
            if description and isinstance(description, str):
                # Find all DP-XXX patterns in the description
                dp_matches = re.findall(r"DP-\d{3}", description)
                found_decision_points.update(dp_matches)

        # Verify all expected decision points are documented
        missing_decision_points = expected_decision_points - found_decision_points
        assert (
            not missing_decision_points
        ), f"Missing decision points in methodology: {sorted(missing_decision_points)}"

        # Verify we found exactly the expected set (no extra DPs)
        extra_decision_points = found_decision_points - expected_decision_points
        assert (
            not extra_decision_points
        ), f"Unexpected decision points in methodology: {sorted(extra_decision_points)}"

    def test_platform_assumptions_section_unchanged(self):
        """Platform Assumptions section structure unchanged after methodology refactor (Design Invariant #1)."""
        wb = openpyxl.Workbook()
        entry = _make_capital_entry(platform="Kraken")
        report = _make_crypto_tax_report(capital_entries=[entry])
        write_assumptions_and_methodology_sheet(wb, capital_entries=report.capital_entries)
        ws = wb["Assumptions & Methodology"]

        # Headers row should be at row 5 (after title at row 1, description at row 2, blank rows 3-4)
        # Or row 6 if "No platform data found" message is present (when no data)
        # With data, headers are at row 5
        headers_row = 5
        expected_headers = [
            "Platform",
            "Operator Entity",
            "Country",
            "Confidence",
            "Review Required",
            "Assumption / Verification Note",
            "Transaction Count",
        ]

        for col_idx, expected_header in enumerate(expected_headers, 1):
            assert ws.cell(headers_row, col_idx).value == expected_header, (
                f"Expected header '{expected_header}' at column {col_idx}"
            )

        # Verify column order with actual data
        assert ws.cell(6, 1).value == "Kraken"
        assert ws.cell(6, 2).value == "Kraken Entity"  # from _make_origin
        assert ws.cell(6, 3).value == "US"
        assert ws.cell(6, 4).value == "high"  # confidence value from make_operator_origin defaults
        assert ws.cell(6, 5).value == "NO"
        assert ws.cell(6, 6).value == ""
        assert ws.cell(6, 7).value == 1

        # Verify red-fill logic for review_required platforms
        bybit_entry = _make_capital_entry(
            platform="Bybit",
            platform_review_required=True,
            operator_assumption="verify region",
        )
        report2 = _make_crypto_tax_report(capital_entries=[bybit_entry])
        wb2 = openpyxl.Workbook()
        write_assumptions_and_methodology_sheet(wb2, capital_entries=report2.capital_entries)
        ws2 = wb2["Assumptions & Methodology"]

        # Bybit row should have red fill
        from tax_reporting.application.persisting.excel_utils import REVIEW_ROW_FILL
        assert ws2.cell(6, 1).fill == REVIEW_ROW_FILL
        assert ws2.cell(6, 5).value == "YES"

    def test_section_headers_bold(self):
        """Methodology section headers have bold font."""
        wb = openpyxl.Workbook()
        entry = _make_capital_entry(platform="Kraken")
        report = _make_crypto_tax_report(capital_entries=[entry])
        write_assumptions_and_methodology_sheet(wb, capital_entries=report.capital_entries)
        ws = wb["Assumptions & Methodology"]

        # Expected section headers
        expected_sections = {
            "Taxable Events",
            "Holding Period & Exemptions",
            "Capital Gains Calculation",
            "Losses",
            "Tax Rates",
            "Other Gains",
            "Implementation",
        }

        for row_idx in range(1, 200):
            cell_value = ws.cell(row_idx, 1).value
            if cell_value in expected_sections:
                assert ws.cell(row_idx, 1).font.bold is True, f"Section '{cell_value}' header must be bold"

    def test_section_spacing(self):
        """Exactly one blank row between methodology sections."""
        wb = openpyxl.Workbook()
        entry = _make_capital_entry(platform="Kraken")
        report = _make_crypto_tax_report(capital_entries=[entry])
        write_assumptions_and_methodology_sheet(wb, capital_entries=report.capital_entries)
        ws = wb["Assumptions & Methodology"]

        # Expected section headers in order
        expected_sections = [
            "Methodology Assumptions",  # First section header
            "Taxable Events",
            "Holding Period & Exemptions",
            "Capital Gains Calculation",
            "Losses",
            "Tax Rates",
            "Other Gains",
            "Implementation",
        ]

        section_rows = {}
        for row_idx in range(1, 200):
            cell_value = ws.cell(row_idx, 1).value
            if cell_value in expected_sections:
                section_rows[cell_value] = row_idx

        # Verify each section after the first has exactly one blank row before it
        prev_row = None
        for section in expected_sections:
            if section not in section_rows:
                continue
            current_row = section_rows[section]
            if prev_row is not None:
                # Check the row immediately before the section header
                blank_row = current_row - 1
                assert (
                    ws.cell(blank_row, 1).value is None or ws.cell(blank_row, 1).value == ""
                ), f"Section '{section}' should have blank row at {blank_row}, found: {ws.cell(blank_row, 1).value}"
            prev_row = current_row

    def test_methodology_includes_derivatives_legal_basis(self):
        """Methodology must cite art. 10(1)(e) for the Derivatives P&L tab and art. 10(1)(k) for the Crypto Gains tab.

        Structural identification (#96): scan rendered cells for the citation
        strings rather than asserting on absolute row positions.
        """
        wb = openpyxl.Workbook()
        entry = _make_capital_entry(platform="Kraken")
        report = _make_crypto_tax_report(capital_entries=[entry])
        write_assumptions_and_methodology_sheet(wb, capital_entries=report.capital_entries)
        ws = wb["Assumptions & Methodology"]

        derivatives_citation_seen = False
        crypto_gains_citation_seen = False
        for row_idx in range(1, 200):
            description = ws.cell(row_idx, 2).value
            if not description or not isinstance(description, str):
                continue
            if "CIRS art. 10(1)(e)" in description and "Derivatives" in description:
                derivatives_citation_seen = True
            if "CIRS art. 10(1)(k)" in description and (
                "Crypto Gains" in description or "crypto" in description.lower()
            ):
                crypto_gains_citation_seen = True

        assert derivatives_citation_seen, (
            "Methodology must include an item citing CIRS art. 10(1)(e) for the Derivatives P&L tab"
        )
        assert crypto_gains_citation_seen, (
            "Methodology must include an item citing CIRS art. 10(1)(k) for the Crypto Gains tab"
        )

    def test_legal_citation_format(self):
        """Legal citations follow expected format patterns."""
        wb = openpyxl.Workbook()
        entry = _make_capital_entry(platform="Kraken")
        report = _make_crypto_tax_report(capital_entries=[entry])
        write_assumptions_and_methodology_sheet(wb, capital_entries=report.capital_entries)
        ws = wb["Assumptions & Methodology"]

        # Expected legal citation patterns (regex)
        legal_citation_regexes = [
            r"CIRS art\. \d+[\(\)\w\s,\.]*",  # CIRS art. X, with optional paragraphs
            r"AT folheto \d{4}-\d{2}-\d{2}",  # AT folheto YYYY-MM-DD
            r"AT PIV \d+",  # AT PIV XXXXX
            r"Ofício Circulado \d+/\d{4}",  # Ofício Circulado N/YYYY
            r"Lei n\.º [\d\-/]+",  # Lei n.º N
            r"Implementation decision",  # Implementation decision
            r"Source:",  # Source: prefix
        ]

        # Collect all methodology item descriptions by structural properties:
        # - Has a label in column 1 (non-empty)
        # - Has a description in column 2 (non-empty string, not a section header)
        # - Column 3 is empty (platform data has multiple columns populated)
        methodology_descriptions = []
        for row_idx in range(1, 200):
            cell_value = ws.cell(row_idx, 1).value
            description = ws.cell(row_idx, 2).value
            column_3_value = ws.cell(row_idx, 3).value

            # Methodology items: label + description present, column 3 empty (not platform data)
            if cell_value and description and isinstance(description, str) and not column_3_value:
                methodology_descriptions.append((cell_value, description))

        # Should have multiple methodology items
        assert len(methodology_descriptions) >= 15, (
            f"Expected at least 15 methodology items, found {len(methodology_descriptions)}"
        )

        # Each description must match at least one legal citation pattern
        for label, description in methodology_descriptions:
            matches_pattern = False
            for pattern in legal_citation_regexes:
                if re.search(pattern, description):
                    matches_pattern = True
                    break
            assert matches_pattern, (
                f"Item '{label}' description must include legal citation matching one of the expected patterns. "
                f"Description: {description}"
            )
