"""Tests for the loan activity sheet writer."""

from decimal import Decimal

import openpyxl
import pytest

from tax_reporting.application.crypto_reporting import (
    CapitalGainPeriodStats,
    CryptoCapitalGainStats,
    CryptoReconciliationSummary,
    CryptoTaxReport,
    LoanActivityEntry,
)
from tax_reporting.application.persisting.loan_activity_sheet import write_loan_activity_sheet
from tax_reporting.domain.constants import (
    LOAN_STATUS_IN_ASSET_INTEREST,
    LOAN_STATUS_NO_EUR_PRICE,
    LOAN_STATUS_OPEN_LOAN,
    LOAN_STATUS_OVERPAID_VERIFY,
    LOAN_STATUS_SETTLED,
)


def _make_loan_activity_entry(**overrides: object) -> LoanActivityEntry:
    defaults = {
        "asset": "SUI",
        "received_count": 2,
        "received_amount": Decimal("300"),
        "received_value_eur": Decimal("1200"),
        "repaid_count": 1,
        "repaid_amount": Decimal("100"),
        "repaid_value_eur": Decimal("400"),
        "balance_amount": Decimal("200"),
        "balance_status": LOAN_STATUS_OPEN_LOAN,
    }
    defaults.update(overrides)
    return LoanActivityEntry(**defaults)  # type: ignore[arg-type]


def _make_crypto_tax_report(
    loan_activity: list[LoanActivityEntry] | None = None,
) -> CryptoTaxReport:
    empty_stats = CapitalGainPeriodStats(
        count=0, cost_total_eur=Decimal("0"), proceeds_total_eur=Decimal("0"), gain_loss_total_eur=Decimal("0")
    )
    grand_total = CapitalGainPeriodStats(
        count=0, cost_total_eur=Decimal("0"), proceeds_total_eur=Decimal("0"), gain_loss_total_eur=Decimal("0")
    )
    stats = CryptoCapitalGainStats(
        short_term=empty_stats, long_term=empty_stats, mixed=empty_stats, unknown=empty_stats, grand_total=grand_total
    )
    reconciliation = CryptoReconciliationSummary(
        capital_rows=0,
        reward_rows=0,
        short_term_rows=0,
        long_term_rows=0,
        mixed_rows=0,
        unknown_rows=0,
        capital_cost_total_eur=Decimal("0"),
        capital_proceeds_total_eur=Decimal("0"),
        capital_gain_total_eur=Decimal("0"),
        reward_total_eur=Decimal("0"),
        opening_holdings=None,
        closing_holdings=None,
    )
    return CryptoTaxReport(
        tax_year=2025,
        capital_entries=[],
        reward_entries=[],
        reconciliation=reconciliation,
        capital_gain_stats=stats,
        loan_activity=loan_activity if loan_activity is not None else [],
    )


def _find_header_row(ws: openpyxl.worksheet.worksheet.Worksheet) -> int:
    for r in range(1, ws.max_row + 1):
        if ws.cell(r, 1).value == "Asset":
            return r
    raise AssertionError("Header row not found")


def _is_light_red_fill(cell: openpyxl.cell.cell.Cell) -> bool:
    fill = cell.fill
    return (
        fill.start_color.rgb == "FFFFCCCC"
        and fill.end_color.rgb == "FFFFCCCC"
        and fill.fill_type == "solid"
    )


def _is_yellow_fill(cell: openpyxl.cell.cell.Cell) -> bool:
    fill = cell.fill
    return (
        fill.start_color.rgb == "FFFFFF00"
        and fill.end_color.rgb == "FFFFFF00"
        and fill.fill_type == "solid"
    )


@pytest.mark.unit
class TestLoanActivitySheetPerAssetBalance:
    """Tests that loan activity data shows per-asset balance."""

    def test_loan_activity_shows_per_asset_balance(self):
        entries = [
            _make_loan_activity_entry(
                asset="SUI",
                received_count=2,
                received_amount=Decimal("300"),
                received_value_eur=Decimal("1200"),
                repaid_count=1,
                repaid_amount=Decimal("100"),
                repaid_value_eur=Decimal("400"),
                balance_amount=Decimal("200"),
                balance_status=LOAN_STATUS_OPEN_LOAN,
            ),
            _make_loan_activity_entry(
                asset="WBTC",
                received_count=0,
                received_amount=Decimal("0"),
                received_value_eur=Decimal("0"),
                repaid_count=3,
                repaid_amount=Decimal("0.031"),
                repaid_value_eur=Decimal("2000"),
                balance_amount=Decimal("-0.031"),
                balance_status=LOAN_STATUS_OVERPAID_VERIFY,
            ),
        ]
        report = _make_crypto_tax_report(loan_activity=entries)
        wb = openpyxl.Workbook()
        write_loan_activity_sheet(wb, report)
        ws = wb["Loan Activity"]
        header_row = _find_header_row(ws)
        sui_row = header_row + 1
        wbtc_row = header_row + 2
        assert ws.cell(sui_row, 1).value == "SUI"
        assert ws.cell(sui_row, 2).value == 2
        assert ws.cell(sui_row, 3).value == Decimal("300")
        assert ws.cell(sui_row, 4).value == Decimal("1200")
        assert ws.cell(sui_row, 5).value == 1
        assert ws.cell(sui_row, 6).value == Decimal("100")
        assert ws.cell(sui_row, 7).value == Decimal("400")
        assert ws.cell(sui_row, 8).value == Decimal("200")
        assert ws.cell(sui_row, 9).value == LOAN_STATUS_OPEN_LOAN
        assert ws.cell(wbtc_row, 1).value == "WBTC"
        assert ws.cell(wbtc_row, 8).value == Decimal("-0.031")
        assert ws.cell(wbtc_row, 9).value == LOAN_STATUS_OVERPAID_VERIFY


@pytest.mark.unit
class TestLoanActivitySheetOverpaid:
    """Tests that overpaid assets get a light-red background."""

    def test_loan_activity_flags_overpaid(self):
        entries = [
            _make_loan_activity_entry(
                asset="WBTC",
                balance_amount=Decimal("-0.031"),
                balance_status=LOAN_STATUS_OVERPAID_VERIFY,
                balance_detail="overshoot 5.0000%",
            ),
        ]
        report = _make_crypto_tax_report(loan_activity=entries)
        wb = openpyxl.Workbook()
        write_loan_activity_sheet(wb, report)
        ws = wb["Loan Activity"]
        header_row = _find_header_row(ws)
        data_row = header_row + 1
        assert ws.cell(data_row, 9).value == LOAN_STATUS_OVERPAID_VERIFY
        assert ws.cell(data_row, 10).value == "overshoot 5.0000%"
        for col in range(1, 11):
            assert _is_light_red_fill(ws.cell(data_row, col)), f"Column {col} should have light-red fill for overpaid"


@pytest.mark.unit
class TestLoanActivitySheetOpenLoan:
    """Tests that open loans show the correct status."""

    def test_loan_activity_flags_open_loan(self):
        entries = [
            _make_loan_activity_entry(
                asset="SUI",
                received_count=3,
                received_amount=Decimal("500"),
                repaid_count=1,
                repaid_amount=Decimal("100"),
                balance_amount=Decimal("400"),
                balance_status=LOAN_STATUS_OPEN_LOAN,
            ),
        ]
        report = _make_crypto_tax_report(loan_activity=entries)
        wb = openpyxl.Workbook()
        write_loan_activity_sheet(wb, report)
        ws = wb["Loan Activity"]
        header_row = _find_header_row(ws)
        data_row = header_row + 1
        assert ws.cell(data_row, 9).value == LOAN_STATUS_OPEN_LOAN
        assert ws.cell(data_row, 8).value == Decimal("400")
        for col in range(1, 11):
            assert not _is_light_red_fill(ws.cell(data_row, col)), (
                f"Column {col} should NOT have light-red fill for open loans"
            )
            assert not _is_yellow_fill(ws.cell(data_row, col)), (
                f"Column {col} should NOT have yellow fill for open loans"
            )
        assert ws.cell(data_row, 10).value is None


@pytest.mark.unit
class TestLoanActivitySheetEmpty:
    """Tests that no loan transactions produces an empty sheet with header only."""

    def test_loan_activity_empty_when_no_loan_transactions(self):
        report = _make_crypto_tax_report(loan_activity=[])
        wb = openpyxl.Workbook()
        write_loan_activity_sheet(wb, report)
        ws = wb["Loan Activity"]
        header_row = _find_header_row(ws)
        next_row = header_row + 1
        assert ws.cell(next_row, 1).value is None


@pytest.mark.unit
class TestLoanActivitySheetSettled:
    """Tests that settled loans show correct status without highlighting."""

    def test_loan_activity_settled_no_highlight(self):
        entries = [
            _make_loan_activity_entry(
                asset="SUI",
                received_count=1,
                received_amount=Decimal("100"),
                received_value_eur=Decimal("500"),
                repaid_count=1,
                repaid_amount=Decimal("100"),
                repaid_value_eur=Decimal("500"),
                balance_amount=Decimal("0"),
                balance_status=LOAN_STATUS_SETTLED,
            ),
        ]
        report = _make_crypto_tax_report(loan_activity=entries)
        wb = openpyxl.Workbook()
        write_loan_activity_sheet(wb, report)
        ws = wb["Loan Activity"]
        header_row = _find_header_row(ws)
        data_row = header_row + 1
        assert ws.cell(data_row, 9).value == LOAN_STATUS_SETTLED
        assert ws.cell(data_row, 8).value == 0.0
        for col in range(1, 11):
            assert not _is_light_red_fill(ws.cell(data_row, col)), (
                f"Column {col} should NOT have light-red fill for settled loans"
            )
            assert not _is_yellow_fill(ws.cell(data_row, col)), (
                f"Column {col} should NOT have yellow fill for settled loans"
            )
        assert ws.cell(data_row, 10).value is None


@pytest.mark.unit
class TestLoanActivitySheetAlwaysCreated:
    """Loan Activity sheet is always created when crypto data is present."""

    def test_loan_activity_sheet_shows_headers_when_no_activity(self):
        """Empty loan_activity produces sheet with headers only, not a missing tab."""
        report = _make_crypto_tax_report(loan_activity=[])
        wb = openpyxl.Workbook()
        write_loan_activity_sheet(wb, report)
        assert "Loan Activity" in wb.sheetnames
        ws = wb["Loan Activity"]
        header_row = _find_header_row(ws)
        assert ws.cell(header_row, 1).value == "Asset"
        assert ws.cell(header_row + 1, 1).value is None


@pytest.mark.unit
class TestLoanActivitySheetNoEurPrice:
    """Tests that LOAN_STATUS_NO_EUR_PRICE assets get a yellow background."""

    def test_loan_activity_flags_no_eur_price_yellow(self):
        entries = [
            _make_loan_activity_entry(
                asset="LBTC",
                balance_amount=Decimal("-0.01"),
                balance_status=LOAN_STATUS_NO_EUR_PRICE,
                balance_detail=None,
            ),
        ]
        report = _make_crypto_tax_report(loan_activity=entries)
        wb = openpyxl.Workbook()
        write_loan_activity_sheet(wb, report)
        ws = wb["Loan Activity"]
        header_row = _find_header_row(ws)
        data_row = header_row + 1
        assert ws.cell(data_row, 9).value == LOAN_STATUS_NO_EUR_PRICE
        assert ws.cell(data_row, 10).value is None
        for col in range(1, 11):
            assert _is_yellow_fill(ws.cell(data_row, col)), (
                f"Column {col} should have yellow fill for no-EUR-price status"
            )


@pytest.mark.unit
class TestLoanActivitySheetInAssetInterest:
    """Tests that LOAN_STATUS_IN_ASSET_INTEREST assets get no fill and render balance_detail."""

    def test_loan_activity_in_asset_interest_no_fill_and_renders_detail(self):
        entries = [
            _make_loan_activity_entry(
                asset="WBTC",
                balance_amount=Decimal("-0.00036"),
                balance_status=LOAN_STATUS_IN_ASSET_INTEREST,
                balance_detail="overshoot 0.5853%",
            ),
        ]
        report = _make_crypto_tax_report(loan_activity=entries)
        wb = openpyxl.Workbook()
        write_loan_activity_sheet(wb, report)
        ws = wb["Loan Activity"]
        header_row = _find_header_row(ws)
        data_row = header_row + 1
        # (ii) the sentinel is rendered in column 9, NOT clobbered by balance_detail.
        assert ws.cell(data_row, 9).value == LOAN_STATUS_IN_ASSET_INTEREST
        # (iii) balance_detail is rendered in the sibling column 10.
        assert ws.cell(data_row, 10).value == "overshoot 0.5853%"
        # (i) no fill is applied across range(1, 11) for any cell.
        for col in range(1, 11):
            assert not _is_light_red_fill(ws.cell(data_row, col)), (
                f"Column {col} should NOT have light-red fill for in-asset-interest"
            )
            assert not _is_yellow_fill(ws.cell(data_row, col)), (
                f"Column {col} should NOT have yellow fill for in-asset-interest"
            )
