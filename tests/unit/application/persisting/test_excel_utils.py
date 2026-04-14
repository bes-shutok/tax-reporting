"""Tests for excel_utils: auto_column_width and safe_remove_file."""

from __future__ import annotations

from pathlib import Path

import openpyxl

from shares_reporting.application.persisting.excel_utils import auto_column_width, safe_remove_file


class TestAutoColumnWidth:
    """Test auto_column_width sets column widths based on non-formula cell content."""

    def test_text_cells_measured(self) -> None:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.cell(1, 1, "Hello")
        ws.cell(2, 1, "World!")

        auto_column_width(ws)

        assert ws.column_dimensions["A"].width == len("World!") + 2

    def test_number_cells_measured(self) -> None:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.cell(1, 1, 12345)

        auto_column_width(ws)

        assert ws.column_dimensions["A"].width == len("12345") + 2

    def test_formula_cells_skipped(self) -> None:
        """When non-formula cells are short (like headers) and formulas are present,
        the column is treated as formula-heavy and gets MIN_DATA_WIDTH."""
        from shares_reporting.application.persisting.excel_utils import MIN_DATA_WIDTH

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.cell(1, 1, "Short")
        ws.cell(2, 1, "=SUM(A1:A100)*1.23456789")

        auto_column_width(ws)

        # "Short" is a header-like label (5 chars < 10 threshold) with a formula,
        # so this is treated as formula-heavy and gets MIN_DATA_WIDTH
        assert ws.column_dimensions["A"].width == MIN_DATA_WIDTH

    def test_empty_column_no_crash(self) -> None:
        """Empty worksheet with no data cells does not crash."""
        wb = openpyxl.Workbook()
        ws = wb.active

        auto_column_width(ws)

        # openpyxl default width for new columns is 13.0; function doesn't crash on empty input
        assert ws.column_dimensions["A"].width == 13.0

    def test_padding_is_two(self) -> None:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.cell(1, 1, "abc")

        auto_column_width(ws)

        assert ws.column_dimensions["A"].width == 5

    def test_multiple_columns_independent(self) -> None:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.cell(1, 1, "Short")
        ws.cell(1, 2, "A longer header value")

        auto_column_width(ws)

        assert ws.column_dimensions["A"].width == len("Short") + 2
        assert ws.column_dimensions["B"].width == len("A longer header value") + 2

    def test_none_cells_ignored(self) -> None:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.cell(1, 1, None)
        ws.cell(2, 1, "Text")

        auto_column_width(ws)

        assert ws.column_dimensions["A"].width == len("Text") + 2

    def test_mixed_formula_and_text_uses_text_only(self) -> None:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.cell(1, 1, "X")
        ws.cell(2, 1, "=B1*C1+D1*E1+F1*G1")
        ws.cell(3, 1, "ABCDEFGHIJ")

        auto_column_width(ws)

        assert ws.column_dimensions["A"].width == len("ABCDEFGHIJ") + 2

class TestAutoColumnWidthWidthBounds:
    """Test auto_column_width applies MAX_CELL_WIDTH cap and MIN_DATA_WIDTH floor."""

    def test_very_long_text_capped_at_max_cell_width(self) -> None:
        """A cell with 150-char value is capped at MAX_CELL_WIDTH + 2, not 152."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.cell(1, 1, "X" * 150)  # 150 characters

        auto_column_width(ws)

        # MAX_CELL_WIDTH = 50, so width should be 50 + 2 = 52
        assert ws.column_dimensions["A"].width == 52

    def test_all_formula_column_gets_min_data_width(self) -> None:
        """A column with only formula cells gets MIN_DATA_WIDTH, not 2."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.cell(1, 1, "=A1+B1")
        ws.cell(2, 1, "=SUM(C1:C100)")

        auto_column_width(ws)

        # MIN_DATA_WIDTH = 12
        assert ws.column_dimensions["A"].width == 12

    def test_header_with_single_formula_gets_min_data_width(self) -> None:
        """Column with header and one formula (1:1 ratio) is treated as formula-heavy.

        This covers the edge case of a single-row IB report where formula_count equals
        the number of non-formula cells. The column should still get MIN_DATA_WIDTH.
        """
        from shares_reporting.application.persisting.excel_utils import MIN_DATA_WIDTH

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.cell(1, 1, "Amt")  # Short header (3 chars)
        ws.cell(2, 1, "=B2*C2")  # Single formula cell

        auto_column_width(ws)

        # With 1:1 formula:non-formula ratio, this is formula-heavy
        # Should get MIN_DATA_WIDTH, not len("Amt") + 2 = 5
        assert ws.column_dimensions["A"].width == MIN_DATA_WIDTH

    def test_mixed_data_and_formula_uses_data_only(self) -> None:
        """Column with both data and formula cells measures data, skips formulas, applies cap."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.cell(1, 1, "Header")
        ws.cell(2, 1, "=A2*B2*1.23")
        ws.cell(3, 1, "ABCDEFGHIJK")  # 11 chars
        ws.cell(4, 1, "=C4*2")

        auto_column_width(ws)

        # Should measure "Header" (6) and "ABCDEFGHIJK" (11), cap at 50, then +2 = 13
        assert ws.column_dimensions["A"].width == 13

    def test_short_data_cells_still_get_padding(self) -> None:
        """Normal short data cells still get len(value) + 2."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.cell(1, 1, "Hi")

        auto_column_width(ws)

        assert ws.column_dimensions["A"].width == 4  # "Hi" is 2 chars + 2 padding


class TestSafeRemoveFile:
    """Test safe_remove_file removes existing files and handles missing files gracefully."""

    def test_removes_existing_file(self, tmp_path: Path) -> None:
        target = tmp_path / "test.xlsx"
        target.write_text("data")

        safe_remove_file(target)

        assert not target.exists()

    def test_handles_missing_file(self, tmp_path: Path) -> None:
        target = tmp_path / "nonexistent.xlsx"

        safe_remove_file(target)

    def test_accepts_string_path(self, tmp_path: Path) -> None:
        target = tmp_path / "test.xlsx"
        target.write_text("data")

        safe_remove_file(str(target))

        assert not target.exists()
