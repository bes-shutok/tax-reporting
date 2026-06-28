"""Tests for the Derivatives P&L sheet writer (CIRS art. 10(1)(e))."""

from decimal import Decimal

import openpyxl
import pytest

from tax_reporting.application.crypto.entities import DerivativesEventType, DerivativesPnLEntry
from tax_reporting.application.crypto_reporting import (
    CapitalGainPeriodStats,
    CryptoCapitalGainStats,
    CryptoReconciliationSummary,
    CryptoTaxReport,
)
from tax_reporting.application.persisting.derivatives_sheet import write_derivatives_sheet
from tests.conftest import build_koinly_jurisdiction

_SHEET_NAME = "Derivatives P&L"
_HEADER_TITLE = "DERIVATIVES P&L (Financial Derivatives: CIRS art. 10(1)(e))"
_COLUMN_HEADERS = [
    "Date",
    "Asset",
    "Platform",
    "Event Type",
    "P&L (EUR)",
    "Operator entity",
    "Operator country",
    "Event count",
    "Notes",
    "Review",
    "Annex",
    "Código",
]
_NUM_COLUMNS = len(_COLUMN_HEADERS)
_EMPTY_STATE_MESSAGE = "No derivatives activity for this jurisdiction"
_LOSS_FOOTNOTE = "Losses are deductible against other Category G gains; carry-forward 5 years per PT-C-016"


def _make_derivatives_entry(**overrides: object) -> DerivativesPnLEntry:
    defaults: dict[str, object] = {
        "date": "2025-01-12",
        "asset": "USDT",
        "platform": "ByBit",
        "pnl_eur": Decimal("140.18"),
        "event_type": DerivativesEventType.PROFIT,
        "source_ref": "OGR:2025-01-12:USDT",
        "legal_category": "CIRS art. 10(1)(e)",
        "review_required": False,
        "review_reason": "",
    }
    defaults.update(overrides)
    return DerivativesPnLEntry(**defaults)  # type: ignore[arg-type]


def _make_crypto_tax_report(derivatives_entries: list[DerivativesPnLEntry] | None = None) -> CryptoTaxReport:
    empty_period = CapitalGainPeriodStats(
        count=0, cost_total_eur=Decimal("0"), proceeds_total_eur=Decimal("0"), gain_loss_total_eur=Decimal("0")
    )
    stats = CryptoCapitalGainStats(
        short_term=empty_period,
        long_term=empty_period,
        mixed=empty_period,
        unknown=empty_period,
        grand_total=empty_period,
    )
    reconciliation = CryptoReconciliationSummary(
        capital_rows=0,
        reward_rows=0,
        short_term_rows=0,
        long_term_rows=0,
        mixed_rows=0,
        unknown_rows=0,
        capital_cost_total_eur=Decimal("0"),
        capital_proceeds_total_eur=Decimal("0"),
        capital_gain_total_eur=Decimal("0"),
        reward_total_eur=Decimal("0"),
        opening_holdings=None,
        closing_holdings=None,
    )
    return CryptoTaxReport(
        tax_year=2025,
        capital_entries=[],
        reward_entries=[],
        reconciliation=reconciliation,
        capital_gain_stats=stats,
        derivatives_entries=derivatives_entries or [],
    )


def _find_row_with_value(ws: openpyxl.worksheet.worksheet.Worksheet, value: str, max_row: int = 50) -> int:
    """Find the first row whose first column equals value. Returns 1-based row index."""
    for r in range(1, min(ws.max_row, max_row) + 1):
        if ws.cell(r, 1).value == value:
            return r
    raise AssertionError(f"Row with value '{value}' not found in column 1")


def _find_row_containing(ws: openpyxl.worksheet.worksheet.Worksheet, substring: str, max_row: int = 60) -> int:
    """Find the first row where any cell value contains substring. Returns 1-based row index."""
    for r in range(1, min(ws.max_row, max_row) + 1):
        for c in range(1, ws.max_column + 1):
            cell_val = ws.cell(r, c).value
            if cell_val is not None and substring in str(cell_val):
                return r
    raise AssertionError(f"Row containing '{substring}' not found")


def _find_row_with_platform(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    platform: str,
    header_row: int,
    max_row: int = 60,
) -> int:
    """Find the data row whose Platform cell (column 3) equals platform.

    Searches data rows starting at header_row + 1 (skipping the header). Returns the
    1-based row index. Used to locate a specific entry's row by its distinguishing
    Platform value so per-row routing cells can be read from THAT row.
    """
    for r in range(header_row + 1, min(ws.max_row, max_row) + 1):
        if ws.cell(r, 3).value == platform:
            return r
    raise AssertionError(f"Data row with Platform '{platform}' not found in column 3")


@pytest.mark.unit
class TestDerivativesSheet:
    """Behavioral coverage for the Derivatives P&L sheet."""

    def test_renders_header_with_legal_basis(self):
        report = _make_crypto_tax_report(derivatives_entries=[_make_derivatives_entry()])
        wb = openpyxl.Workbook()
        write_derivatives_sheet(wb, report, build_koinly_jurisdiction())

        assert _SHEET_NAME in wb.sheetnames
        ws = wb[_SHEET_NAME]
        assert ws.cell(1, 1).value == _HEADER_TITLE

    def test_renders_one_row_per_aggregated_entry(self):
        entries = [
            _make_derivatives_entry(date="2025-01-12", asset="USDT", platform="ByBit"),
            _make_derivatives_entry(date="2025-01-13", asset="USDT", platform="ByBit"),
        ]
        report = _make_crypto_tax_report(derivatives_entries=entries)
        wb = openpyxl.Workbook()
        write_derivatives_sheet(wb, report, build_koinly_jurisdiction())

        ws = wb[_SHEET_NAME]

        header_row = _find_row_with_value(ws, "Date")

        for idx, expected_header in enumerate(_COLUMN_HEADERS, start=1):
            assert ws.cell(header_row, idx).value == expected_header, (
                f"Column {idx} header: expected '{expected_header}', got '{ws.cell(header_row, idx).value}'"
            )

        data_row_1 = header_row + 1
        data_row_2 = header_row + 2

        assert ws.cell(data_row_1, 1).value == "2025-01-12"
        assert ws.cell(data_row_1, 2).value == "USDT"
        assert ws.cell(data_row_1, 3).value == "ByBit"
        assert ws.cell(data_row_1, 4).value == "profit"

        assert ws.cell(data_row_2, 1).value == "2025-01-13"
        assert ws.cell(data_row_2, 2).value == "USDT"
        assert ws.cell(data_row_2, 3).value == "ByBit"

    def test_totals_row(self):
        entries = [
            _make_derivatives_entry(date="2025-01-12", pnl_eur=Decimal("140.18")),
            _make_derivatives_entry(date="2025-01-13", pnl_eur=Decimal("-4.17"), event_type=DerivativesEventType.LOSS),
        ]
        report = _make_crypto_tax_report(derivatives_entries=entries)
        wb = openpyxl.Workbook()
        write_derivatives_sheet(wb, report, build_koinly_jurisdiction())

        ws = wb[_SHEET_NAME]
        totals_row = _find_row_containing(ws, "Total")
        pnl_cell_val = ws.cell(totals_row, 5).value
        assert pnl_cell_val is not None
        assert Decimal(str(pnl_cell_val)) == Decimal("136.01"), (
            f"Expected totals row net P&L = 136.01, got {pnl_cell_val}"
        )

    def test_empty_state_when_no_entries(self):
        report = _make_crypto_tax_report(derivatives_entries=[])
        wb = openpyxl.Workbook()
        write_derivatives_sheet(wb, report, build_koinly_jurisdiction())

        ws = wb[_SHEET_NAME]
        header_row = _find_row_with_value(ws, "Date")
        for idx, expected_header in enumerate(_COLUMN_HEADERS, start=1):
            assert ws.cell(header_row, idx).value == expected_header

        empty_row = _find_row_containing(ws, _EMPTY_STATE_MESSAGE)
        assert empty_row > header_row, "Empty-state row should appear after column headers"

    def test_review_reason_renders_as_yes_with_reason(self):
        entries = [
            _make_derivatives_entry(
                review_required=True,
                review_reason="Ambiguous OGR/CG value mismatch",
            ),
        ]
        report = _make_crypto_tax_report(derivatives_entries=entries)
        wb = openpyxl.Workbook()
        write_derivatives_sheet(wb, report, build_koinly_jurisdiction())

        ws = wb[_SHEET_NAME]
        header_row = _find_row_with_value(ws, "Date")
        data_row = header_row + 1
        # Review lives in column 10 (Annex=11, Código=12 are the last two columns)
        review_cell = ws.cell(data_row, 10).value
        assert review_cell == "YES: Ambiguous OGR/CG value mismatch", (
            f"Review cell must render 'YES: <reason>', got '{review_cell}'"
        )

    def test_loss_deductibility_footnote(self):
        entries = [
            _make_derivatives_entry(date="2025-01-13", pnl_eur=Decimal("-4.17"), event_type=DerivativesEventType.LOSS),
        ]
        report = _make_crypto_tax_report(derivatives_entries=entries)
        wb = openpyxl.Workbook()
        write_derivatives_sheet(wb, report, build_koinly_jurisdiction())

        ws = wb[_SHEET_NAME]
        footnote_row = _find_row_containing(ws, "Losses are deductible")
        cell_text = ""
        for c in range(1, ws.max_column + 1):
            v = ws.cell(footnote_row, c).value
            if v is not None and "Losses are deductible" in str(v):
                cell_text = str(v)
                break
        assert cell_text == _LOSS_FOOTNOTE, f"Footnote text must match exactly, got '{cell_text}'"

    def test_loss_deductibility_footnote_absent_when_no_loss(self):
        entries = [
            _make_derivatives_entry(
                date="2025-01-12",
                pnl_eur=Decimal("140.18"),
                event_type=DerivativesEventType.PROFIT,
            ),
        ]
        report = _make_crypto_tax_report(derivatives_entries=entries)
        wb = openpyxl.Workbook()
        write_derivatives_sheet(wb, report, build_koinly_jurisdiction())

        ws = wb[_SHEET_NAME]
        found = False
        for r in range(1, ws.max_row + 1):
            for c in range(1, ws.max_column + 1):
                v = ws.cell(r, c).value
                if v is not None and "Losses are deductible" in str(v):
                    found = True
                    break
            if found:
                break
        assert not found, "Footnote should NOT appear when there are no loss entries"

    def test_review_row_gets_red_fill(self):
        entries = [
            _make_derivatives_entry(review_required=True, review_reason="Ambiguous OGR/CG value mismatch"),
        ]
        report = _make_crypto_tax_report(derivatives_entries=entries)
        wb = openpyxl.Workbook()
        write_derivatives_sheet(wb, report, build_koinly_jurisdiction())

        ws = wb[_SHEET_NAME]
        header_row = _find_row_with_value(ws, "Date")
        data_row = header_row + 1
        for col in range(1, len(_COLUMN_HEADERS) + 1):
            cell = ws.cell(data_row, col)
            assert cell.fill.patternType == "solid", f"Col {col} should have solid fill, got {cell.fill.patternType}"
            assert cell.fill.fgColor.rgb == "FFFF0000", (
                f"Col {col} should have red fill (FFFF0000), got {cell.fill.fgColor.rgb}"
            )

    # --- 12-column layout with per-row routing columns (Annex=11, Código=12) ---

    def test_header_has_twelve_columns(self):
        report = _make_crypto_tax_report(derivatives_entries=[_make_derivatives_entry()])
        wb = openpyxl.Workbook()
        write_derivatives_sheet(wb, report, build_koinly_jurisdiction())

        ws = wb[_SHEET_NAME]
        header_row = 3
        populated = [
            (c, ws.cell(header_row, c).value)
            for c in range(1, 50)
            if ws.cell(header_row, c).value is not None
        ]
        assert len(populated) == 12, (
            f"Header row should have exactly 12 populated cells, got {len(populated)}: {populated}"
        )
        last_col, last_value = populated[-1]
        assert last_col == 12, f"Last header cell should be at column 12, got column {last_col}"
        assert last_value == "Código", f"Last header cell should be 'Código', got '{last_value}'"

    def test_header_columns_include_operator_event_count_notes(self):
        report = _make_crypto_tax_report(derivatives_entries=[_make_derivatives_entry()])
        wb = openpyxl.Workbook()
        write_derivatives_sheet(wb, report, build_koinly_jurisdiction())

        ws = wb[_SHEET_NAME]
        header_row = _find_row_with_value(ws, "Date")
        header_values = {
            ws.cell(header_row, c).value
            for c in range(1, 50)
            if ws.cell(header_row, c).value is not None
        }
        for required in [
            "Operator entity",
            "Operator country",
            "Event count",
            "Notes",
            "Annex",
            "Código",
        ]:
            assert required in header_values, (
                f"Header row missing required column '{required}'. Headers seen: {sorted(map(str, header_values))}"
            )

    # --- Per-row routing columns (Annex at col 11, Código at col 12) ---
    # The row-2 `_DETAIL_LINE_TEMPLATE` detail line was removed (A4) in favor of
    # per-row routing columns. The detail-line tests below were deleted and replaced
    # by `test_no_single_detail_line_route` plus per-row routing coverage.

    def test_each_row_carries_own_route(self):
        """Each data row carries its OWN Annex (col 11) and Código (col 12), not a
        single detail-line value shared across rows.

        Given two PT-jurisdiction entries with DISTINGUISHING platforms (resident row
        PT-Broker carrying G/Q13 + G51; non-resident row Kraken carrying J/Q9.2.B +
        G30), locate each data row by its Platform cell (column 3) and read THAT
        row's column-11 (Annex) and column-12 (Código) cells. Each row must carry its
        own route, not the other row's values. This defeats column mislabel, field
        swap, and cross-row contamination (set-membership and generic "each row has
        cells" assertions do not).
        """
        entries = [
            _make_derivatives_entry(
                platform="PT-Broker",
                operator_country="PT",
                annex_hint="G/Q13",
                operation_code="G51",
            ),
            _make_derivatives_entry(
                platform="Kraken",
                operator_country="US",
                annex_hint="J/Q9.2.B",
                operation_code="G30",
            ),
        ]
        report = _make_crypto_tax_report(derivatives_entries=entries)
        wb = openpyxl.Workbook()
        write_derivatives_sheet(wb, report, build_koinly_jurisdiction())

        ws = wb[_SHEET_NAME]
        header_row = _find_row_with_value(ws, "Date")
        assert ws.cell(header_row, 11).value == "Annex", (
            f"Header column 11 must be 'Annex', got '{ws.cell(header_row, 11).value}'"
        )
        assert ws.cell(header_row, 12).value == "Código", (
            f"Header column 12 must be 'Código', got '{ws.cell(header_row, 12).value}'"
        )

        # Locate each data row by its Platform cell (column 3), then read THAT row's
        # column-11 (Annex) and column-12 (Código) cells.
        pt_broker_row = _find_row_with_platform(ws, "PT-Broker", header_row)
        kraken_row = _find_row_with_platform(ws, "Kraken", header_row)

        assert ws.cell(pt_broker_row, 11).value == "G/Q13", (
            f"PT-Broker row Annex (col 11) must be 'G/Q13', got '{ws.cell(pt_broker_row, 11).value}'"
        )
        assert ws.cell(pt_broker_row, 12).value == "G51", (
            f"PT-Broker row Código (col 12) must be 'G51', got '{ws.cell(pt_broker_row, 12).value}'"
        )
        assert ws.cell(kraken_row, 11).value == "J/Q9.2.B", (
            f"Kraken row Annex (col 11) must be 'J/Q9.2.B', got '{ws.cell(kraken_row, 11).value}'"
        )
        assert ws.cell(kraken_row, 12).value == "G30", (
            f"Kraken row Código (col 12) must be 'G30', got '{ws.cell(kraken_row, 12).value}'"
        )

        # The legacy `_DETAIL_LINE_TEMPLATE` detail line is GONE so this test cannot
        # be satisfied by reading the detail-line string.
        detail_line_present = any(
            isinstance(ws.cell(r, 1).value, str)
            and "Annex:" in ws.cell(r, 1).value
            and "Código:" in ws.cell(r, 1).value
            for r in range(1, min(ws.max_row, 50) + 1)
        )
        assert not detail_line_present, (
            "The legacy _DETAIL_LINE_TEMPLATE detail line must be gone; found a row "
            "containing 'Annex:' and 'Código:' in column 1."
        )

    def test_mixed_residency_renders_both_routes(self):
        """A mixed-residency set renders BOTH G51 and G30 in the per-row Código column.

        Supplementary to `test_each_row_carries_own_route`: confirms the sheet is not
        collapsing both routes to a single value.
        """
        entries = [
            _make_derivatives_entry(
                platform="PT-Broker",
                operator_country="PT",
                annex_hint="G/Q13",
                operation_code="G51",
            ),
            _make_derivatives_entry(
                platform="Kraken",
                operator_country="US",
                annex_hint="J/Q9.2.B",
                operation_code="G30",
            ),
        ]
        report = _make_crypto_tax_report(derivatives_entries=entries)
        wb = openpyxl.Workbook()
        write_derivatives_sheet(wb, report, build_koinly_jurisdiction())

        ws = wb[_SHEET_NAME]
        header_row = _find_row_with_value(ws, "Date")
        codigos = {
            ws.cell(r, 12).value
            for r in range(header_row + 1, ws.max_row + 1)
            if ws.cell(r, 3).value in {"PT-Broker", "Kraken"}
        }
        assert "G51" in codigos, f"Per-row Código column must contain 'G51', got {codigos}"
        assert "G30" in codigos, f"Per-row Código column must contain 'G30', got {codigos}"

    def test_no_single_detail_line_route(self):
        """The old `entries[0]`-derived `_DETAIL_LINE_TEMPLATE` detail line is gone.

        Per P0, the single detail-line mechanism is removed in favor of per-row
        routing columns. No row in column 1 may carry the legacy
        'Annex: ... | Código: ... | Legal basis: ...' string.
        """
        entries = [
            _make_derivatives_entry(
                platform="PT-Broker",
                operator_country="PT",
                annex_hint="G/Q13",
                operation_code="G51",
                legal_category="CIRS art. 10(1)(e)",
            ),
        ]
        report = _make_crypto_tax_report(derivatives_entries=entries)
        wb = openpyxl.Workbook()
        write_derivatives_sheet(wb, report, build_koinly_jurisdiction())

        ws = wb[_SHEET_NAME]
        for r in range(1, min(ws.max_row, 50) + 1):
            cell_val = ws.cell(r, 1).value
            if isinstance(cell_val, str):
                assert "Legal basis:" not in cell_val, (
                    f"Legacy detail-line string must be gone; row {r} column 1 = {cell_val!r}"
                )

    def test_blank_annex_under_pt_warns(self, caplog):
        """A blank Annex under the PT jurisdiction must surface a warning.

        Pins the dev_lessons #77/#118 'surface invalidity loudly' guarantee that the
        removed detail-line guard previously provided. A row that failed to resolve a
        route (annex_hint == '') under PT must never render silently.
        """
        entries = [
            _make_derivatives_entry(
                platform="PT-Broker",
                operator_country="PT",
                annex_hint="",
                operation_code="",
            ),
        ]
        report = _make_crypto_tax_report(derivatives_entries=entries)

        with caplog.at_level("WARNING", logger="tax_reporting.application.persisting.derivatives_sheet"):
            wb = openpyxl.Workbook()
            write_derivatives_sheet(wb, report, build_koinly_jurisdiction())

        joined = " | ".join(rec.message for rec in caplog.records)
        assert "blank" in joined.lower(), (
            f"Expected a warning mentioning 'blank' for a blank Annex under PT, got: {joined!r}"
        )
        assert "annex" in joined.lower(), (
            f"Expected a warning mentioning 'Annex' for a blank Annex under PT, got: {joined!r}"
        )

    def test_no_blank_annex_warning_when_routes_resolved(self, caplog):
        """Paired negative: when ALL PT entries resolved non-blank annexes, NO warning.

        Forces the blank-Annex guard to be gated behind the real condition (PT + blank
        annex), defeating a trivial unconditional `logger.warning(...)` implementation.
        Mirrors the existing positive/negative convention.
        """
        entries = [
            _make_derivatives_entry(
                platform="PT-Broker",
                operator_country="PT",
                annex_hint="G/Q13",
                operation_code="G51",
            ),
            _make_derivatives_entry(
                platform="Kraken",
                operator_country="US",
                annex_hint="J/Q9.2.B",
                operation_code="G30",
            ),
        ]
        report = _make_crypto_tax_report(derivatives_entries=entries)

        with caplog.at_level("WARNING", logger="tax_reporting.application.persisting.derivatives_sheet"):
            wb = openpyxl.Workbook()
            write_derivatives_sheet(wb, report, build_koinly_jurisdiction())

        blank_annex_warnings = [
            rec for rec in caplog.records
            if "blank" in rec.message.lower() and "annex" in rec.message.lower()
        ]
        assert not blank_annex_warnings, (
            f"Expected no blank-Annex warning when all routes resolved, "
            f"got: {[r.message for r in blank_annex_warnings]}"
        )

    def test_no_blank_annex_warning_under_non_pt(self, caplog):
        """Paired negative: a blank Annex under a NON-PT jurisdiction must NOT warn.

        Non-PT jurisdictions legitimately carry blank annexes (no Modelo 3 hint),
        so the warning is PT-gated. This falsifies the country gate directly: a
        regression that drops the ``jurisdiction.country == PT`` condition (making
        the warning unconditional) would log spurious warnings on every non-PT
        run and fail this test.
        """
        entries = [
            _make_derivatives_entry(
                platform="PT-Broker",
                operator_country="PT",
                annex_hint="",
                operation_code="",
            ),
        ]
        report = _make_crypto_tax_report(derivatives_entries=entries)

        with caplog.at_level("WARNING", logger="tax_reporting.application.persisting.derivatives_sheet"):
            wb = openpyxl.Workbook()
            write_derivatives_sheet(wb, report, build_koinly_jurisdiction(country="DE"))

        blank_annex_warnings = [
            rec for rec in caplog.records
            if "blank" in rec.message.lower() and "annex" in rec.message.lower()
        ]
        assert not blank_annex_warnings, (
            f"Expected no blank-Annex warning under a non-PT jurisdiction, "
            f"got: {[r.message for r in blank_annex_warnings]}"
        )

    def test_row_writes_operator_entity_and_country_in_columns_6_and_7(self):
        entries = [_make_derivatives_entry(operator_entity="ByBit", operator_country="AE")]
        report = _make_crypto_tax_report(derivatives_entries=entries)
        wb = openpyxl.Workbook()
        write_derivatives_sheet(wb, report, build_koinly_jurisdiction())

        ws = wb[_SHEET_NAME]
        header_row = _find_row_with_value(ws, "Date")
        data_row = header_row + 1
        assert ws.cell(data_row, 6).value == "ByBit", (
            f"Column 6 should contain operator_entity 'ByBit', got '{ws.cell(data_row, 6).value}'"
        )
        assert ws.cell(data_row, 7).value == "AE", (
            f"Column 7 should contain operator_country 'AE', got '{ws.cell(data_row, 7).value}'"
        )

    def test_row_writes_event_count_in_column_8(self):
        entries = [_make_derivatives_entry(event_count=3)]
        report = _make_crypto_tax_report(derivatives_entries=entries)
        wb = openpyxl.Workbook()
        write_derivatives_sheet(wb, report, build_koinly_jurisdiction())

        ws = wb[_SHEET_NAME]
        header_row = _find_row_with_value(ws, "Date")
        data_row = header_row + 1
        assert ws.cell(data_row, 8).value == 3, (
            f"Column 8 should contain event_count=3, got '{ws.cell(data_row, 8).value}'"
        )

    def test_row_writes_notes_in_column_9_when_set(self):
        entries = [_make_derivatives_entry(notes="manual annotation")]
        report = _make_crypto_tax_report(derivatives_entries=entries)
        wb = openpyxl.Workbook()
        write_derivatives_sheet(wb, report, build_koinly_jurisdiction())

        ws = wb[_SHEET_NAME]
        header_row = _find_row_with_value(ws, "Date")
        data_row = header_row + 1
        assert ws.cell(data_row, 9).value == "manual annotation", (
            f"Column 9 should contain notes 'manual annotation', got '{ws.cell(data_row, 9).value}'"
        )

    def test_row_writes_notes_in_column_9_default_empty(self):
        report = _make_crypto_tax_report(derivatives_entries=[_make_derivatives_entry()])
        wb = openpyxl.Workbook()
        write_derivatives_sheet(wb, report, build_koinly_jurisdiction())

        ws = wb[_SHEET_NAME]
        header_row = _find_row_with_value(ws, "Date")
        data_row = header_row + 1
        val = ws.cell(data_row, 9).value
        assert val is None or val == "", (
            f"Column 9 should be empty when notes is unset; got '{val!r}'"
        )

    def test_row_writes_review_in_column_10(self):
        entries = [
            _make_derivatives_entry(
                review_required=True,
                review_reason="missing platform mapping",
            ),
        ]
        report = _make_crypto_tax_report(derivatives_entries=entries)
        wb = openpyxl.Workbook()
        write_derivatives_sheet(wb, report, build_koinly_jurisdiction())

        ws = wb[_SHEET_NAME]
        header_row = _find_row_with_value(ws, "Date")
        data_row = header_row + 1
        assert ws.cell(data_row, 10).value == "YES: missing platform mapping", (
            f"Column 10 should render 'YES: <reason>', got '{ws.cell(data_row, 10).value}'"
        )

    def test_review_fill_spans_all_columns(self):
        entries = [
            _make_derivatives_entry(review_required=True, review_reason="missing platform mapping"),
        ]
        report = _make_crypto_tax_report(derivatives_entries=entries)
        wb = openpyxl.Workbook()
        write_derivatives_sheet(wb, report, build_koinly_jurisdiction())

        ws = wb[_SHEET_NAME]
        header_row = _find_row_with_value(ws, "Date")
        data_row = header_row + 1
        for col in range(1, _NUM_COLUMNS + 1):
            cell = ws.cell(data_row, col)
            assert cell.fill.patternType == "solid", (
                f"Col {col} should have solid fill, got {cell.fill.patternType}"
            )
            assert cell.fill.fgColor.rgb == "FFFF0000", (
                f"Col {col} should have red fill (FFFF0000), got {cell.fill.fgColor.rgb}"
            )

    def test_total_row_pnl_in_column_5(self):
        entries = [
            _make_derivatives_entry(date="2025-01-12", pnl_eur=Decimal("140.18")),
            _make_derivatives_entry(date="2025-01-13", pnl_eur=Decimal("60.00")),
        ]
        report = _make_crypto_tax_report(derivatives_entries=entries)
        wb = openpyxl.Workbook()
        write_derivatives_sheet(wb, report, build_koinly_jurisdiction())

        ws = wb[_SHEET_NAME]
        total_row = _find_row_with_value(ws, "Total")
        # Total row label stays at column 1; total P&L value stays at column 5 (unchanged)
        assert ws.cell(total_row, 1).value == "Total", (
            f"Total row label should be in column 1, got '{ws.cell(total_row, 1).value}'"
        )
        pnl_value = ws.cell(total_row, 5).value
        assert pnl_value is not None, "Total row should have a numeric value in column 5"
        assert Decimal(str(pnl_value)) == Decimal("200.18"), (
            f"Total row P&L should be 200.18 in column 5, got {pnl_value}"
        )

    def test_loss_footnote_still_rendered_when_loss_present(self):
        entries = [
            _make_derivatives_entry(
                date="2025-01-13",
                pnl_eur=Decimal("-271.79"),
                event_type=DerivativesEventType.LOSS,
            ),
        ]
        report = _make_crypto_tax_report(derivatives_entries=entries)
        wb = openpyxl.Workbook()
        write_derivatives_sheet(wb, report, build_koinly_jurisdiction())

        ws = wb[_SHEET_NAME]
        total_row = _find_row_with_value(ws, "Total")
        # Footnote must appear AFTER the total row
        footnote_row = _find_row_containing(ws, "Losses are deductible")
        assert footnote_row > total_row, (
            f"Loss footnote (row {footnote_row}) should appear after total row (row {total_row})"
        )
