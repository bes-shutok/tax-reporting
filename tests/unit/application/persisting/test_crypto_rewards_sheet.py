"""Tests for the crypto rewards sheet writer."""

from decimal import Decimal

import openpyxl
import pytest

from shares_reporting.application.crypto_reporting import (
    AggregatedRewardIncomeEntry,
    CapitalGainPeriodStats,
    CryptoCapitalGainStats,
    CryptoReconciliationSummary,
    CryptoRewardIncomeEntry,
    CryptoTaxReport,
    RewardTaxClassification,
)
from shares_reporting.application.persisting.crypto_rewards_sheet import write_crypto_rewards_sheet
from tests.conftest import make_operator_origin


def _make_reward_entry(
    classification: RewardTaxClassification = RewardTaxClassification.TAXABLE_NOW,
    **overrides: object,
) -> CryptoRewardIncomeEntry:
    defaults = {
        "date": "2025-03-15",
        "asset": "ETH",
        "amount": Decimal("0.5"),
        "value_eur": Decimal("1500"),
        "income_label": "Staking Reward",
        "source_type": "staking",
        "wallet": "Kraken",
        "platform": "Kraken",
        "chain": "Ethereum",
        "operator_origin": make_operator_origin(),
        "annex_hint": "J",
        "review_required": False,
        "description": "Staking reward payout",
        "tax_classification": classification,
        "foreign_tax_eur": Decimal("0"),
    }
    defaults.update(overrides)
    return CryptoRewardIncomeEntry(**defaults)  # type: ignore[arg-type]


def _make_aggregated_reward(**overrides: object) -> AggregatedRewardIncomeEntry:
    defaults = {
        "income_code": "401",
        "source_country": "US",
        "gross_income_eur": Decimal("1500"),
        "foreign_tax_eur": Decimal("0"),
        "raw_row_count": 1,
        "chains": ("Ethereum",),
        "description": "Staking income",
    }
    defaults.update(overrides)
    return AggregatedRewardIncomeEntry(**defaults)  # type: ignore[arg-type]


def _make_crypto_tax_report(
    reward_entries: list[CryptoRewardIncomeEntry] | None = None,
) -> CryptoTaxReport:
    entries = reward_entries if reward_entries is not None else [_make_reward_entry()]
    reconciliation = CryptoReconciliationSummary(
        capital_rows=0,
        reward_rows=len(entries),
        short_term_rows=0,
        long_term_rows=0,
        mixed_rows=0,
        unknown_rows=0,
        capital_cost_total_eur=Decimal("0"),
        capital_proceeds_total_eur=Decimal("0"),
        capital_gain_total_eur=Decimal("0"),
        reward_total_eur=sum((e.value_eur for e in entries), start=Decimal("0")),
        opening_holdings=None,
        closing_holdings=None,
    )
    empty_stats = CapitalGainPeriodStats(
        count=0, cost_total_eur=Decimal("0"), proceeds_total_eur=Decimal("0"), gain_loss_total_eur=Decimal("0")
    )
    capital_gain_stats = CryptoCapitalGainStats(
        short_term=empty_stats, long_term=empty_stats, mixed=empty_stats, unknown=empty_stats, grand_total=empty_stats
    )
    return CryptoTaxReport(
        tax_year=2025,
        capital_entries=[],
        reward_entries=entries,
        reconciliation=reconciliation,
        capital_gain_stats=capital_gain_stats,
        pdf_summary=None,
    )


@pytest.mark.unit
class TestCryptoRewardsSheetName:
    """Tests that the sheet is created with the correct name."""

    def test_sheet_named_crypto_rewards(self):
        wb = openpyxl.Workbook()
        report = _make_crypto_tax_report()
        aggregated = [_make_aggregated_reward()]
        write_crypto_rewards_sheet(wb, report, aggregated)
        assert "Crypto Rewards" in wb.sheetnames


@pytest.mark.unit
class TestCryptoRewardsSheetIRSFilingSummary:
    """Tests for section 2: IRS-ready filing summary."""

    IRS_SUMMARY_HEADERS = [
        "Income code",
        "Source country",
        "Reward chain",
        "Gross income (EUR)",
        "Foreign tax (EUR)",
        "Net income (EUR)",
        "Raw rows",
    ]

    def test_section_2_title_written(self):
        wb = openpyxl.Workbook()
        report = _make_crypto_tax_report()
        aggregated = [_make_aggregated_reward()]
        write_crypto_rewards_sheet(wb, report, aggregated)
        ws = wb["Crypto Rewards"]
        found = False
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=1):
            if row[0].value == "2. REWARDS INCOME - IRS-READY FILING SUMMARY":
                found = True
                break
        assert found

    def test_section_2_title_is_bold(self):
        wb = openpyxl.Workbook()
        report = _make_crypto_tax_report()
        aggregated = [_make_aggregated_reward()]
        write_crypto_rewards_sheet(wb, report, aggregated)
        ws = wb["Crypto Rewards"]
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=1):
            if row[0].value == "2. REWARDS INCOME - IRS-READY FILING SUMMARY":
                assert row[0].font.bold is True
                break

    def test_summary_note_is_italic(self):
        wb = openpyxl.Workbook()
        report = _make_crypto_tax_report()
        aggregated = [_make_aggregated_reward()]
        write_crypto_rewards_sheet(wb, report, aggregated)
        ws = wb["Crypto Rewards"]
        found = False
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=1):
            if row[0].value and "taxable immediately" in str(row[0].value):
                assert row[0].font.italic is True
                found = True
                break
        assert found, "Expected italic note about taxable rewards"

    def test_aggregated_headers_written(self):
        wb = openpyxl.Workbook()
        report = _make_crypto_tax_report()
        aggregated = [_make_aggregated_reward()]
        write_crypto_rewards_sheet(wb, report, aggregated)
        ws = wb["Crypto Rewards"]
        header_row = None
        for r in range(1, ws.max_row + 1):
            if ws.cell(r, 1).value == "Income code":
                header_row = r
                break
        assert header_row is not None
        for idx, expected in enumerate(self.IRS_SUMMARY_HEADERS, start=1):
            assert ws.cell(header_row, idx).value == expected

    def test_aggregated_headers_are_bold(self):
        wb = openpyxl.Workbook()
        report = _make_crypto_tax_report()
        aggregated = [_make_aggregated_reward()]
        write_crypto_rewards_sheet(wb, report, aggregated)
        ws = wb["Crypto Rewards"]
        header_row = None
        for r in range(1, ws.max_row + 1):
            if ws.cell(r, 1).value == "Income code":
                header_row = r
                break
        assert header_row is not None
        for idx in range(1, len(self.IRS_SUMMARY_HEADERS) + 1):
            assert ws.cell(header_row, idx).font.bold is True

    def test_aggregated_entry_values(self):
        wb = openpyxl.Workbook()
        report = _make_crypto_tax_report()
        aggregated = [_make_aggregated_reward()]
        write_crypto_rewards_sheet(wb, report, aggregated)
        ws = wb["Crypto Rewards"]
        header_row = None
        for r in range(1, ws.max_row + 1):
            if ws.cell(r, 1).value == "Income code":
                header_row = r
                break
        data_row = header_row + 1
        assert ws.cell(data_row, 1).value == "401"
        assert ws.cell(data_row, 2).value == "US"
        assert ws.cell(data_row, 3).value == "Ethereum"
        assert ws.cell(data_row, 4).value == float(Decimal("1500"))
        assert ws.cell(data_row, 5).value == float(Decimal("0"))
        assert ws.cell(data_row, 6).value == float(Decimal("1500"))
        assert ws.cell(data_row, 7).value == 1

    def test_chains_joined_when_multiple(self):
        agg = _make_aggregated_reward(chains=("Ethereum", "Polygon"))
        wb = openpyxl.Workbook()
        report = _make_crypto_tax_report()
        write_crypto_rewards_sheet(wb, report, [agg])
        ws = wb["Crypto Rewards"]
        header_row = None
        for r in range(1, ws.max_row + 1):
            if ws.cell(r, 1).value == "Income code":
                header_row = r
                break
        data_row = header_row + 1
        assert ws.cell(data_row, 3).value == "Ethereum, Polygon"

    def test_empty_aggregated_rewards_shows_note(self):
        wb = openpyxl.Workbook()
        report = _make_crypto_tax_report()
        write_crypto_rewards_sheet(wb, report, [])
        ws = wb["Crypto Rewards"]
        found = False
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=1):
            val = row[0].value
            if val and "No immediately taxable rewards" in str(val):
                found = True
                break
        assert found


@pytest.mark.unit
class TestCryptoRewardsSheetTaxableNowDetail:
    """Tests for section 2a2: Taxable-now support detail."""

    DETAIL_HEADERS = [
        "Date",
        "Asset",
        "Value (EUR)",
        "Income type",
        "Wallet",
        "Platform",
        "Reward chain",
        "Country",
        "Foreign tax (EUR)",
        "Review flag",
        "Description",
    ]

    def test_section_2a2_title_written(self):
        wb = openpyxl.Workbook()
        report = _make_crypto_tax_report()
        aggregated = [_make_aggregated_reward()]
        write_crypto_rewards_sheet(wb, report, aggregated)
        ws = wb["Crypto Rewards"]
        found = False
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=1):
            if row[0].value == "2a2. TAXABLE-NOW - SUPPORT DETAIL":
                found = True
                break
        assert found

    def test_section_2a2_title_is_bold(self):
        wb = openpyxl.Workbook()
        report = _make_crypto_tax_report()
        aggregated = [_make_aggregated_reward()]
        write_crypto_rewards_sheet(wb, report, aggregated)
        ws = wb["Crypto Rewards"]
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=1):
            if row[0].value == "2a2. TAXABLE-NOW - SUPPORT DETAIL":
                assert row[0].font.bold is True
                break

    def test_detail_headers_written(self):
        wb = openpyxl.Workbook()
        report = _make_crypto_tax_report(reward_entries=[_make_reward_entry()])
        aggregated = [_make_aggregated_reward()]
        write_crypto_rewards_sheet(wb, report, aggregated)
        ws = wb["Crypto Rewards"]
        section_row = None
        for r in range(1, ws.max_row + 1):
            if ws.cell(r, 1).value == "2a2. TAXABLE-NOW - SUPPORT DETAIL":
                section_row = r
                break
        assert section_row is not None
        header_row = None
        for r in range(section_row, ws.max_row + 1):
            if ws.cell(r, 1).value == "Date":
                header_row = r
                break
        assert header_row is not None
        for idx, expected in enumerate(self.DETAIL_HEADERS, start=1):
            assert ws.cell(header_row, idx).value == expected

    def test_taxable_now_entry_values(self):
        entry = _make_reward_entry(
            date="2025-03-15",
            asset="ETH",
            value_eur=Decimal("1500"),
            source_type="staking",
            wallet="Kraken",
            platform="Kraken",
            chain="Ethereum",
        )
        report = _make_crypto_tax_report(reward_entries=[entry])
        aggregated = [_make_aggregated_reward()]
        wb = openpyxl.Workbook()
        write_crypto_rewards_sheet(wb, report, aggregated)
        ws = wb["Crypto Rewards"]
        header_row = None
        for r in range(1, ws.max_row + 1):
            if ws.cell(r, 1).value == "Date" and ws.cell(r, 4).value == "Income type":
                header_row = r
                break
        assert header_row is not None
        data_row = header_row + 1
        assert ws.cell(data_row, 1).value == "2025-03-15"
        assert ws.cell(data_row, 2).value == "ETH"
        assert ws.cell(data_row, 3).value == float(Decimal("1500"))
        assert ws.cell(data_row, 4).value == "staking"
        assert ws.cell(data_row, 5).value == "Kraken"
        assert ws.cell(data_row, 6).value == "Kraken"
        assert ws.cell(data_row, 7).value == "Ethereum"
        assert ws.cell(data_row, 8).value == "US"
        assert ws.cell(data_row, 9).value == float(Decimal("0"))
        assert ws.cell(data_row, 10).value == "NO"
        assert ws.cell(data_row, 11).value == "Staking reward payout"

    def test_taxable_now_review_flag_yes_with_reason(self):
        entry = _make_reward_entry(review_required=True, review_reason="Missing cost basis")
        report = _make_crypto_tax_report(reward_entries=[entry])
        aggregated = [_make_aggregated_reward()]
        wb = openpyxl.Workbook()
        write_crypto_rewards_sheet(wb, report, aggregated)
        ws = wb["Crypto Rewards"]
        header_row = None
        for r in range(1, ws.max_row + 1):
            if ws.cell(r, 1).value == "Date" and ws.cell(r, 4).value == "Income type":
                header_row = r
                break
        data_row = header_row + 1
        assert ws.cell(data_row, 10).value == "YES: Missing cost basis"

    def test_no_taxable_now_entries_shows_note(self):
        deferred = _make_reward_entry(classification=RewardTaxClassification.DEFERRED_BY_LAW)
        report = _make_crypto_tax_report(reward_entries=[deferred])
        wb = openpyxl.Workbook()
        write_crypto_rewards_sheet(wb, report, [])
        ws = wb["Crypto Rewards"]
        found = False
        section_start = None
        for r in range(1, ws.max_row + 1):
            if ws.cell(r, 1).value == "2a2. TAXABLE-NOW - SUPPORT DETAIL":
                section_start = r
                break
        assert section_start is not None
        for r in range(section_start, ws.max_row + 1):
            val = ws.cell(r, 1).value
            if val and "No taxable-now rewards" in str(val):
                found = True
                break
        assert found


@pytest.mark.unit
class TestCryptoRewardsSheetDeferredDetail:
    """Tests for section 2b: Deferred by law support detail."""

    def test_section_2b_title_written(self):
        wb = openpyxl.Workbook()
        report = _make_crypto_tax_report()
        aggregated = [_make_aggregated_reward()]
        write_crypto_rewards_sheet(wb, report, aggregated)
        ws = wb["Crypto Rewards"]
        found = False
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=1):
            if row[0].value == "2b. DEFERRED BY LAW - SUPPORT DETAIL":
                found = True
                break
        assert found

    def test_section_2b_title_is_bold(self):
        wb = openpyxl.Workbook()
        report = _make_crypto_tax_report()
        aggregated = [_make_aggregated_reward()]
        write_crypto_rewards_sheet(wb, report, aggregated)
        ws = wb["Crypto Rewards"]
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=1):
            if row[0].value == "2b. DEFERRED BY LAW - SUPPORT DETAIL":
                assert row[0].font.bold is True
                break

    def test_deferred_note_is_italic(self):
        wb = openpyxl.Workbook()
        report = _make_crypto_tax_report()
        aggregated = [_make_aggregated_reward()]
        write_crypto_rewards_sheet(wb, report, aggregated)
        ws = wb["Crypto Rewards"]
        found = False
        section_start = None
        for r in range(1, ws.max_row + 1):
            if ws.cell(r, 1).value == "2b. DEFERRED BY LAW - SUPPORT DETAIL":
                section_start = r
                break
        assert section_start is not None
        for r in range(section_start, ws.max_row + 1):
            val = ws.cell(r, 1).value
            if val and "deferred until disposal" in str(val):
                assert ws.cell(r, 1).font.italic is True
                found = True
                break
        assert found

    def test_deferred_entry_values(self):
        deferred = _make_reward_entry(
            classification=RewardTaxClassification.DEFERRED_BY_LAW,
            date="2025-04-01",
            asset="BTC",
            value_eur=Decimal("500"),
            source_type="mining",
            wallet="Ledger",
            platform="Ledger",
            chain="Bitcoin",
            description="Mining payout",
        )
        report = _make_crypto_tax_report(reward_entries=[deferred])
        wb = openpyxl.Workbook()
        write_crypto_rewards_sheet(wb, report, [])
        ws = wb["Crypto Rewards"]
        section_start = None
        for r in range(1, ws.max_row + 1):
            if ws.cell(r, 1).value == "2b. DEFERRED BY LAW - SUPPORT DETAIL":
                section_start = r
                break
        assert section_start is not None
        header_row = None
        for r in range(section_start, ws.max_row + 1):
            if ws.cell(r, 1).value == "Date" and ws.cell(r, 2).value == "Asset":
                header_row = r
                break
        assert header_row is not None
        data_row = header_row + 1
        assert ws.cell(data_row, 1).value == "2025-04-01"
        assert ws.cell(data_row, 2).value == "BTC"
        assert ws.cell(data_row, 3).value == float(Decimal("500"))
        assert ws.cell(data_row, 4).value == "mining"
        assert ws.cell(data_row, 10).value == "NO"
        assert ws.cell(data_row, 11).value == "Mining payout"

    def test_no_deferred_entries_shows_note(self):
        taxable = _make_reward_entry(classification=RewardTaxClassification.TAXABLE_NOW)
        report = _make_crypto_tax_report(reward_entries=[taxable])
        aggregated = [_make_aggregated_reward()]
        wb = openpyxl.Workbook()
        write_crypto_rewards_sheet(wb, report, aggregated)
        ws = wb["Crypto Rewards"]
        section_start = None
        for r in range(1, ws.max_row + 1):
            if ws.cell(r, 1).value == "2b. DEFERRED BY LAW - SUPPORT DETAIL":
                section_start = r
                break
        assert section_start is not None
        found = False
        for r in range(section_start, ws.max_row + 1):
            val = ws.cell(r, 1).value
            if val and "No deferred rewards" in str(val):
                found = True
                break
        assert found


@pytest.mark.unit
class TestCryptoRewardsSheetClassificationReconciliation:
    """Tests for section 2c: Rewards classification reconciliation."""

    def test_section_2c_title_written(self):
        wb = openpyxl.Workbook()
        report = _make_crypto_tax_report()
        aggregated = [_make_aggregated_reward()]
        write_crypto_rewards_sheet(wb, report, aggregated)
        ws = wb["Crypto Rewards"]
        found = False
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=1):
            if row[0].value == "2c. REWARDS CLASSIFICATION RECONCILIATION":
                found = True
                break
        assert found

    def test_section_2c_title_is_bold(self):
        wb = openpyxl.Workbook()
        report = _make_crypto_tax_report()
        aggregated = [_make_aggregated_reward()]
        write_crypto_rewards_sheet(wb, report, aggregated)
        ws = wb["Crypto Rewards"]
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=1):
            if row[0].value == "2c. REWARDS CLASSIFICATION RECONCILIATION":
                assert row[0].font.bold is True
                break

    def test_reconciliation_key_value_pairs(self):
        taxable = _make_reward_entry(classification=RewardTaxClassification.TAXABLE_NOW, value_eur=Decimal("100"))
        deferred = _make_reward_entry(
            classification=RewardTaxClassification.DEFERRED_BY_LAW, value_eur=Decimal("200"), asset="BTC"
        )
        report = _make_crypto_tax_report(reward_entries=[taxable, deferred])
        aggregated = [_make_aggregated_reward()]
        wb = openpyxl.Workbook()
        write_crypto_rewards_sheet(wb, report, aggregated)
        ws = wb["Crypto Rewards"]
        section_start = None
        for r in range(1, ws.max_row + 1):
            if ws.cell(r, 1).value == "2c. REWARDS CLASSIFICATION RECONCILIATION":
                section_start = r
                break
        assert section_start is not None
        data_start = section_start + 1
        keys = {}
        for r in range(data_start, data_start + 6):
            key = ws.cell(r, 1).value
            value = ws.cell(r, 2).value
            if key:
                keys[key] = value
        assert keys["Total reward rows (raw)"] == 2
        assert keys["Taxable-now rows (immediately taxable)"] == 1
        assert keys["Deferred-by-law rows (taxation deferred)"] == 1
        assert keys["Taxable-now total value (EUR)"] == float(Decimal("100"))
        assert keys["Deferred total value (EUR)"] == float(Decimal("200"))
        assert keys["Filing-ready lines after aggregation"] == 1

    def test_reconciliation_empty_rewards(self):
        report = _make_crypto_tax_report(reward_entries=[])
        wb = openpyxl.Workbook()
        write_crypto_rewards_sheet(wb, report, [])
        ws = wb["Crypto Rewards"]
        section_start = None
        for r in range(1, ws.max_row + 1):
            if ws.cell(r, 1).value == "2c. REWARDS CLASSIFICATION RECONCILIATION":
                section_start = r
                break
        assert section_start is not None
        data_start = section_start + 1
        assert ws.cell(data_start, 2).value == 0
        assert ws.cell(data_start + 1, 2).value == 0
        assert ws.cell(data_start + 2, 2).value == 0


@pytest.mark.unit
class TestCryptoRewardsSheetAutoWidth:
    """Tests that auto_column_width is called."""

    def test_auto_width_adjusts_columns(self):
        wb = openpyxl.Workbook()
        report = _make_crypto_tax_report()
        aggregated = [_make_aggregated_reward()]
        write_crypto_rewards_sheet(wb, report, aggregated)
        ws = wb["Crypto Rewards"]
        assert ws.column_dimensions["A"].width > 0

    def test_column_widths_respect_max_cell_width_cap(self):
        """Verify no column exceeds MAX_CELL_WIDTH + 2 even with long notes."""
        from shares_reporting.application.persisting.excel_utils import MAX_CELL_WIDTH

        wb = openpyxl.Workbook()
        report = _make_crypto_tax_report()
        aggregated = [_make_aggregated_reward()]
        write_crypto_rewards_sheet(wb, report, aggregated)
        ws = wb["Crypto Rewards"]

        # All column widths should be capped at MAX_CELL_WIDTH + 2
        max_allowed = MAX_CELL_WIDTH + 2
        for col_letter, col_dim in ws.column_dimensions.items():
            if col_dim.width is not None:
                assert (
                    col_dim.width <= max_allowed
                ), f"Column {col_letter} width {col_dim.width} exceeds cap {max_allowed}"

    def test_all_columns_have_reasonable_widths(self):
        """Verify no column is collapsed — all have width >= MIN_DATA_WIDTH floor."""

        wb = openpyxl.Workbook()
        report = _make_crypto_tax_report()
        aggregated = [_make_aggregated_reward()]
        write_crypto_rewards_sheet(wb, report, aggregated)
        ws = wb["Crypto Rewards"]

        for col_idx in range(1, 12):  # Columns A through K
            col_letter = openpyxl.utils.get_column_letter(col_idx)
            width = ws.column_dimensions[col_letter].width
            assert width is not None, f"Column {col_letter} should have a width set"
            assert width >= 4, f"Column {col_letter} width {width} is too small"
