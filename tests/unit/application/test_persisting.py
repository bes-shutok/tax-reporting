"""Tests for export_rollover_file and create_currency_table in persisting.py."""

import csv
from decimal import Decimal
from pathlib import Path

import openpyxl
import pytest

from shares_reporting.application.persisting import create_currency_table, export_rollover_file
from shares_reporting.domain.entities import (
    CurrencyCompany,
    QuantitatedTradeAction,
    TradeAction,
    TradeCycle,
)
from shares_reporting.domain.value_objects import Company, Currency
from shares_reporting.infrastructure.config import Config, ConversionRate


@pytest.mark.unit
class TestExportRolloverFileFee:
    """Test that fees written to the rollover CSV are proportional to the rolled-over quantity."""

    def _make_buy_action(self, quantity: str, fee: str) -> TradeAction:
        return TradeAction(
            company=Company(ticker="AAPL"),
            date_time="2023-05-01, 10:00:00",
            currency=Currency(currency="USD"),
            quantity=quantity,
            price="100.00",
            fee=fee,
        )

    def test_fee_is_proportional_when_partial_quantity_rolls_over(self, tmp_path: Path):
        """Rolled-over quantity 40 of original 100 must write fee 4.0, not full fee 10."""
        original_action = self._make_buy_action(quantity="100", fee="10")
        # Only 40 shares remain unmatched
        leftover_trade = QuantitatedTradeAction(quantity=Decimal("40"), action=original_action)

        currency_company = CurrencyCompany(Currency(currency="USD"), Company(ticker="AAPL"))
        leftover_trades = {currency_company: TradeCycle(bought=[leftover_trade], sold=[])}

        out_file = tmp_path / "leftover.csv"
        export_rollover_file(out_file, leftover_trades)

        rows = list(csv.DictReader(out_file.read_text().splitlines()))
        assert len(rows) == 1
        assert Decimal(rows[0]["Comm/Fee"]) == Decimal("4.0")

    def test_fee_is_unchanged_when_full_quantity_rolls_over(self, tmp_path: Path):
        """When full 100 shares roll over, fee must remain 10 (proportional = 1.0 × 10)."""
        original_action = self._make_buy_action(quantity="100", fee="10")
        leftover_trade = QuantitatedTradeAction(quantity=Decimal("100"), action=original_action)

        currency_company = CurrencyCompany(Currency(currency="USD"), Company(ticker="AAPL"))
        leftover_trades = {currency_company: TradeCycle(bought=[leftover_trade], sold=[])}

        out_file = tmp_path / "leftover.csv"
        export_rollover_file(out_file, leftover_trades)

        rows = list(csv.DictReader(out_file.read_text().splitlines()))
        assert len(rows) == 1
        assert Decimal(rows[0]["Comm/Fee"]) == Decimal("10")


@pytest.mark.unit
class TestCreateCurrencyTable:
    """Test that create_currency_table writes headers to separate columns."""

    def test_both_headers_are_written_to_separate_columns(self):
        """'Base/target' and 'Rate' must land in column_no and column_no+1, not both in column_no."""
        wb = openpyxl.Workbook()
        ws = wb.active
        config = Config(
            base="EUR",
            rates=[ConversionRate(base="EUR", calculated="USD", rate=Decimal("1.10"))],
        )

        create_currency_table(ws, column_no=3, row_no=1, config=config)

        # Row 1 = title "Currency exchange rate", row 2 = sub-headers
        assert ws.cell(2, 3).value == "Base/target"
        assert ws.cell(2, 4).value == "Rate"
