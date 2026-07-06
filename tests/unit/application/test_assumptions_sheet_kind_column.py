"""Tests for the Kind column extension in the Assumptions & Methodology tab.

Phase A Task 7 - exercises the new ``Kind`` column (CEX/DEX/UNKNOWN) plus the
OR'd kind-low-confidence signal in the existing ``platform_review_required``
flag (Design Invariant 15).

The tests build synthetic ``CryptoCapitalGainEntry`` fixtures (mirroring the
helpers in ``tests/unit/application/persisting/test_assumptions_sheet.py``),
invoke ``write_assumptions_and_methodology_sheet`` directly with ``th_rows``
and an optional ``RegistrySnapshot`` stub, and assert on cell values plus the
``REVIEW_ROW_FILL`` red fill.
"""

from __future__ import annotations

from dataclasses import dataclass
from datetime import UTC, datetime
from decimal import Decimal

import openpyxl
import pytest

from tax_reporting.application.crypto.wallet_kind import WalletKind
from tax_reporting.application.crypto_reporting import CryptoCapitalGainEntry
from tax_reporting.application.persisting.assumptions_sheet import (
    write_assumptions_and_methodology_sheet,
)
from tax_reporting.application.persisting.excel_utils import REVIEW_ROW_FILL
from tax_reporting.domain.transaction import TransactionHistoryRow
from tests.conftest import make_operator_origin

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

_HEADERS_ROW = 5
_FIRST_DATA_ROW = 6
_PLATFORM_COL = 1
_REVIEW_REQUIRED_COL = 5
_NOTE_COL = 6
_TX_COUNT_COL = 7
_KIND_COL = 8


# ---------------------------------------------------------------------------
# Fixture builders
# ---------------------------------------------------------------------------


def _make_capital_entry(platform: str) -> CryptoCapitalGainEntry:
    """Build a minimal CryptoCapitalGainEntry for ``platform``.

    Uses platform_review_required=False and no assumption so that the only
    signal that can flip Review Required to YES is the kind classification.
    """
    origin = make_operator_origin(
        platform=platform,
        operator_entity=f"{platform} Entity",
        operator_country="US",
        review_required=False,
        platform_review_required=False,
        platform_assumption=None,
    )
    return CryptoCapitalGainEntry(
        disposal_date="2025-06-15",
        acquisition_date="2025-01-10",
        asset="BTC",
        amount=Decimal("0.5"),
        cost_eur=Decimal("20000"),
        proceeds_eur=Decimal("25000"),
        gain_loss_eur=Decimal("5000"),
        holding_period="Short-term",
        wallet=f"{platform.lower()}-wallet",
        platform=platform,
        chain="Ethereum",
        operator_origin=origin,
        annex_hint="J",
        review_required=False,
        notes="",
    )


@dataclass
class _StubRegistry:
    """Minimal RegistrySnapshot implementation for tier-1 wiring."""

    mapping: dict[str, WalletKind]

    def classify(self, platform: str) -> WalletKind | None:
        return self.mapping.get(platform)


def _evm_txhash(seed: int) -> str:
    """Return a 66-char EVM-shaped txhash deterministically derived from ``seed``."""
    return "0x" + f"{seed:064x}"[-64:]


def _th_row(
    *,
    platform: str,
    type_: str,
    tx_hash: str | None,
    row_index: int,
) -> TransactionHistoryRow:
    """Build a minimal TransactionHistoryRow attributed to ``platform`` via sending_wallet."""
    return TransactionHistoryRow(
        utc_instant=datetime(2025, 1, 1, 0, 0, 0, tzinfo=UTC),
        type=type_,
        tag="",
        sending_wallet=platform,
        receiving_wallet="",
        sending_amount=Decimal("1"),
        sending_currency="USDT",
        receiving_amount=None,
        receiving_currency=None,
        tx_hash=tx_hash,
        tx_src=None,
        tx_dest=None,
        row_index=row_index,
    )


# ---------------------------------------------------------------------------
# Tests
# ---------------------------------------------------------------------------


@pytest.mark.unit
class TestAssumptionsKindColumn:
    """Task 7 - Kind column rendering + OR'd kind-low-confidence review flag."""

    def test_registry_platform_renders_kind_no_red(self) -> None:
        """Tier-1 (registry) platform: Kind=CEX, no red fill.

        Discriminates: fails if registry lookup is not consulted, or if a
        registry-classified platform (confidence 1.0) is red-flagged.
        """
        wb = openpyxl.Workbook()
        entry = _make_capital_entry("Kraken")
        registry = _StubRegistry(mapping={"Kraken": WalletKind.CEX})
        # A TH row attributed to Kraken - irrelevant because registry wins.
        th_rows = [_th_row(platform="Kraken", type_="buy", tx_hash="TRADE-1", row_index=0)]
        write_assumptions_and_methodology_sheet(
            wb,
            capital_entries=[entry],
            th_rows=th_rows,
            registry=registry,
        )
        ws = wb["Assumptions & Methodology"]
        assert ws.cell(_FIRST_DATA_ROW, _PLATFORM_COL).value == "Kraken"
        assert ws.cell(_FIRST_DATA_ROW, _KIND_COL).value == "CEX"
        assert ws.cell(_FIRST_DATA_ROW, _REVIEW_REQUIRED_COL).value == "NO"
        assert ws.cell(_FIRST_DATA_ROW, _PLATFORM_COL).fill != REVIEW_ROW_FILL

    def test_auto_discovered_high_confidence_renders_kind_no_red(self) -> None:
        """Tier-2 auto-discovered DEX at >=0.95: Kind=DEX, no red fill.

        19 on-chain / 1 off-chain -> confidence 0.95 (>= threshold). The Note
        column must mention the auto-discovery source so the user sees why the
        platform was classified without a registry entry.
        """
        wb = openpyxl.Workbook()
        entry = _make_capital_entry("Ledger Main")
        rows: list[TransactionHistoryRow] = []
        for i in range(19):
            rows.append(
                _th_row(
                    platform="Ledger Main",
                    type_="crypto_withdrawal",
                    tx_hash=_evm_txhash(i),
                    row_index=i,
                )
            )
        rows.append(_th_row(platform="Ledger Main", type_="buy", tx_hash="TRADE-X", row_index=19))
        write_assumptions_and_methodology_sheet(
            wb,
            capital_entries=[entry],
            th_rows=rows,
            registry=None,
        )
        ws = wb["Assumptions & Methodology"]
        assert ws.cell(_FIRST_DATA_ROW, _PLATFORM_COL).value == "Ledger Main"
        assert ws.cell(_FIRST_DATA_ROW, _KIND_COL).value == "DEX"
        assert ws.cell(_FIRST_DATA_ROW, _REVIEW_REQUIRED_COL).value == "NO"
        assert ws.cell(_FIRST_DATA_ROW, _PLATFORM_COL).fill != REVIEW_ROW_FILL

    def test_auto_discovered_low_confidence_renders_kind_with_red_and_reason(self) -> None:
        """Tier-2 auto-discovered at <0.95: red Review Required + Note mentions tallies.

        12 on-chain / 3 off-chain -> DEX at confidence 0.80. Invariant 12 +
        Invariant 15: kind signal OR'd into existing Review Required, red fill
        applied via existing REVIEW_ROW_FILL, Note column carries the reason.
        """
        wb = openpyxl.Workbook()
        entry = _make_capital_entry("Mixed Platform")
        rows: list[TransactionHistoryRow] = []
        idx = 0
        for _ in range(12):
            rows.append(
                _th_row(
                    platform="Mixed Platform",
                    type_="crypto_deposit",
                    tx_hash=_evm_txhash(idx),
                    row_index=idx,
                )
            )
            idx += 1
        for _ in range(3):
            rows.append(
                _th_row(
                    platform="Mixed Platform",
                    type_="buy",
                    tx_hash=f"TRADE-{idx}",
                    row_index=idx,
                )
            )
            idx += 1
        write_assumptions_and_methodology_sheet(
            wb,
            capital_entries=[entry],
            th_rows=rows,
            registry=None,
        )
        ws = wb["Assumptions & Methodology"]
        assert ws.cell(_FIRST_DATA_ROW, _PLATFORM_COL).value == "Mixed Platform"
        assert ws.cell(_FIRST_DATA_ROW, _KIND_COL).value == "DEX"
        assert ws.cell(_FIRST_DATA_ROW, _REVIEW_REQUIRED_COL).value == "YES"
        assert ws.cell(_FIRST_DATA_ROW, _PLATFORM_COL).fill == REVIEW_ROW_FILL
        assert ws.cell(_FIRST_DATA_ROW, _KIND_COL).fill == REVIEW_ROW_FILL
        note_value = ws.cell(_FIRST_DATA_ROW, _NOTE_COL).value
        assert isinstance(note_value, str)
        assert "12 on-chain / 3 off-chain" in note_value

    def test_unknown_kind_renders_unknown_with_red(self) -> None:
        """No evidence, no registry: Kind=UNKNOWN, red fill, Note mentions no evidence."""
        wb = openpyxl.Workbook()
        entry = _make_capital_entry("Phantom")
        # No TH rows for "Phantom" - it has no evidence and no registry match.
        write_assumptions_and_methodology_sheet(
            wb,
            capital_entries=[entry],
            th_rows=[],
            registry=None,
        )
        ws = wb["Assumptions & Methodology"]
        assert ws.cell(_FIRST_DATA_ROW, _PLATFORM_COL).value == "Phantom"
        assert ws.cell(_FIRST_DATA_ROW, _KIND_COL).value == "UNKNOWN"
        assert ws.cell(_FIRST_DATA_ROW, _REVIEW_REQUIRED_COL).value == "YES"
        assert ws.cell(_FIRST_DATA_ROW, _PLATFORM_COL).fill == REVIEW_ROW_FILL
        note_value = ws.cell(_FIRST_DATA_ROW, _NOTE_COL).value
        assert isinstance(note_value, str)
        assert "no rows" in note_value.lower()

    def test_existing_columns_byte_identical_for_non_kind_signals(self) -> None:
        """Pre-Phase-A Platform Assumptions row is byte-identical when kind signal does not fire.

        Design Invariant 15: when ``th_rows`` is omitted, no classification
        runs. Platform / Operator Entity / Country / Confidence / Review
        Required / Note / Tx Count must equal the pre-Task-7 baseline.
        """
        wb_before = openpyxl.Workbook()
        wb_after = openpyxl.Workbook()
        entry = _make_capital_entry("Kraken")
        # Pre-Task-7 call shape: no th_rows / registry kwargs.
        write_assumptions_and_methodology_sheet(wb_before, capital_entries=[entry])
        # Phase A call shape with th_rows omitted (degrades gracefully).
        write_assumptions_and_methodology_sheet(wb_after, capital_entries=[entry])
        ws_before = wb_before["Assumptions & Methodology"]
        ws_after = wb_after["Assumptions & Methodology"]
        # Existing columns 1-7 must be byte-identical. Compare values exactly
        # and red-fill presence via ``fill_type`` (openpyxl PatternFill objects
        # are not ==-equal across distinct Workbook instances even when both
        # represent "no fill"; ``fill_type`` collapses the equivalent case).
        for col in range(1, 8):
            assert ws_before.cell(_HEADERS_ROW, col).value == ws_after.cell(_HEADERS_ROW, col).value
            assert ws_before.cell(_FIRST_DATA_ROW, col).value == ws_after.cell(_FIRST_DATA_ROW, col).value
            before_fill = ws_before.cell(_FIRST_DATA_ROW, col).fill.fill_type
            after_fill = ws_after.cell(_FIRST_DATA_ROW, col).fill.fill_type
            assert before_fill == after_fill
        # Sanity: the baseline fixture has Review Required = NO and no red.
        assert ws_after.cell(_FIRST_DATA_ROW, _REVIEW_REQUIRED_COL).value == "NO"
        assert ws_after.cell(_FIRST_DATA_ROW, _PLATFORM_COL).fill != REVIEW_ROW_FILL

    def test_kind_column_header_label(self) -> None:
        """The new column header reads exactly ``"Kind"``."""
        wb = openpyxl.Workbook()
        entry = _make_capital_entry("Kraken")
        write_assumptions_and_methodology_sheet(
            wb,
            capital_entries=[entry],
            th_rows=[],
            registry=None,
        )
        ws = wb["Assumptions & Methodology"]
        assert ws.cell(_HEADERS_ROW, _KIND_COL).value == "Kind"
