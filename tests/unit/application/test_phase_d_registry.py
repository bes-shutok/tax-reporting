"""Phase D Task 1 - production WalletKind registry binding.

Pins the new production registry adapter that sources each platform's
WalletKind from ``resolve_operator_origin`` via the ``wallet_kind`` field
on ``OperatorOrigin``. Closes the 540-row Phase C Binance gap by making
Binance classify as CEX at tier 1 (confidence 1.0) instead of relying on
tier-2 auto-discovery, which left the Kind column blank for every Binance
row in the synthetic corpus.

The registry adapter implements ``RegistrySnapshot`` (Phase A Protocol
defined in ``wallet_kind.py``) and is constructed in
``generate_tax_report`` (``workbook_builder.py``); it flows through the
relaxed gate at ``assumptions_sheet.py`` so the registry-only call path
(no ``th_rows`` kwarg) actually fires the classifier.
"""

from __future__ import annotations

from datetime import UTC, datetime
from decimal import Decimal

import openpyxl

from tax_reporting.application.crypto.operator_origin import (
    _KNOWN_PLATFORM_BRANDS,
    _PLATFORM_KIND,
)
from tax_reporting.application.crypto.wallet_kind import WalletKind
from tax_reporting.application.crypto.wallet_kind_registry import (
    ProductionWalletKindRegistry,
)
from tax_reporting.application.crypto_reporting import CryptoCapitalGainEntry
from tax_reporting.application.persisting.assumptions_kind_column import (
    classify_platforms_for_summaries,
)
from tax_reporting.application.persisting.assumptions_sheet import (
    write_assumptions_and_methodology_sheet,
)
from tax_reporting.domain.transaction import TransactionHistoryRow
from tests.conftest import make_operator_origin

_HEADERS_ROW = 5
_FIRST_DATA_ROW = 6
_PLATFORM_COL = 1
_KIND_COL = 8


def _make_capital_entry(platform: str) -> CryptoCapitalGainEntry:
    """Build a minimal CryptoCapitalGainEntry for ``platform``."""
    origin = make_operator_origin(
        platform=platform,
        operator_entity=f"{platform} Entity",
        operator_country="US",
        review_required=False,
        platform_review_required=False,
        platform_assumption=None,
    )
    return CryptoCapitalGainEntry(
        disposal_date="2025-06-15",
        acquisition_date="2025-01-10",
        asset="BTC",
        amount=Decimal("0.5"),
        cost_eur=Decimal("20000"),
        proceeds_eur=Decimal("25000"),
        gain_loss_eur=Decimal("5000"),
        holding_period="Short-term",
        wallet=f"{platform.lower()}-wallet",
        platform=platform,
        chain="Ethereum",
        operator_origin=origin,
        annex_hint="J",
        review_required=False,
        notes="",
    )


def _th_row(platform: str, row_index: int) -> TransactionHistoryRow:
    return TransactionHistoryRow(
        utc_instant=datetime(2025, 1, 1, 0, 0, 0, tzinfo=UTC),
        type="buy",
        tag="",
        sending_wallet=platform,
        receiving_wallet="",
        sending_amount=Decimal("1"),
        sending_currency="USDT",
        receiving_amount=None,
        receiving_currency=None,
        tx_hash="TRADE-1",
        tx_src=None,
        tx_dest=None,
        row_index=row_index,
    )


class TestPhaseDRegistry:
    """Phase D Task 1 - production WalletKind registry binding."""

    def test_classify_known_cex_platform(self) -> None:
        """Kraken/ByBit/Wirex resolve to CEX via the production registry."""
        registry = ProductionWalletKindRegistry()
        assert registry.classify("Kraken") == WalletKind.CEX
        assert registry.classify("ByBit") == WalletKind.CEX
        assert registry.classify("Wirex") == WalletKind.CEX

    def test_classify_known_dex_platform(self) -> None:
        """Ledger Berachain/SUI/Ledger resolve to DEX via the production registry."""
        registry = ProductionWalletKindRegistry()
        assert registry.classify("Ledger Berachain (BERA)") == WalletKind.DEX
        assert registry.classify("SUI") == WalletKind.DEX
        assert registry.classify("Ledger") == WalletKind.DEX

    def test_classify_binance_resolves_cex(self) -> None:
        """Binance resolves to CEX (closes the 540-row Phase C gap)."""
        registry = ProductionWalletKindRegistry()
        assert registry.classify("Binance") == WalletKind.CEX

    def test_classify_unmapped_returns_none(self) -> None:
        """Unmapped platform returns None so the caller falls through to tier 2."""
        registry = ProductionWalletKindRegistry()
        assert registry.classify("UnknownNewExchange") is None

    def test_assumptions_kind_column_consumes_registry(self) -> None:
        """classify_platforms_for_summaries honors the production registry."""
        registry = ProductionWalletKindRegistry()
        rows = [_th_row("Kraken", 0)]
        result = classify_platforms_for_summaries(["Kraken"], rows, registry)
        assert result["Kraken"].kind == WalletKind.CEX

    def test_registry_only_call_does_not_crash_on_none_th_rows(self) -> None:
        """write_assumptions_and_methodology_sheet with registry= and NO th_rows.

        Pins r7 Blocker #1: forwarding th_rows=None to
        aggregate_platform_evidence crashes with TypeError. The relaxed gate
        must normalize th_rows to () when it is None and registry is supplied.
        """
        wb = openpyxl.Workbook()
        entry = _make_capital_entry("Kraken")
        registry = ProductionWalletKindRegistry()
        write_assumptions_and_methodology_sheet(
            wb,
            capital_entries=[entry],
            reward_entries=[],
            registry=registry,
        )
        ws = wb["Assumptions & Methodology"]
        assert ws.cell(_FIRST_DATA_ROW, _PLATFORM_COL).value == "Kraken"
        assert ws.cell(_FIRST_DATA_ROW, _KIND_COL).value == "CEX"

    def test_every_entity_chain_platform_has_kind(self) -> None:
        """Every platform brand the entity chain recognizes has a non-None kind.

        Pins r7 Medium #7: prevents a platform added to the entity chain but
        missing from _PLATFORM_KIND from silently misclassifying.
        """
        assert _KNOWN_PLATFORM_BRANDS, "_KNOWN_PLATFORM_BRANDS must be non-empty"
        for brand in _KNOWN_PLATFORM_BRANDS:
            assert brand in _PLATFORM_KIND, (
                f"platform '{brand}' is in _KNOWN_PLATFORM_BRANDS but missing from _PLATFORM_KIND"
            )
            assert _PLATFORM_KIND[brand] is not None, (
                f"platform '{brand}' has None kind in _PLATFORM_KIND"
            )
