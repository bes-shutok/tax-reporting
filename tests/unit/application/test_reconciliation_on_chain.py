"""Reconciliation sheet on-chain provenance + delta block (Plan Task 12; F9).

These tests pin the schema extension to ``CryptoReconciliationSummary`` that
surfaces per-wallet source provenance and the Koinly-vs-on-chain delta block
when ``on_chain_th_wallets`` is set (Plan Task 12 of
``docs/history/plans/2026-08-02-on-chain-tx-tagger.md``; review F9: the schema
extension needs a serialization test).

The two new dataclasses
-----------------------

``WalletSourceProvenance`` - one row per wallet, carrying the wallet label, the
source kind (``"koinly"`` or ``"on_chain"``), and the row count the source
contributed to this run's merged Transaction History.

``OnChainDeltaBlock`` - the Koinly-vs-on-chain reconciliation counts for the
opted-in wallets (rows reclassified, rewards added, gas added, LP reclassified)
plus a small list of sample on-chain tx hashes for audit drill-down.

Backward compatibility (load-bearing)
-------------------------------------

When ``on_chain_th_wallets`` is UNSET (the Koinly-only path), both new fields
default to their empty values (``per_wallet_source_provenance=[]`` and
``on_chain_delta=None``), so today's ``CryptoReconciliationSummary``
construction sites keep working and the reconciliation sheet is byte-identical
(Task 1 characterization stays GREEN). The new sections render ONLY when the
fields are populated.
"""

from __future__ import annotations

from dataclasses import asdict
from decimal import Decimal

import openpyxl
import pytest

from tax_reporting.application.crypto.entities import (
    CryptoReconciliationSummary,
    OnChainDeltaBlock,
    WalletSourceProvenance,
)
from tax_reporting.application.crypto_reporting import (
    CapitalGainPeriodStats,
    CryptoCapitalGainStats,
    CryptoTaxReport,
)
from tax_reporting.application.persisting.crypto_reconciliation_sheet import (
    write_crypto_reconciliation_sheet,
)


def _empty_stats() -> CapitalGainPeriodStats:
    return CapitalGainPeriodStats(
        count=0,
        cost_total_eur=Decimal("0"),
        proceeds_total_eur=Decimal("0"),
        gain_loss_total_eur=Decimal("0"),
    )


def _make_report(reconciliation: CryptoReconciliationSummary) -> CryptoTaxReport:
    """Build a minimal ``CryptoTaxReport`` carrying the given reconciliation."""
    empty = _empty_stats()
    capital_gain_stats = CryptoCapitalGainStats(
        short_term=empty, long_term=empty, mixed=empty, unknown=empty, grand_total=empty
    )
    return CryptoTaxReport(
        tax_year=2025,
        capital_entries=[],
        reward_entries=[],
        reconciliation=reconciliation,
        capital_gain_stats=capital_gain_stats,
    )


def _make_reconciliation(
    *,
    per_wallet_source_provenance: list[WalletSourceProvenance] | None = None,
    on_chain_delta: OnChainDeltaBlock | None = None,
) -> CryptoReconciliationSummary:
    """Build a ``CryptoReconciliationSummary`` with the new fields populated.

    Today's construction sites omit the new fields entirely; the defaults
    (empty list / None) preserve byte-identical Koinly-only output. This helper
    exercises the populated path.
    """
    return CryptoReconciliationSummary(
        capital_rows=2,
        reward_rows=1,
        short_term_rows=1,
        long_term_rows=1,
        mixed_rows=0,
        unknown_rows=0,
        capital_cost_total_eur=Decimal("1000"),
        capital_proceeds_total_eur=Decimal("1500"),
        capital_gain_total_eur=Decimal("500"),
        reward_total_eur=Decimal("10"),
        opening_holdings=None,
        closing_holdings=None,
        per_wallet_source_provenance=per_wallet_source_provenance or [],
        on_chain_delta=on_chain_delta,
    )


def _scan_key_values(ws) -> dict[str, object]:
    """Scan the sheet column 1->column 2 (the reconciliation key/value layout)."""
    keys: dict[str, object] = {}
    for r in range(1, ws.max_row + 1):
        key = ws.cell(r, 1).value
        value = ws.cell(r, 2).value
        if key:
            keys[str(key)] = value
    return keys


@pytest.mark.unit
class TestReconciliationOnChain:
    """Plan Task 12: per-wallet source provenance + on-chain delta block."""

    def test_per_wallet_source_provenance(self) -> None:
        """A run with ``on_chain_th_wallets=[BERA]`` -> the reconciliation sheet
        shows BERA=on-chain, other wallets=Koinly, with per-wallet row counts.

        Constructs a ``CryptoReconciliationSummary`` whose
        ``per_wallet_source_provenance`` lists BERA as on-chain (with its row
        count) and another wallet as koinly (with its row count), then asserts
        the writer renders a per-wallet provenance table with both rows.
        """
        provenance = [
            WalletSourceProvenance(
                wallet_label="Ledger Berachain (BERA)",
                source_kind="on_chain",
                row_count=42,
            ),
            WalletSourceProvenance(
                wallet_label="ByBit",
                source_kind="koinly",
                row_count=7,
            ),
        ]
        recon = _make_reconciliation(per_wallet_source_provenance=provenance)
        report = _make_report(recon)
        wb = openpyxl.Workbook()
        write_crypto_reconciliation_sheet(wb, report)
        ws = wb["Crypto Reconciliation"]

        # Locate the provenance table header row.
        header_row = None
        for r in range(1, ws.max_row + 1):
            if ws.cell(r, 1).value == "Wallet" and ws.cell(r, 2).value == "Source":
                header_row = r
                break
        assert header_row is not None, (
            "the per-wallet source provenance table header (Wallet, Source, Rows) "
            "must be rendered when per_wallet_source_provenance is populated"
        )
        # Header columns.
        assert ws.cell(header_row, 3).value == "Rows"
        # BERA = on-chain.
        bera_row = header_row + 1
        assert ws.cell(bera_row, 1).value == "Ledger Berachain (BERA)"
        assert ws.cell(bera_row, 2).value == "on_chain"
        assert ws.cell(bera_row, 3).value == 42
        # ByBit = koinly.
        bybit_row = header_row + 2
        assert ws.cell(bybit_row, 1).value == "ByBit"
        assert ws.cell(bybit_row, 2).value == "koinly"
        assert ws.cell(bybit_row, 3).value == 7

    def test_delta_block_when_flag_on(self) -> None:
        """Flag on -> the sheet renders a delta block listing rows reclassified,
        rewards added, gas added, LP reclassified (Koinly->on-chain), plus
        sample on-chain tx hashes for audit drill-down.
        """
        delta = OnChainDeltaBlock(
            rows_reclassified=12,
            rewards_added=3,
            gas_added=8,
            lp_reclassified=1,
            sample_hashes=["0xaaa111", "0xbbb222", "0xccc333"],
        )
        recon = _make_reconciliation(on_chain_delta=delta)
        report = _make_report(recon)
        wb = openpyxl.Workbook()
        write_crypto_reconciliation_sheet(wb, report)
        ws = wb["Crypto Reconciliation"]
        keys = _scan_key_values(ws)

        assert keys["Rows reclassified (Koinly -> on-chain)"] == 12, (
            f"delta block must list rows_reclassified; keys={keys}"
        )
        assert keys["Rewards added (on-chain)"] == 3
        assert keys["Gas added (on-chain)"] == 8
        assert keys["LP reclassified (Koinly -> on-chain)"] == 1
        # Sample hashes rendered (joined, order-preserving).
        assert "Sample on-chain tx hashes" in keys, (
            f"delta block must list sample hashes; keys={keys}"
        )
        sample_value = keys["Sample on-chain tx hashes"]
        for h in ("0xaaa111", "0xbbb222", "0xccc333"):
            assert h in str(sample_value), (
                f"sample hash {h} must appear in the rendered sample-hashes cell; "
                f"got {sample_value!r}"
            )

    def test_no_delta_block_when_flag_off(self) -> None:
        """Flag unset -> NO delta block (today's behavior; ``on_chain_delta``
        is None and ``per_wallet_source_provenance`` is empty).

        Constructs a ``CryptoReconciliationSummary`` WITHOUT the new fields
        (mirrors today's Koinly-only construction sites) and asserts neither
        the delta block nor the provenance table headers appear.
        """
        recon = CryptoReconciliationSummary(
            capital_rows=1,
            reward_rows=0,
            short_term_rows=1,
            long_term_rows=0,
            mixed_rows=0,
            unknown_rows=0,
            capital_cost_total_eur=Decimal("100"),
            capital_proceeds_total_eur=Decimal("120"),
            capital_gain_total_eur=Decimal("20"),
            reward_total_eur=Decimal("0"),
            opening_holdings=None,
            closing_holdings=None,
        )
        # The defaults are the load-bearing backward-compat contract.
        assert recon.per_wallet_source_provenance == []
        assert recon.on_chain_delta is None

        report = _make_report(recon)
        wb = openpyxl.Workbook()
        write_crypto_reconciliation_sheet(wb, report)
        ws = wb["Crypto Reconciliation"]
        keys = _scan_key_values(ws)

        # None of the delta-block keys may appear.
        for forbidden in (
            "Rows reclassified (Koinly -> on-chain)",
            "Rewards added (on-chain)",
            "Gas added (on-chain)",
            "LP reclassified (Koinly -> on-chain)",
            "Sample on-chain tx hashes",
        ):
            assert forbidden not in keys, (
                f"flag-unset run must NOT render the delta block; key {forbidden!r} "
                f"appeared (keys={keys})"
            )
        # No per-wallet provenance table header either.
        provenance_header_present = any(
            ws.cell(r, 1).value == "Wallet" and ws.cell(r, 2).value == "Source"
            for r in range(1, ws.max_row + 1)
        )
        assert not provenance_header_present, (
            "flag-unset run must NOT render the per-wallet provenance table"
        )

    def test_new_fields_serialize(self) -> None:
        """Review F9: a ``CryptoReconciliationSummary`` with the new fields
        populated survives a round-trip through ``dataclasses.asdict`` (the
        transformation the writer would apply to serialize the schema) and the
        nested ``WalletSourceProvenance`` / ``OnChainDeltaBlock`` values are
        readable back field-by-field.
        """
        provenance = [
            WalletSourceProvenance(
                wallet_label="Ledger Berachain (BERA)",
                source_kind="on_chain",
                row_count=42,
            ),
        ]
        delta = OnChainDeltaBlock(
            rows_reclassified=12,
            rewards_added=3,
            gas_added=8,
            lp_reclassified=1,
            sample_hashes=["0xaaa111", "0xbbb222"],
        )
        recon = _make_reconciliation(
            per_wallet_source_provenance=provenance,
            on_chain_delta=delta,
        )

        # Round-trip: dataclass -> dict -> reconstruct field-by-field.
        serialized = asdict(recon)
        # The new fields are present in the serialized dict.
        assert "per_wallet_source_provenance" in serialized
        assert "on_chain_delta" in serialized

        # Provenance round-trips with full fidelity.
        prov_list = serialized["per_wallet_source_provenance"]
        assert len(prov_list) == 1
        prov0 = prov_list[0]
        assert prov0["wallet_label"] == "Ledger Berachain (BERA)"
        assert prov0["source_kind"] == "on_chain"
        assert prov0["row_count"] == 42
        # Reconstruct and assert equality (the round-trip preserves identity).
        restored_prov = [
            WalletSourceProvenance(
                wallet_label=p["wallet_label"],
                source_kind=p["source_kind"],
                row_count=p["row_count"],
            )
            for p in prov_list
        ]
        assert restored_prov == provenance

        # Delta block round-trips with full fidelity.
        delta_dict = serialized["on_chain_delta"]
        assert delta_dict["rows_reclassified"] == 12
        assert delta_dict["rewards_added"] == 3
        assert delta_dict["gas_added"] == 8
        assert delta_dict["lp_reclassified"] == 1
        assert delta_dict["sample_hashes"] == ["0xaaa111", "0xbbb222"]
        restored_delta = OnChainDeltaBlock(
            rows_reclassified=delta_dict["rows_reclassified"],
            rewards_added=delta_dict["rewards_added"],
            gas_added=delta_dict["gas_added"],
            lp_reclassified=delta_dict["lp_reclassified"],
            sample_hashes=list(delta_dict["sample_hashes"]),
        )
        assert restored_delta == delta
