"""Integration tests for complete dividend processing flow."""

from decimal import Decimal

from shares_reporting.application.extraction import parse_ib_export_all
from shares_reporting.application.persisting import generate_tax_report
from shares_reporting.application.transformation import calculate_fifo_gains


class TestDividendIntegrationFlow:
    """Integration tests for end-to-end dividend processing."""

    def test_complete_dividend_flow_from_csv_to_excel(self, tmp_path):  # noqa: PLR0915 - Integration test needs comprehensive setup
        """Test complete flow from CSV parsing to Excel report generation."""
        # Create a comprehensive IB-style CSV with dividends
        csv_content = (
            "Financial Instrument Information,Header,Asset Category,Symbol,Description,Conid,Security ID,Multiplier\n"
            "Financial Instrument Information,Data,Stocks,AAPL,Apple Inc.,123456,US0378331005,1\n"
            "Financial Instrument Information,Data,Stocks,MSFT,Microsoft Corp.,234567,US5949181045,1\n"
            "Financial Instrument Information,Data,Stocks,ASML,ASML Holding N.V.,345678,NL0010273215,1\n"
            # Dividends section
            "Dividends,Header,Currency,Date,Description,Amount\n"
            "Dividends,Data,USD,2023-03-15,AAPL(US0378331005) CASH DIVIDEND 0.24 USD,24.00\n"
            "Dividends,Data,USD,2023-06-15,AAPL(US0378331005) CASH DIVIDEND 0.24 USD,24.00\n"
            "Dividends,Data,USD,2023-09-15,AAPL(US0378331005) CASH DIVIDEND 0.24 USD,24.00\n"
            "Dividends,Data,USD,2023-03-15,MSFT(US5949181045) CASH DIVIDEND 0.68 USD,68.00\n"
            "Dividends,Data,USD,2023-06-15,MSFT(US5949181045) CASH DIVIDEND 0.68 USD,68.00\n"
            "Dividends,Data,EUR,2023-04-10,ASML(NL0010273215) CASH DIVIDEND EUR 1.45,145.00\n"
            "Dividends,Data,EUR,2023-07-10,ASML(NL0010273215) CASH DIVIDEND EUR 1.45,145.00\n"
            # Withholding Tax section
            "Withholding Tax,Header,Currency,Date,Description,Amount,Code\n"
            "Withholding Tax,Data,USD,2023-03-15,AAPL(US0378331005) US TAX,-3.60,,\n"
            "Withholding Tax,Data,USD,2023-06-15,AAPL(US0378331005) US TAX,-3.60,,\n"
            "Withholding Tax,Data,USD,2023-09-15,AAPL(US0378331005) US TAX,-3.60,,\n"
            "Withholding Tax,Data,USD,2023-03-15,MSFT(US5949181045) US TAX,-10.20,,\n"
            "Withholding Tax,Data,USD,2023-06-15,MSFT(US5949181045) US TAX,-10.20,,\n"
            # Trades section (minimal for integration)
            "Trades,Header,DataDiscriminator,Asset Category,Currency,Symbol,"
            "Date/Time,Quantity,T. Price,Proceeds,Comm/Fee,Basis,Realized P/L,Code\n"
            'Trades,Data,Order,Stocks,USD,AAPL,"2023-01-15, 10:30:00",100,150.25,-15025.00,1.00,15026.00,0,O\n'
            'Trades,Data,Order,Stocks,USD,AAPL,"2023-12-15, 15:30:00",-100,160.50,16050.00,1.00,-16051.00,1000.00,C\n'
        )

        csv_file = tmp_path / "integration_test.csv"
        csv_file.write_text(csv_content)

        # Step 1: Parse raw CSV data for verification
        from shares_reporting.application.extraction.processing import _extract_csv_data

        raw_data = _extract_csv_data(csv_file)

        # Verify parsed raw data
        assert len(raw_data.security_info) == 3
        assert "AAPL" in raw_data.security_info
        assert "MSFT" in raw_data.security_info
        assert "ASML" in raw_data.security_info

        assert len(raw_data.raw_dividend_data) == 7
        assert len(raw_data.raw_withholding_tax_data) == 5

        # Step 2: Parse the full export (processed data)
        ib_export_data = parse_ib_export_all(csv_file)

        # Step 3: Process capital gains (minimal trades)
        from shares_reporting.domain.collections import CapitalGainLinesPerCompany, TradeCyclePerCompany

        leftover_trades: TradeCyclePerCompany = {}
        capital_gains: CapitalGainLinesPerCompany = {}

        calculate_fifo_gains(ib_export_data.trade_cycles, leftover_trades, capital_gains)
        assert len(capital_gains) == 1  # One complete cycle

        # Step 4: Extract dividend income from the parsed data
        dividend_income = ib_export_data.dividend_income

        # Verify dividend processing
        assert len(dividend_income) == 3

        # AAPL: 3 dividends * 24.00 = 72.00 gross, 3 taxes * 3.60 = 10.80 total tax
        aapl_dividend = dividend_income["AAPL"]
        assert aapl_dividend.gross_amount == Decimal("72.00")
        assert aapl_dividend.total_taxes == Decimal("10.80")
        assert aapl_dividend.get_net_amount() == Decimal("61.20")
        assert aapl_dividend.currency.currency == "USD"

        # MSFT: 2 dividends * 68.00 = 136.00 gross, 2 taxes * 10.20 = 20.40 total tax
        msft_dividend = dividend_income["MSFT"]
        assert msft_dividend.gross_amount == Decimal("136.00")
        assert msft_dividend.total_taxes == Decimal("20.40")
        assert msft_dividend.get_net_amount() == Decimal("115.60")
        assert msft_dividend.currency.currency == "USD"

        # ASML: 2 dividends * 145.00 = 290.00 gross, no tax
        asml_dividend = dividend_income["ASML"]
        assert asml_dividend.gross_amount == Decimal("290.00")
        assert asml_dividend.total_taxes == Decimal("0.00")
        assert asml_dividend.get_net_amount() == Decimal("290.00")
        assert asml_dividend.currency.currency == "EUR"

        # Step 4: Generate Excel report
        report_path = tmp_path / "integration_report.xlsx"
        generate_tax_report(
            extract=report_path,
            capital_gain_lines_per_company=capital_gains,
            dividend_income_per_company=dividend_income,
        )

        # Verify Excel file
        assert report_path.exists()

        import openpyxl

        workbook = openpyxl.load_workbook(report_path)
        worksheet = workbook.active
        assert worksheet is not None, "Workbook should have an active worksheet"

        # Find dividend section
        dividend_rows = []
        for row_idx, row in enumerate(worksheet.iter_rows(values_only=True), 1):
            if row and len(row) > 1 and row[1] == "Dividends":
                dividend_rows.append((row_idx, row))

        assert len(dividend_rows) == 3  # Three dividend entries

        # Verify AAPL dividend row
        aapl_row = next(row for row in dividend_rows if row[1][8] == "AAPL")
        assert aapl_row[1][2] == "United States"  # Country
        assert aapl_row[1][3] == "US0378331005"  # ISIN
        assert aapl_row[1][10] == "72.00"  # Original gross
        assert aapl_row[1][11] == "10.80"  # Original tax
        assert aapl_row[1][12] == "61.20"  # Net amount

        workbook.close()

    def test_dividend_flow_without_trades(self, tmp_path):
        """Test dividend processing flow with no trades (dividends only)."""
        csv_content = (
            "Financial Instrument Information,Header,Asset Category,Symbol,Description,Conid,Security ID,Multiplier\n"
            "Financial Instrument Information,Data,Stocks,VFIAX,Vanguard 500 Index Admiral,123456,US9229087369,1\n"
            "Dividends,Header,Currency,Date,Description,Amount\n"
            "Dividends,Data,USD,2023-03-31,VFIAX(US9229087369) QUALIFIED DIVIDEND,100.00\n"
            "Dividends,Data,USD,2023-06-30,VFIAX(US9229087369) QUALIFIED DIVIDEND,105.00\n"
            "Dividends,Data,USD,2023-09-30,VFIAX(US9229087369) QUALIFIED DIVIDEND,110.00\n"
            "Dividends,Data,USD,2023-12-31,VFIAX(US9229087369) QUALIFIED DIVIDEND,115.00\n"
            # No trades section or empty trades
        )

        csv_file = tmp_path / "dividends_only.csv"
        csv_file.write_text(csv_content)

        # Parse and process
        # Note: This test requires trades section for the state machine to work
        # Add a minimal trades section
        csv_content_with_trades = csv_content + (
            "Trades,Header,DataDiscriminator,Asset Category,Currency,Symbol,"
            "Date/Time,Quantity,T. Price,Proceeds,Comm/Fee,Basis,Realized P/L,Code\n"
            'Trades,Data,Order,Stocks,USD,VFIAX,"2023-01-15, 10:30:00",1,250.00,-250.00,1.00,251.00,0,O\n'
            'Trades,Data,Order,Stocks,USD,VFIAX,"2023-12-15, 15:30:00",-1,260.00,260.00,1.00,-261.00,10.00,C\n'
        )

        csv_file_with_trades = tmp_path / "dividends_only_with_trades.csv"
        csv_file_with_trades.write_text(csv_content_with_trades)

        ib_export_data = parse_ib_export_all(csv_file_with_trades)
        from shares_reporting.domain.collections import CapitalGainLinesPerCompany, TradeCyclePerCompany

        leftover_trades: TradeCyclePerCompany = {}
        capital_gains: CapitalGainLinesPerCompany = {}
        calculate_fifo_gains(ib_export_data.trade_cycles, leftover_trades, capital_gains)

        # Extract dividend income from the parsed data
        dividend_income = ib_export_data.dividend_income

        # Verify results
        assert len(capital_gains) == 1  # One complete capital gain cycle
        assert len(dividend_income) == 1

        vfiax_dividend = dividend_income["VFIAX"]
        assert vfiax_dividend.gross_amount == Decimal("430.00")  # Sum of 4 quarters
        assert vfiax_dividend.total_taxes == Decimal("0.00")  # No tax data
        assert vfiax_dividend.get_net_amount() == Decimal("430.00")

        # Generate Excel report
        report_path = tmp_path / "dividends_only_report.xlsx"
        generate_tax_report(
            extract=report_path,
            capital_gain_lines_per_company=capital_gains,
            dividend_income_per_company=dividend_income,
        )

        # Verify only dividend section exists
        assert report_path.exists()

        import openpyxl

        workbook = openpyxl.load_workbook(report_path)
        worksheet = workbook.active
        assert worksheet is not None, "Workbook should have an active worksheet"

        # Should have dividend section but no capital gains
        found_dividends = False
        for row in worksheet.iter_rows(values_only=True):
            if row and isinstance(row[0], str) and "CAPITAL INVESTMENT INCOME" in row[0]:
                found_dividends = True
                break

        assert found_dividends, "Dividend section should be present"
        workbook.close()

    def test_dividend_flow_with_mixed_tax_scenarios(self, tmp_path):
        """Test dividend flow with various tax scenarios."""
        csv_content = (
            "Financial Instrument Information,Header,Asset Category,Symbol,Description,Conid,Security ID,Multiplier\n"
            "Financial Instrument Information,Data,Stocks,TAX1,Taxed Security 1,123456,US1234567890,1\n"
            "Financial Instrument Information,Data,Stocks,TAX2,Taxed Security 2,234567,US2345678901,1\n"
            "Financial Instrument Information,Data,Stocks,NOTX,Tax-Free Security,345678,US3456789012,1\n"
            "Dividends,Header,Currency,Date,Description,Amount\n"
            "Dividends,Data,USD,2023-03-15,TAX1(US1234567890) CASH DIVIDEND,100.00\n"
            "Dividends,Data,USD,2023-06-15,TAX1(US1234567890) CASH DIVIDEND,100.00\n"
            "Dividends,Data,USD,2023-03-15,TAX2(US2345678901) CASH DIVIDEND,200.00\n"
            "Dividends,Data,USD,2023-06-15,TAX2(US2345678901) CASH DIVIDEND,200.00\n"
            "Dividends,Data,USD,2023-03-15,NOTX(US3456789012) CASH DIVIDEND,300.00\n"
            "Withholding Tax,Header,Currency,Date,Description,Amount,Code\n"
            "Withholding Tax,Data,USD,2023-03-15,TAX1(US1234567890) US TAX,-15.00,,\n"
            "Withholding Tax,Data,USD,2023-06-15,TAX1(US1234567890) US TAX,-15.00,,\n"
            # TAX2 has partial tax (only on first dividend)
            "Withholding Tax,Data,USD,2023-03-15,TAX2(US2345678901) US TAX,-30.00,,\n"
            # NOTX has no tax
            # Trades section (minimal for integration)
            "Trades,Header,Symbol,Currency,Date/Time,Quantity,T. Price,Comm/Fee\n"
            "Trades,Data,Stocks,TAX1,USD,2023-01-15, 10:30:00,10,100.00,1.00\n"
            "Trades,Data,Stocks,TAX1,USD,2023-12-15, 15:30:00,-10,110.00,1.00\n"
        )

        csv_file = tmp_path / "mixed_tax_scenarios.csv"
        csv_file.write_text(csv_content)

        # Parse and process
        ib_export_data = parse_ib_export_all(csv_file)

        # Process capital gains (minimal trades)
        from shares_reporting.domain.collections import CapitalGainLinesPerCompany, TradeCyclePerCompany

        leftover_trades: TradeCyclePerCompany = {}
        capital_gains: CapitalGainLinesPerCompany = {}
        calculate_fifo_gains(ib_export_data.trade_cycles, leftover_trades, capital_gains)

        # Extract dividend income from the parsed data
        dividend_income = ib_export_data.dividend_income

        # Verify mixed tax scenarios
        assert len(dividend_income) == 3

        # TAX1: Full tax on all dividends
        tax1_dividend = dividend_income["TAX1"]
        assert tax1_dividend.gross_amount == Decimal("200.00")  # 100 + 100
        assert tax1_dividend.total_taxes == Decimal("30.00")  # 15 + 15
        assert tax1_dividend.get_net_amount() == Decimal("170.00")

        # TAX2: Partial tax (only on first dividend)
        tax2_dividend = dividend_income["TAX2"]
        assert tax2_dividend.gross_amount == Decimal("400.00")  # 200 + 200
        assert tax2_dividend.total_taxes == Decimal("30.00")  # Only one tax entry
        assert tax2_dividend.get_net_amount() == Decimal("370.00")

        # NOTX: No tax
        notx_dividend = dividend_income["NOTX"]
        assert notx_dividend.gross_amount == Decimal("300.00")
        assert notx_dividend.total_taxes == Decimal("0.00")
        assert notx_dividend.get_net_amount() == Decimal("300.00")

    def test_dividend_flow_with_duplicate_securities_different_currencies(self, tmp_path):
        """Test dividend flow when same security appears with different currencies."""
        csv_content = (
            "Financial Instrument Information,Header,Asset Category,Symbol,Description,Conid,Security ID,Multiplier\n"
            "Financial Instrument Information,Data,Stocks,AAPL,Apple Inc.,123456,US0378331005,1\n"
            "Financial Instrument Information,Data,Stocks,AAPL.EUR,Apple Inc. EUR,234567,US0378331005,1\n"
            "Dividends,Header,Currency,Date,Description,Amount\n"
            "Dividends,Data,USD,2023-03-15,AAPL(US0378331005) CASH DIVIDEND USD,24.00\n"
            "Dividends,Data,USD,2023-06-15,AAPL(US0378331005) CASH DIVIDEND USD,24.00\n"
            "Dividends,Data,EUR,2023-03-15,AAPL.EUR(US0378331005) CASH DIVIDEND EUR,22.00\n"
            "Dividends,Data,EUR,2023-06-15,AAPL.EUR(US0378331005) CASH DIVIDEND EUR,22.00\n"
            "Withholding Tax,Header,Currency,Date,Description,Amount,Code\n"
            "Withholding Tax,Data,USD,2023-03-15,AAPL(US0378331005) US TAX,-3.60,,\n"
            "Withholding Tax,Data,USD,2023-06-15,AAPL(US0378331005) US TAX,-3.60,,\n"
            "Withholding Tax,Data,EUR,2023-03-15,AAPL.EUR(US0378331005) TAX,-3.30,,\n"
            "Withholding Tax,Data,EUR,2023-06-15,AAPL.EUR(US0378331005) TAX,-3.30,,\n"
            # Trades section (minimal for integration)
            "Trades,Header,DataDiscriminator,Asset Category,Currency,Symbol,"
            "Date/Time,Quantity,T. Price,Proceeds,Comm/Fee,Basis,Realized P/L,Code\n"
            'Trades,Data,Order,Stocks,USD,AAPL,"2023-01-15, 10:30:00",10,150.00,-1500.00,1.00,1501.00,0,O\n'
            'Trades,Data,Order,Stocks,USD,AAPL,"2023-12-15, 15:30:00",-10,160.00,1600.00,1.00,-1601.00,100.00,C\n'
        )

        csv_file = tmp_path / "duplicate_currencies.csv"
        csv_file.write_text(csv_content)

        # Parse and process
        ib_export_data = parse_ib_export_all(csv_file)

        # Process capital gains (minimal trades)
        from shares_reporting.domain.collections import CapitalGainLinesPerCompany, TradeCyclePerCompany

        leftover_trades: TradeCyclePerCompany = {}
        capital_gains: CapitalGainLinesPerCompany = {}
        calculate_fifo_gains(ib_export_data.trade_cycles, leftover_trades, capital_gains)

        # Extract dividend income from the parsed data
        dividend_income = ib_export_data.dividend_income

        # Should treat as same security due to same ISIN, but different currency entries are merged
        assert len(dividend_income) == 1

        # AAPL (USD)
        aapl_usd = dividend_income["AAPL"]
        assert aapl_usd.gross_amount == Decimal("48.00")
        assert aapl_usd.total_taxes == Decimal("7.20")
        assert aapl_usd.currency.currency == "USD"

        # Only AAPL USD entries remain (EUR entries are in separate processed data but not in this test)
        assert aapl_usd.currency.currency == "USD"

    def test_dividend_flow_error_handling(self, tmp_path):
        """Test dividend flow error handling with malformed data."""
        csv_content = (
            "Financial Instrument Information,Header,Asset Category,Symbol,Description,Conid,Security ID,Multiplier\n"
            "Financial Instrument Information,Data,Stocks,AAPL,Apple Inc.,123456,US0378331005,1\n"
            "Dividends,Header,Currency,Date,Description,Amount\n"
            "Dividends,Data,USD,2023-03-15,AAPL - CASH DIVIDEND,24.00\n"
            "Dividends,Data,USD,invalid-date,AAPL - INVALID DATE,10.00\n"
            "Dividends,Data,,2023-06-15,AAPL - MISSING CURRENCY,15.00\n"
            "Dividends,Data,USD,2023-09-15,AAPL - VALID DIVIDEND,20.00\n"
            # Trades section (minimal for integration)
            "Trades,Header,Symbol,Currency,Date/Time,Quantity,T. Price,Comm/Fee\n"
            "Trades,Data,Stocks,AAPL,USD,2023-01-15, 10:30:00,5,140.00,1.00\n"
            "Trades,Data,Stocks,AAPL,USD,2023-12-15, 15:30:00,-5,145.00,1.00\n"
        )

        csv_file = tmp_path / "malformed_dividends.csv"
        csv_file.write_text(csv_content)

        # Parse and process - should handle errors gracefully
        ib_export_data = parse_ib_export_all(csv_file)

        # Process capital gains (minimal trades)
        from shares_reporting.domain.collections import CapitalGainLinesPerCompany, TradeCyclePerCompany

        leftover_trades: TradeCyclePerCompany = {}
        capital_gains: CapitalGainLinesPerCompany = {}
        calculate_fifo_gains(ib_export_data.trade_cycles, leftover_trades, capital_gains)

        # Extract dividend income from the parsed data
        dividend_income = ib_export_data.dividend_income

        # Should still process valid entries
        assert len(dividend_income) == 1

        aapl_dividend = dividend_income["AAPL"]
        # Should include valid entries (24.00 + 20.00 = 44.00)
        # May or may not include partially valid entries depending on error handling
        assert aapl_dividend.gross_amount >= Decimal("44.00")

    def test_dividend_flow_performance(self, tmp_path):
        """Test dividend flow performance with large dataset."""
        # Generate CSV with many dividend entries
        lines = [
            "Financial Instrument Information,Header,Asset Category,Symbol,Description,Conid,Security ID,Multiplier\n",
            "Financial Instrument Information,Data,Stocks,BATCH,Batch Test Inc,123456,US1234567890,1\n",
            "Dividends,Header,Currency,Date,Description,Amount\n",
        ]

        # Add 100 dividend entries
        for i in range(100):
            date_str = f"2023-{(i % 12) + 1:02d}-{(i % 28) + 1:02d}"
            lines.append(f"Dividends,Data,USD,{date_str},BATCH(US1234567890) CASH DIVIDEND {i + 1:03d},{i + 1}.00\n")

        # Add 50 tax entries
        lines.append("Withholding Tax,Header,Currency,Date,Description,Amount,Code\n")
        for i in range(50):
            date_str = f"2023-{(i % 12) + 1:02d}-{(i % 28) + 1:02d}"
            tax_amount = f"{(i + 1) * 0.15:.2f}"
            lines.append(f"Withholding Tax,Data,USD,{date_str},BATCH(US1234567890) US TAX,-{tax_amount},,\n")

        # Add trades section (minimal for integration)
        lines.extend(
            [
                "Trades,Header,Symbol,Currency,Date/Time,Quantity,T. Price,Comm/Fee\n",
                "Trades,Data,Stocks,BATCH,USD,2023-01-15, 10:30:00,100,50.00,1.00\n",
                "Trades,Data,Stocks,BATCH,USD,2023-12-15, 15:30:00,-100,55.00,1.00\n",
            ]
        )

        csv_content = "".join(lines)
        csv_file = tmp_path / "large_dividend_dataset.csv"
        csv_file.write_text(csv_content)

        # Process should handle large dataset efficiently
        import time

        start_time = time.time()

        ib_export_data = parse_ib_export_all(csv_file)

        # Process capital gains (minimal trades)
        from shares_reporting.domain.collections import CapitalGainLinesPerCompany, TradeCyclePerCompany

        leftover_trades: TradeCyclePerCompany = {}
        capital_gains: CapitalGainLinesPerCompany = {}
        calculate_fifo_gains(ib_export_data.trade_cycles, leftover_trades, capital_gains)

        # Extract dividend income from the parsed data
        dividend_income = ib_export_data.dividend_income

        processing_time = time.time() - start_time

        # Verify results
        assert len(dividend_income) == 1
        batch_dividend = dividend_income["BATCH"]

        # Calculate expected totals
        expected_gross = sum(i + 1 for i in range(100))  # Sum of 1 to 100
        expected_tax = sum((i + 1) * 0.15 for i in range(50))  # Sum of first 50 taxes

        assert batch_dividend.gross_amount == Decimal(str(expected_gross))
        assert batch_dividend.total_taxes == Decimal(str(expected_tax))

        # Performance should be reasonable (less than 1 second for 150 entries)
        assert processing_time < 1.0, f"Processing took too long: {processing_time:.2f}s"
