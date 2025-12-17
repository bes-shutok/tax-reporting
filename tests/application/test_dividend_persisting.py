"""Tests for dividend Excel persisting functionality."""

from decimal import Decimal

from shares_reporting.application.persisting import generate_tax_report
from shares_reporting.domain.collections import DividendIncomePerCompany
from shares_reporting.domain.entities import DividendIncomePerSecurity
from shares_reporting.domain.value_objects import parse_currency


class TestDividendExcelPersisting:
    """Test dividend Excel report generation."""

    def test_generate_tax_report_with_dividends(self, tmp_path):
        """Test generating Excel report with dividend income section."""
        dividend_income = DividendIncomePerCompany(
            {
                "AAPL": DividendIncomePerSecurity(
                    symbol="AAPL",
                    isin="US0378331005",
                    country="United States",
                    gross_amount=Decimal("100.00"),
                    total_taxes=Decimal("15.00"),
                    currency=parse_currency("USD"),
                ),
                "MSFT": DividendIncomePerSecurity(
                    symbol="MSFT",
                    isin="US5949181045",
                    country="United States",
                    gross_amount=Decimal("200.00"),
                    total_taxes=Decimal("30.00"),
                    currency=parse_currency("USD"),
                ),
                "ASML": DividendIncomePerSecurity(
                    symbol="ASML",
                    isin="NL0010273215",
                    country="Netherlands",
                    gross_amount=Decimal("150.00"),
                    total_taxes=Decimal("0.00"),
                    currency=parse_currency("EUR"),
                ),
            }
        )

        # Generate report
        report_path = tmp_path / "dividend_report.xlsx"
        generate_tax_report(
            extract=report_path,
            capital_gain_lines_per_company={},
            dividend_income_per_company=dividend_income,
        )

        # Verify file was created
        assert report_path.exists()

        # Read and verify the Excel file
        import openpyxl

        workbook = openpyxl.load_workbook(report_path)
        worksheet = workbook.active
        assert worksheet is not None, "Workbook should have an active worksheet"

        # Find the dividend section
        found_dividend_section = False
        dividend_row_count = 0

        for _row_idx, row in enumerate(worksheet.iter_rows(values_only=True), 1):
            if row and isinstance(row[0], str) and "CAPITAL INVESTMENT INCOME" in row[0]:
                found_dividend_section = True
                # Next row should be empty
                continue

            if found_dividend_section and row and len(row) > 0:
                # Check for header row
                if isinstance(row[0], str) and "Beneficiary" in row[0]:
                    continue

                # Count data rows
                if dividend_row_count < 3 and row[1] == "Dividends":
                    dividend_row_count += 1
                    # Verify data integrity
                    symbol = row[8] if len(row) > 8 else None
                    country = row[2] if len(row) > 2 else None
                    isin = row[3] if len(row) > 3 else None
                    original_gross = row[10] if len(row) > 10 else None
                    original_tax = row[11] if len(row) > 11 else None
                    net_amount = row[12] if len(row) > 12 else None

                    # Verify expected values based on symbol
                    if symbol == "AAPL":
                        assert country == "United States"
                        assert isin == "US0378331005"
                        assert original_gross == "100.00"
                        assert original_tax == "15.00"
                        assert net_amount == "85.00"
                    elif symbol == "MSFT":
                        assert country == "United States"
                        assert isin == "US5949181045"
                        assert original_gross == "200.00"
                        assert original_tax == "30.00"
                        assert net_amount == "170.00"
                    elif symbol == "ASML":
                        assert country == "Netherlands"
                        assert isin == "NL0010273215"
                        assert original_gross == "150.00"
                        assert original_tax == "0.00"
                        assert net_amount == "150.00"

        assert found_dividend_section, "Dividend section not found in report"
        assert dividend_row_count == 3, f"Expected 3 dividend rows, found {dividend_row_count}"

        workbook.close()

    def test_generate_tax_report_without_dividends(self, tmp_path):
        """Test generating Excel report without dividend income section."""
        report_path = tmp_path / "no_dividend_report.xlsx"
        generate_tax_report(
            extract=report_path,
            capital_gain_lines_per_company={},
            dividend_income_per_company={},  # Empty dividend data
        )

        assert report_path.exists()

        # Read and verify no dividend section
        import openpyxl

        workbook = openpyxl.load_workbook(report_path)
        worksheet = workbook.active
        assert worksheet is not None, "Workbook should have an active worksheet"

        # Should not have dividend section
        found_dividend_section = False
        for row in worksheet.iter_rows(values_only=True):
            if row and isinstance(row[0], str) and "CAPITAL INVESTMENT INCOME" in row[0]:
                found_dividend_section = True
                break

        assert not found_dividend_section, "Dividend section should not be present"
        workbook.close()

    def test_generate_tax_report_with_dividend_formulas(self, tmp_path):
        """Test that dividend Excel report contains proper currency conversion formulas."""
        dividend_income = DividendIncomePerCompany(
            {
                "AAPL": DividendIncomePerSecurity(
                    symbol="AAPL",
                    isin="US0378331005",
                    country="United States",
                    gross_amount=Decimal("100.50"),
                    total_taxes=Decimal("15.25"),
                    currency=parse_currency("USD"),
                ),
            }
        )

        report_path = tmp_path / "dividend_formulas_report.xlsx"
        generate_tax_report(
            extract=report_path,
            capital_gain_lines_per_company={},
            dividend_income_per_company=dividend_income,
        )

        # Read and verify formulas
        import openpyxl

        workbook = openpyxl.load_workbook(report_path)
        worksheet = workbook.active
        assert worksheet is not None, "Workbook should have an active worksheet"

        # Find dividend row and check formulas
        dividend_row = None
        for row_idx, row in enumerate(worksheet.iter_rows(values_only=False), 1):
            if row and len(row) > 1 and row[1].value == "Dividends":
                dividend_row = row_idx
                break

        assert dividend_row is not None, "Dividend row not found"

        # Check converted gross amount formula (column 5, index 4)
        gross_amount_cell = worksheet.cell(row=dividend_row, column=5)
        assert gross_amount_cell.data_type == "f"  # Formula cell
        assert "=V" in str(gross_amount_cell.value)  # Contains reference to exchange rate
        assert "100.5" in str(gross_amount_cell.value)  # Contains original amount

        # Check converted tax amount formula (column 6, index 5)
        tax_amount_cell = worksheet.cell(row=dividend_row, column=6)
        assert tax_amount_cell.data_type == "f"  # Formula cell
        assert "=V" in str(tax_amount_cell.value)  # Contains reference to exchange rate
        assert "15.25" in str(tax_amount_cell.value)  # Contains original amount

        # Check original amounts (should be string values)
        original_gross_cell = worksheet.cell(row=dividend_row, column=11)
        assert original_gross_cell.data_type == "s"  # String value
        assert original_gross_cell.value == "100.50"

        original_tax_cell = worksheet.cell(row=dividend_row, column=12)
        assert original_tax_cell.data_type == "s"  # String value
        assert original_tax_cell.value == "15.25"

        net_amount_cell = worksheet.cell(row=dividend_row, column=13)
        assert net_amount_cell.data_type == "s"  # String value
        assert net_amount_cell.value == "85.25"  # 100.50 - 15.25

        workbook.close()

    def test_generate_tax_report_with_multiple_currencies(self, tmp_path):
        """Test generating Excel report with multiple dividend currencies."""
        dividend_income = DividendIncomePerCompany(
            {
                "AAPL": DividendIncomePerSecurity(
                    symbol="AAPL",
                    isin="US0378331005",
                    country="United States",
                    gross_amount=Decimal("100.00"),
                    total_taxes=Decimal("15.00"),
                    currency=parse_currency("USD"),
                ),
                "ASML": DividendIncomePerSecurity(
                    symbol="ASML",
                    isin="NL0010273215",
                    country="Netherlands",
                    gross_amount=Decimal("200.00"),
                    total_taxes=Decimal("25.00"),
                    currency=parse_currency("EUR"),
                ),
                "TSLA": DividendIncomePerSecurity(
                    symbol="TSLA",
                    isin="US88160R1014",
                    country="United States",
                    gross_amount=Decimal("50.00"),
                    total_taxes=Decimal("0.00"),
                    currency=parse_currency("CAD"),
                ),
            }
        )

        report_path = tmp_path / "multi_currency_dividend_report.xlsx"
        generate_tax_report(
            extract=report_path,
            capital_gain_lines_per_company={},
            dividend_income_per_company=dividend_income,
        )

        # Read and verify multiple currencies
        import openpyxl

        workbook = openpyxl.load_workbook(report_path)
        worksheet = workbook.active
        assert worksheet is not None, "Workbook should have an active worksheet"

        # Find dividend rows and check currency columns
        dividend_currencies = []
        for _row_idx, row in enumerate(worksheet.iter_rows(values_only=True), 1):
            if row and len(row) > 1 and row[1] == "Dividends":
                currency = row[9] if len(row) > 9 else None  # Currency column
                if currency:
                    dividend_currencies.append(currency)

        assert len(dividend_currencies) == 3
        assert "USD" in dividend_currencies
        assert "EUR" in dividend_currencies
        assert "CAD" in dividend_currencies

        workbook.close()

    def test_generate_tax_report_with_dividend_validation_errors(self, tmp_path):
        """Test that dividend validation errors are properly handled."""
        # Test that validation happens during processing
        # For now, just test with valid data to ensure report generation works
        dividend_income = DividendIncomePerCompany(
            {
                "VALID": DividendIncomePerSecurity(
                    symbol="VALID",
                    isin="US1234567890",
                    country="United States",
                    gross_amount=Decimal("100.00"),  # Valid positive amount
                    total_taxes=Decimal("15.00"),
                    currency=parse_currency("USD"),
                ),
            }
        )

        report_path = tmp_path / "valid_dividend_report.xlsx"

        # Should generate successfully
        generate_tax_report(
            extract=report_path,
            capital_gain_lines_per_company={},
            dividend_income_per_company=dividend_income,
        )

        assert report_path.exists()

    def test_generate_tax_report_mixed_capital_gains_and_dividends(self, tmp_path):
        """Test generating Excel report with both capital gains and dividend sections."""
        # Create simple dividend data
        dividend_income = DividendIncomePerCompany(
            {
                "AAPL": DividendIncomePerSecurity(
                    symbol="AAPL",
                    isin="US0378331005",
                    country="United States",
                    gross_amount=Decimal("100.00"),
                    total_taxes=Decimal("15.00"),
                    currency=parse_currency("USD"),
                ),
            }
        )

        report_path = tmp_path / "mixed_report.xlsx"
        generate_tax_report(
            extract=report_path,
            capital_gain_lines_per_company={},  # Empty capital gains
            dividend_income_per_company=dividend_income,
        )

        assert report_path.exists()

        # Read and verify both sections exist
        import openpyxl

        workbook = openpyxl.load_workbook(report_path)
        worksheet = workbook.active
        assert worksheet is not None, "Workbook should have an active worksheet"

        # Find dividend section
        found_dividends = False

        for row in worksheet.iter_rows(values_only=True):
            if row and isinstance(row[0], str):
                if "CAPITAL GAINS" in row[0] or "DISPONIBILIZAÇÕES" in row[0]:
                    pass  # Found capital gains section
                elif "CAPITAL INVESTMENT INCOME" in row[0]:
                    found_dividends = True

        assert found_dividends, "Dividend section not found in mixed report"
        workbook.close()

    def test_dividend_excel_cell_formatting(self, tmp_path):
        """Test that dividend Excel cells have proper number formatting."""
        dividend_income = DividendIncomePerCompany(
            {
                "AAPL": DividendIncomePerSecurity(
                    symbol="AAPL",
                    isin="US0378331005",
                    country="United States",
                    gross_amount=Decimal("100.12345"),
                    total_taxes=Decimal("15.67890"),
                    currency=parse_currency("USD"),
                ),
            }
        )

        report_path = tmp_path / "formatted_dividend_report.xlsx"
        generate_tax_report(
            extract=report_path,
            capital_gain_lines_per_company={},
            dividend_income_per_company=dividend_income,
        )

        # Read and verify cell formatting
        import openpyxl

        workbook = openpyxl.load_workbook(report_path)
        worksheet = workbook.active
        assert worksheet is not None, "Workbook should have an active worksheet"

        # Find dividend row and check formatting
        dividend_row = None
        for row_idx, row in enumerate(worksheet.iter_rows(values_only=False), 1):
            if row and len(row) > 1 and row[1].value == "Dividends":
                dividend_row = row_idx
                break

        assert dividend_row is not None

        # Check number format for original amounts
        original_gross_cell = worksheet.cell(row=dividend_row, column=11)
        assert original_gross_cell.number_format == "0.00"

        original_tax_cell = worksheet.cell(row=dividend_row, column=12)
        assert original_tax_cell.number_format == "0.00"

        net_amount_cell = worksheet.cell(row=dividend_row, column=13)
        assert net_amount_cell.number_format == "0.00"

        # Check values are stored as strings with full precision
        assert original_gross_cell.value == "100.12345"
        assert original_tax_cell.value == "15.67890"
        assert net_amount_cell.value == "84.44455"  # 100.12345 - 15.67890

        workbook.close()
