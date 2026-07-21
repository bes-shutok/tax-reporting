"""End-to-end data-trace verification for the CRG-021 + CRG-022 reward partitions.

Verifies (per docs/history/plans/2026-07-18-crypto-dust-partition-fee-skip.md
Task 7 + docs/history/plans/2026-07-19-deferred-reward-dust-skip.md Task 6)
that the committed synthetic Koinly 2025 example fixture in
``resources/source/example/2025/koinly/dust-partition/`` drives the production
``load_koinly_crypto_report`` -> ``generate_tax_report`` pipeline such that:

- (CRG-021, taxable-now side) the Crypto Supplementary tab renders the
  per-(asset, wallet) dust summary block for zero-value taxable-now rewards on
  priced assets, while zero-value rewards on assets with NO priced row anywhere
  in the export keep their per-row ``YES`` review flag in the Section 2 detail
  table.
- (CRG-022, deferred side) zero-value DEFERRED_BY_LAW reward rows are routed at
  parse time into ``skipped_zero_value_deferred_rewards`` and rendered in the
  single Section 3 "Suppressed zero-value deferred rewards" block, with a
  per-row reason distinguishing ``dust`` (priced-asset rounding artifact) from
  ``unpriced`` (no Koinly price feed).

These are full-pipeline pins of the shared priced-asset discriminator
(``_priced_assets_in_export``): they exercise ``write_crypto_supplementary_sheet``
end-to-end through the same ``generate_tax_report`` entry point the production
CLI uses. There is NO RED phase for the e2e tier because the production pipeline
is already GREEN from the unit-tier TDD tasks; the e2e tests pass directly.

Fixture asset choice. The CRG-021 plan's Task 7 prose names ``BTC`` (priced
asset) and ``OSBGT`` (unpriced asset) as the motivating examples, mirroring the
unit fixtures in ``test_crypto_supplementary_sheet.py``. The taxable-now dust
partition operates exclusively on ``taxable_now_entries``, so only
fiat-denominated rewards can reach it through the full pipeline (every
crypto-denominated ticker routes to DEFERRED_BY_LAW via
``_classify_reward_tax_status``). The CRG-021 taxable-now fixture therefore
uses the fiat currency codes ``EUR`` (priced asset) and ``USD`` (unpriced
asset). The CRG-022 deferred-side extension (Task 6) adds crypto-denominated
rows directly: ``SOL`` (priced deferred - has a non-zero SOL row so its
zero-value row routes to dust) and ``OSBGT`` (unpriced deferred - no priced
OSBGT row anywhere so its zero-value row routes to unpriced). Both SOL and
OSBGT are in ``popular_crypto_tokens.json``, so they survive the parse-time
``is_known`` gate and reach the deferred-skip path.

Fixture contents (one scenario, six income rows + one CG USD row):
- Income row 1: ``EUR`` priced (``Value (EUR) = 3,00``, Wirex) -> TAXABLE_NOW,
  Detail path (priced asset, not zero).
- Income row 2: ``EUR`` zero-value (``Value (EUR) = 0,00``, Wirex) ->
  TAXABLE_NOW, Dust path (EUR has the priced row above; zero collapses to the
  per-(EUR, Wirex) dust summary line).
- Income row 3: ``USD`` zero-value (``Value (EUR) = 0,00``, Wirex) ->
  TAXABLE_NOW, per-row YES path (USD has NO priced income row anywhere; the
  priced USD CG row keeps USD in ``known_assets`` so the zero-value income
  reward is ``is_known`` and not skipped at parse, but the dust discriminator
  only consults ``reward_entries``, where USD has no priced row).
- Income row 4: ``SOL`` priced (``Value (EUR) = 5,00``, Wirex) ->
  DEFERRED_BY_LAW, retained in ``reward_entries`` (value_eur > 0, not skipped);
  this is the priced anchor that makes SOL "priced" so the zero-value SOL row
  routes to dust.
- Income row 5: ``SOL`` zero-value (``Value (EUR) = 0,00``, Wirex) ->
  DEFERRED_BY_LAW, parse-time skip into ``skipped_zero_value_deferred_rewards``
  with reason "dust" (SOL is priced by row 4).
- Income row 6: ``OSBGT`` zero-value (``Value (EUR) = 0,00``, Wirex) ->
  DEFERRED_BY_LAW, parse-time skip into
  ``skipped_zero_value_deferred_rewards`` with reason "unpriced" (no priced
  OSBGT row anywhere in the export).
- CG row: one ``USD`` disposal (cost 5,00, proceeds 6,00, gain 1,00, Wirex) so
  USD enters ``known_assets`` via the CG scan in ``_collect_known_asset_tickers``
  and the zero-value USD income reward is retained (not routed to
  ``skipped_zero_value_tokens``).

Wallets are limited to ``SYNTHETIC_WALLET_ALLOWLIST`` (``Wirex``); all
TxHash/TxSrc/TxDest cells empty (synthetic-data hygiene, scoped by
``test_example_data_is_synthetic``).
"""

from __future__ import annotations

import re
from decimal import Decimal
from pathlib import Path

import openpyxl
import pytest

from tax_reporting.application.crypto_reporting import (
    RewardTaxClassification,
    load_koinly_crypto_report,
)
from tax_reporting.application.persisting import generate_tax_report
from tests.conftest import build_koinly_jurisdiction

pytestmark = pytest.mark.e2e

_FIXTURE_DIR = Path("resources/source/example/2025/koinly/dust-partition")

# Phase E Task 5: mirror test_example_report_generation._EXAMPLE_JURISDICTION.
# ``exclude_loan_repayment_gains=False`` avoids the loan-asset FIFO rebuild
# (no loan-affected assets in this fixture; the rebuild would emit a benign
# warning but no entries). Income-code classification stays ON so the
# taxable-now EUR/USD interest rewards resolve to the official Tabela V E25
# code (a mandatory Quadro 8A field; without it the aggregation step
# fail-closes).
_JURISDICTION = build_koinly_jurisdiction(exclude_loan_repayment_gains=False)

# Fixture-derived constants (CRG-021 Invariant 1: the partition is a view; the
# expected values are derived directly from the committed fixture rows, NOT
# from a "partition-disabled baseline" which does not exist). Constant derivation
# order (r3 finding #3): the fixture CSV was extended FIRST, then the constants
# below were derived from the literal ``Value (EUR)`` cells (NOT from
# ``report.reconciliation.reward_total_eur`` at test time).
_PRICED_ASSET = "EUR"
_UNPRICED_ASSET = "USD"
# Deferred-side fixtures (CRG-022): SOL is a priced deferred asset (has a
# non-zero SOL row in the export) so its zero-value row routes to DUST; OSBGT
# has NO priced row anywhere so its zero-value row routes to UNPRICED.
_DEFERRED_PRICED_ASSET = "SOL"
_DEFERRED_UNPRICED_ASSET = "OSBGT"
_PRICED_VALUE_EUR = Decimal("3.00")  # EUR priced row Value (EUR) cell
_SOL_PRICED_VALUE_EUR = Decimal("5.00")  # SOL priced row Value (EUR) cell
# CG row that keeps USD in known_assets (proceeds > 0 in the CG scan).
_CG_USD_PROCEEDS_EUR = Decimal("6.00")
# reward_total_eur: EUR priced (3.00) + EUR zero (0.00) + USD zero (0.00) +
# SOL priced (5.00) = 8.00. The zero-value rows (EUR, USD, SOL zero, OSBGT)
# contribute nothing. Updated by Task 6 (was 3.00 in the predecessor plan
# because the fixture had no SOL priced row).
_EXPECTED_REWARD_TOTAL_EUR = _PRICED_VALUE_EUR + _SOL_PRICED_VALUE_EUR  # 8.00
# taxable-now total = sum over taxable-now rows only (EUR/USD are fiat). SOL
# is DEFERRED_BY_LAW so its 5.00 EUR value stays out of taxable-now. This is
# the constant the existing taxable-now test (CRG-021) is pinned to; introduced
# as a separate constant (r3 finding #8) so the deferred-skip test cannot
# silently shift the taxable-now pin.
_EXPECTED_TAXABLE_NOW_TOTAL_EUR = _PRICED_VALUE_EUR  # 3.00
# reward_entries length: 4 (EUR priced + EUR zero + USD zero + SOL priced).
# Zero-value SOL and OSBGT are routed to ``skipped_zero_value_deferred_rewards``
# at parse time (CRG-022) and do NOT appear in ``reward_entries``.
_EXPECTED_REWARD_ENTRIES_LEN = 4  # EUR priced + EUR zero + USD zero + SOL priced
_EXPECTED_DETAIL_ROWS = 2  # EUR priced + USD unpriced-YES (zero-EUR is dust)
_EXPECTED_DUST_ROWS = 1  # EUR zero-value (EUR priced elsewhere)
# Deferred-skip audit-list length: zero SOL (dust) + zero OSBGT (unpriced) = 2.
# ``_EXPECTED_SKIPPED_DEFERRED_LEN`` pins Invariant 1 at the e2e tier.
_EXPECTED_SKIPPED_DEFERRED_LEN = 2
# Separate constant for the deferred-skip tax-math pin (r3 finding #8 - NOT the
# same as ``_EXPECTED_REWARD_TOTAL_EUR``; the deferred-skip Invariant 2 pin
# asserts reward_total_eur == 8.00 because the parse-time skip removes only
# zero-value rows that contributed nothing).
_EXPECTED_DEFERRED_SKIP_REWARD_TOTAL_EUR = _EXPECTED_REWARD_TOTAL_EUR  # 8.00
_EXPECTED_DEFERRED_SKIP_TAXABLE_NOW_TOTAL_EUR = _EXPECTED_TAXABLE_NOW_TOTAL_EUR  # 3.00
_EXPECTED_DEFERRED_SKIP_DEFERRED_TOTAL_EUR = _SOL_PRICED_VALUE_EUR  # 5.00
# Reporting OTHER CAPITAL INVESTMENT INCOME line pinned value: only the 3.00 EUR
# priced taxable-now row feeds the aggregate (the zero-value EUR/USD taxable-now
# rows and the priced SOL deferred row contribute nothing to aggregate_taxable_rewards
# - zeros add nothing, and deferred rewards are filtered out). Wirex resolves to
# operator entity ``GB`` so the income aggregates to the GB row with the official
# E25 description.
_EXPECTED_DEFERRED_SKIP_REPORTING_GROSS_EUR = _PRICED_VALUE_EUR  # 3.00
_EXPECTED_DEFERRED_SKIP_REPORTING_NET_EUR = _PRICED_VALUE_EUR  # 3.00 (no foreign tax)
_EXPECTED_DEFERRED_SKIP_REPORTING_RAW_ROW_COUNT = 3  # EUR priced + EUR zero + USD zero


def _load_report():
    """Load the synthetic dust-partition fixture through the production pipeline."""
    report = load_koinly_crypto_report(_FIXTURE_DIR, jurisdiction=_JURISDICTION)
    assert report is not None, (
        f"load_koinly_crypto_report returned None for {_FIXTURE_DIR}; "
        "the committed fixture is missing or the three-file presence guard failed."
    )
    return report


def _generate_workbook(tmp_path: Path) -> openpyxl.Workbook:
    """Run the full production pipeline and return the loaded workbook.

    ``generate_tax_report`` is the same entry point the production CLI uses
    (``write_crypto_supplementary_sheet`` is invoked from ``workbook_builder``
    when ``crypto_tax_report`` is non-None). Empty capital-gains and dividend
    inputs are passed because the dust-partition tests are rewards-only.
    """
    report = _load_report()
    output_path = tmp_path / "dust_partition.xlsx"
    generate_tax_report(output_path, {}, None, crypto_tax_report=report)
    assert output_path.exists(), "generate_tax_report did not produce a workbook"
    return openpyxl.load_workbook(output_path)


def _supplementary_first_column(wb: openpyxl.Workbook) -> list[str | None]:
    """Return the first-column cell values from the Crypto Supplementary tab."""
    ws = wb["Crypto Supplementary"]
    return [
        (row[0] if row else None)
        for row in ws.iter_rows(values_only=True)
    ]


def _section2_detail_rows(wb: openpyxl.Workbook) -> list[dict[str, object]]:
    """Return the Section 2 reward-detail data rows keyed by header.

    Scans the Crypto Supplementary tab for the ``2. TAXABLE-NOW - SUPPORT
    DETAIL`` section header, then reads the header row and each subsequent
    data row until the Dust summary block or the next section starts. Returns a
    list of dicts keyed by the 11 reward-detail headers
    (``_REWARD_DETAIL_HEADERS`` in ``crypto_supplementary_sheet``).
    """
    ws = wb["Crypto Supplementary"]
    rows = list(ws.iter_rows(values_only=True))
    in_section = False
    header: list[str | None] = []
    detail: list[dict[str, object]] = []
    for row in rows:
        first = row[0] if row else None
        if isinstance(first, str) and first.startswith("2. TAXABLE-NOW"):
            in_section = True
            continue
        if not in_section:
            continue
        # Within Section 2 the detail block runs from the header row up to the
        # "Dust summary:" line (or the next section, whichever comes first).
        if isinstance(first, str) and (first.startswith("3. ") or first == "Dust summary:"):
            break
        if not header:
            # First non-empty row after the section header + note is the column
            # header row. Skip the italic note (single-cell row 1 value).
            if first == "Date" or (isinstance(first, str) and "Date" in first):
                header = [str(c) if c is not None else "" for c in row]
            continue
        # Data row: first cell is an ISO date.
        if isinstance(first, str) and re.match(r"^\d{4}-\d{2}-\d{2}", first):
            entry: dict[str, object] = {}
            for idx, col_name in enumerate(header):
                if idx < len(row) and col_name:
                    entry[col_name] = row[idx]
            detail.append(entry)
    return detail


def _reconciliation_map(wb: openpyxl.Workbook) -> dict[str, object]:
    """Return the Section 4 reconciliation key-value pairs as a dict."""
    ws = wb["Crypto Supplementary"]
    rows = list(ws.iter_rows(values_only=True))
    result: dict[str, object] = {}
    in_section = False
    for row in rows:
        first = row[0] if row else None
        if isinstance(first, str) and first.startswith("4. REWARDS CLASSIFICATION RECONCILIATION"):
            in_section = True
            continue
        if not in_section:
            continue
        if isinstance(first, str) and first.startswith("5. "):
            break
        if isinstance(first, str) and first:
            value = row[1] if len(row) > 1 else None
            result[first] = value
    return result


def _suppressed_deferred_table_rows(
    wb: openpyxl.Workbook, asset: str
) -> list[dict[str, object]]:
    """Return the rows of the Section 3 "Suppressed zero-value deferred rewards"
    column tables whose Asset column matches ``asset``.

    The block has two column tables (Deferred dust / Deferred unpriced), each
    with columns: Asset | Wallet | Rows | Summed amount | Category. This helper
    walks both tables and returns the rows for ``asset`` (regardless of which
    bucket - callers that care about the bucket check ``row["category"]``).

    Each returned row is a dict with keys: asset, wallet, rows, summed_amount
    (a string formatted :.8f), category (``"dust"`` or ``"unpriced"``).
    """
    ws = wb["Crypto Supplementary"]
    out: list[dict[str, object]] = []
    in_block = False
    in_table = False
    for row in ws.iter_rows(values_only=True):
        first = row[0] if row else None
        if not isinstance(first, str):
            in_table = False
            continue
        if first == "Suppressed zero-value deferred rewards":
            in_block = True
            continue
        if not in_block:
            continue
        # End the block at the next section header.
        if first.startswith("4. ") or first.startswith("5. "):
            break
        # Sub-headers toggle which table we're in; the row after a sub-header
        # is the column-table header ("Asset" | ...) - skip it.
        if first.startswith(("Deferred dust", "Deferred unpriced")):
            in_table = True
            continue
        if not in_table:
            continue
        # The column-table header row reads "Asset" in column A - skip it.
        if first == "Asset":
            continue
        if first == asset:
            out.append({
                "asset": row[0],
                "wallet": row[1],
                "rows": row[2],
                "summed_amount": row[3],
                "category": row[4],
            })
    return out


def _dust_summary_table_rows(
    wb: openpyxl.Workbook, asset: str
) -> list[dict[str, object]]:
    """Return the rows of the Section 2 "Dust summary:" column table whose
    Asset column matches ``asset``.

    The taxable-now dust block has one column table with columns:
    Asset | Wallet | Rows | Summed Value (EUR) | Category. Taxable-now sums
    ``value_eur`` (EUR), not native amount (asymmetric with the deferred side).

    Each returned row is a dict with keys: asset, wallet, rows,
    summed_value_eur (a float), category (always ``"dust"`` for taxable-now).
    """
    ws = wb["Crypto Supplementary"]
    out: list[dict[str, object]] = []
    in_block = False
    in_table = False
    for row in ws.iter_rows(values_only=True):
        first = row[0] if row else None
        if not isinstance(first, str):
            in_table = False
            continue
        if first == "Dust summary:":
            in_block = True
            continue
        if not in_block:
            continue
        # End the block at the next section header.
        if first.startswith(("2. ", "3. ", "4. ", "5. ")):
            break
        # The sub-header "Taxable-now dust (...)" marks the start of the table;
        # the row after it is the column-table header ("Asset" | ...) - skip it.
        if first.startswith("Taxable-now dust"):
            in_table = True
            continue
        if not in_table:
            continue
        if first == "Asset":
            continue
        if first == asset:
            out.append({
                "asset": row[0],
                "wallet": row[1],
                "rows": row[2],
                "summed_value_eur": row[3],
                "category": row[4],
            })
    return out


def _reporting_other_capital_investment_income_row(
    wb: openpyxl.Workbook,
) -> tuple[object, ...] | None:
    """Return the Reporting worksheet's OTHER CAPITAL INVESTMENT INCOME data row.

    Mirrors the line-locator pattern at
    ``test_example_report_generation.py:390-399``: scan the Reporting sheet for
    the ``"OTHER CAPITAL INVESTMENT INCOME"`` subsection header, then return the
    single aggregated reward data row that follows it (row at subsection_idx+1).
    Returns ``None`` when the subsection or its data row is absent.
    """
    ws = wb["Reporting"]
    rows = list(ws.iter_rows(values_only=True))
    for idx, row in enumerate(rows):
        if row and isinstance(row[0], str) and "OTHER CAPITAL INVESTMENT INCOME" in row[0]:
            if idx + 1 < len(rows):
                return rows[idx + 1]
            return None
    return None


def _crypto_reconciliation_map(wb: openpyxl.Workbook) -> dict[str, object]:
    """Return the Crypto Reconciliation sheet's key-value pairs as a dict.

    Scans the Reconciliation sheet (Section 3) for label/value pairs; stops at
    the SKIPPED ZERO VALUE TOKENS section header. Used to assert the
    cross-sheet ``Skipped zero-value deferred rewards (audit)`` count.
    """
    ws = wb["Crypto Reconciliation"]
    result: dict[str, object] = {}
    for row in ws.iter_rows(values_only=True):
        first = row[0] if row else None
        if not isinstance(first, str) or not first:
            continue
        if first.startswith("4. SKIPPED"):
            break
        if first.startswith("3. RECONCILIATION"):
            continue
        # The skipped-tokens table header is "Source section" / "Asset"; stop
        # before those rows.
        if first in {"Source section"}:
            break
        value = row[1] if len(row) > 1 else None
        result[first] = value
    return result


class TestCryptoDustPartitionE2E:
    """E2E pin for the CRG-021 dust partition via the full production pipeline."""

    def test_priced_asset_zero_collapses_to_dust_via_full_pipeline(self, tmp_path: Path) -> None:
        """Zero-value taxable-now reward on a priced asset collapses to the dust summary block.

        Given the fixture with one priced EUR reward (3.00 EUR) and one
        zero-value EUR reward, the Crypto Supplementary tab MUST render a
        ``Dust summary:`` block containing exactly one EUR row in its column
        table, and the Section 4 reconciliation MUST split into detail/dust
        counts (the zero-value EUR row counts as dust, not detail).
        """
        wb = _generate_workbook(tmp_path)
        first_col = _supplementary_first_column(wb)

        # The dust summary block header is present and the EUR row appears in
        # its column table with Rows=1, Summed Value (EUR)=0.0, Category=dust.
        assert "Dust summary:" in first_col, (
            "Dust summary header missing from Crypto Supplementary; "
            "the zero-value priced-asset reward did not route to dust."
        )
        eur_dust_rows = _dust_summary_table_rows(wb, asset=_PRICED_ASSET)
        assert len(eur_dust_rows) == 1, (
            f"Expected one {_PRICED_ASSET} row in the dust summary table; "
            f"got {eur_dust_rows}; first column: {first_col}"
        )
        row = eur_dust_rows[0]
        assert row["rows"] == 1, f"Rows count wrong: {row}"
        assert float(row["summed_value_eur"]) == 0.0, f"Summed Value (EUR) wrong: {row}"
        assert row["category"] == "dust", f"Category wrong: {row}"

        # Section 4 reconciliation shows the split counts.
        recon = _reconciliation_map(wb)
        assert recon.get("Taxable-now detail rows") == _EXPECTED_DETAIL_ROWS, (
            f"Expected Taxable-now detail rows = {_EXPECTED_DETAIL_ROWS}, "
            f"got {recon.get('Taxable-now detail rows')!r}; recon: {recon}"
        )
        assert recon.get("Taxable-now dust rows (suppressed from detail)") == _EXPECTED_DUST_ROWS, (
            f"Expected Taxable-now dust rows = {_EXPECTED_DUST_ROWS}, "
            f"got {recon.get('Taxable-now dust rows (suppressed from detail)')!r}; recon: {recon}"
        )
        wb.close()

    def test_unpriced_asset_zero_keeps_per_row_yes_via_full_pipeline(self, tmp_path: Path) -> None:
        """Zero-value taxable-now reward on an asset with no priced income row keeps per-row YES.

        Given the fixture's zero-value USD reward (USD has NO priced row in
        ``reward_entries``; the priced USD CG row keeps USD in
        ``known_assets`` so the income reward is retained, not skipped), the
        USD row MUST appear in the Section 2 detail table with its ``Review
        flag`` cell starting ``YES:`` (full-pipeline version of the r9 headline
        fix pinned at the helper boundary by
        ``TestPartitionTaxableNow#test_unpriced_asset_zero_stays_in_real_rows``).
        """
        wb = _generate_workbook(tmp_path)
        detail = _section2_detail_rows(wb)

        usd_rows = [r for r in detail if str(r.get("Asset", "")) == _UNPRICED_ASSET]
        assert usd_rows, (
            f"Zero-value {_UNPRICED_ASSET} reward should appear in the Section 2 detail "
            f"table (unpriced asset -> per-row YES, NOT dust); detail rows: {detail}"
        )
        review_flag = str(usd_rows[0].get("Review flag", ""))
        assert review_flag.startswith("YES:"), (
            f"Unpriced-asset zero-value reward review flag must start 'YES:' (r9 headline fix); "
            f"got {review_flag!r}. USD detail row: {usd_rows[0]}"
        )

        # Sanity: the priced EUR row also stays in the detail table with NO
        # review flag (it is the priced anchor that makes EUR "priced").
        eur_rows = [r for r in detail if str(r.get("Asset", "")) == _PRICED_ASSET]
        assert eur_rows, f"Priced {_PRICED_ASSET} reward should appear in Section 2 detail; rows: {detail}"
        assert str(eur_rows[0].get("Review flag", "")) == "NO", (
            f"Priced {_PRICED_ASSET} reward should carry review flag 'NO' (no review needed); "
            f"got {eur_rows[0].get('Review flag')!r}"
        )
        wb.close()

    def test_reward_entries_unchanged_by_partition(self, tmp_path: Path) -> None:
        """Invariant 1 (taxable-now side): the partition does not mutate reward_entries or totals.

        Given the fixture, the production pipeline's ``reward_entries`` length
        and the reconciliation totals MUST match the values derived directly
        from the fixture rows. There is no runtime flag to disable the
        partition, so there is no "partition-disabled baseline" to compare
        against (CRG-021 A2); the invariant is that the taxable-now partition
        is a VIEW over the taxable-now subset of ``reward_entries`` and never
        mutates the list, the ``taxable_now_total_eur``, or the
        ``reward_total_eur``. NOTE (Task 6 / CRG-022): the deferred side is
        NOT a view - zero-value deferred rows are skipped at parse time into
        ``skipped_zero_value_deferred_rewards``. This test stays scoped to the
        taxable-now pin; the deferred-skip pin lives in
        ``test_tax_math_unchanged_by_deferred_skip``.
        """
        report = _load_report()

        # reward_entries length: exactly the FOUR income rows that are NOT
        # routed to the deferred-skip list (EUR priced + EUR zero + USD zero +
        # SOL priced; the priced CG USD row is a capital gain, not a reward).
        assert len(report.reward_entries) == _EXPECTED_REWARD_ENTRIES_LEN, (
            f"Expected reward_entries length = {_EXPECTED_REWARD_ENTRIES_LEN} "
            f"(EUR priced + EUR zero + USD zero + SOL priced), got {len(report.reward_entries)}. "
            f"Entries: {[(e.asset, e.value_eur, e.tax_classification) for e in report.reward_entries]}"
        )

        # reward_total_eur: sum of value_eur over ALL reward_entries (including
        # the SOL priced row). The taxable-now partition never mutates totals
        # (Invariant 1). The zero-value deferred rows (SOL zero, OSBGT) were
        # skipped at parse time and contribute nothing; the priced SOL row (5.00)
        # does contribute, so the total is 3.00 + 5.00 = 8.00.
        expected_reward_total = sum((e.value_eur for e in report.reward_entries), start=Decimal("0"))
        assert expected_reward_total == _EXPECTED_REWARD_TOTAL_EUR, (
            "Self-check: derived reward_total_eur does not match the expected constant; "
            f"derived={expected_reward_total}, expected={_EXPECTED_REWARD_TOTAL_EUR}"
        )
        assert report.reconciliation.reward_total_eur == _EXPECTED_REWARD_TOTAL_EUR, (
            f"Pipeline reconciliation.reward_total_eur = {report.reconciliation.reward_total_eur} "
            f"must equal the fixture-derived total {_EXPECTED_REWARD_TOTAL_EUR} (Invariant 1: "
            "the partition does not mutate totals)."
        )

        # taxable-now total: sum of value_eur over taxable-now rows only
        # (EUR/USD are fiat → TAXABLE_NOW; SOL is DEFERRED_BY_LAW). Pinned to a
        # SEPARATE constant (r3 finding #8) so a future edit to the deferred
        # side cannot silently shift this taxable-now pin.
        taxable_now = [
            e for e in report.reward_entries
            if e.tax_classification == RewardTaxClassification.TAXABLE_NOW
        ]
        expected_taxable_now_total = sum((e.value_eur for e in taxable_now), start=Decimal("0"))
        assert expected_taxable_now_total == _EXPECTED_TAXABLE_NOW_TOTAL_EUR, (
            "Self-check: taxable-now total (EUR priced 3.00 + EUR zero 0.00 + USD zero 0.00) "
            "must equal the taxable-now pin; SOL is deferred and excluded. "
            f"got {expected_taxable_now_total}, expected {_EXPECTED_TAXABLE_NOW_TOTAL_EUR}"
        )

        # Generate the workbook and confirm Section 4 reconciliation counts.
        wb = _generate_workbook(tmp_path)
        recon = _reconciliation_map(wb)
        detail_count = recon.get("Taxable-now detail rows")
        dust_count = recon.get("Taxable-now dust rows (suppressed from detail)")
        assert detail_count == _EXPECTED_DETAIL_ROWS, (
            f"Taxable-now detail rows = {detail_count!r}, expected {_EXPECTED_DETAIL_ROWS}; recon: {recon}"
        )
        assert dust_count == _EXPECTED_DUST_ROWS, (
            f"Taxable-now dust rows = {dust_count!r}, expected {_EXPECTED_DUST_ROWS}; recon: {recon}"
        )
        # Invariant 1 corollary: detail + dust MUST equal the taxable-now row count
        # (complementary-filter partition; the count-equality guard was dropped per
        # r4 Monitor #4 as tautological at the helper boundary, but it is a useful
        # presentation-layer sanity check at the e2e level).
        assert detail_count + dust_count == len(taxable_now), (
            f"detail ({detail_count}) + dust ({dust_count}) must equal taxable-now row count "
            f"({len(taxable_now)}); recon: {recon}"
        )
        # The taxable-now total value row must equal the fixture-derived taxable-now total
        # (NOT reward_total - SOL's 5.00 EUR is deferred and stays out of taxable-now).
        assert recon.get("Taxable-now total value (EUR)") == float(_EXPECTED_TAXABLE_NOW_TOTAL_EUR), (
            f"Taxable-now total value = {recon.get('Taxable-now total value (EUR)')!r}, "
            f"expected {float(_EXPECTED_TAXABLE_NOW_TOTAL_EUR)}; recon: {recon}"
        )
        wb.close()

    # ------------------------------------------------------------------
    # CRG-022 deferred-side e2e pins (Task 6). The fixture CSV now carries
    # one non-zero SOL row (DEFERRED_BY_LAW, value_eur=5.00 → stays in
    # reward_entries), one zero-value SOL row (DEFERRED_BY_LAW, value_eur=0
    # → skipped into skipped_zero_value_deferred_rewards; SOL is priced so
    # dust reason), and one zero-value OSBGT row (DEFERRED_BY_LAW, value_eur=0
    # → skipped; OSBGT has NO priced row so unpriced reason). These 5 tests
    # pin the parse-time skip (Invariant 1: list preservation), the single
    # suppressed-rewards block render (Invariant 4), and the tax-math-
    # unchanged Invariant 2 against PINNED reference values.
    # ------------------------------------------------------------------

    def test_zero_value_deferred_priced_asset_in_suppressed_block_via_full_pipeline(
        self, tmp_path: Path
    ) -> None:
        """Zero-value deferred reward on a priced asset (SOL) routes to the suppressed block with a 'dust' reason.

        Given the fixture's non-zero SOL row (5.00 EUR) plus a zero-value SOL
        row, the zero-value SOL row MUST be skipped at parse time into
        ``skipped_zero_value_deferred_rewards`` (NOT in ``reward_entries``),
        and the Section 3 "Suppressed zero-value deferred rewards" block MUST
        render a SOL line carrying the ``dust`` reason substring (SOL is
        priced because of the 5.00 EUR row in the same export).
        """
        report = _load_report()
        # The zero-value SOL row must NOT appear in reward_entries.
        sol_in_reward_entries = [
            e for e in report.reward_entries
            if e.asset == _DEFERRED_PRICED_ASSET and e.value_eur == Decimal("0")
        ]
        assert not sol_in_reward_entries, (
            f"Zero-value {_DEFERRED_PRICED_ASSET} row must be skipped at parse time "
            f"(CRG-022), not retained in reward_entries; found: {sol_in_reward_entries}"
        )
        # And it MUST appear in the skipped audit list.
        sol_skipped = [
            e for e in report.skipped_zero_value_deferred_rewards
            if e.asset == _DEFERRED_PRICED_ASSET
        ]
        assert sol_skipped, (
            f"Zero-value {_DEFERRED_PRICED_ASSET} row must be in "
            f"skipped_zero_value_deferred_rewards; got: {report.skipped_zero_value_deferred_rewards}"
        )

        wb = _generate_workbook(tmp_path)
        sol_rows = _suppressed_deferred_table_rows(wb, asset=_DEFERRED_PRICED_ASSET)
        sol_dust_rows = [r for r in sol_rows if r["category"] == "dust"]
        assert sol_dust_rows, (
            f"Expected one {_DEFERRED_PRICED_ASSET} row with category 'dust' in the suppressed "
            f"block; rows: {sol_rows}"
        )
        # Pin the summed-amount :.8f format AND the exact value (Invariants 4 + r4 #3).
        # A format-only regex would let a value_eur-vs-amount swap (always 0 for
        # skipped rows) pass as "0.00000000"; the exact value pin catches that.
        sol_row = sol_dust_rows[0]
        assert sol_row["summed_amount"] == "0.01000000", (
            f"SOL summed amount must equal the fixture's '0,01000000' native amount "
            f"formatted :.8f; got {sol_row!r}"
        )
        wb.close()

    def test_zero_value_deferred_unpriced_asset_in_suppressed_block_via_full_pipeline(
        self, tmp_path: Path
    ) -> None:
        """Zero-value deferred reward on an unpriced asset (OSBGT) routes to the suppressed block.

        Given the fixture's zero-value OSBGT row (OSBGT has NO priced row
        anywhere in the export), the row MUST be skipped at parse time into
        ``skipped_zero_value_deferred_rewards`` (NOT in ``reward_entries``),
        and the Section 3 suppressed block MUST render an OSBGT line carrying
        the ``unpriced`` reason substring.
        """
        report = _load_report()
        osbgt_in_reward_entries = [
            e for e in report.reward_entries if e.asset == _DEFERRED_UNPRICED_ASSET
        ]
        assert not osbgt_in_reward_entries, (
            f"Zero-value {_DEFERRED_UNPRICED_ASSET} row must be skipped at parse time "
            f"(CRG-022), not retained in reward_entries; found: {osbgt_in_reward_entries}"
        )
        osbgt_skipped = [
            e for e in report.skipped_zero_value_deferred_rewards
            if e.asset == _DEFERRED_UNPRICED_ASSET
        ]
        assert osbgt_skipped, (
            f"Zero-value {_DEFERRED_UNPRICED_ASSET} row must be in "
            f"skipped_zero_value_deferred_rewards; got: {report.skipped_zero_value_deferred_rewards}"
        )

        wb = _generate_workbook(tmp_path)
        osbgt_rows = _suppressed_deferred_table_rows(wb, asset=_DEFERRED_UNPRICED_ASSET)
        osbgt_unpriced_rows = [r for r in osbgt_rows if r["category"] == "unpriced"]
        assert osbgt_unpriced_rows, (
            f"Expected one {_DEFERRED_UNPRICED_ASSET} row with category 'unpriced' in the "
            f"suppressed block; rows: {osbgt_rows}"
        )
        # Pin the summed-amount :.8f format AND the exact value. The exact pin
        # catches a value_eur-vs-amount swap (always 0 for skipped rows) that a
        # format-only regex would let through as "0.00000000".
        osbgt_row = osbgt_unpriced_rows[0]
        assert osbgt_row["summed_amount"] == "1.00000000", (
            f"OSBGT summed amount must equal the fixture's '1,00000000' native amount "
            f"formatted :.8f; got {osbgt_row!r}"
        )
        wb.close()

    def test_skipped_zero_value_deferred_rewards_list_preserved(self) -> None:
        """Invariant 1 (list preservation, user's hard requirement) at the e2e tier.

        Given the extended fixture, ``len(skipped_zero_value_deferred_rewards)``
        MUST equal the number of zero-value deferred rows in the fixture
        (= 2: zero SOL + zero OSBGT). Each skipped row is retained as a FULL
        ``CryptoRewardIncomeEntry`` (NOT count-only) with asset, wallet,
        platform, and amount preserved. A regression that silently drops
        rows or projects to a count-only shape fails this test.
        """
        report = _load_report()

        assert len(report.skipped_zero_value_deferred_rewards) == _EXPECTED_SKIPPED_DEFERRED_LEN, (
            f"Expected skipped list length = {_EXPECTED_SKIPPED_DEFERRED_LEN} "
            f"(zero SOL + zero OSBGT), got {len(report.skipped_zero_value_deferred_rewards)}. "
            f"Skipped assets: {sorted(e.asset for e in report.skipped_zero_value_deferred_rewards)}"
        )

        # Each skipped entry retains full fidelity. The fixture's skipped rows
        # carry wallet="Wirex", platform="Wirex", and the exact native amounts
        # written into the CSV (SOL 0.01000000, OSBGT 1.00000000).
        skipped_by_asset = {e.asset: e for e in report.skipped_zero_value_deferred_rewards}
        assert _DEFERRED_PRICED_ASSET in skipped_by_asset, (
            f"Skipped list missing {_DEFERRED_PRICED_ASSET}; got {sorted(skipped_by_asset)}"
        )
        assert _DEFERRED_UNPRICED_ASSET in skipped_by_asset, (
            f"Skipped list missing {_DEFERRED_UNPRICED_ASSET}; got {sorted(skipped_by_asset)}"
        )

        sol_skipped = skipped_by_asset[_DEFERRED_PRICED_ASSET]
        assert sol_skipped.wallet == "Wirex", f"SOL skipped wallet lost; got {sol_skipped.wallet!r}"
        assert sol_skipped.platform == "Wirex", f"SOL skipped platform lost; got {sol_skipped.platform!r}"
        assert sol_skipped.amount == Decimal("0.01000000"), (
            f"SOL skipped amount not preserved; got {sol_skipped.amount!r}"
        )
        assert sol_skipped.value_eur == Decimal("0"), (
            f"SOL skipped value_eur must be ZERO; got {sol_skipped.value_eur!r}"
        )
        assert sol_skipped.tax_classification == RewardTaxClassification.DEFERRED_BY_LAW, (
            f"SOL skipped classification must be DEFERRED_BY_LAW; got {sol_skipped.tax_classification}"
        )

        osbgt_skipped = skipped_by_asset[_DEFERRED_UNPRICED_ASSET]
        assert osbgt_skipped.wallet == "Wirex"
        assert osbgt_skipped.platform == "Wirex"
        assert osbgt_skipped.amount == Decimal("1.00000000"), (
            f"OSBGT skipped amount not preserved; got {osbgt_skipped.amount!r}"
        )
        assert osbgt_skipped.value_eur == Decimal("0")
        assert osbgt_skipped.tax_classification == RewardTaxClassification.DEFERRED_BY_LAW

    def test_tax_math_unchanged_by_deferred_skip(self, tmp_path: Path) -> None:
        """Invariant 2 (tax math unchanged) pinned against ALL FIVE reference values.

        Given the extended fixture, the parse-time skip removes zero-value
        deferred rows (SOL zero + OSBGT) which contributed nothing to tax
        math. ALL FIVE Invariant-2 quantities MUST match PINNED reference
        constants derived from the literal ``Value (EUR)`` cells (NOT
        recomputed from the fixture at test time):

        - ``reward_total_eur`` = 8.00 (3.00 EUR priced + 5.00 SOL priced)
        - ``taxable_now_total_eur`` = 3.00 (only EUR/USD; SOL is deferred)
        - ``deferred_total_eur`` = 5.00 (non-zero SOL only; zero SOL and
          OSBGT were skipped at parse time and contribute nothing)
        - Section 4 ``Taxable-now total value (EUR)`` = 3.0
        - Section 4 ``Deferred total value (EUR)`` = 5.0
        - Reporting OTHER CAPITAL INVESTMENT INCOME line gross = 3.00,
          net = 3.00, raw_row_count = 3 (the EUR priced + EUR zero + USD
          zero taxable-now rows aggregate to the Wirex/GB E25 line).
        """
        report = _load_report()

        # (1) reward_total_eur
        assert report.reconciliation.reward_total_eur == _EXPECTED_DEFERRED_SKIP_REWARD_TOTAL_EUR, (
            f"reward_total_eur = {report.reconciliation.reward_total_eur} != "
            f"{_EXPECTED_DEFERRED_SKIP_REWARD_TOTAL_EUR} (skip must not change reward total)"
        )

        # (2) taxable_now_total_eur and (3) deferred_total_eur via Section 4
        wb = _generate_workbook(tmp_path)
        recon = _reconciliation_map(wb)

        taxable_now_value = recon.get("Taxable-now total value (EUR)")
        deferred_value = recon.get("Deferred total value (EUR)")
        assert taxable_now_value == float(_EXPECTED_DEFERRED_SKIP_TAXABLE_NOW_TOTAL_EUR), (
            f"Section 4 Taxable-now total value = {taxable_now_value!r}, expected "
            f"{float(_EXPECTED_DEFERRED_SKIP_TAXABLE_NOW_TOTAL_EUR)}; recon: {recon}"
        )
        assert deferred_value == float(_EXPECTED_DEFERRED_SKIP_DEFERRED_TOTAL_EUR), (
            f"Section 4 Deferred total value = {deferred_value!r}, expected "
            f"{float(_EXPECTED_DEFERRED_SKIP_DEFERRED_TOTAL_EUR)}; recon: {recon}"
        )

        # Also pin the cross-sheet Reconciliation audit line (r3 finding #2):
        # the Crypto Reconciliation sheet's ``Reward rows`` count and the
        # sibling ``Skipped zero-value deferred rewards (audit)`` line.
        rc_map = _crypto_reconciliation_map(wb)
        assert rc_map.get("Reward rows") == _EXPECTED_REWARD_ENTRIES_LEN, (
            f"Reconciliation 'Reward rows' = {rc_map.get('Reward rows')!r}, expected "
            f"{_EXPECTED_REWARD_ENTRIES_LEN} (raw count post-skip); recon: {rc_map}"
        )
        assert rc_map.get("Skipped zero-value deferred rewards (audit)") == _EXPECTED_SKIPPED_DEFERRED_LEN, (
            f"Reconciliation 'Skipped zero-value deferred rewards (audit)' = "
            f"{rc_map.get('Skipped zero-value deferred rewards (audit)')!r}, expected "
            f"{_EXPECTED_SKIPPED_DEFERRED_LEN}; recon: {rc_map}"
        )
        assert rc_map.get("Rewards total (EUR)") == float(_EXPECTED_DEFERRED_SKIP_REWARD_TOTAL_EUR), (
            f"Reconciliation 'Rewards total (EUR)' = {rc_map.get('Rewards total (EUR)')!r}, "
            f"expected {float(_EXPECTED_DEFERRED_SKIP_REWARD_TOTAL_EUR)}; recon: {rc_map}"
        )

        # (4 + 5) Reporting OTHER CAPITAL INVESTMENT INCOME line. Mirrors the
        # line-locator pattern at test_example_report_generation.py:390-399.
        reporting_row = _reporting_other_capital_investment_income_row(wb)
        assert reporting_row is not None, (
            "Reporting OTHER CAPITAL INVESTMENT INCOME subsection + data row not found"
        )
        # Column layout (mirrors test_example_report_generation.py:406-415):
        #   [0] income-type description (E25 PT description)
        #   [1] source_country (Tabela X) - Wirex resolves to "GB"
        #   [3] gross_income_eur
        #   [4] foreign_tax_eur
        #   [11] net_income_eur
        #   [12] source label ("Koinly")
        #   [13] raw_row_count
        gross = reporting_row[3]
        foreign_tax = reporting_row[4]
        net = reporting_row[11]
        source_label = reporting_row[12]
        raw_row_count = reporting_row[13]
        assert float(gross) == float(_EXPECTED_DEFERRED_SKIP_REPORTING_GROSS_EUR), (
            f"Reporting reward gross = {gross!r}, expected "
            f"{float(_EXPECTED_DEFERRED_SKIP_REPORTING_GROSS_EUR)}"
        )
        assert float(foreign_tax) == 0.0, f"Reporting reward foreign tax = {foreign_tax!r}, expected 0"
        assert float(net) == float(_EXPECTED_DEFERRED_SKIP_REPORTING_NET_EUR), (
            f"Reporting reward net = {net!r}, expected "
            f"{float(_EXPECTED_DEFERRED_SKIP_REPORTING_NET_EUR)}"
        )
        assert source_label == "Koinly", f"Reporting reward source = {source_label!r}, expected 'Koinly'"
        assert raw_row_count == _EXPECTED_DEFERRED_SKIP_REPORTING_RAW_ROW_COUNT, (
            f"Reporting reward raw_row_count = {raw_row_count!r}, expected "
            f"{_EXPECTED_DEFERRED_SKIP_REPORTING_RAW_ROW_COUNT} "
            f"(EUR priced + EUR zero + USD zero taxable-now rows)"
        )
        wb.close()

    def test_taxable_now_partition_still_works(self, tmp_path: Path) -> None:
        """Regression guard: the extended fixture's taxable-now rows still render per CRG-021 (Part 7).

        The deferred-side additions (non-zero SOL, zero SOL, zero OSBGT) MUST
        NOT silently break the taxable-now partition. The fixture's three
        taxable-now rows (EUR priced, EUR zero, USD zero) still drive the
        same CRG-021 behavior: EUR zero → dust summary block (EUR priced
        elsewhere); USD zero → per-row YES (USD has no priced income row).
        """
        wb = _generate_workbook(tmp_path)
        first_col = _supplementary_first_column(wb)

        # CRG-021 taxable-now Dust summary block still renders exactly one
        # EUR row (the zero-value EUR row; EUR priced by the 3.00 row).
        assert "Dust summary:" in first_col, (
            "CRG-021 Dust summary header missing; deferred-side additions broke the taxable-now block."
        )
        eur_dust_rows = _dust_summary_table_rows(wb, asset=_PRICED_ASSET)
        assert len(eur_dust_rows) == 1, (
            f"Expected one {_PRICED_ASSET} row in the taxable-now dust table; "
            f"got {eur_dust_rows}; first column: {first_col}"
        )
        eur_row = eur_dust_rows[0]
        assert eur_row["rows"] == 1, f"Rows count wrong: {eur_row}"
        assert float(eur_row["summed_value_eur"]) == 0.0, (
            f"Summed Value (EUR) wrong: {eur_row}"
        )
        assert eur_row["category"] == "dust", f"Category wrong: {eur_row}"

        # CRG-021 USD unpriced-YES row still appears in the Section 2 detail
        # table with a YES: review flag (USD has no priced income row).
        detail = _section2_detail_rows(wb)
        usd_rows = [r for r in detail if str(r.get("Asset", "")) == _UNPRICED_ASSET]
        assert usd_rows, (
            f"Zero-value {_UNPRICED_ASSET} reward should still appear in Section 2 detail "
            f"(taxable-now partition unchanged); detail rows: {detail}"
        )
        assert str(usd_rows[0].get("Review flag", "")).startswith("YES:"), (
            f"Unpriced-asset zero-value reward review flag must still start 'YES:'; "
            f"got {usd_rows[0].get('Review flag')!r}"
        )

        # The deferred additions do not bleed into Section 2: no SOL or OSBGT
        # row should appear there (they are deferred-side, rendered only in
        # Section 3 detail / suppressed block).
        sol_in_section2 = [r for r in detail if str(r.get("Asset", "")) == _DEFERRED_PRICED_ASSET]
        assert not sol_in_section2, (
            f"Priced {_DEFERRED_PRICED_ASSET} (deferred) should NOT appear in Section 2 "
            f"taxable-now detail; got {sol_in_section2}"
        )
        osbgt_in_section2 = [r for r in detail if str(r.get("Asset", "")) == _DEFERRED_UNPRICED_ASSET]
        assert not osbgt_in_section2, (
            f"Zero-value {_DEFERRED_UNPRICED_ASSET} (deferred + skipped) should NOT appear in "
            f"Section 2 taxable-now detail; got {osbgt_in_section2}"
        )

        # The priced SOL deferred row DOES appear in Section 3 DEFERRED detail
        # (value_eur > 0, not skipped); the zero SOL + OSBGT rows appear ONLY
        # in the Suppressed block, not in the deferred detail table.
        sol_suppressed_rows = _suppressed_deferred_table_rows(wb, asset=_DEFERRED_PRICED_ASSET)
        assert sol_suppressed_rows, (
            "SOL suppressed-block row missing (the zero-SOL row must render in the block)."
        )
        wb.close()
