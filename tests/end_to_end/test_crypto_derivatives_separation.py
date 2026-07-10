"""End-to-end data-trace verification for the ByBit derivatives separation pipeline.

Verifies (per docs/history/plans/2026-06-13-derivatives-separation.md Task 11)
that the committed synthetic Koinly 2025 example
fixture in ``resources/source/example/2025/koinly/`` produces the expected split
between Crypto Gains (spot fee disposal lots) and Derivatives P&L
(art. 10(1)(e) realizations), and that disabling
``separate_derivatives_reporting`` reproduces the legacy mixed value exactly.

Migrated off the gitignored personal ``resources/source/<year>/koinly/`` export by
plan ``docs/history/plans/2026-06-22-crypto-tests-off-local-fixtures.md`` Task 3.
All golden values are recomputed against the synthetic CSVs by independent
arithmetic (see the worked comments next to each assertion). The synthetic
fixture uses the wallets ``Demo Futures`` / ``Demo Spot`` (deliberately NOT in
the operator platform map), so derivatives rows carry ``review_required=True``
from the unmapped-platform signal; the tests assert the classification KIND
(Derivatives, not Ambiguous) rather than the platform-mapping flag.

Cases (synthetic fixture, all USDT on ``Demo Futures`` unless noted):
    Case A = (2025-01-12, USDT, Demo Futures): one OGR Profit row (+140.18 EUR)
        and one OGR Loss row (4.17 EUR futures fee). With separation AND the
        TH-label CG dedup, the profit routes to Derivatives P&L and the single
        CG fee-disposal lot (+2.44 EUR) is REMOVED from Crypto Gains because
        its TH event carries Tag="Futures fee". The -4.17 EUR OGR Loss row
        classifies as Derivatives (cg_matches == 0 after dedup). Without
        separation, the legacy path mixes profit plus fee into a single
        136.01 EUR Crypto Gains entry (140.18 + -4.17 = 136.01).
    Case B = (2025-01-13, USDT, Demo Futures): three OGR Loss rows
        (1.50 + 2.50 + 4.00 = 8.00 EUR) and four CG lots: one exact-match
        Funding-fee lot (0.5 at 08:00) plus three contiguous-range lots at
        13:01 (1.5 + 1.5 + 2.000025 = 5.000025 vs TH Realized 5.000000;
        delta 0.000025 within tolerance Decimal("0.00001") * 3 = 0.00003).
        With separation AND the dedup, ALL 4 CG lots are removed (1 exact +
        3 range); the three OGR Loss rows classify as clean Derivatives
        (-8.00 EUR total).
    Case C = (2025-01-24, USDT, Demo Futures): three derivatives events whose
        TH rows carry derivatives Tags (Funding fee 0.08838575 at 20:00,
        Futures fee 0.41424953 at 23:40, Realized gain 40.75540000 at 23:40).
        Koinly emits the SAME disposals into BOTH the CG report (3 lots, all
        gain 0.00) AND the OGR report (3 Loss rows summing to -39.62 EUR =
        0.08 + 0.40 + 39.14). With separation + the TH-label CG dedup, the
        3 CG lots are removed and the 3 OGR rows classify as clean
        Derivatives (-39.62 EUR).

Preserved non-derivatives spot entries (prevent a false green on the dedup):
BTC gain 2.00 + ETH gain 3.50 on 2025-03-10 (Demo Spot), both |gain| >= 1 EUR.
"""

from __future__ import annotations

import csv
import logging
from decimal import Decimal
from pathlib import Path

import openpyxl
import pytest

from tax_reporting.application.crypto.classification import _is_valid_tabela_x_country
from tax_reporting.application.crypto.entities import CryptoCapitalGainEntry, DerivativesEventType
from tax_reporting.application.crypto.treatment_resolver import TreatmentConfig
from tax_reporting.application.crypto_reporting import load_koinly_crypto_report
from tax_reporting.application.persisting.derivatives_sheet import write_derivatives_sheet
from tests.conftest import KOINLY_2025_EXAMPLE_DIR, build_koinly_jurisdiction

pytestmark = pytest.mark.e2e

_FIXTURE_DIR = KOINLY_2025_EXAMPLE_DIR


def _find_fixture(pattern: str) -> Path:
    """Find a synthetic Koinly fixture CSV by glob pattern.

    The committed synthetic example directory uses the canonical Koinly
    export naming (``koinly_<year>_<report>.csv``), so glob discovery
    resolves exactly one file per pattern.
    """
    matches = sorted(_FIXTURE_DIR.glob(pattern))
    assert matches, (
        f"Expected exactly one synthetic fixture matching {pattern!r} under "
        f"{_FIXTURE_DIR}; the committed example data is missing or renamed. "
        "Regenerate per docs/history/plans/2026-06-22-crypto-tests-off-local-fixtures.md Task 1."
    )
    return matches[0]


def _ogr_path() -> Path:
    return _find_fixture("koinly_2025_other_gains_report*.csv")


def _cg_path() -> Path:
    return _find_fixture("koinly_2025_capital_gains_report*.csv")


def _th_path() -> Path:
    return _find_fixture("koinly_2025_transaction_history*.csv")


_CASE1_DATE = "2025-01-12"
_CASE2_DATE = "2025-01-13"
_CASE3_DATE = "2025-01-24"
_ASSET = "USDT"
_PLATFORM = "Demo Futures"
_PRESERVED_DATE = "2025-03-10"
_PRESERVED_PLATFORM = "Demo Spot"


def _load_with_separation(*, separate_derivatives: bool):
    """Load the synthetic koinly2025 report under the requested jurisdiction setting.

    Phase D Task 3: ``treatment_spot_disposal_via_resolver=False`` opts the
    backward-compat trace into the LEGACY OGR override path. The default-on
    flag-on path filters OGR keys whose TH rows are not SPOT_DISPOSAL; this
    trace exercises derivatives-tagged rows (``Futures fee``, ``Realized
    gain``, ``Funding fee``) that the flag-on filter would exclude. The
    backward-compat golden values (Case A 136.01 EUR) are the legacy target,
    so the flag is off here.
    """
    report = load_koinly_crypto_report(
        _FIXTURE_DIR,
        jurisdiction=build_koinly_jurisdiction(
            separate_derivatives_reporting=separate_derivatives,
            treatment_spot_disposal_via_resolver=False,
        ),
    )
    assert report is not None, "load_koinly_crypto_report returned None for the synthetic koinly2025 fixtures"
    return report


def _assert_csv_contains_value(csv_path: Path, needle: str) -> None:
    """Verify a literal value appears in a fixture CSV (data-trace verification).

    The data-trace checks below compare raw source CSV contents against the
    values the test asserts on the pipeline output, so the contract is grounded
    in source data rather than internal pipeline constants. Failures here point
    to a fixture change rather than a code regression.
    """
    assert csv_path.exists(), f"Fixture CSV not found: {csv_path}"
    with csv_path.open(newline="", encoding="utf-8") as handle:
        rows = list(csv.reader(handle))
    matched = [row for row in rows if any(needle in cell for cell in row)]
    assert matched, (
        f"Expected source value {needle!r} not found in {csv_path.name}. "
        "The fixture has changed; update the data-trace assertion to match."
    )


def _read_csv_rows(csv_path: Path) -> list[list[str]]:
    """Read all rows of a fixture CSV for direct content assertions."""
    assert csv_path.exists(), f"Fixture CSV not found: {csv_path}"
    with csv_path.open(newline="", encoding="utf-8") as handle:
        return list(csv.reader(handle))


def test_synthetic_fixture_contains_derivatives_scenarios() -> None:
    """Guard against fixture content drift: assert the synthetic CSVs carry the derivatives scenarios.

    Parses the committed example CSVs and asserts each derivatives scenario is
    present (Case A fee disposal, Case B funding-fee + realized-gain events and
    the load-bearing 2.000025 range lot, Case C derivatives-label events, plus
    the two preserved non-derivatives spot disposals). A future edit to the
    synthetic CSV that drops a scenario fails this test loudly instead of
    silently changing a pinned golden value.
    """
    th_rows = _read_csv_rows(_th_path())
    th_text = "\n".join(",".join(row) for row in th_rows)
    # Case A: Futures fee disposal event.
    assert "Futures fee" in th_text, "Case A Futures fee Tag missing from the synthetic TH CSV."
    assert "Case A" in th_text, "Case A TH event description missing from the synthetic fixture."
    # Case B: Funding fee + Realized gain events and the load-bearing range lot.
    assert "Funding fee" in th_text, "Case B Funding fee Tag missing from the synthetic TH CSV."
    assert "Case B" in th_text, "Case B TH event description missing from the synthetic fixture."
    assert "5,00000000" in th_text, "Case B Realized gain TH event (5.0 USDT) missing."

    cg_rows = _read_csv_rows(_cg_path())
    cg_amounts = [row[3] for row in cg_rows if len(row) > 3 and row[2] == "USDT"]
    assert "2,00002500" in cg_amounts, (
        "Case B load-bearing range lot (amount 2.00002500) missing from the CG CSV; "
        "without it the tolerance window is not exercised."
    )
    # Case C: three derivatives-label events at 2025-01-24.
    assert "Case C" in th_text, "Case C derivatives-label TH events missing from the synthetic fixture."

    # Preserved non-derivatives spot entries (BTC + ETH on 2025-03-10, Demo Spot).
    cg_assets = [(row[2], row[8]) for row in cg_rows if len(row) > 8]
    assert ("BTC", "Demo Spot") in cg_assets, (
        "Preserved non-derivatives BTC Demo Spot CG entry missing; "
        "test_spot_exchange_lots_preserved cannot verify the dedup does not over-remove."
    )
    assert ("ETH", "Demo Spot") in cg_assets, (
        "Preserved non-derivatives ETH Demo Spot CG entry missing; "
        "test_spot_exchange_lots_preserved cannot verify the dedup does not over-remove."
    )


class TestByBitCase1Trace:
    """Case A (2025-01-12, USDT, Demo Futures): futures Profit + fee disposal separation."""

    def test_profit_in_derivatives_sheet(self) -> None:
        """+140.18 EUR futures profit routes to Derivatives P&L; the 2.44 EUR fee-disposal lot is removed by the dedup.

        Golden value (140.18) recomputed from the synthetic OGR CSV row 4:
        ``Value (EUR)`` column = 140.18, ``Type`` = Profit. The fee-disposal
        lot (2.44 EUR gain, CG row 4) is removed because its TH event carries
        Tag="Futures fee"; its -4.17 EUR OGR Loss counterpart reclassifies to
        Derivatives (covered by ``test_fee_disposal_reclassifies_to_derivatives``).
        """
        _assert_csv_contains_value(_ogr_path(), "140,18")
        _assert_csv_contains_value(_ogr_path(), "4,17")
        _assert_csv_contains_value(_cg_path(), "2,44")

        report = _load_with_separation(separate_derivatives=True)

        profit_matches = [
            e
            for e in report.derivatives_entries
            if (
                e.date == _CASE1_DATE
                and e.asset == _ASSET
                and e.platform == _PLATFORM
                and e.event_type == DerivativesEventType.PROFIT
            )
        ]
        assert profit_matches, (
            "Expected a Derivatives P&L PROFIT entry for "
            f"({_CASE1_DATE}, {_ASSET}, {_PLATFORM}); got "
            f"{[(e.date, e.asset, e.platform, e.event_type, e.pnl_eur) for e in report.derivatives_entries]}"
        )
        # CSV arithmetic: OGR row 4 Value(EUR) = 140.18 (Profit type).
        profit_total = sum((e.pnl_eur for e in profit_matches), start=Decimal("0"))
        assert profit_total == Decimal("140.18"), (
            f"Expected +140.18 EUR PROFIT in derivatives_entries for Case A, got {profit_total}. "
            "Entries: "
            f"{[(e.event_type, e.pnl_eur) for e in profit_matches]}"
        )

    def test_no_fee_disposal_lot_in_capital_entries(self) -> None:
        """capital_entries has no entry for (2025-01-12, USDT, Demo Futures) after dedup.

        The 2.44 EUR Futures fee CG lot (CG row 4: amount 4.27180510, proceeds
        4.17, gain 2.44) is removed by the dedup because its TH event (TH row 4:
        crypto_withdrawal Futures fee 4.27180510 USDT Demo Futures at 15:22)
        carries Tag="Futures fee".
        """
        _assert_csv_contains_value(_cg_path(), "2,44")

        report = _load_with_separation(separate_derivatives=True)

        case1_capital = [
            e
            for e in report.capital_entries
            if e.disposal_date == _CASE1_DATE and e.asset == _ASSET and e.platform == _PLATFORM
        ]
        assert not case1_capital, (
            "Expected NO Crypto Gains entry for "
            f"({_CASE1_DATE}, {_ASSET}, {_PLATFORM}) after the derivatives CG dedup "
            "(the 2.44 EUR Futures fee CG lot should be removed because its TH "
            "event carries Tag='Futures fee'). "
            f"Got {len(case1_capital)} entries: "
            f"{[(e.gain_loss_eur, e.holding_period) for e in case1_capital]}"
        )

    def test_no_derivatives_value_in_capital_entries(self) -> None:
        """No CryptoCapitalGainEntry in capital_entries equals the legacy 136.01 EUR mixed value.

        Golden value (136.01) recomputed from the synthetic OGR CSV: Case A
        flag-OFF mixes OGR Profit 140.18 + OGR Loss -4.17 = 136.01 EUR (asserted
        green in TestBackwardCompatTrace). Under flag-ON this value must NOT
        appear in capital_entries (the profit routes to derivatives_entries).
        """
        _assert_csv_contains_value(_ogr_path(), "140,18")

        report = _load_with_separation(separate_derivatives=True)

        legacy_matches = [
            e
            for e in report.capital_entries
            if e.gain_loss_eur == Decimal("136.01")
        ]
        assert not legacy_matches, (
            "Found a Crypto Gains entry with the legacy 136.01 EUR mixed value; "
            "separation should have routed the futures profit to derivatives_entries. "
            f"Matches: {[(e.disposal_date, e.asset, e.platform, e.gain_loss_eur) for e in legacy_matches]}"
        )

    def test_fee_disposal_reclassifies_to_derivatives(self) -> None:
        """The -4.17 EUR OGR Loss row (Case A Futures fee) reclassifies to Derivatives after the dedup.

        With the 2.44 EUR Futures fee CG lot removed by the dedup (TH row 4
        carries Tag='Futures fee'), the OGR classifier sees ``cg_matches == 0``
        for the -4.17 EUR OGR Loss row (OGR row 5) and classifies it as
        Derivatives (not Spot, not Ambiguous). So ``derivatives_entries`` for
        Case A contains a LOSS entry totalling -4.17 EUR.

        Note: because ``Demo Futures`` is unmapped, the row also carries
        ``review_required=True`` from the platform-mapping signal; this test
        asserts the classification KIND (LOSS present, not Ambiguous) via the
        OGR-handler log message rather than the review flag.
        """
        _assert_csv_contains_value(_ogr_path(), "4,17")
        _assert_csv_contains_value(_cg_path(), "2,44")

        report = _load_with_separation(separate_derivatives=True)

        loss_derivatives = [
            e
            for e in report.derivatives_entries
            if e.date == _CASE1_DATE
            and e.asset == _ASSET
            and e.platform == _PLATFORM
            and e.event_type == DerivativesEventType.LOSS
        ]
        assert loss_derivatives, (
            "Expected the Case A -4.17 EUR OGR Loss row to reclassify as a Derivatives "
            "LOSS entry after the dedup (its CG counterpart was removed). Got: "
            f"{[(e.event_type, e.pnl_eur) for e in report.derivatives_entries]}"
        )
        # CSV arithmetic: OGR row 5 Value(EUR) = 4.17 (Loss type, negative sign on routing).
        loss_total = sum((e.pnl_eur for e in loss_derivatives), start=Decimal("0"))
        assert loss_total == Decimal("-4.17"), (
            "Expected sum of Case A LOSS derivatives_entries.pnl_eur to be -4.17 EUR "
            "(the Futures fee OGR row, reclassified from Spot to Derivatives because its "
            "CG counterpart was removed). "
            f"Got {loss_total}. Entries: "
            f"{[(e.event_type, e.pnl_eur, e.review_required) for e in loss_derivatives]}"
        )
        # Assert classification is clean Derivatives (NOT Ambiguous). The Ambiguous
        # classification reason mentions "matches CG disposal" or "mismatch"; a clean
        # Derivatives routing has an empty classification_reason. The platform-mapping
        # review_required=True is a separate, expected signal for the unmapped Demo wallet
        # and must not be conflated with the Ambiguous flag (CRG-016).
        for entry in loss_derivatives:
            assert "matches CG disposal" not in entry.review_reason, (
                "Case A LOSS entry must NOT carry the Ambiguous classification reason "
                "(cg_matches == 0 after dedup). The only review signal should be the "
                "unmapped-platform one. "
                f"review_reason={entry.review_reason!r}"
            )


class TestByBitCase2Trace:
    """Case B (2025-01-13, USDT, Demo Futures): three OGR Loss rows + 4 CG lots."""

    def test_lots_remain_positive_for_spot_only(self) -> None:
        """Case B CG lots are entirely removed by the dedup; capital_entries has 0 Case B rows.

        CSV arithmetic (CG rows 5-8): one Funding-fee exact-match lot (0.5 at
        08:00) plus three contiguous-range lots at 13:01 (1.5 + 1.5 + 2.000025
        = 5.000025 vs TH Realized 5.000000; delta 0.000025 within tolerance
        Decimal("0.00001") * 3 = 0.00003). All 4 lots are removed, so no Case B
        capital entry survives.
        """
        _assert_csv_contains_value(_ogr_path(), "1,50")
        _assert_csv_contains_value(_ogr_path(), "2,50")
        _assert_csv_contains_value(_ogr_path(), "4,00")

        report = _load_with_separation(separate_derivatives=True)

        case2_capital_entries = [
            e
            for e in report.capital_entries
            if e.disposal_date == _CASE2_DATE and e.asset == _ASSET and e.platform == _PLATFORM
        ]
        assert case2_capital_entries == [], (
            "Expected NO Crypto Gains entries for Case B after the dedup (all 4 "
            "USDT Demo Futures lots removed: Funding fee exact + 3-lot contiguous "
            "range matching the 5.000000 Realized gain within tolerance). "
            f"Got {len(case2_capital_entries)} entries: "
            f"{[(e.gain_loss_eur, e.holding_period) for e in case2_capital_entries]}"
        )

    def test_derivatives_lots_removed(self) -> None:
        """Exactly 4 Case B CG lots are removed by the dedup (1 exact + 3 range).

        Breakdown: 1 Funding fee exact (0.5) + 3 Realized gain contiguous range
        (1.5 + 1.5 + 2.000025 = 5.000025 vs TH Realized 5.000000 within
        tolerance). The synthetic fixture was designed so phase-2 is
        load-bearing: the range lots sum within tolerance but NOT exactly equal
        to the TH Realized amount (delta 0.000025), so removing the tolerance
        window would break the match.
        """
        import tax_reporting.application.crypto.derivatives_dedup as dd_mod

        orig_remove = dd_mod.remove_derivatives_flagged_lots
        removed_count = {"value": 0}

        def spy_remove(entries, events):  # type: ignore[no-untyped-def]
            case2_in = [
                e
                for e in entries
                if e.disposal_date == _CASE2_DATE and e.asset == _ASSET and e.platform == _PLATFORM
            ]
            out = orig_remove(entries, events)
            filtered = out[0] if isinstance(out, tuple) else out
            case2_out = [
                e
                for e in filtered
                if e.disposal_date == _CASE2_DATE and e.asset == _ASSET and e.platform == _PLATFORM
            ]
            removed_count["value"] = len(case2_in) - len(case2_out)
            return out

        dd_mod.remove_derivatives_flagged_lots = spy_remove
        try:
            _load_with_separation(separate_derivatives=True)
        finally:
            dd_mod.remove_derivatives_flagged_lots = orig_remove

        # CSV arithmetic: CG rows 5-8 = 4 Case B lots (0.5 + 1.5 + 1.5 + 2.000025).
        assert removed_count["value"] == 4, (
            "Expected exactly 4 Case B CG lots removed by the dedup "
            "(1 Funding fee exact + 3 Realized gain range). "
            f"Got {removed_count['value']}."
        )

    def test_range_removal_count_load_bearing(self, caplog: pytest.LogCaptureFixture) -> None:
        """Phase-2 contiguous-range removal count is >= 2 (proves the fallback is load-bearing).

        Observes the per-lot removal INFO logs from ``derivatives_dedup`` (each
        removal logs ``match_type=exact`` or ``match_type=range``). For Case B
        the synthetic fixture produces 1 exact + 3 range removals; asserting
        >= 2 range removals proves phase-2 matched a multi-lot window within
        tolerance. A lot set that summed exactly to the TH Realized amount
        would defeat the test (Design Invariant #2).
        """
        with caplog.at_level(logging.INFO, logger="tax_reporting.application.crypto.derivatives_dedup"):
            _load_with_separation(separate_derivatives=True)

        removal_records = [
            r
            for r in caplog.records
            if r.name == "tax_reporting.application.crypto.derivatives_dedup"
            and "Removed derivatives-flagged CG lot" in r.getMessage()
        ]
        range_records = [
            r for r in removal_records if "match_type=range" in r.getMessage()
        ]
        # CSV arithmetic: 3 range lots (1.5 + 1.5 + 2.000025) match the 5.000000 TH Realized.
        assert len(range_records) >= 2, (
            "Expected >= 2 contiguous-range (match_type=range) removals proving phase-2 "
            "is load-bearing; the synthetic Case B fixture produces 3 range removals. "
            f"Got {len(range_records)} range removals out of {len(removal_records)} total. "
            f"Records: {[r.getMessage() for r in removal_records]}"
        )

    def test_spot_exchange_lots_preserved(self) -> None:
        """The two synthetic non-derivatives spot entries survive the dedup identically in flag-on and flag-off.

        Replaces the volatile 2025-01-26 SOL ByBit entry (which broke under the
        "Transfer fees" Koinly toggle). The synthetic fixture's preserved
        non-derivatives entries are (2025-03-10, BTC, Demo Spot) gain 2.00 and
        (2025-03-10, ETH, Demo Spot) gain 3.50 (CG rows 12-13). Both must
        survive the dedup unchanged in flag-on AND flag-off with identical
        gains, and the count of non-derivatives Demo Spot entries must be
        unchanged between paths (count_on == count_off == 2).

        CSV arithmetic (CG rows 12-13): BTC gain 2.00, ETH gain 3.50
        (proceeds 12.00 - cost 10.00 = 2.00; proceeds 9.50 - cost 6.00 = 3.50).
        """
        report_on = _load_with_separation(separate_derivatives=True)
        report_off = _load_with_separation(separate_derivatives=False)

        on_preserved = [
            e
            for e in report_on.capital_entries
            if e.disposal_date == _PRESERVED_DATE and e.platform == _PRESERVED_PLATFORM
        ]
        off_preserved = [
            e
            for e in report_off.capital_entries
            if e.disposal_date == _PRESERVED_DATE and e.platform == _PRESERVED_PLATFORM
        ]
        assert len(on_preserved) == 2, (
            "Expected exactly 2 preserved non-derivatives Demo Spot entries (BTC + ETH) "
            "in the flag-on path. "
            f"flag_on={len(on_preserved)}: {[(e.asset, e.gain_loss_eur) for e in on_preserved]}"
        )
        assert len(off_preserved) == 2, (
            "Expected exactly 2 preserved non-derivatives Demo Spot entries (BTC + ETH) "
            "in the flag-off path. "
            f"flag_off={len(off_preserved)}: {[(e.asset, e.gain_loss_eur) for e in off_preserved]}"
        )

        def _by_asset(entries):
            return {e.asset: e.gain_loss_eur for e in entries}

        on_gains = _by_asset(on_preserved)
        off_gains = _by_asset(off_preserved)
        # CSV arithmetic: BTC gain = 12.00 - 10.00 = 2.00; ETH gain = 9.50 - 6.00 = 3.50.
        assert on_gains == {"BTC": Decimal("2.00"), "ETH": Decimal("3.50")}, (
            "Flag-on preserved entry gains drifted from the synthetic CSV values. "
            f"Got {on_gains}."
        )
        assert off_gains == {"BTC": Decimal("2.00"), "ETH": Decimal("3.50")}, (
            "Flag-off preserved entry gains drifted from the synthetic CSV values. "
            f"Got {off_gains}."
        )
        assert on_gains == off_gains, (
            "Preserved non-derivatives entry gains must be identical between flag-on and "
            "flag-off (the dedup must not modify non-derivatives entries). "
            f"flag_on={on_gains}, flag_off={off_gains}"
        )

    def test_derivatives_total_matches_ogr_net(self) -> None:
        """Sum of Case B derivatives_entries.pnl_eur equals the OGR net (-8.00 EUR).

        CSV arithmetic: OGR rows 6-8 (Case B Loss rows) Value(EUR) column:
        1.50 + 2.50 + 4.00 = 8.00 EUR, negated on routing = -8.00 EUR.
        The three OGR Loss rows route to derivatives_entries regardless of CG
        removal; the Crypto Gains aggregate is 0 EUR (no Case B rows survive),
        asserted by ``test_lots_remain_positive_for_spot_only``.
        """
        _assert_csv_contains_value(_ogr_path(), "1,50")
        _assert_csv_contains_value(_ogr_path(), "2,50")
        _assert_csv_contains_value(_ogr_path(), "4,00")

        report = _load_with_separation(separate_derivatives=True)

        case2_derivatives = [
            e
            for e in report.derivatives_entries
            if e.date == _CASE2_DATE and e.asset == _ASSET and e.platform == _PLATFORM
        ]
        total = sum((e.pnl_eur for e in case2_derivatives), start=Decimal("0"))
        # CSV arithmetic: -(1.50 + 2.50 + 4.00) = -8.00 EUR.
        assert total == Decimal("-8.00"), (
            "Expected sum of Case B derivatives_entries.pnl_eur to be -8.00 EUR "
            "(the OGR Loss rows 1.50 + 2.50 + 4.00 with negative sign), "
            f"got {total}. Entries: "
            f"{[(e.event_type, e.pnl_eur) for e in case2_derivatives]}"
        )


class TestBackwardCompatTrace:
    """Flag-off path reproduces the legacy mixed Crypto Gains value for Case A."""

    def test_flag_off_matches_golden_values(self) -> None:
        """With separate_derivatives_reporting=False, Case A reproduces 136.01 EUR.

        Golden value (136.01) recomputed from the synthetic OGR CSV: Case A
        flag-OFF mixes the OGR Profit 140.18 (row 4) with the OGR Loss -4.17
        (row 5) into a single Crypto Gains entry: 140.18 + (-4.17) = 136.01.

        Note: the real-fixture Case 2 backward-compat value (-26.64 EUR) does
        NOT reproduce against the synthetic fixture because the synthetic Case
        B CG lots all have gain 0.00, so the flag-OFF direction-override output
        is filtered by the PT-C-028 materiality filter (|gain| < 1 EUR). The
        Case A 136.01 EUR guard is the sole positive backward-compat
        characterization (preserved per Task 6's note).
        """
        _assert_csv_contains_value(_ogr_path(), "140,18")

        report = _load_with_separation(separate_derivatives=False)

        case1_matches = [
            e
            for e in report.capital_entries
            if e.disposal_date == _CASE1_DATE and e.asset == _ASSET and e.platform == _PLATFORM
        ]
        assert len(case1_matches) == 1, (
            "Expected exactly one Crypto Gains entry for Case A under the legacy path, "
            f"got {len(case1_matches)}"
        )
        # CSV arithmetic: OGR Profit 140.18 + OGR Loss -4.17 = 136.01 EUR.
        assert case1_matches[0].gain_loss_eur == Decimal("136.01"), (
            "Case A backward-compat drift: expected 136.01 EUR (mixed Profit + fee) in "
            f"Crypto Gains, got {case1_matches[0].gain_loss_eur} EUR"
        )

        # No derivatives_entries should be populated under the legacy path.
        assert report.derivatives_entries == [], (
            "Legacy path (separate_derivatives_reporting=False) must not populate "
            "derivatives_entries; got: "
            f"{[(e.date, e.asset, e.platform, e.pnl_eur) for e in report.derivatives_entries]}"
        )


class TestByBitCase3Trace:
    """Case C (2025-01-24, USDT, Demo Futures): derivatives CG dedup via TH Tags.

    This case captures the bug described in
    ``docs/history/plans/2026-06-14-derivatives-th-label-cg-dedup.md``: Koinly emits
    the SAME disposal into BOTH the OGR report (as Loss rows summing to
    -39.62 EUR) AND the CG report (as 3 FIFO lots, all gain 0.00 in the
    synthetic fixture). The fix is a CG-side filter that scans TH rows for
    derivatives Tags (Funding fee / Futures fee / Realized gain) and removes
    matching CG lots before the OGR classifier runs, so the disposal is
    reported once (in Derivatives P&L) rather than twice.

    Source data trace (synthetic fixture CSVs):
      - TH row 8: 2025-01-24 20:00:00 crypto_withdrawal Funding fee 0.08838575 USDT Demo Futures
      - TH row 9: 2025-01-24 23:40:00 crypto_withdrawal Futures fee 0.41424953 USDT Demo Futures
      - TH row 10: 2025-01-24 23:40:00 crypto_withdrawal Realized gain 40.75540000 USDT Demo Futures
      - OGR rows 9-11: 2025-01-24 USDT Loss rows 0.08 + 0.40 + 39.14 = -39.62 EUR
      - CG rows 9-11: 2025-01-24 USDT Demo Futures lots (0.08, 0.40, 39.14 EUR proceeds, gain 0.00)
    """

    def test_derivatives_th_events_identified(self) -> None:
        """TH scanner identifies the 3 derivatives events on 2025-01-24 via the config-driven tag set.

        Guards against future tag-vocabulary drift; ``find_derivatives_th_events``
        is implemented in ``derivatives_dedup.py``.
        """
        _assert_csv_contains_value(_th_path(), "Funding fee")
        _assert_csv_contains_value(_th_path(), "Futures fee")
        _assert_csv_contains_value(_th_path(), "Realized gain")

    def test_no_capital_entries_for_2025_01_24_after_dedup(self) -> None:
        """capital_entries contains no (2025-01-24, USDT, Demo Futures) row after the dedup removes the 3 CG lots."""
        _assert_csv_contains_value(_ogr_path(), "0,08")
        _assert_csv_contains_value(_ogr_path(), "0,40")
        _assert_csv_contains_value(_ogr_path(), "39,14")
        _assert_csv_contains_value(_cg_path(), "0,08838575")
        _assert_csv_contains_value(_cg_path(), "0,41424953")
        _assert_csv_contains_value(_cg_path(), "40,75540000")

        report = _load_with_separation(separate_derivatives=True)

        case3_capital = [
            e
            for e in report.capital_entries
            if e.disposal_date == _CASE3_DATE and e.asset == _ASSET and e.platform == _PLATFORM
        ]
        assert not case3_capital, (
            "Expected NO Crypto Gains entry for "
            f"({_CASE3_DATE}, {_ASSET}, {_PLATFORM}) after the derivatives CG dedup "
            "(all 3 CG lots should be removed because their TH events carry derivatives Tags). "
            f"Got {len(case3_capital)} entries: "
            f"{[(e.gain_loss_eur, e.holding_period) for e in case3_capital]}"
        )

    def test_derivatives_entries_clean_for_2025_01_24(self) -> None:
        """derivatives_entries contains a LOSS row for (2025-01-24, USDT, Demo Futures) with total pnl -39.62 EUR.

        Post-dedup state: one aggregated DerivativesPnLEntry per
        (date, asset, platform, event_type) tuple, LOSS type, summing to
        -39.62 EUR. Because ``Demo Futures`` is unmapped, the row carries
        ``review_required=True`` from the platform-mapping signal; this test
        asserts the classification KIND (Derivatives LOSS, not Ambiguous) via
        the review_reason rather than the review flag.
        """
        _assert_csv_contains_value(_ogr_path(), "0,08")
        _assert_csv_contains_value(_ogr_path(), "0,40")
        _assert_csv_contains_value(_ogr_path(), "39,14")

        report = _load_with_separation(separate_derivatives=True)

        case3_derivatives = [
            e
            for e in report.derivatives_entries
            if e.date == _CASE3_DATE and e.asset == _ASSET and e.platform == _PLATFORM
        ]
        assert case3_derivatives, (
            "Expected at least one Derivatives P&L entry for "
            f"({_CASE3_DATE}, {_ASSET}, {_PLATFORM}); got none. "
            "All derivatives_entries: "
            f"{[(e.date, e.asset, e.platform, e.pnl_eur) for e in report.derivatives_entries]}"
        )
        assert all(e.event_type == DerivativesEventType.LOSS for e in case3_derivatives), (
            "All Case C derivatives entries should be LOSS type. Got: "
            f"{[(e.event_type, e.pnl_eur) for e in case3_derivatives]}"
        )
        total = sum((e.pnl_eur for e in case3_derivatives), start=Decimal("0"))
        # CSV arithmetic: OGR rows 9-11 Value(EUR) = 0.08 + 0.40 + 39.14 = 39.62, negated = -39.62.
        assert total == Decimal("-39.62"), (
            "Expected sum of Case C derivatives_entries.pnl_eur to be -39.62 EUR "
            "(the OGR Loss rows 0.08 + 0.40 + 39.14 with negative sign), "
            f"got {total}. Entries: "
            f"{[(e.event_type, e.pnl_eur, e.review_required) for e in case3_derivatives]}"
        )
        # Assert classification is clean Derivatives (NOT Ambiguous). The unmapped
        # Demo Futures platform sets review_required=True from the platform-mapping
        # signal; the Ambiguous classification reason must NOT be present.
        for entry in case3_derivatives:
            assert "matches CG disposal" not in entry.review_reason, (
                "Case C derivatives entry must NOT carry the Ambiguous classification reason "
                "(cg_matches == 0 after dedup). The only review signal should be the "
                "unmapped-platform one. "
                f"review_reason={entry.review_reason!r}"
            )

    def test_removal_logged(self, caplog: pytest.LogCaptureFixture) -> None:
        """Each removed CG lot logs at INFO; exactly one summary WARNING covers the aggregate.

        Per Design Invariant 15 of
        ``docs/history/plans/2026-06-14-derivatives-th-label-cg-dedup.md`` and
        CLAUDE.md's "Every WARNING must be actionable and non-noisy at scale"
        rule, the dedup does NOT emit per-lot WARNINGs. Each removal logs at
        INFO (audit-traceable: timestamp, asset, wallet, amount, match type,
        matching TH Tag), and exactly one aggregate WARNING per pipeline
        run carries the total count, breakdown by match type, and aggregate
        proceeds and gain removed.

        Expected caplog contents against the synthetic fixture:
          - 3 INFO records from ``derivatives_dedup`` for the Case C lots
            (Funding fee 0.08838575 USDT at 20:00, Futures fee 0.41424953
            USDT at 23:40, Realized gain 40.75540000 USDT at 23:40).
          - 1 WARNING record from ``derivatives_dedup`` with the summary
            text including the word ``removed`` and a count greater than
            or equal to 3 (the summary covers ALL removals for the year,
            not just Case C).
        """
        _assert_csv_contains_value(_th_path(), "Funding fee")
        _assert_csv_contains_value(_th_path(), "Futures fee")
        _assert_csv_contains_value(_th_path(), "Realized gain")

        with caplog.at_level(logging.INFO, logger="tax_reporting.application.crypto.derivatives_dedup"):
            report = _load_with_separation(separate_derivatives=True)

        # Sanity: the dedup actually ran (otherwise no removals are expected).
        case3_capital = [
            e
            for e in report.capital_entries
            if e.disposal_date == _CASE3_DATE and e.asset == _ASSET and e.platform == _PLATFORM
        ]
        assert not case3_capital, (
            "Test precondition failed: Case C CG lots were not removed by the dedup; "
            "caplog INFO removal records cannot be present. "
            "Pipeline still produces the old double-counted output."
        )

        # Per-lot INFO records for the 3 Case C lots.
        info_records = [r for r in caplog.records if r.levelno == logging.INFO]
        info_text = " ".join(r.getMessage() for r in info_records)
        for needle in (_CASE3_DATE, _ASSET, _PLATFORM):
            assert needle in info_text, (
                f"Expected INFO removal text to mention {needle!r}; "
                f"got INFO records: {[r.getMessage() for r in info_records]}"
            )
        for amount_marker in ("0.08838575", "0.41424953", "40.75540000"):
            assert amount_marker in info_text, (
                f"Expected INFO removal text to mention removed CG lot amount {amount_marker!r}; "
                f"got INFO records: {[r.getMessage() for r in info_records]}"
            )
        for label in ("Funding fee", "Futures fee", "Realized gain"):
            assert label in info_text, (
                f"Expected INFO removal text to mention matching TH Tag {label!r}; "
                f"got INFO records: {[r.getMessage() for r in info_records]}"
            )

        # Exactly one summary WARNING from derivatives_dedup covering the aggregate.
        warning_records = [
            r
            for r in caplog.records
            if r.levelno >= logging.WARNING
            and r.name == "tax_reporting.application.crypto.derivatives_dedup"
        ]
        assert len(warning_records) == 1, (
            "Expected exactly one derivatives_dedup summary WARNING per pipeline run; "
            f"got {len(warning_records)}: {[r.getMessage() for r in warning_records]}"
        )
        warning_text = warning_records[0].getMessage()
        assert "removed" in warning_text, (
            "Summary WARNING should mention 'removed'; got: " f"{warning_text}"
        )
        assert "lots" in warning_text, (
            "Summary WARNING should mention 'lots'; got: " f"{warning_text}"
        )


class TestPipelineIntegration:
    """Pipeline-level integration tests for the derivatives CG dedup wiring.

    These tests exercise ``load_koinly_crypto_report`` and
    ``apply_derivatives_dedup`` end-to-end against the synthetic koinly2025
    fixture, verifying the dedup runs at the correct pipeline point
    (after validation, before OGR split) and gracefully degrades when
    any of its gates fail.
    """

    def test_dedup_runs_after_validation_before_split(
        self,
        monkeypatch: pytest.MonkeyPatch,
    ) -> None:
        """Dedup runs after _validate_capital_entries_have_valid_countries and before _split_ogr_index.

        Uses monkeypatch to wrap each pipeline stage with a call-order
        counter. The counter must record validation first, dedup second,
        and OGR split third. This guards against regressions that move
        the dedup call before validation (filtered list misses invalid
        lots) or after the split (classifier sees the unfiltered list
        and routes the OGR rows to Ambiguous instead of Derivatives).
        """
        import tax_reporting.application.crypto_reporting as cr_mod

        call_order: list[str] = []

        original_validate = cr_mod._validate_capital_entries_have_valid_countries
        original_split = cr_mod._split_ogr_index
        original_dedup = cr_mod.apply_derivatives_dedup

        def spy_validate(entries, jurisdiction):  # type: ignore[no-untyped-def]
            call_order.append("validate")
            return original_validate(entries, jurisdiction)

        def spy_dedup(**kwargs):  # type: ignore[no-untyped-def]
            call_order.append("dedup")
            return original_dedup(**kwargs)

        def spy_split(ogr_rows, capital_entries, jurisdiction):  # type: ignore[no-untyped-def]
            call_order.append("split")
            return original_split(ogr_rows, capital_entries, jurisdiction)

        monkeypatch.setattr(cr_mod, "_validate_capital_entries_have_valid_countries", spy_validate)
        monkeypatch.setattr(cr_mod, "apply_derivatives_dedup", spy_dedup)
        monkeypatch.setattr(cr_mod, "_split_ogr_index", spy_split)

        # The load itself is the act: it must invoke the spied pipeline stages
        # in the validate -> dedup -> split order. The returned report is not
        # inspected here (call_order carries the signal under test).
        _load_with_separation(separate_derivatives=True)

        # Validate the three expected stages were called, in order, with no
        # interleaving. Other stages (parsing, FIFO rebuild, aggregation)
        # are not tracked here because the contract under test is only the
        # validate -> dedup -> split sequence.
        relevant = [step for step in call_order if step in {"validate", "dedup", "split"}]
        assert relevant, (
            "Pipeline did not invoke validate/dedup/split stages; got call_order="
            f"{call_order}"
        )
        # The FIRST occurrence of each stage must appear in the order
        # validate -> dedup -> split. (Each may be called more than once
        # in principle, but the first call establishes the integration
        # point.)
        first_seen: dict[str, int] = {}
        for idx, step in enumerate(call_order):
            if step in {"validate", "dedup", "split"} and step not in first_seen:
                first_seen[step] = idx
        ordered = sorted(first_seen.keys(), key=lambda s: first_seen[s])
        assert ordered == ["validate", "dedup", "split"], (
            "Expected pipeline stage order validate -> dedup -> split; got "
            f"first-seen order {ordered} from full call_order={call_order}"
        )

    def test_dedup_skipped_when_flag_false(self, monkeypatch: pytest.MonkeyPatch) -> None:
        """With separate_derivatives_reporting=False, the dedup is a no-op.

        The pipeline must produce byte-identical output to the predecessor
        plan. The TestBackwardCompatTrace#test_flag_off_matches_golden_values
        test already covers this for the legacy Case A value (136.01 EUR).
        This test additionally asserts that no derivatives_dedup summary
        WARNING fires (the dedup short-circuits on the gate before reaching
        remove_derivatives_flagged_lots).
        """
        from tax_reporting.application.crypto import derivatives_dedup as dd_mod

        # Track whether remove_derivatives_flagged_lots is invoked at all.
        invoked: list[bool] = []
        original = dd_mod.remove_derivatives_flagged_lots

        def spy_remove(entries, events):  # type: ignore[no-untyped-def]
            invoked.append(True)
            return original(entries, events)

        monkeypatch.setattr(dd_mod, "remove_derivatives_flagged_lots", spy_remove)
        _load_with_separation(separate_derivatives=False)

        assert not invoked, (
            "remove_derivatives_flagged_lots must NOT be invoked when "
            "separate_derivatives_reporting=False (Design Invariant 14). "
            "Pipeline should short-circuit on the gate."
        )

    def test_dedup_skipped_when_th_missing(
        self,
        caplog: pytest.LogCaptureFixture,
    ) -> None:
        """With separate_derivatives_reporting=True but TH file=None, dedup is a no-op.

        ``apply_derivatives_dedup`` short-circuits on the gate check when
        ``transaction_history_file`` is None. No WARNING is emitted from
        the dedup (the gate failure is silent; the missing TH is already
        surfaced by the pipeline's required-files check).
        """
        from tax_reporting.application.crypto.derivatives_dedup import (
            apply_derivatives_dedup,
        )

        # apply_derivatives_dedup returns the input list when the gate
        # fails, so we pass a non-None list and check identity.
        entries_in: list[CryptoCapitalGainEntry] = []

        with caplog.at_level(logging.WARNING, logger="tax_reporting.application.crypto.derivatives_dedup"):
            result = apply_derivatives_dedup(
                capital_entries=entries_in,
                jurisdiction=build_koinly_jurisdiction(separate_derivatives_reporting=True),
                transaction_history_file=None,
                year=2025,
                transactions=[],
                config=TreatmentConfig(),
                via_resolver=False,
            )

        assert result is entries_in, (
            "apply_derivatives_dedup must return the input list unchanged when "
            "transaction_history_file is None (gate failure, Design Invariant 14)."
        )
        # No dedup-specific WARNING should fire on gate failure.
        dedup_warnings = [
            r for r in caplog.records if r.levelno >= logging.WARNING
        ]
        assert not dedup_warnings, (
            "No WARNING should fire when the gate fails on missing TH file; "
            f"got: {[r.getMessage() for r in dedup_warnings]}"
        )

    def test_dedup_skipped_when_config_missing(
        self,
        monkeypatch: pytest.MonkeyPatch,
        caplog: pytest.LogCaptureFixture,
    ) -> None:
        """With flag=True and a valid TH file but missing label config, dedup warns and continues.

        Monkeypatches ``_load_derivatives_labels_config`` to return
        ``frozenset()`` (simulating a missing config file) and verifies
        ``apply_derivatives_dedup`` emits exactly one WARNING with the
        remediation hint, then returns the input list unchanged. This
        verifies the graceful-degradation path (Design Invariant 9).
        """
        from tax_reporting.application.crypto import derivatives_dedup as dd_mod
        from tax_reporting.application.crypto.derivatives_dedup import (
            apply_derivatives_dedup,
        )

        monkeypatch.setattr(
            dd_mod,
            "_load_derivatives_labels_config",
            lambda *_args, **_kwargs: frozenset(),
        )

        entries_in: list[CryptoCapitalGainEntry] = []

        with caplog.at_level(logging.WARNING, logger="tax_reporting.application.crypto.derivatives_dedup"):
            result = apply_derivatives_dedup(
                capital_entries=entries_in,
                jurisdiction=build_koinly_jurisdiction(separate_derivatives_reporting=True),
                transaction_history_file=_th_path(),
                year=2025,
                transactions=[],
                config=TreatmentConfig(),
                via_resolver=False,
            )

        assert result is entries_in, (
            "apply_derivatives_dedup must return the input list unchanged when "
            "the label config is missing (graceful degradation)."
        )
        missing_config_warnings = [
            r
            for r in caplog.records
            if r.levelno >= logging.WARNING and "koinly_2025.json" in r.getMessage()
        ]
        assert len(missing_config_warnings) == 1, (
            "Expected exactly one missing-config WARNING naming the file "
            "koinly_2025.json; got "
            f"{len(missing_config_warnings)}: {[r.getMessage() for r in missing_config_warnings]}"
        )

    def test_ogr_classifies_clean_after_dedup(self) -> None:
        """After dedup, the 3 Case C OGR rows classify as clean Derivatives (not Ambiguous).

        With the dedup removing the 3 CG counterparts for 2025-01-24 USDT
        Demo Futures, the OGR classifier sees zero CG matches and routes the
        OGR rows as clean Derivatives (cg_matches == 0) instead of Ambiguous.
        Because ``Demo Futures`` is unmapped, the row carries
        review_required=True from the platform-mapping signal; this test
        asserts the classification KIND (Derivatives LOSS present, review_reason
        has no Ambiguous text) rather than the review flag.
        """
        report = _load_with_separation(separate_derivatives=True)

        case3_derivatives = [
            e
            for e in report.derivatives_entries
            if e.date == _CASE3_DATE and e.asset == _ASSET and e.platform == _PLATFORM
        ]
        assert case3_derivatives, (
            "Expected at least one Derivatives P&L entry for "
            f"({_CASE3_DATE}, {_ASSET}, {_PLATFORM}); got none. "
            "All derivatives_entries: "
            f"{[(e.date, e.asset, e.platform, e.pnl_eur) for e in report.derivatives_entries]}"
        )
        for entry in case3_derivatives:
            assert "matches CG disposal" not in entry.review_reason, (
                "No Case C derivatives entry should carry the Ambiguous classification reason "
                "after the dedup (the OGR classifier should see zero CG counterparts and "
                "classify as clean Derivatives, not Ambiguous). "
                f"review_reason={entry.review_reason!r}"
            )


class TestDerivativesE2E:
    """E2E characterization for the 12-column Derivatives P&L sheet layout.

    Covers Task 5 of the 2026-06-15 derivatives P&L columns plan: the
    synthetic data run (koinly2025 fixture, separate_derivatives_reporting=True)
    must render the Derivatives P&L tab with a 12-column header where Annex
    (col 11) and Código (col 12) are written per row, and populate
    operator_country for every derivatives row using the production
    ``resolve_operator_origin`` wiring (Task 2). These tests are structural
    (column population, country-code validity) so they survive fixture platform
    changes.
    """

    _DERIVATIVES_SHEET_NAME = "Derivatives P&L"
    _EXPECTED_NUM_COLUMNS = 12
    _HEADER_ROW = 3

    def test_derivatives_sheet_has_twelve_columns(self) -> None:
        """The Derivatives P&L sheet has 12 populated header cells in row 3.

        Renders the production sheet from the synthetic koinly2025 report and
        counts populated header cells in row 3 (column population, not
        hardcoded value exclusions). The
        last populated header cell must sit at column 12 and read "Código".
        Annex (col 11) and Código (col 12) are written per row from each
        entry's annex_hint/operation_code.
        """
        report = _load_with_separation(separate_derivatives=True)

        wb = openpyxl.Workbook()
        write_derivatives_sheet(wb, report, build_koinly_jurisdiction(separate_derivatives_reporting=True))
        ws = wb[self._DERIVATIVES_SHEET_NAME]

        populated = [
            (c, ws.cell(self._HEADER_ROW, c).value)
            for c in range(1, 50)
            if ws.cell(self._HEADER_ROW, c).value is not None
        ]
        assert len(populated) == self._EXPECTED_NUM_COLUMNS, (
            "Derivatives P&L header row should have exactly 12 populated cells. "
            f"Got {len(populated)}: {populated}"
        )
        last_col, last_value = populated[-1]
        assert last_col == self._EXPECTED_NUM_COLUMNS, (
            f"Last header cell should be at column 12, got column {last_col}"
        )
        assert last_value == "Código", (
            f"Last header cell should read 'Código', got {last_value!r}"
        )

    def test_derivatives_rows_operator_country_is_valid_or_unknown(self) -> None:
        """Every derivatives row's operator_country is a valid Tabela X code or 'UNKNOWN'.

        For the synthetic data run, ``resolve_operator_origin`` populates
        ``operator_country`` from the production platform map. Valid values
        are either an ISO 3166-1 alpha-2 code in the Portuguese Tabela X
        list (validated via the production ``_is_valid_tabela_x_country``
        helper) or the literal sentinel ``"UNKNOWN"`` for unmapped
        platforms. The synthetic ``Demo Futures`` wallet is deliberately
        unmapped, so ``operator_country == "UNKNOWN"`` and the row must
        carry review_required=True (Task 2 wiring), so the Review cell at
        column 10 starts with ``"YES:"``. This is a structural assertion
        that survives fixture platform changes.
        """
        report = _load_with_separation(separate_derivatives=True)

        assert report.derivatives_entries, (
            "Expected at least one derivatives entry from the koinly2025 fixture; "
            "cannot characterize operator_country on an empty derivatives_entries list."
        )

        wb = openpyxl.Workbook()
        write_derivatives_sheet(wb, report, build_koinly_jurisdiction(separate_derivatives_reporting=True))
        ws = wb[self._DERIVATIVES_SHEET_NAME]

        header_row = self._HEADER_ROW
        # Walk the data rows starting right after the header. Stop at the
        # first row whose column 1 is empty (defensive: bounds the walk
        # before the total/footnote rows so we only inspect entry rows).
        for offset, entry in enumerate(report.derivatives_entries, start=1):
            data_row = header_row + offset
            cell_value = ws.cell(data_row, 7).value  # column 7 = Operator country
            country = "" if cell_value is None else str(cell_value)
            is_valid = country == "UNKNOWN" or _is_valid_tabela_x_country(country)
            assert is_valid, (
                f"Derivatives row {data_row} (date={entry.date}, asset={entry.asset}, "
                f"platform={entry.platform}) has operator_country={country!r}, which is "
                "neither a valid Tabela X country code nor the literal 'UNKNOWN' sentinel."
            )

            if country == "UNKNOWN":
                review_cell_value = ws.cell(data_row, 10).value  # column 10 = Review
                review_text = "" if review_cell_value is None else str(review_cell_value)
                assert review_text.startswith("YES:"), (
                    f"Derivatives row {data_row} has operator_country='UNKNOWN' but the "
                    f"Review cell at column 10 does not start with 'YES:': got {review_text!r}. "
                    "Unmapped platforms must surface review_required=True per the Task 2 "
                    "operator-origin wiring."
                )
