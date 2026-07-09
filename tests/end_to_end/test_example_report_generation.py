"""End-to-end test verifying the repository can generate a report from committed example inputs.

The example data under resources/source/example/ is fully synthetic and exercises every
major feature: shares capital gains, dividends, rollover/leftover integration, crypto
capital events, crypto rewards, and Token origin resolution via transaction history.
"""

from __future__ import annotations

import csv
import io
import re
from decimal import Decimal
from pathlib import Path

import openpyxl
import pytest

from tax_reporting.application.crypto_reporting import (
    RewardTaxClassification,
    load_koinly_crypto_report,
)
from tax_reporting.application.extraction import parse_ib_export_all
from tax_reporting.application.persisting import generate_tax_report
from tax_reporting.application.persisting.tax_constants import get_income_code_description
from tax_reporting.application.transformation import calculate_fifo_gains
from tax_reporting.domain.collections import TradeCyclePerCompany

EXAMPLE_DIR = Path("resources", "source", "example")
EXAMPLE_IB_EXPORT = EXAMPLE_DIR / "ib_export.csv"
EXAMPLE_KOINLY_DIR = EXAMPLE_DIR / "2024" / "koinly"

# Directories created by the crypto-tests-off-local-fixtures plan (2026-06-22) that ship
# committed, fully-synthetic Koinly exports. The synthetic-data hygiene checks below are
# scoped to these directories only: the legacy example/2024/koinly/ files predate the
# synthetic-data convention and use real-looking wallet names (Kraken, Binance) and
# non-empty TxHash/TxSrc/TxDest values, so applying the checks globally would false-fail on
# pre-existing legacy data (scope a new-convention validator to new work, or
# accept-list the legacy token).
SYNTHETIC_KOINLY_2025_DIRS = [
    EXAMPLE_DIR / "2025" / "koinly",
    EXAMPLE_DIR / "2025" / "koinly" / "zero_basis",
    EXAMPLE_DIR / "2025" / "koinly" / "payment",
]
# The path structure (example/<year>/koinly/[<scenario>/]) is the synthetic marker;
# no filename suffix is required. Suffix retained empty for backward-compat with any
# downstream tooling that referenced the constant.
SYNTHETIC_FILENAME_SUFFIXES = (".csv",)
# Blockchain/personal-detail columns that must be empty in synthetic data.
SENSITIVE_COLUMN_NAMES = ("TxHash", "TxSrc", "TxDest")
# Wallet-bearing columns whose values must come from a small synthetic allowlist.
WALLET_COLUMN_NAMES = ("Sending Wallet", "Receiving Wallet", "Wallet Name")
# Wallet labels permitted in synthetic example data (Design Invariant #1). Empty string is
# allowed because Koinly leaves these blank for non-wallet-scoped rows (e.g. fiat legs).
SYNTHETIC_WALLET_ALLOWLIST = {"Demo Spot", "Demo Futures", "Demo Payment", "Wirex", ""}


def _read_koinly_csv(path: Path) -> tuple[list[str], list[dict[str, str]]]:
    """Read a Koinly CSV, skipping the leading title/blank lines before the header row.

    Koinly exports start with a title line (e.g. "Capital gains report 2025") and, in some
    report variants, a trailing blank line before the column header. The header row is the
    first line that both contains a comma and mentions a known column keyword
    (Date/Asset/Type). Returns (fieldnames, data_rows) using csv.DictReader semantics.
    """
    with path.open(newline="") as f:
        lines = f.readlines()
    header_idx: int | None = None
    for i, line in enumerate(lines):
        stripped = line.strip()
        if stripped and "," in stripped and any(kw in stripped for kw in ("Date", "Asset", "Type")):
            header_idx = i
            break
    if header_idx is None:
        return [], []
    reader = csv.DictReader(io.StringIO("".join(lines[header_idx:])))
    return list(reader.fieldnames or []), list(reader)


@pytest.mark.e2e
def test_example_ib_export_parses_successfully():
    ib_data = parse_ib_export_all(EXAMPLE_IB_EXPORT)
    assert len(ib_data.trade_cycles) >= 2
    assert len(ib_data.dividend_income) >= 2


@pytest.mark.e2e
def test_example_leftover_integrated():
    ib_data = parse_ib_export_all(EXAMPLE_IB_EXPORT)
    leftover_path = EXAMPLE_DIR / "shares-leftover.csv"
    assert leftover_path.exists()
    acme_cycle = next(
        (
            cycle
            for currency_company, cycle in ib_data.trade_cycles.items()
            if currency_company.currency.currency == "USD" and currency_company.company.ticker == "ACME"
        ),
        None,
    )
    assert acme_cycle is not None
    assert len(acme_cycle.bought) == 2
    leftover_buy = acme_cycle.bought[0]
    assert leftover_buy.quantity == 20
    assert leftover_buy.action.price == Decimal("15.00")


@pytest.mark.e2e
def test_example_shares_capital_gains():
    ib_data = parse_ib_export_all(EXAMPLE_IB_EXPORT)
    leftover_trades: TradeCyclePerCompany = {}
    capital_gains = {}
    calculate_fifo_gains(ib_data.trade_cycles, leftover_trades, capital_gains)
    total_cg_lines = sum(len(lines) for lines in capital_gains.values())
    assert total_cg_lines >= 2
    assert len(leftover_trades) >= 1


@pytest.mark.e2e
def test_example_crypto_report_loads():
    crypto = load_koinly_crypto_report(EXAMPLE_KOINLY_DIR)
    assert crypto is not None
    assert len(crypto.capital_entries) >= 2
    assert len(crypto.reward_entries) >= 2


@pytest.mark.e2e
def test_example_crypto_token_origin_is_resolved():
    crypto = load_koinly_crypto_report(EXAMPLE_KOINLY_DIR)
    assert crypto is not None
    non_blank = [e for e in crypto.capital_entries if e.token_swap_history]
    assert len(non_blank) >= 1, (
        "At least one capital entry should have non-blank Token origin from transaction history"
    )
    for entry in non_blank:
        assert "confidence" in entry.token_swap_history, (
            f"Token origin for {entry.asset} should include confidence level, got '{entry.token_swap_history}'"
        )
        assert not re.match(r"^[\d.,]+$", entry.token_swap_history.split()[0]), (
            f"Token origin for {entry.asset} starts with numeric value, "
            f"likely a CSV column alignment issue: '{entry.token_swap_history}'"
        )


@pytest.mark.e2e
def test_example_full_pipeline_generates_excel(tmp_path: Path):
    ib_data = parse_ib_export_all(EXAMPLE_IB_EXPORT)
    leftover_trades: TradeCyclePerCompany = {}
    capital_gains = {}
    calculate_fifo_gains(ib_data.trade_cycles, leftover_trades, capital_gains)
    crypto = load_koinly_crypto_report(EXAMPLE_KOINLY_DIR)
    output_path = tmp_path / "extract.xlsx"
    crypto_sheet_created = generate_tax_report(
        output_path,
        capital_gains,
        ib_data.dividend_income,
        crypto_tax_report=crypto,
    )
    assert crypto_sheet_created
    assert output_path.exists()
    wb = openpyxl.load_workbook(output_path)
    assert "Reporting" in wb.sheetnames
    assert "Crypto Gains" in wb.sheetnames
    assert "Crypto Supplementary" in wb.sheetnames
    assert "Crypto Reconciliation" in wb.sheetnames
    assert "Loan Activity" in wb.sheetnames
    assert "Assumptions & Methodology" in wb.sheetnames

    # Verify Loan Activity sheet has expected structure
    loan_ws = wb["Loan Activity"]
    header_values = [cell.value for cell in list(loan_ws.iter_rows(min_row=3, max_row=3))[0]]
    assert "Asset" in header_values
    assert "Received Count" in header_values
    assert "Received Amount" in header_values
    assert "Repaid Count" in header_values
    assert "Repaid Amount" in header_values
    assert "Balance Status" in header_values

    wb.close()


@pytest.mark.e2e
def test_example_crypto_sheet_has_resolved_token_origin(tmp_path: Path):
    ib_data = parse_ib_export_all(EXAMPLE_IB_EXPORT)
    leftover_trades: TradeCyclePerCompany = {}
    capital_gains = {}
    calculate_fifo_gains(ib_data.trade_cycles, leftover_trades, capital_gains)
    crypto = load_koinly_crypto_report(EXAMPLE_KOINLY_DIR)
    output_path = tmp_path / "extract.xlsx"
    generate_tax_report(output_path, capital_gains, ib_data.dividend_income, crypto_tax_report=crypto)
    wb = openpyxl.load_workbook(output_path)
    ws = wb["Crypto Gains"]
    token_origin_col = None
    header_row_num = None
    for row in ws.iter_rows(min_row=1, max_row=20):
        for cell in row:
            if cell.value == "Token origin":
                token_origin_col = cell.column
                header_row_num = cell.row
                break
        if token_origin_col:
            break
    assert token_origin_col is not None, "Token origin column header not found"
    assert header_row_num is not None
    resolved_count = 0
    for row in ws.iter_rows(min_row=header_row_num + 1, max_row=ws.max_row):
        origin_val = row[token_origin_col - 1].value if token_origin_col <= len(row) else None
        asset = row[2].value if len(row) > 2 else None
        if (
            asset
            and isinstance(asset, str)
            and asset.strip()
            and origin_val
            and isinstance(origin_val, str)
            and "confidence" in origin_val
        ):
            resolved_count += 1
    assert resolved_count >= 1, "At least one row should have resolved Token origin with confidence level"
    wb.close()


@pytest.mark.e2e
def test_example_data_is_synthetic():
    ib_content = EXAMPLE_IB_EXPORT.read_text()
    assert "Demo Taxpayer" in ib_content
    assert "Demo Broker LLC" in ib_content
    assert "U9999999" in ib_content
    koinly_capital = sorted(EXAMPLE_KOINLY_DIR.glob("*capital_gains_report*.csv"))[0]
    koinly_filename = koinly_capital.name
    assert "xY9kLm2pQr" in koinly_filename
    koinly_income = sorted(EXAMPLE_KOINLY_DIR.glob("*income_report*.csv"))[0]
    assert "aB3cDn5oEf" in koinly_income.name

    # --- Synthetic-data hygiene for the new committed example/2025/koinly/ fixtures ---
    # (Design Invariant #1: synthesis, not sanitization; Evaluation Criteria
    # "Synthetic-data hygiene"). Scoped to the directories created by the
    # crypto-tests-off-local-fixtures plan: the legacy example/2024/koinly/ files use a
    # 10-char token, real wallet names (Kraken/Binance), and non-empty TxHash/TxSrc/TxDest,
    # so a global check would false-fail on pre-existing legacy data.
    scanned_csvs: list[Path] = []
    for synth_dir in SYNTHETIC_KOINLY_2025_DIRS:
        scanned_csvs.extend(sorted(synth_dir.glob("koinly_*.csv")))
    assert scanned_csvs, (
        "Pre-condition failed: no koinly_*.csv files found under the synthetic "
        "example/2025/koinly/ directories; the hygiene checks cannot run."
    )

    # Sub-check 1: filename shape. Synthetic CSVs must follow the canonical Koinly export
    # naming (koinly_<year>_<report>.csv), not the legacy 10-char mixed-case token form
    # used by example/2024/koinly/. The path structure (under example/<year>/koinly/) is
    # the synthetic-vs-real marker; no _synth suffix is required.
    bad_filenames = [
        csv_path
        for csv_path in scanned_csvs
        if not re.match(r"^koinly_\d{4}_.*\.csv$", csv_path.name)
    ]
    assert not bad_filenames, (
        "Synthetic example CSVs must follow the koinly_<year>_<report>.csv naming; "
        f"offending files: {[p.name for p in bad_filenames]}"
    )

    # Sub-checks 2 and 3 run per-file. Both gracefully skip columns that are absent from a
    # given report type (only transaction_history carries TxHash/TxSrc/TxDest; wallet
    # columns differ between report types).
    sensitive_violations: list[str] = []
    wallet_violations: list[str] = []
    for csv_path in scanned_csvs:
        columns, rows = _read_koinly_csv(csv_path)
        present_sensitive = [c for c in SENSITIVE_COLUMN_NAMES if c in columns]
        present_wallets = [c for c in WALLET_COLUMN_NAMES if c in columns]

        for row_num, row in enumerate(rows, start=1):
            # Sub-check 2: sensitive blockchain/personal columns must be empty for every row.
            for col in present_sensitive:
                value = (row.get(col) or "").strip()
                if value:
                    sensitive_violations.append(
                        f"{csv_path} row {row_num} column {col!r} must be empty, got {value!r}"
                    )
            # Sub-check 3: wallet columns must only use the synthetic allowlist.
            for col in present_wallets:
                value = (row.get(col) or "").strip()
                if value not in SYNTHETIC_WALLET_ALLOWLIST:
                    wallet_violations.append(
                        f"{csv_path} row {row_num} column {col!r} has non-synthetic wallet "
                        f"{value!r}; allowed {sorted(SYNTHETIC_WALLET_ALLOWLIST - {''})}"
                    )

    assert not sensitive_violations, (
        "Synthetic example data must not carry real blockchain/personal details; "
        "TxHash/TxSrc/TxDest must be empty for every row. Violations:\n  - "
        + "\n  - ".join(sensitive_violations)
    )
    assert not wallet_violations, (
        "Synthetic example data must use only synthetic wallet labels "
        f"({sorted(SYNTHETIC_WALLET_ALLOWLIST - {''})}); real exchange/wallet names are "
        "forbidden. Violations:\n  - " + "\n  - ".join(wallet_violations)
    )


def _count_csv_data_rows(path: Path) -> int:
    """Count data rows in a Koinly CSV (skipping title, blank, and header lines)."""
    with path.open() as f:
        return max(0, sum(1 for line in f if line.strip()) - 2)


@pytest.mark.e2e
def test_example_crypto_source_has_high_volume_rows():
    capital_csv = sorted(EXAMPLE_KOINLY_DIR.glob("*capital_gains_report*.csv"))[0]
    income_csv = sorted(EXAMPLE_KOINLY_DIR.glob("*income_report*.csv"))[0]
    capital_rows = _count_csv_data_rows(capital_csv)
    income_rows = _count_csv_data_rows(income_csv)
    total_rows = capital_rows + income_rows
    assert total_rows >= 1000, (
        f"Example crypto source should contain at least 1000 rows, got {total_rows} "
        f"(capital={capital_rows}, income={income_rows})"
    )


@pytest.mark.e2e
def test_example_crypto_capital_gains_aggregate_to_few_lines():
    crypto = load_koinly_crypto_report(EXAMPLE_KOINLY_DIR)
    assert crypto is not None
    assert len(crypto.capital_entries) <= 5, (
        f"Aggregated capital entries should be at most 5, got {len(crypto.capital_entries)}"
    )
    capital_csv = sorted(EXAMPLE_KOINLY_DIR.glob("*capital_gains_report*.csv"))[0]
    raw_capital_rows = _count_csv_data_rows(capital_csv)
    compression_ratio = raw_capital_rows / len(crypto.capital_entries)
    assert compression_ratio >= 50, (
        f"Capital gains compression ratio should be >= 50x, got {compression_ratio:.1f}x "
        f"({raw_capital_rows} raw -> {len(crypto.capital_entries)} aggregated)"
    )
    total_gain = sum(float(e.gain_loss_eur) for e in crypto.capital_entries)
    assert 100 <= total_gain <= 200, (
        f"Total capital gains should be modest (100-200 EUR), got {total_gain:.2f} EUR"
    )


@pytest.mark.e2e
def test_example_crypto_rewards_are_many_but_classified():
    crypto = load_koinly_crypto_report(EXAMPLE_KOINLY_DIR)
    assert crypto is not None
    assert len(crypto.reward_entries) == 160, (
        f"Raw reward entries should be exactly 160, got {len(crypto.reward_entries)}"
    )
    total_reward_value = sum(float(e.value_eur) for e in crypto.reward_entries)
    assert total_reward_value <= 1000, (
        f"Total reward value should be modest (<1000 EUR), got {total_reward_value:.2f} EUR"
    )
    deferred = [e for e in crypto.reward_entries if e.tax_classification == RewardTaxClassification.DEFERRED_BY_LAW]
    taxable = [e for e in crypto.reward_entries if e.tax_classification == RewardTaxClassification.TAXABLE_NOW]
    assert len(taxable) == 10, (
        f"Example data should have exactly 10 taxable-now (fiat-denominated) rewards, got {len(taxable)}"
    )
    assert len(deferred) == 150, (
        f"Example data should have exactly 150 deferred-by-law (crypto-denominated) rewards, got {len(deferred)}"
    )
    assert len(taxable) + len(deferred) == len(crypto.reward_entries), (
        f"Taxable ({len(taxable)}) + deferred ({len(deferred)}) should equal total ({len(crypto.reward_entries)})"
    )


@pytest.mark.e2e
def test_example_taxable_now_rewards_are_reported_on_reporting_sheet(tmp_path: Path):
    """Given the example report data, immediately taxable fiat rewards are visible on Reporting
    as aggregate rows with exact gross, tax, net, and raw-row-count values, and still
    traceable on Crypto Supplementary."""
    ib_data = parse_ib_export_all(EXAMPLE_IB_EXPORT)
    leftover_trades: TradeCyclePerCompany = {}
    capital_gains = {}
    calculate_fifo_gains(ib_data.trade_cycles, leftover_trades, capital_gains)
    crypto = load_koinly_crypto_report(EXAMPLE_KOINLY_DIR)
    output_path = tmp_path / "extract.xlsx"
    generate_tax_report(output_path, capital_gains, ib_data.dividend_income, crypto_tax_report=crypto)

    wb = openpyxl.load_workbook(output_path)
    reporting_ws = wb["Reporting"]
    rows = list(reporting_ws.iter_rows(values_only=True))

    # Locate the OTHER CAPITAL INVESTMENT INCOME subsection header, then the
    # single aggregated reward data row that follows it. The example's taxable
    # EUR rewards are interest-type, which resolve to the official Tabela V code
    # E25 under PT, so the row is located structurally from the subsection header.
    subsection_idx = None
    for idx, row in enumerate(rows):
        if row and isinstance(row[0], str) and "OTHER CAPITAL INVESTMENT INCOME" in row[0]:
            subsection_idx = idx
            break
    assert subsection_idx is not None, "OTHER CAPITAL INVESTMENT INCOME subsection not found in Reporting"
    reward_row = rows[subsection_idx + 1]
    assert reward_row is not None, "Taxable fiat reward row not found in Reporting"

    # Interest rewards resolve to income_code="E25" under PT -> the income-type
    # cell carries the official E25 description (a complete filing row), not a
    # review marker.
    assert isinstance(reward_row[0], str)
    assert reward_row[0] == get_income_code_description("E25"), (
        f"Expected E25 income-type description, got {reward_row[0]!r}"
    )
    assert reward_row[1] == "IE"
    assert float(reward_row[3]) == pytest.approx(30.00), f"Expected gross 30.00, got {reward_row[3]}"
    assert float(reward_row[4]) == pytest.approx(0.0), f"Expected foreign tax 0, got {reward_row[4]}"
    assert float(reward_row[11]) == pytest.approx(30.00), f"Expected net 30.00, got {reward_row[11]}"
    assert reward_row[13] == 10, f"Expected raw_row_count 10, got {reward_row[13]}"
    assert reward_row[12] == "Koinly"

    rewards_ws = wb["Crypto Supplementary"]
    rewards_labels = []
    for row in rewards_ws.iter_rows(values_only=True):
        first_cell = row[0] if row else None
        if isinstance(first_cell, str):
            rewards_labels.append(first_cell)
    assert "2. TAXABLE-NOW - SUPPORT DETAIL" in rewards_labels, (
        "Crypto Supplementary should still have support detail for taxable-now rewards"
    )

    wb.close()


@pytest.mark.e2e
def test_example_high_volume_crypto_sheet_is_compact(tmp_path: Path):
    ib_data = parse_ib_export_all(EXAMPLE_IB_EXPORT)
    leftover_trades: TradeCyclePerCompany = {}
    capital_gains = {}
    calculate_fifo_gains(ib_data.trade_cycles, leftover_trades, capital_gains)
    crypto = load_koinly_crypto_report(EXAMPLE_KOINLY_DIR)
    output_path = tmp_path / "extract.xlsx"
    generate_tax_report(output_path, capital_gains, ib_data.dividend_income, crypto_tax_report=crypto)
    wb = openpyxl.load_workbook(output_path)
    ws = wb["Crypto Gains"]
    asset_col = None
    for row in ws.iter_rows(min_row=1, max_row=20):
        for cell in row:
            if cell.value == "Asset":
                asset_col = cell.column
                break
        if asset_col:
            break
    assert asset_col is not None, "Asset column header not found in Crypto Gains sheet"
    capital_data_rows = 0
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        first_val = row[0].value if row else None
        if isinstance(first_val, str) and "CAPITAL GAINS STATISTICS" in first_val:
            break
        asset_val = row[asset_col - 1].value if asset_col <= len(row) else None
        if asset_val and isinstance(asset_val, str) and asset_val.strip():
            capital_data_rows += 1
    assert capital_data_rows <= 10, (
        f"Capital gains data rows in Crypto Gains sheet should be at most 10, got {capital_data_rows}"
    )
    wb.close()
